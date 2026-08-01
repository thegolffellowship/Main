"""
Transaction Email Tracker — Flask application.

Automatically checks your email inbox for transaction/receipt emails,
parses purchase data with AI (Claude), and displays it in a web dashboard.
Includes a webhook connector for external integrations and a daily email report.
"""

import os
import re
import json
import secrets
import logging
import shutil
import sqlite3
import time
import threading
from datetime import datetime, timedelta
from functools import wraps

import anthropic as _anthropic
from flask import Flask, Response, jsonify, redirect, render_template, request, send_file, session
from dotenv import load_dotenv
from apscheduler.schedulers.background import BackgroundScheduler

from email_parser.database import (
    init_db,
    get_all_items,
    get_item,
    get_known_email_uids,
    get_expense_seen_uids,
    mark_expense_email_seen,
    get_item_stats,
    get_audit_report,
    get_data_snapshot,
    save_items,
    update_item,
    delete_item,
    delete_manual_player,
    credit_item,
    transfer_item,
    reverse_credit,
    wd_item,
    payout_credit,
    create_event,
    seed_events,
    add_player_to_event,
    add_payment_to_event,
    get_add_payment_quote,
    upgrade_rsvp_to_paid,
    autofix_side_games,
    autofix_all,
    undo_autofix,
    normalize_tee_choices,
    sync_events_from_items,
    get_all_events,
    update_event,
    delete_event,
    merge_events,
    get_orphaned_items,
    resolve_orphaned_items,
    save_parse_warnings,
    get_parse_warnings,
    dismiss_parse_warning,
    resolve_parse_warning,
    get_all_event_aliases,
    add_event_alias,
    delete_event_alias,
    get_known_rsvp_uids,
    save_rsvps,
    get_rsvps_for_event,
    get_all_rsvps,
    get_all_rsvps_bulk,
    get_rsvp_stats,
    rematch_rsvps,
    audit_event_rsvps,
    manual_match_rsvp,
    unmatch_rsvp,
    get_rsvp_overrides,
    set_rsvp_override,
    get_rsvp_email_overrides,
    set_rsvp_email_override,
    merge_customers,
    update_customer_info,
    create_customer,
    create_customer_from_rsvp,
    link_rsvp_to_customer,
    import_roster,
    preview_roster_import,
    get_customer_aliases,
    add_customer_alias,
    delete_customer_alias,
    parse_names_ai,
    validate_email,
    validate_phone,
    add_custom_field,
    save_feedback,
    get_all_feedback,
    update_feedback_status,
    get_message_templates,
    get_message_template,
    create_message_template,
    update_message_template,
    delete_message_template,
    log_message,
    get_message_log,
    get_all_handicap_players,
    get_handicap_rounds,
    import_handicap_rounds,
    delete_handicap_round,
    delete_all_handicap_rounds_for_player,
    get_handicap_settings,
    update_handicap_settings,
    get_handicap_export_data,
    build_handicap_card_data,
    build_handicap_card_html,
    relink_all_unlinked_players,
    mark_email_processed,
    clear_failed_processed,
    refund_item,
    set_event_status,
    can_restore_event,
    get_cancellation_players,
    get_player_credits,
    get_rsvp_credit_info,
    get_event_rsvp_credit_map,
    mark_rsvp_credit_notified,
    apply_credit_to_rsvp,
    create_rsvp_only_item,
    reverse_credit_application,
    get_app_setting,
    set_app_setting,
    # Accounting module
    get_all_acct_entities,
    create_acct_entity,
    update_acct_entity,
    get_acct_categories,
    create_acct_category,
    update_acct_category,
    delete_acct_category,
    get_acct_accounts,
    create_acct_account,
    update_acct_account,
    get_acct_account_balances,
    get_acct_transactions,
    get_acct_transaction,
    _create_acct_ledger_entry,
    update_acct_transaction,
    delete_acct_transaction,
    reconcile_acct_transaction,
    get_acct_tags,
    create_acct_tag,
    delete_acct_tag,
    get_acct_summary,
    get_acct_monthly_totals,
    get_acct_category_breakdown,
    preview_acct_csv,
    import_acct_csv,
    get_acct_recurring,
    create_acct_recurring,
    delete_acct_recurring,
    auto_categorize_transactions,
    get_acct_review_queue,
    get_acct_categorization_stats,
    reset_acct_data,
    get_acct_account_rules,
    set_acct_account_rule,
    get_all_acct_account_rules,
    calculate_order_allocation,
    get_acct_allocations,
    get_event_financial_summary,
    backfill_financial_entries,
    backfill_acct_transactions,
    backfill_missing_godaddy_orders,
    repair_orphan_pay_children,
    capture_email_aliases_from_items,
    heal_items_from_customers,
    migrate_item_to_order_entries,
    cleanup_duplicate_godaddy_entries,
    backup_database,
    scan_price_games_mismatches,
    reconcile_orphan_venmo_payments,
    save_expense_transaction,
    get_expense_transactions,
    get_unified_transactions,
    update_expense_transaction,
    get_blocked_merchants,
    block_merchant,
    dismiss_bank_deposit,
    record_internal_transfer,
    save_action_item,
    get_action_items,
    update_action_item,
    get_pending_review_count,
    get_coo_financial_snapshot,
    get_coo_review_queue,
    get_all_coo_manual_values,
    set_coo_manual_value,
    get_chart_of_accounts,
    get_ledger_entries,
    import_bank_statement,
    run_bank_reconciliation,
    close_period,
    get_reconciliation_summary,
    # Bank deposit reconciliation (new)
    get_bank_accounts,
    import_bank_deposits,
    import_venmo_statement,
    run_deposit_auto_match,
    manual_match_deposit,
    batch_match_deposit,
    merge_transactions,
    get_match_suggestions,
    unmatch_deposit,
    get_bank_deposits,
    get_unreconciled_transactions,
    get_reconciliation_dashboard,
    get_monthly_reconciliation,
    get_event_reconciliation_status,
    get_cashflow_data,
    get_coo_agents,
    get_agent_action_log,
    batch_dismiss_action_items,
    consolidate_action_items,
    # Keyword rules
    get_acct_keyword_rules,
    create_acct_keyword_rule,
    update_acct_keyword_rule,
    delete_acct_keyword_rule,
    # Batch categorization preview + promotion
    get_expense_batch_preview,
    batch_approve_expenses,
    # Create ledger entry from orphaned bank deposit
    create_entry_from_deposit,
    # Liabilities Dashboard
    get_accounting_liabilities,
    # Month Close
    get_month_close_status,
    # Contractor tracking
    get_contractor_payouts,
    get_contractor_managers,
    add_contractor_payout,
    update_contractor_payout,
    delete_contractor_payout,
    # TGF Payouts
    get_tgf_data,
    add_tgf_event,
    add_tgf_golfer,
    import_tgf_golfers,
    update_tgf_event,
    delete_tgf_event,
    # MVP linking
    get_mvp_unlinked_events,
    set_mvp_unlink,
    # COO Chat persistence
    get_chat_sessions,
    get_chat_session,
    create_chat_session,
    add_chat_message,
    update_chat_session_title,
    update_chat_session_summary,
    get_chat_master_context,
    build_coo_full_context,
    delete_chat_session,
    # Pairings
    get_event_pairings,
    save_event_pairings,
    delete_event_pairings,
    generate_event_pairings,
    get_pairing_history_counts,
    _pairing_time_slots,
    # Duplicate Detective
    find_duplicate_candidates,
    dismiss_duplicate_pair,
    get_duplicate_detective_mode,
    set_duplicate_detective_mode,
    get_duplicate_merge_audit,
    reverse_duplicate_merge,
)
from email_parser.database import DB_PATH, get_connection
from email_parser.timezone_utils import now_central, today_central_str
from email_parser.fetcher import (
    fetch_transaction_emails, fetch_all_emails, fetch_email_by_id,
    send_mail_graph, render_msg_template, send_bulk_emails,
)
from email_parser.parser import parse_email, parse_emails, _strip_html
from email_parser.expense_parser import (
    classify_email, parse_chase_alert, parse_venmo_payment, parse_p2p_payment,
    parse_expense_receipt, parse_action_required,
    match_event_from_memo, match_customer_from_name,
    match_event_from_customer,
    get_merchant_context,
)
from email_parser.coo_email import build_coo_email_html
from email_parser.report import send_daily_report
from email_parser.rsvp_parser import fetch_rsvp_emails, parse_rsvp_emails

load_dotenv()

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Player identity resolution
# ---------------------------------------------------------------------------
# Customer Info card and the manager's Edit dialog are the source of truth
# for a player's identity (email, phone, name, chapter, status). items.*
# columns are historical snapshots from each order and can carry typos or
# stale values that the manager has since corrected. Every customer-facing
# read MUST resolve through these canonical helpers — never read items.*
# directly. See database.py for the per-field implementations and the
# documented priority chain.
from email_parser.database import (
    resolve_player_email as _resolve_player_email_db,
    resolve_player_phone as _resolve_player_phone,
    resolve_player_name as _resolve_player_name,
    resolve_player_chapter as _resolve_player_chapter,
    resolve_player_status as _resolve_player_status,
)


def _resolve_player_email(item: dict, conn=None) -> str:
    """Thin wrapper that preserves the existing app.py call signature."""
    return _resolve_player_email_db(item, conn=conn)


# ---------------------------------------------------------------------------
# Simple in-memory rate limiter for login endpoint
# ---------------------------------------------------------------------------
_login_attempts: dict[str, list[float]] = {}  # IP → list of timestamps
_LOGIN_MAX_ATTEMPTS = 10
_LOGIN_WINDOW_SECONDS = 15 * 60  # 15 minutes


def _check_login_rate_limit() -> bool:
    """Return True if the request IP is within rate limits, False if exceeded."""
    # Key on the LAST X-Forwarded-For hop: Railway's edge proxy APPENDS the
    # real client IP, while everything left of it is client-supplied. Keying
    # on the first entry let an attacker bypass the limiter by rotating a
    # fake header (or poison an arbitrary IP's bucket to lock someone out).
    fwd = request.headers.get("X-Forwarded-For", "")
    if fwd:
        ip = fwd.split(",")[-1].strip() or "unknown"
    else:
        ip = request.remote_addr or "unknown"
    now = time.time()
    cutoff = now - _LOGIN_WINDOW_SECONDS
    # Clean old entries
    attempts = [t for t in _login_attempts.get(ip, []) if t > cutoff]
    _login_attempts[ip] = attempts
    if len(attempts) >= _LOGIN_MAX_ATTEMPTS:
        return False
    attempts.append(now)
    return True


app = Flask(__name__)
# Gzip-compress text responses (HTML/CSS/JS/JSON) so large API payloads
# like /api/handicaps/rounds for high-volume players shrink from ~80 KB
# to ~10 KB on the wire. flask-compress only compresses responses with
# a compressible MIME type and at or above its default min length, and
# it respects the client's Accept-Encoding header so non-supporting
# clients still get a plain response.
from flask_compress import Compress
Compress(app)
_secret_key = os.getenv("SECRET_KEY")
if not _secret_key:
    raise RuntimeError(
        "SECRET_KEY environment variable is not set. "
        "Generate one with: python -c \"import secrets; print(secrets.token_hex(32))\" "
        "and add SECRET_KEY=<value> to your .env file or Railway environment variables."
    )
app.secret_key = _secret_key


@app.after_request
def _no_store_api_responses(resp):
    """Live operational data must never be served from the browser's HTTP
    cache. Safari/iOS caches GET XHRs that carry no Cache-Control header,
    so the Payouts PWA kept re-serving a stale /api/tgf payload for many
    minutes after the DB changed (Kerry 2026-07-20: 'Taking forever for
    the BARNA and Cedar Creek thing to resolve' — the server was long
    since correct). Applies to /api/* only; static assets keep their
    defaults."""
    if request.path.startswith("/api/") and "Cache-Control" not in resp.headers:
        resp.headers["Cache-Control"] = "no-store"
    # HTML pages too (Kerry 2026-07-20 mobile): the iOS PWA served a
    # cached /events page whose OLD inline JS kept failing long after the
    # fix deployed — the modal even showed pre-fix error text. Pages must
    # always be fetched fresh; /static assets keep their defaults.
    elif (not request.path.startswith("/static/")
          and "Cache-Control" not in resp.headers
          and resp.mimetype == "text/html"):
        resp.headers["Cache-Control"] = "no-store"
    return resp


@app.context_processor
def _inject_shell_flag():
    """Nav Shell v2 kill switch (nav-shell-070926, Kerry-ratified #58).

    SHELL_V2 env var, default ON. Flip to 0/false/off on Railway and
    restart to instantly revert every page to its legacy header — no
    redeploy needed (rollback structure per mailbox #53).
    """
    return {"shell_v2": os.environ.get("SHELL_V2", "1").strip().lower() not in ("0", "false", "off")}


@app.errorhandler(500)
def handle_500(e):
    """Return JSON instead of HTML for unhandled server errors."""
    logger.exception("Unhandled server error: %s", e)
    return jsonify({"error": "Internal server error"}), 500


@app.route("/health")
def health_check():
    """Health check endpoint for Railway / monitoring. No auth required."""
    try:
        conn = get_connection()
        conn.execute("SELECT 1")
        conn.close()
        return jsonify({"status": "ok", "db": "ok"}), 200
    except Exception as e:
        logger.error("Health check failed: %s", e)
        return jsonify({"status": "error", "db": "error", "detail": str(e)}), 500


# ---------------------------------------------------------------------------
# Email check job (with background tracking)
# ---------------------------------------------------------------------------
_inbox_check_lock = threading.Lock()
_inbox_check_status = {
    "running": False,
    "error": None,
    "emails_fetched": 0,
    "emails_parsed": 0,
    "items_saved": 0,
    "message": None,
}


def check_inbox():
    """Fetch new transaction emails, parse them with AI, and save to DB."""
    tenant_id = os.getenv("AZURE_TENANT_ID")
    client_id = os.getenv("AZURE_CLIENT_ID")
    client_secret = os.getenv("AZURE_CLIENT_SECRET")
    address = os.getenv("EMAIL_ADDRESS")

    if not all([tenant_id, client_id, client_secret, address]):
        logger.warning("Azure AD / email credentials not configured — skipping inbox check")
        _inbox_check_status["message"] = "Azure AD / email credentials not configured"
        return

    logger.info("Checking inbox for %s ...", address)
    emails = fetch_transaction_emails(
        tenant_id=tenant_id,
        client_id=client_id,
        client_secret=client_secret,
        email_address=address,
        since_date=datetime.now() - timedelta(days=7),
    )

    _inbox_check_status["emails_fetched"] = len(emails)

    if not emails:
        logger.info("No new transaction emails found")
        _inbox_check_status["message"] = "No transaction emails matched filters in the last 7 days"
        return

    # Skip emails already parsed — avoids burning AI credits on duplicates
    known_uids = get_known_email_uids()
    new_emails = [e for e in emails if e.get("uid") not in known_uids]
    logger.info(
        "Fetched %d transaction emails, %d already parsed, %d new to process",
        len(emails), len(emails) - len(new_emails), len(new_emails),
    )

    if not new_emails:
        _inbox_check_status["message"] = f"All {len(emails)} emails already parsed — nothing new"
        return

    # Parse and save one email at a time so items appear on the dashboard
    # incrementally instead of waiting for the entire batch to finish.
    _inbox_check_status["emails_fetched"] = len(new_emails)
    total_saved = 0
    total_parsed = 0
    for i, email_data in enumerate(new_emails, 1):
        try:
            rows = parse_email(email_data)
            total_parsed += 1
            _inbox_check_status["emails_parsed"] = total_parsed
            if rows:
                _dup_alerts: list = []
                count = save_items(rows, _alerts_out=_dup_alerts)
                total_saved += count
                _inbox_check_status["items_saved"] = total_saved
                logger.info("Email %d/%d: saved %d items", i, len(new_emails), count)
                for _alert in _dup_alerts:
                    _send_dup_reg_alert(_alert)
                # Persist any parse warnings (e.g. item_name is just a course name)
                try:
                    save_parse_warnings(rows)
                except Exception:
                    logger.exception("Failed to save parse warnings (non-fatal)")
            else:
                logger.info("Email %d/%d: no items extracted", i, len(new_emails))
            # Always mark as processed so we don't re-parse next cycle
            mark_email_processed(email_data.get("uid", ""), len(rows))
        except (_anthropic.BadRequestError, _anthropic.AuthenticationError) as e:
            logger.error(
                "Stopping at email %d/%d — Anthropic API fatal error: %s",
                i, len(new_emails), e.message,
            )
            raise
        except Exception:
            logger.exception("Failed to parse email %d/%d uid=%s", i, len(new_emails), email_data.get("uid"))

    # Auto-sync: create event entries for any new event-like items
    if total_saved > 0:
        try:
            sync_result = sync_events_from_items()
            if sync_result.get("inserted"):
                logger.info("Auto-synced %d new events from incoming transactions", sync_result["inserted"])
        except Exception:
            logger.exception("Auto-sync events failed (non-fatal)")

        # Auto-sync: link season contest payments to enrollments
        try:
            from email_parser.database import sync_season_contests_from_items
            sc_result = sync_season_contests_from_items()
            if sc_result.get("enrolled") or sc_result.get("linked"):
                logger.info(
                    "Auto-synced season contests: %d new enrollments, %d payments linked",
                    sc_result.get("enrolled", 0), sc_result.get("linked", 0),
                )
        except Exception:
            logger.exception("Auto-sync season contests failed (non-fatal)")

    _inbox_check_status["message"] = f"Done — saved {total_saved} items from {len(new_emails)} new emails ({len(emails)} total scanned)"
    logger.info("Done — saved %d total new items from %d new emails", total_saved, len(new_emails))


def send_coo_daily_email():
    """Send the daily COO briefing email. Runs compliance checks first."""
    from email_parser.database import run_compliance_checks, log_agent_action
    try:
        checks = run_compliance_checks()
        if checks:
            logger.info("Compliance checks created %d action items: %s", len(checks), checks)
            log_agent_action("Compliance Agent", "daily_compliance_run",
                             f"Created {len(checks)} items: {json.dumps(checks)}",
                             outcome="completed")
    except Exception:
        logger.exception("Compliance checks failed (non-fatal)")

    coo_to = os.getenv("COO_EMAIL_TO")
    if not coo_to:
        logger.info("COO_EMAIL_TO not set — skipping daily email")
        return False

    tenant_id = os.getenv("AZURE_TENANT_ID")
    client_id = os.getenv("AZURE_CLIENT_ID")
    client_secret = os.getenv("AZURE_CLIENT_SECRET")
    from_address = os.getenv("EMAIL_ADDRESS")

    if not all([tenant_id, client_id, client_secret, from_address]):
        logger.warning("Azure AD / email not configured — skipping COO email")
        return False

    try:
        subject, html_body = build_coo_email_html()
        ok = send_mail_graph(
            tenant_id=tenant_id, client_id=client_id,
            client_secret=client_secret, from_address=from_address,
            to_address=coo_to, subject=subject, html_body=html_body,
        )
        if ok:
            logger.info("COO daily email sent to %s", coo_to)
        else:
            logger.error("COO daily email failed to send to %s", coo_to)
        return ok
    except Exception:
        logger.exception("COO daily email error")
        return False


def check_expense_inbox(force=False, days_back=None):
    """Classify and extract data from non-order emails (Chase alerts, Venmo, receipts).

    Cost note: every email the classifier touches is recorded in
    `expense_seen_emails`, so each email is classified (and billed to Anthropic)
    at most once — regardless of how often the scheduler runs. Polling frequency
    is therefore decoupled from cost; the lookback window only bounds a cheap
    Microsoft Graph fetch.

    Args:
        force: If True, reprocess expense/action emails (ignore expense dedup).
        days_back: Explicit lookback in days (admin/manual runs). When None
            (the scheduled call), the window is 48h steady-state
            (EXPENSE_LOOKBACK_HOURS) — or a one-time wider backfill
            (EXPENSE_BACKFILL_DAYS, default 14) on a cold start, i.e. a fresh
            DB or a Railway volume that was wiped.
    """
    tenant_id = os.getenv("AZURE_TENANT_ID")
    client_id = os.getenv("AZURE_CLIENT_ID")
    client_secret = os.getenv("AZURE_CLIENT_SECRET")
    # Expense emails (Chase/Venmo) go to kerry@, not admin@
    address = os.getenv("EXPENSE_EMAIL_ADDRESS") or os.getenv("RSVP_EMAIL_ADDRESS") or os.getenv("EMAIL_ADDRESS")

    if not all([tenant_id, client_id, client_secret, address]):
        return {"error": "Azure AD credentials not configured"}

    # Decide the lookback window. Dedup makes a wide window free (already-seen
    # uids are filtered before any AI call), so the window is sized for SAFETY
    # — wide enough to clear a Railway redeploy gap — not minimized.
    seen_uids = get_expense_seen_uids()
    if days_back is not None:
        since_date = datetime.now() - timedelta(days=days_back)
    elif not seen_uids:
        backfill_days = int(os.getenv("EXPENSE_BACKFILL_DAYS", "14"))
        since_date = datetime.now() - timedelta(days=backfill_days)
        logger.warning(
            "Expense classifier cold start (no prior seen-email records) — "
            "one-time backfill of %d days. If this logs on every redeploy, the "
            "Railway persistent volume is not configured (see CLAUDE.md).",
            backfill_days,
        )
    else:
        lookback_hours = int(os.getenv("EXPENSE_LOOKBACK_HOURS", "48"))
        since_date = datetime.now() - timedelta(hours=lookback_hours)

    try:
        emails = fetch_all_emails(
            tenant_id=tenant_id, client_id=client_id, client_secret=client_secret,
            email_address=address, since_date=since_date,
            max_emails=300,
            include_subfolders=["2025 Chase", "2025 Venmo", "Payouts", "Invoices"],
        )
    except Exception:
        logger.exception("Failed to fetch emails for expense classification")
        return {"error": "Failed to fetch emails"}

    if not emails:
        return {"fetched": 0, "new": 0, "processed": 0}

    # Skip already-processed emails (check both processed_emails and expense_transactions)
    known_uids = get_known_email_uids()
    conn = get_connection()
    try:
        expense_uids = {r["email_uid"] for r in conn.execute(
            "SELECT email_uid FROM expense_transactions WHERE email_uid IS NOT NULL"
        ).fetchall()}
        action_uids = {r["email_uid"] for r in conn.execute(
            "SELECT email_uid FROM action_items WHERE email_uid IS NOT NULL"
        ).fetchall()}
    finally:
        conn.close()

    # seen_uids (expense_seen_emails) is the comprehensive gate going forward;
    # expense_uids/action_uids are kept as defense-in-depth for rows that
    # predate this table.
    all_known = known_uids | expense_uids | action_uids | seen_uids
    if force:
        # In force mode, only skip order emails (processed_emails) — reprocess expense/action
        new_emails = [e for e in emails if e.get("uid") not in known_uids]
    else:
        new_emails = [e for e in emails if e.get("uid") not in all_known]

    if not new_emails:
        return {"fetched": len(emails), "new": 0, "processed": 0,
                "note": f"All {len(emails)} fetched emails already processed"}

    logger.info("Classifying %d new emails for expense processing", len(new_emails))
    processed = 0
    conn = get_connection()

    for email_data in new_emails:
        try:
            body_text = _strip_html(email_data.get("html") or email_data.get("text", ""))
            classification = classify_email(
                email_data.get("subject", ""),
                email_data.get("from", ""),
                body_text,
            )
            email_type = classification["type"]
            confidence = classification["confidence"]

            # Record the email as seen the moment it is classified — BEFORE the
            # per-type branches (several of which `continue`). This is the fix
            # that stops `unknown`/order/RSVP emails from being re-classified
            # and re-billed to Anthropic on every scheduler cycle. A rare
            # classify_email() exception falls through to the except handler
            # below and is left unmarked so it can retry next cycle (matches
            # the GoDaddy parser's behavior).
            mark_expense_email_seen(email_data.get("uid", ""), email_type)

            if email_type == "godaddy_order" or email_type == "golf_genius_rsvp":
                continue  # Handled by existing parsers

            if email_type == "unknown":
                continue

            if email_type == "chase_transaction_alert":
                merchant_ctx = None
                extracted = parse_chase_alert(
                    email_data.get("subject", ""),
                    email_data.get("from", ""),
                    body_text, merchant_ctx,
                )
                if extracted.get("confidence", 0) > 0:
                    merchant_name = (extracted.get("merchant") or "").strip()
                    # Auto-ignore Chase Credit Card Statement notifications
                    # (monthly statement alerts, not actual transactions)
                    if merchant_name.upper() in (
                        "CHASE CREDIT CARD STATEMENT",
                        "CREDIT CARD STATEMENT",
                        "CHASE CREDIT CRD AUTOPAY",
                    ):
                        review_status = "ignored"
                        notes = "Auto-ignored: monthly statement notification"
                    else:
                        review_status = "approved" if extracted["confidence"] >= 95 else "pending"
                        notes = None
                    save_expense_transaction({
                        "email_uid": email_data["uid"],
                        "source_type": "chase_alert",
                        "merchant": merchant_name,
                        "amount": extracted.get("amount"),
                        "transaction_date": extracted.get("transaction_date"),
                        "account_last4": extracted.get("account_last4"),
                        "account_name": extracted.get("account_name"),
                        "transaction_type": extracted.get("transaction_type", "expense"),
                        "confidence": extracted["confidence"],
                        "review_status": review_status,
                        "notes": notes,
                        "raw_extract": json.dumps(extracted),
                    })
                    processed += 1

            elif email_type in ("venmo_payment", "paypal_payment",
                                 "cashapp_payment", "zelle_payment"):
                # One handler for all peer-to-peer providers (Venmo / PayPal /
                # Cash App / Zelle). They carry the same fields — recipient,
                # amount, and the typed note that holds the true payee + event
                # — so matching is provider-agnostic; only the source label and
                # a couple of Venmo-only extras differ (v2.84.0, Kerry).
                _prov = {"venmo_payment": "venmo", "paypal_payment": "paypal",
                         "cashapp_payment": "cashapp", "zelle_payment": "zelle"}[email_type]
                _prov_label = {"venmo": "Venmo", "paypal": "PayPal",
                               "cashapp": "Cash App", "zelle": "Zelle"}[_prov]
                extracted = parse_p2p_payment(
                    email_data.get("subject", ""),
                    email_data.get("from", ""),
                    body_text,
                    provider=_prov_label,
                )
                if extracted.get("confidence", 0) > 0:
                    memo_txt = extracted.get("memo", "") or ""
                    event_name = match_event_from_memo(memo_txt, conn)
                    # For outbound payouts TGF types the real payee into the
                    # memo ("Matt Griffin - Winnings for s9.17 …"). That name
                    # outranks the account's display name, which can belong to
                    # someone else (Matt's Venmo shows "robert griffin";
                    # Don's PayPal is his partner's, etc.). Try the memo payee
                    # prefix first; fall back to the recipient display name.
                    customer_id = None
                    # Optional leading "For " — Cash App renders the note as
                    # "For <Name> - Winnings for …".
                    _pm = re.match(r"\s*(?:for\s+)?(.+?)\s+-\s+winnings\s+for\b", memo_txt, re.I)
                    if _pm:
                        customer_id = match_customer_from_name(_pm.group(1).strip(), conn)
                    if not customer_id:
                        customer_id = match_customer_from_name(extracted.get("recipient_name", ""), conn)
                    # Fallback: if no event from memo but customer was found, check their registrations
                    if not event_name and customer_id:
                        event_name = match_event_from_customer(customer_id, conn)
                    review_status = "approved" if extracted["confidence"] >= 95 else "pending"
                    p2p_email_date = (email_data.get("date") or "")[:10]
                    # Venmo embeds the other party's @handle in link URLs the
                    # LLM never sees; PayPal/Cash App don't, so this is
                    # Venmo-only.
                    other_handle = None
                    if _prov == "venmo":
                        try:
                            from email_parser.expense_parser import extract_venmo_other_party_handle
                            other_handle = extract_venmo_other_party_handle(email_data.get("html") or "")
                        except Exception:
                            logger.warning("venmo handle extraction failed", exc_info=True)
                    saved = save_expense_transaction({
                        "email_uid": email_data["uid"],
                        "source_type": _prov,
                        "merchant": extracted.get("recipient_name"),
                        "amount": extracted.get("amount"),
                        "transaction_date": extracted.get("transaction_date") or p2p_email_date or None,
                        "transaction_type": extracted.get("transaction_type", "payout"),
                        "event_name": event_name,
                        "customer_id": customer_id,
                        "confidence": extracted["confidence"],
                        "review_status": review_status,
                        "notes": extracted.get("memo"),
                        "raw_extract": json.dumps(extracted),
                        "other_party_handle": other_handle,
                    })
                    processed += 1
                    # Stamp the player's Venmo @handle on their customer record if not set
                    if _prov == "venmo" and saved and saved.get("id") and saved.get("customer_id"):
                        try:
                            from email_parser.database import capture_venmo_handle_for_customer
                            capture_venmo_handle_for_customer(saved["id"])
                        except Exception:
                            logger.warning("capture_venmo_handle_for_customer failed for exp %s",
                                           saved.get("id"), exc_info=True)
                    # Auto-match incoming payments against open balance-due credit-transfers
                    # (runs for both approved and pending — matcher will auto-approve on match)
                    if saved and saved.get("transaction_type") == "received":
                        try:
                            from email_parser.database import auto_match_venmo_inbound_to_balance_due
                            auto_match_venmo_inbound_to_balance_due([saved["id"]])
                        except Exception:
                            logger.warning("inbound balance-due auto-match failed for exp %s",
                                           saved.get("id"), exc_info=True)
                        # Overpayment returns ride the same inbound receipt —
                        # "Overpaid winnings for <code>" memos close the open
                        # tgf_overpayments row the REQUEST button created
                        # (v2.141.0, the Hogue a9.19 CTP correction)
                        try:
                            from email_parser.database import recover_tgf_overpayments
                            recover_tgf_overpayments([saved["id"]])
                        except Exception:
                            logger.warning("overpayment recovery failed for exp %s",
                                           saved.get("id"), exc_info=True)
                    # Auto-confirm outbound winnings payments against pending
                    # tgf_payouts (Kerry, 2026-07-08 — PAYOUTS tab flips to
                    # PAID as soon as the receipt email lands; now Venmo,
                    # PayPal, and Cash App)
                    if saved and saved.get("transaction_type") == "payout":
                        try:
                            from email_parser.database import auto_match_venmo_payouts_to_tgf
                            auto_match_venmo_payouts_to_tgf([saved["id"]])
                        except Exception:
                            logger.warning("payout auto-match failed for exp %s",
                                           saved.get("id"), exc_info=True)
                        # Refund watches ride the same receipt, AFTER the
                        # winnings matcher (which gets first claim) —
                        # verifies + records player-credit refunds paid
                        # via the red Refund buttons (Kerry 2026-07-15)
                        try:
                            from email_parser.database import auto_match_refund_watches
                            rw = auto_match_refund_watches([saved["id"]])
                            if rw.get("verified"):
                                logger.info("Refund watch verified: %s", rw["matches"])
                        except Exception:
                            logger.warning("refund-watch auto-match failed for exp %s",
                                           saved.get("id"), exc_info=True)

            elif email_type == "expense_receipt":
                raw_email_date = (email_data.get("date") or "")[:10]
                extracted = parse_expense_receipt(
                    email_data.get("subject", ""),
                    email_data.get("from", ""),
                    body_text,
                    email_date=raw_email_date or None,
                )
                if extracted.get("confidence", 0) > 0:
                    review_status = "approved" if extracted["confidence"] >= 95 else "pending"
                    save_expense_transaction({
                        "email_uid": email_data["uid"],
                        "source_type": "receipt",
                        "merchant": extracted.get("merchant"),
                        "amount": extracted.get("amount"),
                        "transaction_date": extracted.get("transaction_date") or raw_email_date or None,
                        "account_last4": extracted.get("account_last4"),
                        "category": extracted.get("category"),
                        "entity": extracted.get("entity", "TGF"),
                        "confidence": extracted["confidence"],
                        "review_status": review_status,
                        "notes": extracted.get("description"),
                        "raw_extract": json.dumps(extracted),
                    })
                    processed += 1

            elif email_type == "action_required":
                extracted = parse_action_required(
                    email_data.get("subject", ""),
                    email_data.get("from", ""),
                    body_text,
                )
                if extracted.get("confidence", 0) > 0:
                    save_action_item({
                        "email_uid": email_data["uid"],
                        "subject": extracted.get("subject", email_data.get("subject")),
                        "from_name": extracted.get("from_name"),
                        "from_email": extracted.get("from_email", email_data.get("from")),
                        "summary": extracted.get("summary"),
                        "urgency": extracted.get("urgency", "medium"),
                        "category": extracted.get("category", "other"),
                        "email_date": (email_data.get("date") or "")[:10],
                        "confidence": extracted["confidence"],
                    })
                    processed += 1

        except Exception:
            logger.exception("Error processing email uid=%s for expense classification",
                             email_data.get("uid"))

    conn.close()
    if processed:
        logger.info("Expense email processing: %d items saved from %d emails", processed, len(new_emails))
    return {"fetched": len(emails), "new": len(new_emails), "processed": processed}


def check_rsvp_inbox():
    """Fetch new RSVP emails from Golf Genius, parse them, and save to DB."""
    rsvp_address = os.getenv("RSVP_EMAIL_ADDRESS")
    if not rsvp_address:
        logger.info("RSVP_EMAIL_ADDRESS not configured — skipping RSVP check")
        return

    logger.info("Checking RSVP inbox for %s ...", rsvp_address)
    try:
        emails = fetch_rsvp_emails(
            since_date=datetime.now() - timedelta(days=7),
        )
    except Exception as e:
        logger.exception("Failed to fetch RSVP emails: %s", e)
        return

    if not emails:
        logger.info("No RSVP emails found")
        return

    # Skip already-processed
    known_uids = get_known_rsvp_uids()
    new_emails = [e for e in emails if e.get("uid") not in known_uids]
    logger.info(
        "RSVP: fetched %d emails, %d already processed, %d new",
        len(emails), len(emails) - len(new_emails), len(new_emails),
    )

    if not new_emails:
        return

    parsed = parse_rsvp_emails(new_emails)
    if parsed:
        saved = save_rsvps(parsed)
        logger.info("RSVP: saved %d new RSVPs from %d emails", saved, len(new_emails))

    # Also re-run matching for any previously unmatched RSVPs
    rematch_rsvps()

    # Self-healing match hygiene (Kerry 2026-07-15): rematch_rsvps only
    # fills UNMATCHED rsvps — it never clears a bad first-name match
    # (the Daniel South → Daniel Lehan class). Run the full audit
    # (clear mismatches + rematch) across upcoming events after every
    # ingest, so bad matches never outlive one inbox cycle.
    try:
        from email_parser.database import audit_upcoming_event_rsvps
        res = audit_upcoming_event_rsvps()
        if res.get("cleared") or res.get("rematched"):
            logger.info("RSVP auto-audit after ingest: %s", res)
    except Exception:
        logger.exception("RSVP auto-audit failed (non-fatal)")

    # Check for credited players who just RSVPd and send admin alert emails
    _send_rsvp_credit_alerts()


def _send_dup_reg_alert(alert: dict) -> None:
    """Send an immediate email alert for a duplicate or guest-purchase registration."""
    tenant_id = os.getenv("AZURE_TENANT_ID")
    client_id = os.getenv("AZURE_CLIENT_ID")
    client_secret = os.getenv("AZURE_CLIENT_SECRET")
    from_address = os.getenv("EMAIL_ADDRESS")
    coo_to = os.getenv("COO_EMAIL_TO") or from_address
    if not all([tenant_id, client_id, client_secret, from_address, coo_to]):
        logger.warning("Email not configured — skipping dup-reg alert")
        return
    try:
        atype = alert["type"]
        customer = alert["customer"]
        event_name = alert["event_name"]
        order_id = alert.get("order_id", "")
        new_id = alert.get("new_item_id", "")
        prior_id = alert.get("prior_item_id", "")
        prior_order = alert.get("prior_order_id", "")

        if atype == "guest_purchase":
            subject = f"⚠ Guest purchase needs name — {customer} / {event_name}"
            headline = "Guest Purchase Detected"
            color = "#f59e0b"
            details = (
                f"<strong>{customer}</strong> purchased two spots for <strong>{event_name}</strong> "
                f"in order <strong>{order_id}</strong>. The second spot appears to be for a guest "
                f"but no guest name was provided."
            )
            action_label = "Assign Guest Name"
        else:
            subject = f"⚠ Duplicate registration — {customer} / {event_name}"
            headline = "Duplicate Registration Detected"
            color = "#ef4444"
            details = (
                f"<strong>{customer}</strong> now has two active registrations for "
                f"<strong>{event_name}</strong>: "
                f"item #{prior_id} (order {prior_order}) and item #{new_id} (order {order_id})."
            )
            action_label = "Review &amp; Credit Duplicate"

        base_url = "https://tgf-tracker.up.railway.app"
        events_url = f"{base_url}/events"
        html_body = f"""<!DOCTYPE html>
<html><body style="font-family:sans-serif;max-width:600px;margin:0 auto;padding:20px;">
<div style="border-left:4px solid {color};padding:12px 16px;background:#fafafa;margin-bottom:16px;">
  <h2 style="margin:0 0 8px;color:{color};">{headline}</h2>
  <p style="margin:0;">{details}</p>
</div>
<p><strong>Action required:</strong> {action_label}</p>
<p><a href="{events_url}" style="background:{color};color:#fff;padding:8px 16px;
   border-radius:4px;text-decoration:none;display:inline-block;">View Events Tab &rarr;</a></p>
<p style="color:#6b7280;font-size:0.85em;">{alert.get('summary','')}</p>
</body></html>"""
        ok = send_mail_graph(
            tenant_id=tenant_id, client_id=client_id,
            client_secret=client_secret, from_address=from_address,
            to_address=coo_to, subject=subject, html_body=html_body,
        )
        if ok:
            logger.info("Dup-reg alert sent: %s / %s (%s)", customer, event_name, atype)
        else:
            logger.warning("Dup-reg alert send failed: %s / %s", customer, event_name)
    except Exception:
        logger.exception("_send_dup_reg_alert failed")


def _send_rsvp_credit_alerts():
    """Find newly matched RSVPs for credited players and send admin alert emails."""
    from email_parser.database import _connect as _db_connect
    from email_parser.fetcher import send_mail_graph

    tenant_id = os.getenv("AZURE_TENANT_ID")
    client_id = os.getenv("AZURE_CLIENT_ID")
    client_secret = os.getenv("AZURE_CLIENT_SECRET")
    from_address = os.getenv("EMAIL_ADDRESS")
    # All credit alerts go here during testing — remove override to go live
    alert_to = os.getenv("CREDIT_ALERT_EMAIL_OVERRIDE", "kerry@thegolffellowship.com")

    if not all([tenant_id, client_id, client_secret, from_address]):
        logger.warning("Email credentials not configured — skipping credit alerts")
        return

    # Find RSVPs that: (a) are matched to an event, (b) haven't been notified yet
    with _db_connect() as conn:
        pending = conn.execute(
            """SELECT id FROM rsvps
               WHERE matched_event IS NOT NULL
                 AND response = 'PLAYING'
                 AND credit_notified_at IS NULL""",
        ).fetchall()

    for row in pending:
        rsvp_id = row["id"]
        try:
            info = get_rsvp_credit_info(rsvp_id)
            if not info:
                # No credits — stamp it anyway so we skip next time
                mark_rsvp_credit_notified(rsvp_id)
                continue

            # Build email
            player = info["player_name"]
            event_name = info["event_name"]
            event_date = info["event_date"]
            course = info["course"]
            total_credit = info["total_credit"]
            # Use subtotal (pre-tx-fee) — balance is paid via Venmo with no tx fee
            new_price = info.get("new_event_subtotal") or info["new_event_price"]
            amount_owed = info["amount_owed"]
            can_calc = info["can_calculate"]
            sel = info["selections"]

            credit_lines = "".join(
                f"<li>${c['credit_amount']:.2f} from <em>{c['event_name']}</em></li>"
                for c in info["credits"]
            )

            if can_calc:
                scenario = (
                    f"<strong style='color:#dc2626;'>Balance due: ${amount_owed:.2f}</strong>"
                    if amount_owed > 0
                    else f"<strong style='color:#16a34a;'>Excess credit: ${abs(amount_owed):.2f}</strong>"
                    if amount_owed < 0
                    else "<strong style='color:#6b7280;'>Credit covers exactly</strong>"
                )
                price_table = f"""
                <table style="border-collapse:collapse; margin:0.75rem 0; font-size:0.9rem;">
                  <tr><td style="padding:3px 12px 3px 0; color:#6b7280;">New event cost</td>
                      <td style="padding:3px 0;"><strong>${new_price:.2f}</strong></td></tr>
                  <tr><td style="padding:3px 12px 3px 0; color:#6b7280;">Credit applied</td>
                      <td style="padding:3px 0;">−${total_credit:.2f}</td></tr>
                  <tr style="border-top:1px solid #e5e7eb;">
                      <td style="padding:6px 12px 3px 0;">Result</td>
                      <td style="padding:6px 0;">{scenario}</td></tr>
                </table>"""
                action_note = (
                    f"<p><strong>Action:</strong> Player owes ${amount_owed:.2f} — "
                    f"apply credit in the admin UI and request Venmo payment of "
                    f"<strong>${amount_owed:.2f} to @tgf-payments</strong>.</p>"
                    if amount_owed > 0
                    else f"<p><strong>Action:</strong> Credit more than covers this event. "
                    f"Apply credit in admin UI. Excess ${abs(amount_owed):.2f} stays on account.</p>"
                    if amount_owed < 0
                    else "<p><strong>Action:</strong> Credit covers exactly — apply in admin UI.</p>"
                )
            else:
                price_table = "<p style='color:#d97706;'>⚠ Event pricing not configured — calculate manually.</p>"
                action_note = "<p>Apply credit manually via the Events admin page.</p>"

            subject = (
                f"[CREDIT ALERT] {player} RSVPd for {event_name}"
                + (" — owes ${:.2f}".format(amount_owed) if (can_calc and amount_owed > 0) else
                   " — credit covers it" if can_calc else " — pricing unknown")
            )

            html_body = f"""
<p>Hi Kerry,</p>
<p>A credited player has RSVPd for an upcoming event.</p>
<table style="border-collapse:collapse; background:#f9fafb; border:1px solid #e5e7eb;
              border-radius:6px; padding:0.75rem; margin:0.75rem 0; font-size:0.9rem; width:100%;">
  <tr><td style="padding:3px 12px 3px 0; color:#6b7280; width:130px;">Player</td>
      <td><strong>{player}</strong></td></tr>
  <tr><td style="padding:3px 12px 3px 0; color:#6b7280;">Event</td>
      <td><strong>{event_name}</strong></td></tr>
  <tr><td style="padding:3px 12px 3px 0; color:#6b7280;">Date / Course</td>
      <td>{event_date} &bull; {course}</td></tr>
  <tr><td style="padding:3px 12px 3px 0; color:#6b7280;">Previous selections</td>
      <td>{sel['holes']}h &bull; {sel['side_games']} &bull; {sel['user_status']}</td></tr>
</table>
<p><strong>Credits on account:</strong></p>
<ul>{credit_lines}</ul>
<p><strong>Total credit: ${total_credit:.2f}</strong></p>
{price_table}
{action_note}
<p style="font-size:0.8rem; color:#9ca3af;">
  This alert was sent to kerry@thegolffellowship.com instead of the player for testing.
  Remove CREDIT_ALERT_EMAIL_OVERRIDE env var to route to players.
</p>
<p>— TGF System</p>"""

            ok = send_mail_graph(
                tenant_id=tenant_id,
                client_id=client_id,
                client_secret=client_secret,
                from_address=from_address,
                to_address=alert_to,
                subject=subject,
                html_body=html_body,
            )
            if ok:
                logger.info("Credit alert sent for RSVP %s (%s)", rsvp_id, player)
                from email_parser.database import log_message
                log_message({
                    "event_name": event_name,
                    "channel": "email",
                    "recipient_name": player,
                    "recipient_address": alert_to,
                    "subject": subject,
                    "body_preview": f"Credit alert: {player} RSVPd for {event_name}. Credit: ${total_credit:.2f}",
                    "status": "sent",
                    "sent_by": "system",
                })
            else:
                logger.warning("Credit alert email failed for RSVP %s", rsvp_id)

            mark_rsvp_credit_notified(rsvp_id)
        except Exception:
            logger.warning("Credit alert check failed for RSVP %s", rsvp_id, exc_info=True)
            mark_rsvp_credit_notified(rsvp_id)  # stamp it to avoid retry loops


def _check_inbox_background():
    """Wrapper that runs check_inbox in a background thread with status tracking."""
    _inbox_check_status["emails_fetched"] = 0
    _inbox_check_status["emails_parsed"] = 0
    _inbox_check_status["items_saved"] = 0
    _inbox_check_status["message"] = None
    try:
        check_inbox()
        _inbox_check_status["error"] = None
    except Exception as e:
        logger.exception("Background inbox check failed")
        msg = str(e)
        from email_parser.ops_alerts import maybe_alert_anthropic_billing
        maybe_alert_anthropic_billing(e)
        # Provide user-friendly messages for known Anthropic API errors
        if "credit balance is too low" in msg.lower():
            _inbox_check_status["error"] = (
                "Anthropic API credit balance is too low. "
                "Please visit console.anthropic.com to add credits."
            )
        elif isinstance(e, _anthropic.AuthenticationError):
            _inbox_check_status["error"] = (
                "Anthropic API key is invalid or expired. "
                "Please check your ANTHROPIC_API_KEY in the .env file."
            )
        else:
            _inbox_check_status["error"] = msg
    finally:
        _inbox_check_status["running"] = False


# ---------------------------------------------------------------------------
# Connector API-key auth helper
# ---------------------------------------------------------------------------
def require_connector_key(f):
    """Decorator that validates the X-API-Key header against CONNECTOR_API_KEY."""
    @wraps(f)
    def decorated(*args, **kwargs):
        expected = os.getenv("CONNECTOR_API_KEY")
        if not expected:
            return jsonify({"error": "CONNECTOR_API_KEY not configured on server."}), 500
        provided = request.headers.get("X-API-Key", "")
        if not secrets.compare_digest(provided, expected):
            return jsonify({"error": "Invalid or missing API key."}), 401
        return f(*args, **kwargs)
    return decorated


# ---------------------------------------------------------------------------
# Input validation helper
# ---------------------------------------------------------------------------
MAX_STRING_LENGTH = 1000


def validate_json_fields(data: dict, required: list[str] = None,
                         max_len: int = MAX_STRING_LENGTH) -> str | None:
    """Validate JSON input fields. Returns an error message or None if valid."""
    if required:
        for field in required:
            if not data.get(field):
                return f"'{field}' is required."
    for key, value in data.items():
        if isinstance(value, str) and len(value) > max_len:
            return f"'{key}' exceeds maximum length of {max_len} characters."
    return None


# ---------------------------------------------------------------------------
# Role-based access helpers
# ---------------------------------------------------------------------------
# Role hierarchy: each rank includes the capabilities of the ranks below it.
_ROLE_RANK = {"member": 0, "view-only": 1, "manager": 2, "admin": 3}


def require_role(role):
    """Decorator that checks the session for a minimum role level.

    Roles are ranked view-only < manager < admin; a route declaring
    @require_role("manager") admits manager and admin sessions only.
    Until v2.16.15 only "admin" was actually enforced — a view-only
    session passed every manager-declared endpoint.
    """
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            # "member" is the PUBLIC read tier (v2.53.0, Kerry): endpoints
            # declaring it serve anonymous visitors — the pinless /member
            # pages (Season Contests + Handicaps). Only PII-free GET reads
            # may declare it.
            if _ROLE_RANK.get(role, 1) <= 0:
                return f(*args, **kwargs)
            user_role = session.get("role")
            if not user_role:
                return jsonify({"error": "Not authenticated. Please log in."}), 401
            if _ROLE_RANK.get(user_role, 0) < _ROLE_RANK.get(role, 0):
                label = "Admin" if role == "admin" else "Manager"
                return jsonify({"error": f"{label} access required."}), 403
            return f(*args, **kwargs)
        return decorated
    return decorator


# ---------------------------------------------------------------------------
# Scheduler
# ---------------------------------------------------------------------------
scheduler = BackgroundScheduler(daemon=True)


def start_scheduler():
    # Email-dedup memory (processed_emails / expense_seen_emails) lives in
    # SQLite. On Railway without a persistent volume the DB is wiped on every
    # redeploy, which resets dedup and re-bills the entire expense backfill
    # window of Anthropic calls on the next run. Make a misconfig loud.
    if not os.getenv("DATABASE_PATH"):
        logger.warning(
            "DATABASE_PATH is NOT set — SQLite DB at %s is EPHEMERAL on Railway. "
            "Every redeploy wipes email-dedup memory and re-bills the expense "
            "backfill window. Configure a persistent volume + DATABASE_PATH "
            "(see CLAUDE.md -> Railway Persistent Volume).", DB_PATH,
        )
    else:
        logger.info("DB persistence OK — DATABASE_PATH=%s", os.getenv("DATABASE_PATH"))

    interval = int(os.getenv("CHECK_INTERVAL_MINUTES", "5"))
    scheduler.add_job(
        check_inbox,
        "interval",
        minutes=interval,
        id="inbox_check",
        replace_existing=True,
    )

    # RSVP inbox check — same interval as transaction inbox
    if os.getenv("RSVP_EMAIL_ADDRESS"):
        scheduler.add_job(
            check_rsvp_inbox,
            "interval",
            minutes=interval,
            id="rsvp_inbox_check",
            replace_existing=True,
        )
        logger.info("RSVP scheduler: checking %s every %d minutes",
                     os.getenv("RSVP_EMAIL_ADDRESS"), interval)

    # Expense email classifier — runs MORE often than the order inbox so
    # Venmo payout receipts are caught (and the payout flips to PAID)
    # within a couple minutes even when the admin pays outside the app and
    # never taps an in-app Pay button (which is what schedules the fast
    # 75s/180s sweep). This is cost-neutral: the expense_seen_emails dedup
    # bills each email to the AI exactly once regardless of poll frequency
    # (see docs/claude/expense-workflow.md → Dedup & Cost Control), so a
    # tighter interval only adds free Microsoft Graph calls. Default 2 min;
    # override with EXPENSE_CHECK_INTERVAL_MINUTES (Kerry 2026-07-13:
    # confirmations "took way longer than I'd like").
    expense_interval = int(os.getenv("EXPENSE_CHECK_INTERVAL_MINUTES", "2"))
    scheduler.add_job(
        check_expense_inbox,
        "interval",
        minutes=expense_interval,
        id="expense_inbox_check",
        replace_existing=True,
    )
    logger.info("Expense email classifier scheduled every %d minutes", expense_interval)

    # COO daily email — runs at 7:00 AM US/Central
    coo_email_to = os.getenv("COO_EMAIL_TO")
    if coo_email_to:
        scheduler.add_job(
            send_coo_daily_email,
            "cron",
            hour=7,
            minute=0,
            timezone="US/Central",
            id="coo_daily_email",
            replace_existing=True,
        )
        logger.info("COO daily email scheduled for 07:00 US/Central → %s", coo_email_to)

    # Daily digest — runs at 6:00 AM US/Central by default
    report_hour = int(os.getenv("DAILY_REPORT_HOUR", "6"))
    report_tz = os.getenv("DAILY_REPORT_TZ", "US/Central")
    if os.getenv("DAILY_REPORT_TO"):
        scheduler.add_job(
            send_daily_report,
            "cron",
            hour=report_hour,
            minute=0,
            timezone=report_tz,
            id="daily_report",
            replace_existing=True,
        )
        logger.info("Daily digest scheduled for %02d:00 %s → %s",
                     report_hour, report_tz, os.getenv("DAILY_REPORT_TO"))

    # Auto payment reminders — every other day at 6:00 AM US/Central
    reminder_tz = os.getenv("DAILY_REPORT_TZ", "US/Central")
    scheduler.add_job(
        send_auto_payment_reminders,
        "cron",
        day="*/2",
        hour=6,
        minute=0,
        timezone=reminder_tz,
        id="auto_payment_reminders",
        replace_existing=True,
    )
    logger.info("Auto payment reminders scheduled every other day at 06:00 %s",
                reminder_tz)

    # Membership renewal reminders — daily at 9:00 AM US/Central. Sends per-window
    # notices (T-30 / T-7 / T-0 / T+14 lapsed), thank-you confirmations on detected
    # renewals, and the no-response admin digest.
    scheduler.add_job(
        run_membership_reminders,
        "cron",
        hour=9,
        minute=0,
        timezone=reminder_tz,
        id="membership_reminders",
        replace_existing=True,
    )
    logger.info("Membership renewal reminders scheduled daily at 09:00 %s", reminder_tz)

    # Golf Genius nightly sync: intentionally NOT scheduled (admin decision,
    # 2026-07). The screen-scraping upload never established a reliable
    # connection to GG; handicaps flow via the manual CSV export
    # (/api/handicaps/export-csv) which the admin uploads in the GG UI.
    # The on-demand POST /api/handicaps/sync-golf-genius endpoint remains
    # for explicit admin-triggered attempts.
    logger.info("Golf Genius nightly sync disabled — handicaps are exported manually via CSV")

    # Monthly points snapshot — daily GG refresh so the Contests MONTHLY
    # tab serves instantly from the DB; the Refresh button covers
    # same-day needs. 05:30 Central sits between the digest jobs.
    def refresh_monthly_points_job():
        from email_parser.database import (refresh_monthly_points_snapshot,
                                           record_monthly_points_payouts)
        try:
            refresh_monthly_points_snapshot()
            logger.info("Monthly points snapshot refreshed from Golf Genius")
        except Exception:
            logger.exception("Monthly points snapshot refresh failed (non-fatal)")
        # Record completed months' winners as SEASON CONTEST payout accounts
        # (v2.51.0, Kerry) — idempotent; new months appear once complete
        try:
            res = record_monthly_points_payouts()
            if res.get("recorded"):
                logger.info("Monthly points payouts recorded: %s",
                            [r["code"] for r in res["recorded"]])
        except Exception:
            logger.exception("Monthly points payout recording failed (non-fatal)")

    scheduler.add_job(
        refresh_monthly_points_job,
        "cron",
        hour=5,
        minute=30,
        timezone="US/Central",
        id="monthly_points_snapshot",
        replace_existing=True,
    )
    logger.info("Monthly points snapshot refresh scheduled daily at 05:30 US/Central")

    # ── Auto GG results sync (v2.40.0) ──────────────────────────────
    # Kerry closes an event in GG → results + payouts appear in the
    # Tracker without any manual import. Hourly noon-11pm Central (events
    # finish evenings); pure HTTP against the public portals, no AI spend.
    # Disable with AUTO_GG_SYNC=0.
    def auto_gg_results_sync_job():
        from email_parser.database import auto_gg_results_sync
        try:
            res = auto_gg_results_sync()
            rec = (res.get("payouts") or {}).get("recorded") or []
            logger.info("Auto GG results sync done: %d portal(s), %d event payout refresh(es)",
                        len(res.get("portals") or []), len(rec))
        except Exception:
            logger.exception("Auto GG results sync failed (non-fatal)")

    if os.getenv("AUTO_GG_SYNC", "1") != "0":
        scheduler.add_job(
            auto_gg_results_sync_job,
            "cron",
            hour="12-23",
            minute=10,
            timezone="US/Central",
            id="auto_gg_results_sync",
            replace_existing=True,
            max_instances=1,
            coalesce=True,
        )
        logger.info("Auto GG results sync scheduled hourly 12:10-23:10 US/Central")

    # ── Auto pairings grab (v2.99.0, Kerry) ─────────────────────────
    # Nightly walk of both portals' FINAL tee sheets into pairing
    # history + past-event PAIRINGS tabs. Idempotent (replace per
    # event); rounds without a published sheet are skipped, so the
    # team-board fallback data stands. Disable with AUTO_PAIRINGS_GRAB=0.
    def auto_pairings_grab_job():
        from email_parser.database import import_gg_teesheets_all
        for portal in ("sa", "austin"):
            try:
                res = import_gg_teesheets_all(portal, apply=True,
                                              budget_seconds=280)
                logger.info("Auto pairings grab %s: %d/%d rounds applied",
                            portal, res.get("ok", 0),
                            res.get("rounds_total", 0))
            except Exception:
                logger.exception("Auto pairings grab failed for %s "
                                 "(non-fatal)", portal)

    if os.getenv("AUTO_PAIRINGS_GRAB", "1") != "0":
        scheduler.add_job(
            auto_pairings_grab_job,
            "cron",
            hour=3,
            minute=20,
            timezone="US/Central",
            id="auto_pairings_grab",
            replace_existing=True,
            max_instances=1,
            coalesce=True,
        )
        logger.info("Auto pairings grab scheduled daily at 03:20 US/Central")

    # ── Nightly RSVP match audit (Kerry 2026-07-15) ──────────────────
    # Belt-and-braces sweep behind the per-ingest and on-open audits:
    # clear email-mismatched matches + rematch across upcoming events.
    # Disable with AUTO_RSVP_AUDIT=0.
    def auto_rsvp_audit_job():
        from email_parser.database import audit_upcoming_event_rsvps
        try:
            res = audit_upcoming_event_rsvps()
            logger.info("Nightly RSVP audit: %s", res)
        except Exception:
            logger.exception("Nightly RSVP audit failed (non-fatal)")

    if os.getenv("AUTO_RSVP_AUDIT", "1") != "0":
        scheduler.add_job(
            auto_rsvp_audit_job,
            "cron",
            hour=3,
            minute=35,
            timezone="US/Central",
            id="auto_rsvp_audit",
            replace_existing=True,
            max_instances=1,
            coalesce=True,
        )
        logger.info("Nightly RSVP audit scheduled daily at 03:35 US/Central")

    # First boot after this ships (or a fresh volume): populate the
    # snapshot in the background so the first MONTHLY open doesn't wait
    try:
        from email_parser.database import load_gg_snapshot
        if load_gg_snapshot("monthly_points") is None:
            scheduler.add_job(refresh_monthly_points_job,
                              id="monthly_points_bootstrap", replace_existing=True)
            logger.info("No monthly points snapshot found — bootstrap fetch queued")
    except Exception:
        logger.exception("Monthly points bootstrap check failed (non-fatal)")

    # Weekly cleanup: prune old processed_emails records (>90 days)
    scheduler.add_job(
        prune_processed_emails,
        "cron",
        day_of_week="sun",
        hour=3,
        minute=0,
        timezone="US/Central",
        id="prune_processed_emails",
        replace_existing=True,
    )
    logger.info("Processed emails pruning scheduled Sundays at 03:00 US/Central")

    # Run one-time startup prune
    try:
        pruned = prune_processed_emails()
        logger.info("Startup prune: deleted %d processed_emails older than 90 days", pruned)
    except Exception:
        logger.exception("Startup prune failed (non-fatal)")

    scheduler.start()
    logger.info("Scheduler started — checking inbox every %d minutes", interval)


def prune_processed_emails(days=90):
    """Delete processed_emails records older than N days.

    These records only exist to prevent re-parsing the same email. After 90 days
    the email is long gone from the inbox anyway, so the record is dead weight.
    Returns the number of rows deleted.
    """
    from email_parser.database import _connect
    with _connect() as conn:
        cursor = conn.execute(
            "DELETE FROM processed_emails WHERE processed_at < datetime('now', ?)",
            (f"-{days} days",),
        )
        conn.commit()
        deleted = cursor.rowcount
    if deleted:
        logger.info("Pruned %d processed_emails older than %d days", deleted, days)
    return deleted


def send_auto_payment_reminders():
    """Send payment reminders to all RSVP-only/gg_rsvp players for upcoming events."""
    tenant_id = os.getenv("AZURE_TENANT_ID")
    client_id = os.getenv("AZURE_CLIENT_ID")
    client_secret = os.getenv("AZURE_CLIENT_SECRET")
    from_address = os.getenv("EMAIL_ADDRESS")
    if not all([tenant_id, client_id, client_secret, from_address]):
        logger.warning("Auto reminders: email credentials not configured, skipping")
        return

    today = today_central_str()
    events = get_all_events()
    items = get_all_items()

    total_sent = 0
    total_failed = 0

    for ev in events:
        event_date = ev.get("event_date") or ""
        # Skip past events
        if event_date and event_date < today:
            continue
        event_name = ev.get("item_name") or ""
        if not event_name:
            continue

        # Find RSVP-only / gg_rsvp players for this event (case-insensitive)
        rsvp_players = [
            i for i in items
            if (i.get("item_name") or "").lower() == event_name.lower()
            and (i.get("transaction_status") or "active") in ("rsvp_only", "gg_rsvp")
        ]
        if not rsvp_players:
            continue

        for player in rsvp_players:
            to_email = _resolve_player_email(player)
            if not to_email:
                continue
            player_name = player.get("customer") or "Player"
            subject = f"Payment Reminder — {event_name}"
            html_body = (
                f"<p>Hi {player_name},</p>"
                f"<p>This is a friendly reminder that we have you down for "
                f"<strong>{event_name}</strong>, but we haven't received your payment yet.</p>"
                f"<p>Please complete your registration at your earliest convenience.</p>"
                f"<p>Thanks,<br>The Golf Fellowship</p>"
            )
            try:
                ok = send_mail_graph(
                    tenant_id=tenant_id, client_id=client_id,
                    client_secret=client_secret, from_address=from_address,
                    to_address=to_email, subject=subject, html_body=html_body,
                )
                if ok:
                    total_sent += 1
                else:
                    total_failed += 1
            except Exception:
                logger.exception("Auto reminder failed for %s", to_email)
                total_failed += 1

    logger.info("Auto payment reminders: %d sent, %d failed", total_sent, total_failed)


# ---------------------------------------------------------------------------
# Membership renewal reminders
# ---------------------------------------------------------------------------

def _membership_send_email(to_address: str, subject: str, html_body: str) -> bool:
    """Thin wrapper that membership notices use to fire emails through Graph.

    Two automatic admin-copy rules:

    1. **CC** the address(es) in `MEMBERSHIP_ADMIN_CC` (default
       ``admin@thegolffellowship.com``) when the TO is admin@thegolffellowship.com
       — i.e. internal admin notifications (roster opt-in/out, no-response digest).
    2. **BCC** the address(es) in `MEMBERSHIP_MEMBER_BCC` (default
       ``admin@thegolffellowship.com``) when the TO is **not** admin@ — i.e.
       member-facing reminders. The member doesn't see admin@ on their headers.

    Both env vars accept comma-separated lists. Set either to ``""`` to disable.
    Both lists are de-duplicated against the TO line so we never send a redundant
    copy if the same address ends up on TO and CC/BCC.

    Returns False (without raising) if email credentials aren't configured —
    the caller treats that as "skipped" and won't stamp the notice column,
    so the next scheduler run will retry.
    """
    tenant_id = os.getenv("AZURE_TENANT_ID")
    client_id = os.getenv("AZURE_CLIENT_ID")
    client_secret = os.getenv("AZURE_CLIENT_SECRET")
    from_address = os.getenv("EMAIL_ADDRESS")
    if not all([tenant_id, client_id, client_secret, from_address]):
        logger.warning("Membership send: email credentials missing, skipping send to %s", to_address)
        return False

    to_set = {a.strip().lower() for a in (to_address or "").split(",") if a.strip()}
    is_internal_admin = "admin@thegolffellowship.com" in (to_address or "").lower()

    def _resolve_extra(env_var: str, default: str) -> str | None:
        raw = os.getenv(env_var)
        if raw is None:
            raw = default
        raw = raw.strip()
        if not raw:
            return None
        # Drop addresses already on the TO line (avoid the redundant copy).
        filtered = ",".join(
            a.strip() for a in raw.split(",")
            if a.strip() and a.strip().lower() not in to_set
        )
        return filtered or None

    cc_address = _resolve_extra("MEMBERSHIP_ADMIN_CC", "admin@thegolffellowship.com") if is_internal_admin else None
    bcc_address = _resolve_extra("MEMBERSHIP_MEMBER_BCC", "admin@thegolffellowship.com") if not is_internal_admin else None

    try:
        return send_mail_graph(
            tenant_id, client_id, client_secret, from_address,
            to_address, subject, html_body,
            cc_address=cc_address,
            bcc_address=bcc_address,
        )
    except Exception:
        logger.exception("Membership send: graph failure for %s", to_address)
        return False


def run_membership_reminders():
    """Wrapper for the scheduler — runs the daily membership job."""
    try:
        from email_parser.memberships import daily_membership_job
        counts = daily_membership_job(_membership_send_email)
        logger.info("Membership reminders run complete: %s", counts)
        return counts
    except Exception:
        logger.exception("Membership reminders run failed")
        return None


# ---------------------------------------------------------------------------
# Input validation
# ---------------------------------------------------------------------------
_MAX_FIELD_LEN = 1000  # max characters per field value in update requests


def _validate_update_fields(data: dict) -> str | None:
    """Return an error message if any field value is invalid, else None."""
    for key, value in data.items():
        if not isinstance(key, str):
            return f"Field name must be a string, got {type(key).__name__}"
        if isinstance(value, str) and len(value) > _MAX_FIELD_LEN:
            return f"Field '{key}' exceeds max length ({_MAX_FIELD_LEN} chars)"
    return None


# ---------------------------------------------------------------------------
# Routes — Pages
# ---------------------------------------------------------------------------
@app.route("/")
def index():
    # EVENTS is the landing page (Kerry, 2026-07-08). Old transaction
    # deep-links (/?txn=123) keep working by continuing to the
    # Transactions page with their query intact.
    if request.args:
        return redirect("/transactions?" + request.query_string.decode())
    return redirect("/events")


@app.route("/transactions")
def transactions_page():
    if session.get("role") == "view-only":
        return redirect("/events")
    return render_template("index.html")


# ---------------------------------------------------------------------------
# Routes — API
# ---------------------------------------------------------------------------
@app.route("/api/items")
@require_role("view-only")
def api_items():
    """Return all item rows as JSON."""
    items = get_all_items()
    return jsonify(items)


@app.route("/api/stats")
@require_role("view-only")
def api_stats():
    """Return summary statistics."""
    stats = get_item_stats()
    return jsonify(stats)


@app.route("/api/audit")
@require_role("view-only")
def api_audit():
    """Data-quality report: field fill-rates, missing-data flags, value distributions."""
    report = get_audit_report()
    return jsonify(report)


@app.route("/api/migrate-customers-preview")
@require_role("admin")
def api_migrate_customers_preview():
    """Dry-run preview of customer migration — read-only, no inserts."""
    from migrate_customers import dry_run_json
    from email_parser.database import get_connection
    conn = get_connection()
    try:
        return jsonify(dry_run_json(conn))
    finally:
        conn.close()


@app.route("/api/migrate-customers", methods=["POST"])
@require_role("admin")
def api_migrate_customers():
    """Run the customer migration (idempotent)."""
    from migrate_customers import migrate
    from email_parser.database import get_connection
    conn = get_connection()
    try:
        stats = migrate(conn)
        return jsonify(stats)
    finally:
        conn.close()


@app.route("/api/data-snapshot")
@require_role("view-only")
def api_data_snapshot():
    """Quick snapshot of recent items + stats for inspection."""
    limit = request.args.get("limit", 50, type=int)
    return jsonify(get_data_snapshot(limit=limit))


@app.route("/api/items/<int:item_id>", methods=["PATCH"])
@require_role("admin")
def api_update_item(item_id):
    """Update specific fields on an item row (for inline editing). Admin only."""
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "Request body must be JSON."}), 400
    err = validate_json_fields(data)
    if err:
        return jsonify({"error": err}), 400

    # Identity coherence: renaming the customer on an item must re-resolve
    # customer_id — otherwise the row keeps the OLD person's id and every
    # cid-keyed feature (credits, winnings, memberships, balance-due
    # matching) still attributes the transaction to them. An explicit
    # customer_id in the payload wins (assign-member sends one).
    if (data.get("customer") or "").strip() and "customer_id" not in data:
        from email_parser.database import _resolve_or_create_customer
        conn = get_connection()
        try:
            row = conn.execute(
                "SELECT customer FROM items WHERE id = ?", (item_id,)
            ).fetchone()
            if row and (row["customer"] or "").strip().lower() != data["customer"].strip().lower():
                cid = _resolve_or_create_customer(
                    conn,
                    customer_name=data["customer"].strip(),
                    customer_email=(data.get("customer_email") or "").strip() or None,
                )
                conn.commit()
                if cid:
                    data["customer_id"] = cid
        finally:
            conn.close()

    updated = update_item(item_id, data)
    if updated:
        return jsonify({"status": "ok"})
    return jsonify({"error": "not found or no valid fields"}), 404


@app.route("/api/items/<int:item_id>/assign-guest", methods=["POST"])
@require_role("manager")
def api_assign_guest(item_id):
    """Assign the actual guest player name to a GUEST registration.

    When a member buys a guest registration, both items initially show the
    buyer as the customer. This endpoint swaps the customer to the actual
    guest and records the buyer in a 'Purchased by' note.
    """
    data = request.get_json(silent=True)
    guest_name = (data.get("guest_name") or "").strip() if data else ""
    if not guest_name:
        return jsonify({"error": "guest_name is required."}), 400
    err = validate_json_fields(data)
    if err:
        return jsonify({"error": err}), 400

    conn = get_connection()
    try:
        item = conn.execute("SELECT * FROM items WHERE id = ?", (item_id,)).fetchone()
        if not item:
            return jsonify({"error": "Item not found."}), 404
        item = dict(item)
        buyer = item["customer"] or ""

        from email_parser.parser import _normalize_customer_name
        normalized = _normalize_customer_name(guest_name)

        # Resolve (or create) the guest's own customer record — leaving
        # customer_id NULL made the guest's registration invisible to every
        # cid-keyed feature and left a row the boot backfill had to guess at.
        guest_email = ((data or {}).get("guest_email") or "").strip() or None
        guest_phone = ((data or {}).get("guest_phone") or "").strip() or None
        from email_parser.database import _resolve_or_create_customer
        guest_cid = _resolve_or_create_customer(
            conn, customer_name=normalized, customer_email=guest_email,
            phone=guest_phone, chapter=item.get("chapter"),
            user_status=item.get("user_status"),
        )
        conn.commit()

        changes = {
            "customer": normalized,
            "guest_name": normalized,
            "notes": f"Purchased by {buyer}",
            "customer_email": guest_email,
            "customer_phone": guest_phone,
            "customer_id": guest_cid or None,
        }
        update_item(item_id, changes)

        out = {"status": "ok", "customer": normalized, "buyer": buyer,
               "customer_id": guest_cid}
        if not guest_cid:
            # Never report success on the identity when it did not happen —
            # a NULL customer_id is invisible to every cid-keyed feature.
            out["warning"] = ("No customer record could be created for this "
                              "name — the registration is still unlinked.")

        # WHO BROUGHT THEM. The buyer is known at this exact moment, so the
        # referral is derived rather than asked for (Principle 1). This is a
        # RELATIONSHIP only — referral FEES arise solely from a redeemed
        # coupon or a payout receipt, and recording this must never mint one.
        if guest_cid and item.get("customer_id"):
            from email_parser.database import set_referred_by
            ref = set_referred_by(guest_cid, item["customer_id"])
            if "error" not in ref:
                out["referred_by_customer_id"] = ref["referred_by_customer_id"]
                out["referred_by_name"] = ref["referred_by_name"]

        # Optional STARTING handicap (Kerry 2026-07-30): a guest has no TGF
        # rounds, so no index, so they cannot be flighted and read "—" on
        # every roster. Stamping one here is the natural moment. Never a
        # handicap round — a real computed index always supersedes it.
        sh = (data or {}).get("starting_handicap_18")
        if sh not in (None, "") and guest_cid:
            from email_parser.database import set_starting_handicap
            res = set_starting_handicap(
                guest_cid, sh, set_by=f"assign-guest:{session.get('role')}",
                note=f"Guest of {buyer}" if buyer else "Guest registration")
            if "error" in res:
                # The name assignment already succeeded — report the handicap
                # problem without pretending the whole action failed.
                out["starting_handicap_error"] = res["error"]
            else:
                out["starting_handicap_18"] = res["starting_handicap_18"]
        return jsonify(out)
    finally:
        conn.close()


@app.route("/api/customers/<int:customer_id>/starting-handicap",
           methods=["POST"])
@require_role("manager")
def api_set_starting_handicap(customer_id):
    """Set or clear a player's placeholder handicap.

    For anyone with no established TGF index — guests, but first timers too
    (a MEMBER first timer has the same gap). Body: `starting_handicap_18`
    (null clears), optional `note`. Never creates a handicap round.
    """
    from email_parser.database import set_starting_handicap
    body = request.get_json(silent=True) or {}
    res = set_starting_handicap(
        customer_id, body.get("starting_handicap_18"),
        set_by=f"manual:{session.get('role')}",
        note=(body.get("note") or "").strip() or None)
    return (jsonify(res), 400) if "error" in res else jsonify(res)


@app.route("/api/items/<int:item_id>/assign-member", methods=["POST"])
@require_role("manager")
def api_assign_member(item_id):
    """Re-assign a registration to an existing member by customer_id.

    Used when a buyer purchased a spot for someone else and the parser
    couldn't identify the actual player (e.g. a family member who is a
    known member). Links the item to the canonical customer record so
    email, handicap, and credit lookups all work correctly.
    """
    data = request.get_json(silent=True) or {}
    customer_id = data.get("customer_id")
    if not customer_id:
        return jsonify({"error": "customer_id is required."}), 400

    conn = get_connection()
    try:
        item = conn.execute("SELECT * FROM items WHERE id = ?", (item_id,)).fetchone()
        if not item:
            return jsonify({"error": "Item not found."}), 404
        item = dict(item)
        buyer = item["customer"] or ""

        cust = conn.execute(
            """SELECT c.customer_id,
                      TRIM(COALESCE(NULLIF(c.company_name,''),
                           NULLIF(TRIM(c.first_name || ' ' || c.last_name
                                  || COALESCE(' ' || NULLIF(TRIM(c.suffix),''), '')),''))) AS full_name,
                      c.current_player_status,
                      c.chapter,
                      c.phone,
                      ce.email AS primary_email
               FROM customers c
               LEFT JOIN customer_emails ce ON ce.customer_id = c.customer_id AND ce.is_primary = 1
               WHERE c.customer_id = ?""",
            (customer_id,),
        ).fetchone()
        if not cust:
            return jsonify({"error": "Customer not found."}), 404
        cust = dict(cust)

        from email_parser.parser import _normalize_customer_name
        normalized = _normalize_customer_name(cust["full_name"] or "")

        changes = {
            "customer": normalized,
            "customer_id": cust["customer_id"],
            "customer_email": cust.get("primary_email") or None,
            "customer_phone": cust.get("phone") or None,
            "user_status": cust.get("current_player_status") or item.get("user_status"),
            "chapter": cust.get("chapter") or item.get("chapter"),
            "notes": f"Purchased by {buyer}" if buyer != normalized else item.get("notes"),
        }
        update_item(item_id, changes)

        return jsonify({"status": "ok", "customer": normalized, "buyer": buyer})
    finally:
        conn.close()


@app.route("/api/items/<int:item_id>", methods=["DELETE"])
@require_role("admin")
def api_delete_item(item_id):
    """Delete an item row by ID. Admin only."""
    deleted = delete_item(item_id)
    if deleted:
        return jsonify({"status": "ok"})
    return jsonify({"error": "not found"}), 404


@app.route("/api/events/delete-manual-player/<int:item_id>", methods=["DELETE"])
@require_role("manager")
def api_delete_manual_player(item_id):
    """Delete a manually added player. Only works for manual entries."""
    if delete_manual_player(item_id):
        return jsonify({"status": "ok"})
    return jsonify({"error": "Not found or not a manually added player."}), 400


@app.route("/api/check-now", methods=["POST"])
@require_role("manager")
def api_check_now():
    """Manually trigger an inbox check (runs in background to avoid timeout)."""
    tenant_id = os.getenv("AZURE_TENANT_ID")
    client_id = os.getenv("AZURE_CLIENT_ID")
    client_secret = os.getenv("AZURE_CLIENT_SECRET")
    address = os.getenv("EMAIL_ADDRESS")
    api_key = os.getenv("ANTHROPIC_API_KEY")

    if not all([tenant_id, client_id, client_secret, address]):
        return jsonify({"error": "Azure AD credentials not configured. Create a .env file from .env.example."}), 400

    if not api_key:
        return jsonify({"error": "ANTHROPIC_API_KEY not configured. Add it to your .env file."}), 400

    with _inbox_check_lock:
        if _inbox_check_status["running"]:
            return jsonify({"status": "already_running"})
        _inbox_check_status["running"] = True
        _inbox_check_status["error"] = None

    thread = threading.Thread(target=_check_inbox_background, daemon=True)
    thread.start()

    return jsonify({"status": "started"})


@app.route("/api/check-status")
@require_role("view-only")
def api_check_status():
    """Poll this endpoint to check if the background inbox check is done."""
    running = _inbox_check_status["running"]
    error = _inbox_check_status["error"]

    progress = {
        "emails_fetched": _inbox_check_status["emails_fetched"],
        "emails_parsed": _inbox_check_status["emails_parsed"],
        "items_saved": _inbox_check_status["items_saved"],
    }

    if running:
        return jsonify({"status": "running", "progress": progress})

    if error:
        return jsonify({"status": "error", "error": error, "progress": progress})

    stats = get_item_stats()
    return jsonify({
        "status": "done",
        "stats": stats,
        "progress": progress,
        "message": _inbox_check_status.get("message"),
    })


@app.route("/admin/backup")
@require_role("admin")
def admin_backup():
    """Stream the SQLite database file as a download. Admin-only."""
    db_path = str(DB_PATH)
    if not os.path.isfile(db_path):
        return jsonify({"error": "Database file not found"}), 404
    # Copy to a temp file to avoid streaming a locked WAL-mode DB
    backup_path = db_path + ".backup"
    shutil.copy2(db_path, backup_path)
    # Also checkpoint WAL into the backup
    try:
        conn = sqlite3.connect(backup_path)
        conn.execute("PRAGMA wal_checkpoint(TRUNCATE)")
        conn.close()
    except Exception:
        logger.debug("WAL checkpoint on backup copy failed (non-fatal)", exc_info=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return send_file(
        backup_path,
        mimetype="application/x-sqlite3",
        as_attachment=True,
        download_name=f"tgf_transactions_{timestamp}.db",
    )


@app.route("/api/health")
def api_health():
    """Diagnostic endpoint for Railway troubleshooting."""
    db_path = str(DB_PATH)
    db_exists = os.path.isfile(db_path)
    db_dir_exists = os.path.isdir(os.path.dirname(db_path))
    try:
        conn = get_connection()
        row = conn.execute("SELECT COUNT(*) as cnt FROM items").fetchone()
        item_count = row["cnt"]
        conn.close()
        db_readable = True
    except Exception:
        item_count = 0
        db_readable = False
    env_keys = ["AZURE_TENANT_ID", "AZURE_CLIENT_ID", "AZURE_CLIENT_SECRET",
                "EMAIL_ADDRESS", "ANTHROPIC_API_KEY", "DATABASE_PATH",
                "SECRET_KEY", "ADMIN_PIN", "MANAGER_PIN", "RSVP_EMAIL_ADDRESS"]
    env_status = {k: ("set" if os.getenv(k) else "missing") for k in env_keys}
    return jsonify({
        "status": "ok" if db_readable else "error",
        "database_path": db_path,
        "database_exists": db_exists,
        "database_dir_exists": db_dir_exists,
        "database_readable": db_readable,
        "item_count": item_count,
        "env_vars": env_status,
    })


@app.route("/api/config-status")
@require_role("view-only")
def api_config_status():
    """Check whether email, AI, and connector credentials are configured."""
    email_ok = all([
        os.getenv("AZURE_TENANT_ID"),
        os.getenv("AZURE_CLIENT_ID"),
        os.getenv("AZURE_CLIENT_SECRET"),
        os.getenv("EMAIL_ADDRESS"),
    ])
    ai_ok = bool(os.getenv("ANTHROPIC_API_KEY"))
    connector_ok = bool(os.getenv("CONNECTOR_API_KEY"))
    report_ok = bool(os.getenv("DAILY_REPORT_TO"))
    rsvp_ok = bool(os.getenv("RSVP_EMAIL_ADDRESS"))
    return jsonify({
        "configured": email_ok and ai_ok,
        "email": email_ok,
        "ai": ai_ok,
        "connector": connector_ok,
        "daily_report": report_ok,
        "rsvp": rsvp_ok,
    })


# ---------------------------------------------------------------------------
# Routes — Connector / Webhook
# ---------------------------------------------------------------------------
@app.route("/api/connector/ingest", methods=["POST"])
@require_connector_key
def api_connector_ingest():
    """
    Webhook endpoint for external systems to push order data.

    Accepts JSON with one of two formats:

    1. Pre-structured items (direct insert):
       {
         "items": [
           { "email_uid": "ext-123", "item_index": 0, "merchant": "...",
             "customer": "...", "item_name": "...", ... }
         ]
       }

    2. Raw email text (parsed by AI):
       {
         "raw_email": {
           "uid": "ext-123",
           "subject": "New Order #...",
           "from": "noreply@store.com",
           "text": "... full email body ..."
         }
       }
    """
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "Request body must be JSON."}), 400

    # Format 1: pre-structured items
    if "items" in data:
        items = data["items"]
        if not isinstance(items, list) or not items:
            return jsonify({"error": "'items' must be a non-empty array."}), 400
        _dup_alerts_c: list = []
        count = save_items(items, _alerts_out=_dup_alerts_c)
        for _al in _dup_alerts_c:
            _send_dup_reg_alert(_al)
        return jsonify({"status": "ok", "inserted": count, "received": len(items)})

    # Format 2: raw email for AI parsing
    if "raw_email" in data:
        raw = data["raw_email"]
        if not isinstance(raw, dict) or not raw.get("text"):
            return jsonify({"error": "'raw_email' must have at least a 'text' field."}), 400

        api_key = os.getenv("ANTHROPIC_API_KEY")
        if not api_key:
            return jsonify({"error": "ANTHROPIC_API_KEY not configured — cannot parse raw email."}), 500

        rows = parse_emails([raw])
        if rows:
            _dup_alerts_r: list = []
            count = save_items(rows, _alerts_out=_dup_alerts_r)
            for _al in _dup_alerts_r:
                _send_dup_reg_alert(_al)
            return jsonify({"status": "ok", "inserted": count, "parsed_items": len(rows)})
        return jsonify({"status": "ok", "inserted": 0, "message": "No items could be parsed from the email."}), 200

    return jsonify({"error": "Request must contain 'items' or 'raw_email'."}), 400


@app.route("/api/connector/info")
def api_connector_info():
    """Return connector configuration (whether key is set, not the key itself)."""
    key_set = bool(os.getenv("CONNECTOR_API_KEY"))
    return jsonify({
        "enabled": key_set,
        "endpoint": "/api/connector/ingest",
        "methods": ["POST"],
        "auth": "X-API-Key header",
        "formats": ["pre-structured items", "raw email for AI parsing"],
    })


@app.route("/api/audit/emails")
@require_role("admin")
def api_audit_emails():
    """
    Fetch raw emails from inbox AND the corresponding parsed DB records,
    returning them side-by-side so the user can verify extraction accuracy.

    Query params:
        limit       — max emails to return (default 25)
        days        — how far back to look (default 7); ignored if start_date is set
        start_date  — optional YYYY-MM-DD; overrides `days` lower bound
        end_date    — optional YYYY-MM-DD (exclusive upper bound)
    """
    tenant_id = os.getenv("AZURE_TENANT_ID")
    client_id = os.getenv("AZURE_CLIENT_ID")
    client_secret = os.getenv("AZURE_CLIENT_SECRET")
    address = os.getenv("EMAIL_ADDRESS")

    if not all([tenant_id, client_id, client_secret, address]):
        return jsonify({"error": "Azure AD / email credentials not configured."}), 400

    limit = request.args.get("limit", 25, type=int)
    days = request.args.get("days", 7, type=int)
    start_date_str = request.args.get("start_date", "").strip()
    end_date_str = request.args.get("end_date", "").strip()

    since_dt: datetime
    until_dt: datetime | None = None
    if start_date_str:
        try:
            since_dt = datetime.strptime(start_date_str, "%Y-%m-%d")
        except ValueError:
            return jsonify({"error": "start_date must be YYYY-MM-DD"}), 400
        if end_date_str:
            try:
                # end_date is inclusive in the UI, exclusive in the OData filter
                until_dt = datetime.strptime(end_date_str, "%Y-%m-%d") + timedelta(days=1)
            except ValueError:
                return jsonify({"error": "end_date must be YYYY-MM-DD"}), 400
    else:
        since_dt = datetime.now() - timedelta(days=days)

    try:
        emails = fetch_transaction_emails(
            tenant_id=tenant_id,
            client_id=client_id,
            client_secret=client_secret,
            email_address=address,
            since_date=since_dt,
            until_date=until_dt,
        )
    except Exception as e:
        logger.exception("Audit: failed to fetch emails")
        return jsonify({"error": f"Failed to fetch emails: {e}"}), 500

    # Build lookups of DB items keyed by email_uid AND order_id.
    # The order_id index covers the case where Microsoft Graph re-keys an
    # already-imported email under a brand-new message id — save_items()
    # dedups by (order_id, item_index) but the row keeps the OLD uid, so a
    # uid-only lookup falsely reports the email as "Not Parsed".
    all_items = get_all_items()
    db_by_uid: dict[str, list[dict]] = {}
    db_by_order: dict[str, list[dict]] = {}
    for item in all_items:
        uid = item.get("email_uid", "")
        if uid:
            db_by_uid.setdefault(uid, []).append(item)
        oid = (item.get("order_id") or "").strip()
        if oid:
            db_by_order.setdefault(oid, []).append(item)

    _order_id_re = re.compile(r"#(R\d+)")

    comparisons = []
    for email in emails[:limit]:
        uid = email.get("uid", "")
        body_text = email.get("text", "")
        if not body_text and email.get("html"):
            body_text = _strip_html(email["html"])

        # Truncate body for transport (keep first 2000 chars for review)
        body_preview = body_text[:2000] if body_text else "(empty)"

        db_rows = db_by_uid.get(uid, [])
        if not db_rows:
            # Fallback: pull order_id from subject (e.g. "New Order #R805080852")
            # and look up by order_id to catch cross-uid re-keys.
            m = _order_id_re.search(email.get("subject", "") or "")
            if m:
                db_rows = db_by_order.get(m.group(1), [])

        # Determine audit status
        if not db_rows:
            status = "missing"
            status_detail = "Email was fetched but no items were parsed/saved"
        else:
            # Check for missing critical fields
            issues = []
            for row in db_rows:
                missing = []
                for f in ["customer", "order_id", "item_name", "item_price"]:
                    if not row.get(f):
                        missing.append(f)
                if missing:
                    issues.append({"item_index": row.get("item_index", 0), "missing": missing})
            if issues:
                status = "incomplete"
                status_detail = f"{len(issues)} item(s) have missing fields"
            else:
                status = "ok"
                status_detail = f"{len(db_rows)} item(s) parsed successfully"

        comparisons.append({
            "email_uid": uid,
            "subject": email.get("subject", ""),
            "from": email.get("from", ""),
            "date": email.get("date", ""),
            "body_preview": body_preview,
            "status": status,
            "status_detail": status_detail,
            "parsed_items": db_rows,
        })

    # Summary counts
    total = len(comparisons)
    ok_count = sum(1 for c in comparisons if c["status"] == "ok")
    incomplete_count = sum(1 for c in comparisons if c["status"] == "incomplete")
    missing_count = sum(1 for c in comparisons if c["status"] == "missing")

    return jsonify({
        "total_emails": total,
        "ok": ok_count,
        "incomplete": incomplete_count,
        "missing": missing_count,
        "comparisons": comparisons,
    })


@app.route("/matrix")
def matrix_page():
    # Admin-only page
    if session.get("role") != "admin":
        return redirect("/events")
    matrix9, matrix18 = _load_matrix()
    return render_template("matrix.html", matrix9=matrix9, matrix18=matrix18)


@app.route("/changelog")
def changelog_page():
    # Admin-only page — managers are redirected to home
    if session.get("role") != "admin":
        return redirect("/events")
    return render_template("changelog.html")


@app.route("/audit")
def audit_page():
    # Admin-only page — managers are redirected to home
    if session.get("role") != "admin":
        return redirect("/events")
    return render_template("audit.html")


@app.route("/database")
def database_page():
    if session.get("role") != "admin":
        return redirect("/events")
    return render_template("database.html")


# ---------------------------------------------------------------------------
# Duplicate Detective (admin)
# ---------------------------------------------------------------------------
def _dd_build_summary(candidates):
    """Build the summary block embedded in the page banner + report exports."""
    high = sum(1 for c in candidates if c["confidence"] >= 0.90)
    medium = sum(1 for c in candidates if 0.70 <= c["confidence"] < 0.90)
    low = sum(1 for c in candidates if c["confidence"] < 0.70)
    by_pattern = {"A": 0, "B": 0, "C": 0, "D": 0}
    for c in candidates:
        by_pattern[c["pattern"]] = by_pattern.get(c["pattern"], 0) + 1
    variance = sum(c["variance_impact"] for c in candidates)
    fk_warned = sum(1 for c in candidates if c.get("fk_warnings"))
    return {
        "total": len(candidates),
        "high": high,
        "medium": medium,
        "low": low,
        "by_pattern": by_pattern,
        "variance_recovery": round(variance, 2),
        "fk_warned": fk_warned,
    }


@app.route("/admin/duplicate-detective")
def duplicate_detective_page():
    if session.get("role") != "admin":
        return redirect("/events")
    candidates = find_duplicate_candidates()
    bootstrap = {
        "candidates": candidates,
        "mode": get_duplicate_detective_mode(),
        "summary": _dd_build_summary(candidates),
    }
    return render_template(
        "duplicate_detective.html",
        bootstrap_json=json.dumps(bootstrap, default=str),
    )


@app.route("/admin/duplicate-detective/api/candidates")
@require_role("admin")
def api_duplicate_detective_candidates():
    include_dismissed = request.args.get("include_dismissed", "0") in ("1", "true", "yes")
    pattern = request.args.getlist("pattern") or None
    try:
        min_conf = float(request.args.get("min_confidence", "0"))
    except ValueError:
        min_conf = 0.0
    cands = find_duplicate_candidates(
        include_dismissed=include_dismissed,
        pattern_filter=pattern,
        min_confidence=min_conf,
    )
    return jsonify({"candidates": cands, "summary": _dd_build_summary(cands)})


@app.route("/admin/duplicate-detective/set-mode", methods=["POST"])
@require_role("admin")
def api_duplicate_detective_set_mode():
    body = request.get_json(silent=True) or {}
    mode = body.get("mode")
    try:
        set_duplicate_detective_mode(mode)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    return jsonify({"status": "ok", "mode": mode})


@app.route("/admin/duplicate-detective/merge-batch", methods=["POST"])
@require_role("admin")
def api_duplicate_detective_merge_batch():
    """Auto-merge every candidate with confidence >= 0.90 and no FK
    warnings. Only allowed in auto_high_confidence mode. Each pair runs in
    its own transaction (merge_duplicate_pair) so a single failure does
    not abort the batch."""
    from email_parser.database import (
        merge_duplicate_pair,
        DuplicateMergeError,
    )

    mode = get_duplicate_detective_mode()
    if mode != "auto_high_confidence":
        return jsonify({
            "error": "Batch merge requires mode=auto_high_confidence "
                     f"(current mode: {mode})",
        }), 409

    body = request.get_json(silent=True) or {}
    threshold = float(body.get("min_confidence", 0.90))
    candidates = find_duplicate_candidates(min_confidence=threshold)
    qualifying = [c for c in candidates if not c.get("fk_warnings")]
    skipped_fk = [c for c in candidates if c.get("fk_warnings")]

    merged_by = session.get("user") or "kerry"
    merged_results = []
    errors = []
    for c in qualifying:
        try:
            r = merge_duplicate_pair(
                int(c["suggested_survivor_id"]),
                int(c["suggested_merged_id"]),
                confidence=c["confidence"],
                reason=f"auto-batch: pattern {c['pattern']} | {c['rationale']}",
                merged_by=merged_by,
                allow_fk_hard_error=False,
            )
            merged_results.append({
                "candidate_id": c["candidate_id"],
                "audit_id": r["audit_id"],
                "variance_impact": c["variance_impact"],
                "noop": r["noop"],
            })
        except DuplicateMergeError as e:
            errors.append({
                "candidate_id": c["candidate_id"],
                "error": str(e),
            })
        except Exception as e:
            logger.exception("Auto-batch merge failed for candidate %s", c["candidate_id"])
            errors.append({
                "candidate_id": c["candidate_id"],
                "error": str(e),
            })

    return jsonify({
        "status": "ok",
        "mode": mode,
        "threshold": threshold,
        "qualifying": len(qualifying),
        "merged": len([r for r in merged_results if not r["noop"]]),
        "noop": len([r for r in merged_results if r["noop"]]),
        "errors": errors,
        "skipped_fk_warnings": [c["candidate_id"] for c in skipped_fk],
        "variance_recovered": round(
            sum(r["variance_impact"] for r in merged_results if not r["noop"]), 2
        ),
        "results": merged_results,
    })


@app.route("/admin/duplicate-detective/merge/<candidate_id>", methods=["POST"])
@require_role("admin")
def api_duplicate_detective_merge(candidate_id):
    """Merge a single candidate pair. Refused in dry_run_only mode."""
    from email_parser.database import (
        merge_duplicate_pair,
        DuplicateMergeError,
    )

    mode = get_duplicate_detective_mode()
    if mode == "dry_run_only":
        return jsonify({
            "error": "Mode is dry_run_only — merges are disabled. "
                     "Switch to review_each or auto_high_confidence first.",
        }), 409

    body = request.get_json(silent=True) or {}
    survivor = body.get("surviving_txn_id")
    merged = body.get("merged_txn_id")
    if not survivor or not merged:
        return jsonify({
            "error": "surviving_txn_id and merged_txn_id required",
        }), 400
    try:
        result = merge_duplicate_pair(
            int(survivor),
            int(merged),
            confidence=body.get("confidence"),
            reason=body.get("reason") or f"manual merge via candidate {candidate_id}",
            merged_by=session.get("user") or "kerry",
            allow_fk_hard_error=bool(body.get("allow_fk_hard_error")),
        )
    except DuplicateMergeError as e:
        return jsonify({"error": str(e)}), 422
    except Exception as e:
        logger.exception("Duplicate merge failed")
        return jsonify({"error": str(e)}), 500
    return jsonify({"status": "ok", **result})


@app.route("/admin/duplicate-detective/dismiss/<candidate_id>", methods=["POST"])
@require_role("admin")
def api_duplicate_detective_dismiss(candidate_id):
    body = request.get_json(silent=True) or {}
    txn_a = body.get("txn_a_id")
    txn_b = body.get("txn_b_id")
    if not txn_a or not txn_b:
        return jsonify({"error": "txn_a_id and txn_b_id required"}), 400
    reason = (body.get("reason") or "").strip()
    inserted = dismiss_duplicate_pair(int(txn_a), int(txn_b), reason)
    return jsonify({"status": "ok", "inserted": inserted})


@app.route("/admin/duplicate-detective/audit")
def duplicate_detective_audit_page():
    if session.get("role") != "admin":
        return redirect("/events")
    rows = get_duplicate_merge_audit(limit=500)
    return render_template(
        "duplicate_detective_audit.html",
        audit_json=json.dumps(rows, default=str),
    )


@app.route("/admin/duplicate-detective/reverse/<int:audit_id>", methods=["POST"])
@require_role("admin")
def api_duplicate_detective_reverse(audit_id):
    from email_parser.database import DuplicateMergeError
    try:
        r = reverse_duplicate_merge(
            audit_id, reversed_by=session.get("user") or "kerry"
        )
    except DuplicateMergeError as e:
        return jsonify({"error": str(e)}), 422
    except Exception as e:
        logger.exception("Duplicate merge reverse failed")
        return jsonify({"error": str(e)}), 500
    return jsonify({"status": "ok", **r})


@app.route("/admin/duplicate-detective/export.csv")
@require_role("admin")
def api_duplicate_detective_export_csv():
    """CSV: one row per candidate pair, all detection fields."""
    import csv
    import io

    cands = find_duplicate_candidates()
    fieldnames = [
        "candidate_id", "pattern", "confidence", "match_kind",
        "date_delta_days", "amount_delta", "variance_impact",
        "suggested_survivor_id", "suggested_merged_id",
        "txn_a_id", "txn_a_date", "txn_a_source", "txn_a_source_ref",
        "txn_a_customer", "txn_a_customer_id", "txn_a_amount",
        "txn_a_description",
        "txn_b_id", "txn_b_date", "txn_b_source", "txn_b_source_ref",
        "txn_b_customer", "txn_b_customer_id", "txn_b_amount",
        "txn_b_description",
        "rationale", "fk_warnings",
    ]
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=fieldnames)
    w.writeheader()
    for c in cands:
        a, b = c["txn_a"], c["txn_b"]
        w.writerow({
            "candidate_id": c["candidate_id"],
            "pattern": c["pattern"],
            "confidence": c["confidence"],
            "match_kind": c.get("match_kind") or "",
            "date_delta_days": c["date_delta_days"],
            "amount_delta": c["amount_delta"],
            "variance_impact": c["variance_impact"],
            "suggested_survivor_id": c["suggested_survivor_id"],
            "suggested_merged_id": c["suggested_merged_id"],
            "txn_a_id": a["id"], "txn_a_date": a.get("date"),
            "txn_a_source": a.get("source"), "txn_a_source_ref": a.get("source_ref"),
            "txn_a_customer": a.get("customer"), "txn_a_customer_id": a.get("customer_id"),
            "txn_a_amount": a.get("amount") if a.get("amount") is not None else a.get("total_amount"),
            "txn_a_description": (a.get("description") or "")[:200],
            "txn_b_id": b["id"], "txn_b_date": b.get("date"),
            "txn_b_source": b.get("source"), "txn_b_source_ref": b.get("source_ref"),
            "txn_b_customer": b.get("customer"), "txn_b_customer_id": b.get("customer_id"),
            "txn_b_amount": b.get("amount") if b.get("amount") is not None else b.get("total_amount"),
            "txn_b_description": (b.get("description") or "")[:200],
            "rationale": c["rationale"],
            "fk_warnings": " | ".join(c.get("fk_warnings") or []),
        })
    fname = f"duplicate_detective_report_{datetime.now().strftime('%Y-%m-%d')}.csv"
    return Response(
        buf.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@app.route("/admin/duplicate-detective/export.md")
@require_role("admin")
def api_duplicate_detective_export_md():
    """Markdown summary suitable for sharing in chat/email."""
    cands = find_duplicate_candidates()
    summary = _dd_build_summary(cands)
    today = today_central_str()
    lines = [
        f"# Duplicate Detective — Dry-Run Report",
        f"Generated: {today}",
        "",
        "## Summary",
        f"- Total probable duplicates: **{summary['total']}**",
        f"- Estimated variance recovery: **${summary['variance_recovery']:.2f}**",
        f"- Confidence: high ≥0.90: **{summary['high']}** · "
        f"medium 0.70–0.90: **{summary['medium']}** · "
        f"low <0.70: **{summary['low']}**",
        f"- Pattern A (Venmo CSV ↔ exp-promoted): **{summary['by_pattern']['A']}**",
        f"- Pattern B (in-app ↔ exp-promoted): **{summary['by_pattern']['B']}**",
        f"- Pattern C (in-app ↔ Venmo CSV): **{summary['by_pattern']['C']}**",
        f"- Pattern D (manual fallback): **{summary['by_pattern']['D']}**",
        f"- Pairs with FK warnings (manual review): **{summary['fk_warned']}**",
        "",
        "## Top 10 highest-impact pairs",
    ]
    top = sorted(cands, key=lambda c: -abs(c["variance_impact"]))[:10]
    if not top:
        lines.append("_None._")
    else:
        for c in top:
            a, b = c["txn_a"], c["txn_b"]
            survivor_id = c["suggested_survivor_id"]
            merged_id = c["suggested_merged_id"]
            cust = a.get("customer") or b.get("customer") or "(unknown)"
            lines.append(
                f"- **${c['variance_impact']:.2f}** · {cust} · "
                f"Pattern {c['pattern']} · confidence {int(c['confidence']*100)}% · "
                f"keep txn #{survivor_id}, merge txn #{merged_id} "
                f"({a.get('source_ref') or '(no ref)'} vs "
                f"{b.get('source_ref') or '(no ref)'})"
            )
    if summary["fk_warned"]:
        lines.append("")
        lines.append("## Pairs with FK warnings (require manual review)")
        for c in cands:
            if not c.get("fk_warnings"):
                continue
            lines.append(
                f"- txn #{c['suggested_survivor_id']} ↔ txn #{c['suggested_merged_id']}: "
                + " | ".join(c["fk_warnings"])
            )
    body = "\n".join(lines) + "\n"
    fname = f"duplicate_detective_summary_{today}.md"
    return Response(
        body,
        mimetype="text/markdown",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@app.route("/api/database/tables")
@require_role("admin")
def api_database_tables():
    """List all user tables and their row counts."""
    conn = get_connection()
    try:
        tables = conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' "
            "AND name NOT LIKE 'sqlite_%' ORDER BY name"
        ).fetchall()
        result = []
        for t in tables:
            name = t["name"]
            count = conn.execute(f'SELECT COUNT(*) AS c FROM "{name}"').fetchone()["c"]
            result.append({"name": name, "row_count": count})
        return jsonify(result)
    finally:
        conn.close()


@app.route("/api/database/table/<table_name>")
@require_role("admin")
def api_database_table(table_name):
    """Return rows from a specific table with pagination."""
    conn = get_connection()
    try:
        # Validate table name exists (prevent SQL injection)
        valid = conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
            (table_name,),
        ).fetchone()
        if not valid:
            return jsonify({"error": "Table not found"}), 404

        limit = request.args.get("limit", 100, type=int)
        offset = request.args.get("offset", 0, type=int)
        search = request.args.get("search", "").strip()
        sort_col = request.args.get("sort", "").strip()
        sort_dir = request.args.get("dir", "asc").strip().lower()

        # Get column names (and full schema for the empty-table view in
        # /database — name/type/notnull/default/pk so the admin can still see
        # the table structure when no rows exist yet).
        cols_info = conn.execute(f'PRAGMA table_info("{table_name}")').fetchall()
        columns = [c["name"] for c in cols_info]
        schema = [
            {
                "name": c["name"],
                "type": c["type"],
                "notnull": bool(c["notnull"]),
                "default": c["dflt_value"],
                "pk": bool(c["pk"]),
            }
            for c in cols_info
        ]

        # Build query
        where_clause = ""
        params: list = []
        if search:
            # Search across all text columns
            conditions = [f'CAST("{col}" AS TEXT) LIKE ?' for col in columns]
            where_clause = "WHERE " + " OR ".join(conditions)
            params = [f"%{search}%"] * len(columns)

        # Total count (with search filter)
        total = conn.execute(
            f'SELECT COUNT(*) AS c FROM "{table_name}" {where_clause}', params
        ).fetchone()["c"]

        # Sort
        order_clause = ""
        if sort_col and sort_col in columns:
            direction = "DESC" if sort_dir == "desc" else "ASC"
            order_clause = f'ORDER BY "{sort_col}" {direction}'
        else:
            order_clause = "ORDER BY rowid DESC"

        rows = conn.execute(
            f'SELECT * FROM "{table_name}" {where_clause} {order_clause} LIMIT ? OFFSET ?',
            params + [limit, offset],
        ).fetchall()

        return jsonify({
            "table": table_name,
            "columns": columns,
            "schema": schema,
            "rows": [dict(r) for r in rows],
            "total": total,
            "limit": limit,
            "offset": offset,
        })
    finally:
        conn.close()


def _load_matrix_from_file() -> tuple[dict, dict]:
    """Parse the static games-matrix.js file and return (matrix9, matrix18)."""
    matrix_path = os.path.join(
        os.path.dirname(__file__), "static", "js", "games-matrix.js"
    )
    with open(matrix_path, "r") as f:
        content = f.read()
    m9 = re.search(r"window\.GAMES_MATRIX_9\s*=\s*(\{.*?\});", content, re.DOTALL)
    m18 = re.search(r"window\.GAMES_MATRIX_18\s*=\s*(\{.*?\});", content, re.DOTALL)
    return json.loads(m9.group(1)), json.loads(m18.group(1))


def _load_matrix() -> tuple[dict, dict]:
    """Load matrices from DB if saved, otherwise from the static JS file."""
    db9 = get_app_setting("games_matrix_9")
    db18 = get_app_setting("games_matrix_18")
    if db9 and db18:
        return json.loads(db9), json.loads(db18)
    return _load_matrix_from_file()


@app.route("/api/matrix", methods=["GET"])
@require_role("view-only")
def api_matrix_get():
    """Return the current games matrix (from DB if edited, else from static file)."""
    matrix9, matrix18 = _load_matrix()
    return jsonify({"matrix9": matrix9, "matrix18": matrix18})


@app.route("/api/matrix", methods=["PUT"])
@require_role("admin")
def api_matrix_save():
    """Save edits to the side-games matrix (persisted in DB)."""
    try:
        data = request.get_json(force=True)
        changes = data.get("changes", {})
        if not changes:
            return jsonify({"error": "No changes provided"}), 400

        matrix9, matrix18 = _load_matrix()

        for change_key, new_val in changes.items():
            parts = change_key.split(":", 2)
            if len(parts) != 3:
                continue
            holes, pc, field_key = parts
            matrix = matrix9 if holes == "9" else matrix18
            if pc not in matrix:
                continue
            entry = matrix[pc]

            if field_key.startswith("skins."):
                idx = int(field_key.split(".")[1])
                if "skins" not in entry:
                    entry["skins"] = []
                while len(entry["skins"]) <= idx:
                    entry["skins"].append(None)
                entry["skins"][idx] = new_val
            else:
                entry[field_key] = new_val

        # Recalculate skins values for any entry where skinsTotal or skinsFlights changed
        for change_key in changes:
            parts = change_key.split(":", 2)
            if len(parts) != 3:
                continue
            holes, pc, field_key = parts
            if field_key in ("skinsTotal", "skinsFlights"):
                m = matrix9 if holes == "9" else matrix18
                entry = m.get(pc)
                if entry:
                    st = entry.get("skinsTotal") or 0
                    sf = entry.get("skinsFlights") or 0
                    if st and sf:
                        if "skins" not in entry:
                            entry["skins"] = [None] * 9
                        while len(entry["skins"]) < 9:
                            entry["skins"].append(None)
                        for i in range(9):
                            entry["skins"][i] = round(st / sf / (i + 1), 2)

        # Persist to database (survives Railway redeploys)
        set_app_setting("games_matrix_9", json.dumps(matrix9))
        set_app_setting("games_matrix_18", json.dumps(matrix18))

        # Also update the static file as a cache (best-effort, may be read-only)
        try:
            matrix_path = os.path.join(
                os.path.dirname(__file__), "static", "js", "games-matrix.js"
            )
            new_content = "// Auto-generated from 25-SideGame-PrizeMatrix.xlsx\n"
            new_content += "// Last edited via Matrix UI\n\n"
            new_content += "window.GAMES_MATRIX_9 = "
            new_content += json.dumps(matrix9, indent=2)
            new_content += ";\n\n"
            new_content += "window.GAMES_MATRIX_18 = "
            new_content += json.dumps(matrix18, indent=2)
            new_content += ";\n"
            with open(matrix_path, "w") as f:
                f.write(new_content)
        except Exception:
            logger.debug("Could not update static matrix file (non-fatal)")

        return jsonify({"status": "ok", "matrix9": matrix9, "matrix18": matrix18})
    except Exception as e:
        logger.exception("Matrix save failed")
        return jsonify({"error": str(e)}), 500


@app.route("/api/audit/autofix-side-games", methods=["POST"])
@require_role("admin")
def api_autofix_side_games():
    """Fix side_games misplacement in existing DB rows."""
    try:
        result = autofix_side_games()
        return jsonify({"status": "ok", **result})
    except Exception as e:
        logger.exception("Autofix failed")
        return jsonify({"error": str(e)}), 500


@app.route("/api/audit/autofix-all", methods=["POST"])
@require_role("admin")
def api_autofix_all():
    """Run all autofixes: side_games, customer names, course names, tee choices."""
    try:
        result = autofix_all()
        tee_fixed = normalize_tee_choices()
        result["tee_choices_fixed"] = tee_fixed
        return jsonify({"status": "ok", **result})
    except Exception as e:
        logger.exception("Autofix-all failed")
        return jsonify({"error": str(e)}), 500


@app.route("/api/audit/undo-autofix", methods=["POST"])
@require_role("admin")
def api_undo_autofix():
    """Revert a previous autofix using the saved details."""
    try:
        data = request.get_json(force=True)
        details = data.get("details", [])
        if not details:
            return jsonify({"error": "No details provided"}), 400
        result = undo_autofix(details)
        return jsonify({"status": "ok", **result})
    except Exception as e:
        logger.exception("Undo autofix failed")
        return jsonify({"error": str(e)}), 500


@app.route("/api/audit/autofix-tee-choices", methods=["POST"])
@require_role("admin")
def api_autofix_tee_choices():
    """Normalize all tee_choice values to standard: <50, 50-64, 65+, Forward."""
    try:
        updated = normalize_tee_choices()
        return jsonify({"status": "ok", "tee_choices_fixed": updated})
    except Exception as e:
        logger.exception("Autofix tee choices failed")
        return jsonify({"error": str(e)}), 500


@app.route("/api/audit/re-extract-fields", methods=["POST"])
@require_role("admin")
def api_re_extract_fields():
    """Re-parse existing transaction emails to backfill new fields.

    Fetches original emails from Graph API, re-runs AI extraction,
    and updates backfill fields (partner_request, fellowship, notes, holes,
    address, transaction_fees) on existing items. Also overwrites item_name
    if the AI returns an improved value.
    """
    tenant_id = os.getenv("AZURE_TENANT_ID")
    client_id = os.getenv("AZURE_CLIENT_ID")
    client_secret = os.getenv("AZURE_CLIENT_SECRET")
    email_address = os.getenv("EMAIL_ADDRESS")

    if not all([tenant_id, client_id, client_secret, email_address]):
        return jsonify({"error": "Azure AD credentials not configured"}), 400

    BACKFILL_FIELDS = ["partner_request", "fellowship", "notes", "holes",
                       "address", "address2", "city", "state", "zip",
                       "transaction_fees", "coupon_code", "coupon_amount",
                       "guest_name"]
    # Fields where re-extract should overwrite existing (possibly wrong) values
    OVERWRITE_FIELDS = {"item_name"}

    items = get_all_items()
    # Find items missing any backfill fields; since transaction_fees is new,
    # this will pick up virtually all existing items, and the OVERWRITE_FIELDS
    # logic below will also correct item_name where the AI now returns better data.
    candidates = [
        it for it in items
        if it.get("transaction_status") in (None, "active")
        and not it.get("email_uid", "").startswith("manual-")
        and not all(it.get(f) for f in BACKFILL_FIELDS)
    ]

    total = len(candidates)
    updated = 0
    skipped = 0
    errors = 0

    # Group by email_uid to avoid re-fetching the same email multiple times
    uid_groups = {}
    for it in candidates:
        uid = it.get("email_uid", "")
        if uid:
            uid_groups.setdefault(uid, []).append(it)

    for uid, group_items in uid_groups.items():
        try:
            email_data = fetch_email_by_id(
                tenant_id, client_id, client_secret, email_address, uid
            )
            if not email_data:
                skipped += len(group_items)
                continue

            parsed_rows = parse_email(email_data)
            if not parsed_rows:
                skipped += len(group_items)
                continue

            for it in group_items:
                idx = it.get("item_index", 0) or 0
                if idx < len(parsed_rows):
                    parsed = parsed_rows[idx]
                else:
                    # Try to find by matching item name
                    parsed = next(
                        (p for p in parsed_rows
                         if p.get("item_name") == it.get("item_name")),
                        parsed_rows[0] if len(parsed_rows) == 1 else None,
                    )

                if not parsed:
                    skipped += 1
                    continue

                changes = {}
                for field in BACKFILL_FIELDS:
                    new_val = parsed.get(field)
                    if new_val and not it.get(field):
                        changes[field] = new_val
                # Overwrite fields — update even if existing value is present
                for field in OVERWRITE_FIELDS:
                    new_val = parsed.get(field)
                    if new_val and new_val != it.get(field):
                        changes[field] = new_val

                # Guest-swap: if parser promoted the guest to customer,
                # overwrite customer + guest_name on the existing item
                parsed_customer = (parsed.get("customer") or "").strip()
                current_customer = (it.get("customer") or "").strip()
                parsed_notes = parsed.get("notes") or ""
                if (parsed_customer and parsed_customer.lower() != current_customer.lower()
                        and "Purchased by" in parsed_notes):
                    changes["customer"] = parsed_customer
                    changes["guest_name"] = parsed.get("guest_name") or parsed_customer
                    changes["notes"] = parsed_notes
                    changes["customer_email"] = None
                    changes["customer_phone"] = None
                    changes["customer_id"] = None

                if changes:
                    update_item(it["id"], changes)
                    updated += 1
                else:
                    skipped += 1

        except Exception:
            logger.exception("Re-extract failed for email_uid=%s", uid)
            errors += len(group_items)

    return jsonify({
        "status": "ok",
        "total_candidates": total,
        "updated": updated,
        "skipped": skipped,
        "errors": errors,
    })


@app.route("/api/audit/reextract-order", methods=["POST"])
@require_role("admin")
def api_reextract_order():
    """Re-parse a single order's email to backfill coupon (or any missing) fields.

    Accepts JSON body: {"order_id": "R854482675"}
    Re-fetches the original email, re-runs AI extraction, and updates
    coupon_code and coupon_amount (plus other backfill fields) on all rows
    sharing that order_id.
    """
    data = request.get_json(force=True) or {}
    order_id = data.get("order_id", "").strip()
    if not order_id:
        return jsonify({"error": "order_id is required"}), 400

    tenant_id = os.getenv("AZURE_TENANT_ID")
    client_id = os.getenv("AZURE_CLIENT_ID")
    client_secret = os.getenv("AZURE_CLIENT_SECRET")
    email_address = os.getenv("EMAIL_ADDRESS")
    if not all([tenant_id, client_id, client_secret, email_address]):
        return jsonify({"error": "Azure AD credentials not configured"}), 400

    BACKFILL_FIELDS = ["coupon_code", "coupon_amount", "transaction_fees",
                       "partner_request", "fellowship", "notes", "holes",
                       "address", "address2", "city", "state", "zip"]
    # Fields where re-extract should overwrite even if existing value differs
    FORCE_UPDATE_FIELDS = ["side_games", "holes", "item_price", "fall_net_points_race"]

    items = get_all_items()
    order_items = [it for it in items if it.get("order_id") == order_id]
    if not order_items:
        return jsonify({"error": f"No items found for order_id={order_id}"}), 404

    # All rows in an order share the same email_uid
    uid = order_items[0].get("email_uid", "")
    if not uid or uid.startswith("manual-"):
        return jsonify({"error": "Order has no parseable email (manual entry)"}), 400

    try:
        email_data = fetch_email_by_id(
            tenant_id, client_id, client_secret, email_address, uid
        )
        if not email_data:
            return jsonify({"error": f"Could not fetch email {uid} from Graph API"}), 404

        parsed_rows = parse_email(email_data)
        if not parsed_rows:
            return jsonify({"error": "AI extraction returned no results"}), 500

        updated = 0
        changes_detail = []
        for it in order_items:
            # Skip manual entries (Add Payment, comp, etc.)
            uid = it.get("email_uid") or ""
            if uid.startswith("manual-"):
                continue

            idx = it.get("item_index", 0) or 0
            if idx < len(parsed_rows):
                parsed = parsed_rows[idx]
            else:
                # Fallback: match by item_name (case-insensitive)
                parsed = next(
                    (p for p in parsed_rows
                     if (p.get("item_name") or "").lower() == (it.get("item_name") or "").lower()),
                    parsed_rows[0] if len(parsed_rows) == 1 else None,
                )
            if not parsed:
                continue

            changes = {}
            for field in BACKFILL_FIELDS:
                new_val = parsed.get(field)
                if new_val and not it.get(field):
                    changes[field] = new_val

            # Force-update fields: overwrite if parsed value differs
            for field in FORCE_UPDATE_FIELDS:
                new_val = parsed.get(field)
                if new_val and str(new_val).strip().upper() != str(it.get(field) or "").strip().upper():
                    changes[field] = new_val

            # Guest-swap: if parser promoted the guest to customer,
            # overwrite customer on the existing item (not guest_name — that
            # field belongs to whoever the player requested as a partner)
            parsed_customer = (parsed.get("customer") or "").strip()
            current_customer = (it.get("customer") or "").strip()
            parsed_notes = parsed.get("notes") or ""
            if (parsed_customer and parsed_customer.lower() != current_customer.lower()
                    and "Purchased by" in parsed_notes):
                changes["customer"] = parsed_customer
                changes["notes"] = parsed_notes
                changes["customer_email"] = None
                changes["customer_phone"] = None
                changes["customer_id"] = None

            if changes:
                update_item(it["id"], changes)
                updated += 1
                changes_detail.append({"id": it["id"], "fields": list(changes.keys())})

        return jsonify({
            "status": "ok",
            "order_id": order_id,
            "items_in_order": len(order_items),
            "items_updated": updated,
            "changes": changes_detail,
            "parsed_count": len(parsed_rows),
        })

    except Exception as exc:
        logger.exception("reextract-order failed for %s", order_id)
        return jsonify({"error": str(exc)}), 500


@app.route("/api/audit/reimport-order", methods=["POST"])
@require_role("admin")
def api_reimport_order():
    """Re-fetch and re-parse an email by uid, INSERTing the resulting items.

    Use case: an order's items were deleted (e.g. to clean up a parser
    mis-extraction) and need to be re-imported from scratch. The standard
    re-extract endpoint only UPDATEs existing rows, so when the items have
    been removed it has nothing to operate on. This endpoint bypasses that:
    fetch the email, run the parser, and call save_items().

    The save_items() pipeline already has the cross-email-uid dedup gate
    plus the UNIQUE(email_uid, item_index) constraint, so no duplicates are
    created — if rows already exist for this order they will be skipped.

    Body: {"email_uid": "..."}
    """
    data = request.get_json(force=True) or {}
    email_uid = (data.get("email_uid") or "").strip()
    if not email_uid:
        return jsonify({"error": "email_uid is required"}), 400
    if email_uid.startswith("manual-"):
        return jsonify({"error": "Manual entries cannot be re-imported"}), 400

    tenant_id = os.getenv("AZURE_TENANT_ID")
    client_id = os.getenv("AZURE_CLIENT_ID")
    client_secret = os.getenv("AZURE_CLIENT_SECRET")
    email_address = os.getenv("EMAIL_ADDRESS")
    if not all([tenant_id, client_id, client_secret, email_address]):
        return jsonify({"error": "Azure AD credentials not configured"}), 400

    try:
        email_data = fetch_email_by_id(
            tenant_id, client_id, client_secret, email_address, email_uid
        )
        if not email_data:
            return jsonify({"error": f"Could not fetch email {email_uid} from Graph API"}), 404

        parsed_rows = parse_email(email_data)
        if not parsed_rows:
            return jsonify({
                "status": "ok",
                "email_uid": email_uid,
                "parsed_count": 0,
                "inserted": 0,
                "message": "Parser returned 0 items — email body may be malformed.",
            })

        _dup_alerts_ri: list = []
        inserted = save_items(parsed_rows, _alerts_out=_dup_alerts_ri)
        for _al in _dup_alerts_ri:
            _send_dup_reg_alert(_al)
        return jsonify({
            "status": "ok",
            "email_uid": email_uid,
            "parsed_count": len(parsed_rows),
            "inserted": inserted,
            "skipped": len(parsed_rows) - inserted,
        })
    except Exception as exc:
        logger.exception("reimport-order failed for email_uid=%s", email_uid)
        return jsonify({"error": str(exc)}), 500


@app.route("/api/audit/retry-failed", methods=["POST"])
@require_role("admin")
def api_retry_failed():
    """Re-process emails that previously parsed 0 items.

    Clears 0-item entries from processed_emails, then re-runs check_inbox
    which will pick them up as 'new' emails and re-parse them.
    """
    cleared = clear_failed_processed()
    if cleared == 0:
        return jsonify({"status": "ok", "cleared": 0, "message": "No failed emails to retry"})

    # Now re-run inbox check to pick up the cleared emails
    try:
        check_inbox()
    except Exception:
        logger.exception("Retry failed: inbox check error")
        return jsonify({"status": "partial", "cleared": cleared,
                        "message": f"Cleared {cleared} entries but inbox check failed"}), 500

    return jsonify({
        "status": "ok",
        "cleared": cleared,
        "message": f"Cleared {cleared} failed entries and re-processed inbox",
    })


@app.route("/api/audit/delete-phantom-duplicates", methods=["POST"])
@require_role("admin")
def api_delete_phantom_duplicates():
    """Delete phantom duplicate item rows.

    For each group of items sharing (order_id, customer, item_name, item_price)
    with COUNT > 1, keep the lowest-id row (the original parse — earliest
    created_at) and DELETE every later row. Skips any row that has downstream
    references that deletion would orphan:

    - acct_allocations.item_id pointing at the row
    - acct_transactions.item_id pointing at the row
    - items.transferred_from_id / items.transferred_to_id pointing at the row
    - items.parent_item_id pointing at the row (child +PAY rows)

    Idempotent: after one run no groups have COUNT > 1, so re-running is a
    no-op.

    Query params:
      since=YYYY-MM-DD   restrict to orders with order_date >= this date
                         (default: 2026-04-26 to scope this to the May-3
                         backfill incident; pass since=1900-01-01 for ALL)
      dry_run=1          report what would be deleted without deleting
    """
    since = (request.args.get("since") or "2026-04-26").strip()
    dry_run = request.args.get("dry_run") in ("1", "true", "yes")
    conn = get_connection()
    try:
        groups = conn.execute(
            """
            SELECT order_id, customer, item_name, item_price,
                   GROUP_CONCAT(id ORDER BY id) as ids
            FROM items
            WHERE COALESCE(transaction_status, 'active') = 'active'
              AND order_id IS NOT NULL AND order_id != ''
              AND customer IS NOT NULL AND customer != ''
              AND item_name IS NOT NULL AND item_name != ''
              AND order_date >= ?
            GROUP BY order_id, LOWER(customer), LOWER(item_name), item_price
            HAVING COUNT(*) > 1
            """,
            (since,),
        ).fetchall()

        deleted = 0
        skipped = 0
        skip_reasons: list[str] = []
        deleted_ids: list[int] = []

        for g in groups:
            ids = [int(x) for x in (g["ids"] or "").split(",") if x]
            if len(ids) < 2:
                continue
            # ids is sorted ASC by GROUP_CONCAT (... ORDER BY id);
            # the first is the original — keep it.
            for dup_id in ids[1:]:
                # Acct allocations
                alloc_count = 0
                try:
                    alloc_count = conn.execute(
                        "SELECT COUNT(*) as cnt FROM acct_allocations WHERE item_id = ?",
                        (dup_id,),
                    ).fetchone()["cnt"]
                except Exception:
                    pass
                if alloc_count > 0:
                    skipped += 1
                    skip_reasons.append(f"#{dup_id}: {alloc_count} acct_allocations")
                    continue

                # Acct transactions
                txn_count = 0
                try:
                    txn_count = conn.execute(
                        "SELECT COUNT(*) as cnt FROM acct_transactions WHERE item_id = ?",
                        (dup_id,),
                    ).fetchone()["cnt"]
                except Exception:
                    pass
                if txn_count > 0:
                    skipped += 1
                    skip_reasons.append(f"#{dup_id}: {txn_count} acct_transactions")
                    continue

                # Other items referencing this id (transfer chain or +PAY children)
                ref_count = conn.execute(
                    """SELECT COUNT(*) as cnt FROM items
                       WHERE transferred_from_id = ?
                          OR transferred_to_id = ?
                          OR parent_item_id = ?""",
                    (dup_id, dup_id, dup_id),
                ).fetchone()["cnt"]
                if ref_count > 0:
                    skipped += 1
                    skip_reasons.append(f"#{dup_id}: {ref_count} item back-refs")
                    continue

                if dry_run:
                    deleted += 1
                    deleted_ids.append(dup_id)
                    continue

                try:
                    conn.execute("DELETE FROM items WHERE id = ?", (dup_id,))
                    deleted += 1
                    deleted_ids.append(dup_id)
                except Exception as e:
                    logger.warning("delete-phantom-duplicates: failed on id=%s: %s", dup_id, e)
                    skipped += 1
                    skip_reasons.append(f"#{dup_id}: delete error: {e}")

        if not dry_run:
            conn.commit()

        return jsonify({
            "status": "ok",
            "since": since,
            "dry_run": dry_run,
            "groups_processed": len(groups),
            "deleted": deleted,
            "deleted_ids": deleted_ids[:200],
            "skipped": skipped,
            "skip_reasons": skip_reasons[:50],
        })
    finally:
        conn.close()


@app.route("/api/audit/membership-mashup-scan", methods=["GET"])
@require_role("admin")
def api_membership_mashup_scan():
    """Scan for TGF MEMBERSHIP rows that look like victims of the
    membership+event mash-up parser bug.

    Background: Haiku (the parser model) sometimes returns a single row
    for an order that actually contained both a TGF MEMBERSHIP and an
    event line item — using the membership's name with the event's
    price / holes / side_games / tee_choice. The membership row should
    have NULL for those event-side fields. Any membership row with
    non-null holes, side_games, or tee_choice is suspect.

    Read-only. Returns one entry per suspect row with the fields needed
    to decide whether to delete + re-import that order.
    """
    conn = get_connection()
    try:
        rows = conn.execute(
            """
            SELECT i.id, i.order_id, i.email_uid, i.customer, i.item_name,
                   i.item_price, i.holes, i.side_games, i.tee_choice,
                   i.user_status, i.order_date, i.total_amount,
                   i.transaction_fees, i.transaction_status, i.created_at
            FROM items i
            WHERE i.item_name = 'TGF MEMBERSHIP'
              AND COALESCE(i.transaction_status, 'active') = 'active'
              AND (
                   (i.holes IS NOT NULL AND i.holes != '')
                OR (i.side_games IS NOT NULL AND i.side_games != ''
                    AND UPPER(i.side_games) != 'NONE')
                OR (i.tee_choice IS NOT NULL AND i.tee_choice != '')
              )
            ORDER BY i.order_date DESC, i.id DESC
            """
        ).fetchall()

        suspects = [dict(r) for r in rows]
        return jsonify({
            "status": "ok",
            "count": len(suspects),
            "rows": suspects,
        })
    finally:
        conn.close()


@app.route("/api/audit/credited-duplicate-scan", methods=["GET"])
@require_role("admin")
def api_credited_duplicate_scan():
    """Find orders where the same item appears as both active AND credited/refunded.

    These are likely cases where a credit was issued on one copy but a phantom
    duplicate active copy was left behind (like the Chalfant Canyon Springs case).
    The active copy should be deleted.

    Returns list of problem groups with item ids so the admin can delete via UI.
    """
    conn = get_connection()
    try:
        rows = conn.execute("""
            SELECT
                order_id,
                LOWER(item_name) AS item_key,
                item_name,
                item_price,
                customer,
                GROUP_CONCAT(id || ':' || COALESCE(transaction_status,'active') ORDER BY id) AS id_statuses,
                SUM(CASE WHEN COALESCE(transaction_status,'active') = 'active' THEN 1 ELSE 0 END) AS active_cnt,
                SUM(CASE WHEN transaction_status IN ('credited','refunded') THEN 1 ELSE 0 END) AS credited_cnt,
                GROUP_CONCAT(CASE WHEN COALESCE(transaction_status,'active') = 'active' THEN id END) AS active_ids,
                GROUP_CONCAT(CASE WHEN transaction_status IN ('credited','refunded') THEN id END) AS credited_ids
            FROM items
            WHERE order_id IS NOT NULL AND order_id NOT LIKE 'manual-%'
              AND item_name IS NOT NULL AND item_name != ''
            GROUP BY order_id, LOWER(item_name), item_price
            HAVING active_cnt > 0 AND credited_cnt > 0
            ORDER BY order_id
        """).fetchall()

        results = []
        for r in rows:
            active_ids = [int(x) for x in (r["active_ids"] or "").split(",") if x]
            credited_ids = [int(x) for x in (r["credited_ids"] or "").split(",") if x]
            results.append({
                "order_id": r["order_id"],
                "item_name": r["item_name"],
                "item_price": r["item_price"],
                "customer": r["customer"],
                "id_statuses": r["id_statuses"],
                "active_ids": active_ids,
                "credited_ids": credited_ids,
                "note": "Delete the active_ids — they are phantom duplicates of the credited copy.",
            })

        return jsonify({"count": len(results), "problems": results})
    finally:
        conn.close()


@app.route("/api/audit/duplicate-items-diagnostic", methods=["GET"])
@require_role("admin")
def api_duplicate_items_diagnostic():
    """Read-only diagnostic for duplicate item rows.

    Groups items by (order_id, customer, item_name, item_price) — order_id
    instead of email_uid so we catch both same-email-uid duplicates AND
    same-order-different-email-uid duplicates. Returns per-row metadata to
    classify the root cause:

    - Same email_uid, different item_index, same created_at → AI returned
      multiple items in one parse (hallucination at extraction time).
    - Same email_uid, different item_index, different created_at → some
      later script appended rows under the same email_uid (e.g. re-extract).
    - Different email_uid, same order_id → same order ingested via two
      different emails or two parser runs with different uids.
    - Manual entries → manual-entry duplicates.

    Query params:
      since=YYYY-MM-DD   filter to orders with order_date >= this date
                         (default: 2026-04-26 to focus on the May-3 backfill
                         incident; pass since=1900-01-01 to see ALL dupes)

    Skips rows where order_id is NULL (manual + RSVP-only rows often lack one).
    Read-only — does not mutate any rows.
    """
    since = (request.args.get("since") or "2026-04-26").strip()
    conn = get_connection()
    try:
        groups = conn.execute(
            """
            SELECT order_id, customer, item_name, item_price,
                   COUNT(*) as cnt,
                   GROUP_CONCAT(id ORDER BY id) as ids
            FROM items
            WHERE COALESCE(transaction_status, 'active') = 'active'
              AND customer IS NOT NULL AND customer != ''
              AND item_name IS NOT NULL AND item_name != ''
              AND order_id IS NOT NULL AND order_id != ''
              AND order_date >= ?
            GROUP BY order_id, LOWER(customer), LOWER(item_name), item_price
            HAVING COUNT(*) > 1
            ORDER BY MIN(id) DESC
            """,
            (since,),
        ).fetchall()

        from datetime import datetime
        result_groups = []
        same_uid_diff_idx_same_created = 0
        same_uid_diff_idx_diff_created = 0
        different_uid = 0
        manual_uid = 0
        max_created_gap_seconds = 0.0

        for g in groups:
            ids = [int(x) for x in (g["ids"] or "").split(",") if x]
            rows = conn.execute(
                f"""
                SELECT id, email_uid, item_index, order_id, order_date,
                       created_at, transaction_status, customer_id, merchant
                FROM items
                WHERE id IN ({",".join("?" * len(ids))})
                ORDER BY id
                """,
                tuple(ids),
            ).fetchall()
            row_dicts = [dict(r) for r in rows]

            uids = {r["email_uid"] for r in row_dicts}
            indexes = sorted({r["item_index"] for r in row_dicts})
            is_manual = any((r["email_uid"] or "").startswith("manual-") for r in row_dicts)

            created_times = []
            for r in row_dicts:
                try:
                    created_times.append(datetime.fromisoformat((r["created_at"] or "").replace("Z", "+00:00")))
                except Exception:
                    pass
            gap_seconds = 0.0
            if len(created_times) >= 2:
                gap_seconds = (max(created_times) - min(created_times)).total_seconds()
                if gap_seconds > max_created_gap_seconds:
                    max_created_gap_seconds = gap_seconds

            if is_manual:
                pattern = "manual"
                manual_uid += 1
            elif len(uids) > 1:
                pattern = "different_uid_same_order"
                different_uid += 1
            elif gap_seconds < 60:
                pattern = "same_uid_diff_idx_same_created"
                same_uid_diff_idx_same_created += 1
            else:
                pattern = "same_uid_diff_idx_diff_created"
                same_uid_diff_idx_diff_created += 1

            result_groups.append({
                "customer": g["customer"],
                "item_name": g["item_name"],
                "item_price": g["item_price"],
                "order_id": g["order_id"],
                "count": g["cnt"],
                "pattern": pattern,
                "distinct_email_uids": len(uids),
                "item_indexes": indexes,
                "created_at_gap_seconds": round(gap_seconds, 2),
                "rows": row_dicts,
            })

        return jsonify({
            "status": "ok",
            "since": since,
            "total_duplicate_groups": len(result_groups),
            "total_extra_rows": sum(g["count"] - 1 for g in result_groups),
            "patterns": {
                "same_uid_diff_idx_same_created": same_uid_diff_idx_same_created,
                "same_uid_diff_idx_diff_created": same_uid_diff_idx_diff_created,
                "different_uid_same_order": different_uid,
                "manual": manual_uid,
            },
            "max_created_at_gap_seconds": round(max_created_gap_seconds, 2),
            "groups": result_groups,
        })
    finally:
        conn.close()


@app.route("/api/audit/expand-quantities", methods=["POST"])
@require_role("admin")
def api_expand_quantities():
    """Find items with quantity > 1 and create missing partner rows.

    For each item with qty > 1, checks if partner rows already exist
    (same email_uid, consecutive item_index). If not, creates them using
    the partner_request name or 'Guest of <buyer>'.

    This is a one-time backfill for orders placed before quantity expansion
    was added to the parser.
    """
    from email_parser.parser import _normalize_customer_name
    from email_parser.database import _resolve_or_create_customer

    conn = get_connection()
    try:
        # Find items with quantity > 1
        qty_items = conn.execute(
            "SELECT * FROM items WHERE quantity > 1 ORDER BY id"
        ).fetchall()

        created = 0
        skipped = 0
        details = []

        for item in qty_items:
            item = dict(item)
            qty = item["quantity"]
            email_uid = item["email_uid"]
            base_index = item["item_index"]
            buyer = item["customer"] or "Unknown"

            # Check how many rows already exist for this email_uid
            existing = conn.execute(
                "SELECT item_index FROM items WHERE email_uid = ? ORDER BY item_index",
                (email_uid,),
            ).fetchall()
            existing_indices = {r["item_index"] for r in existing}

            # Find the next available item_index
            max_idx = max(existing_indices) if existing_indices else -1

            partner_name = (item.get("partner_request") or "").strip()

            for extra_i in range(1, qty):
                new_idx = max_idx + extra_i
                if new_idx in existing_indices:
                    skipped += 1
                    continue

                # Build the partner row from the original
                partner_row = dict(item)
                partner_row["item_index"] = new_idx
                partner_row["quantity"] = 1
                partner_row["customer_email"] = None
                partner_row["customer_phone"] = None
                partner_row["address"] = None
                partner_row["address2"] = None
                partner_row["city"] = None
                partner_row["state"] = None
                partner_row["zip"] = None
                # Remove DB-generated fields
                partner_row.pop("id", None)
                partner_row.pop("created_at", None)

                # partner_row started as a copy of the buyer's row, so it
                # carries the BUYER's customer_id — the partner row must get
                # the partner's own identity (or none for an unnamed guest),
                # not inherit the buyer's.
                if extra_i == 1 and partner_name:
                    partner_row["customer"] = _normalize_customer_name(partner_name)
                    partner_row["partner_request"] = None
                    partner_row["notes"] = f"Purchased by {buyer}"
                    partner_row["customer_id"] = _resolve_or_create_customer(
                        conn, customer_name=partner_row["customer"],
                        customer_email=None,
                    )
                else:
                    partner_row["customer"] = f"Guest of {buyer}"
                    partner_row["notes"] = f"Purchased by {buyer}"
                    partner_row["customer_id"] = None

                # Insert the partner row
                cols = [c for c in partner_row.keys() if c not in ("id", "created_at")]
                placeholders = ", ".join("?" for _ in cols)
                col_names = ", ".join(cols)
                values = tuple(partner_row.get(c) for c in cols)

                try:
                    conn.execute(
                        f"INSERT OR IGNORE INTO items ({col_names}) VALUES ({placeholders})",
                        values,
                    )
                    created += 1
                    details.append(f"{buyer} → {partner_row['customer']} ({item['item_name']})")
                except Exception as e:
                    logger.warning("Failed to create partner row: %s", e)
                    skipped += 1

            # Update original row quantity to 1
            conn.execute(
                "UPDATE items SET quantity = 1 WHERE id = ?", (item["id"],)
            )

        conn.commit()
    finally:
        conn.close()

    return jsonify({
        "status": "ok",
        "found_qty_items": len(qty_items),
        "created": created,
        "skipped": skipped,
        "details": details,
    })


@app.route("/api/audit/fix-guest-customers", methods=["POST"])
@require_role("admin")
def api_fix_guest_customers():
    """Fix GUEST items where the customer is still the buyer instead of the guest.

    Finds items with user_status containing 'GUEST' and a non-empty guest_name
    that differs from the current customer. Swaps the customer to the guest_name
    and adds a 'Purchased by <buyer>' note.
    """
    from email_parser.parser import _normalize_customer_name

    conn = get_connection()
    try:
        guests = conn.execute(
            """SELECT id, customer, guest_name, notes
               FROM items
               WHERE user_status LIKE '%GUEST%'
                 AND guest_name IS NOT NULL AND guest_name != ''
                 AND COALESCE(transaction_status, 'active') = 'active'"""
        ).fetchall()

        fixed = 0
        details = []

        for row in guests:
            row = dict(row)
            guest = _normalize_customer_name(row["guest_name"])
            buyer = (row["customer"] or "").strip()
            if not guest or guest.lower() == buyer.lower():
                continue

            conn.execute(
                """UPDATE items
                   SET customer = ?, notes = ?,
                       customer_email = NULL, customer_phone = NULL,
                       address = NULL, address2 = NULL,
                       city = NULL, state = NULL, zip = NULL,
                       customer_id = NULL
                   WHERE id = ?""",
                (guest, f"Purchased by {buyer}", row["id"]),
            )
            fixed += 1
            details.append(f"{buyer} → {guest} (id={row['id']})")

        conn.commit()
    finally:
        conn.close()

    return jsonify({
        "status": "ok",
        "fixed": fixed,
        "details": details,
    })


# ---------------------------------------------------------------------------
# Routes — Events
# ---------------------------------------------------------------------------
@app.route("/events")
def events_page():
    matrix9, matrix18 = _load_matrix()
    return render_template("events.html", matrix9=matrix9, matrix18=matrix18)


@app.route("/customers")
def customers_page():
    if session.get("role") == "view-only":
        return redirect("/events")
    return render_template("customers.html")


@app.route("/api/customers")
@require_role("view-only")
def api_customers_canonical():
    """Return all customer records from the canonical customers + customer_emails tables.

    This is the source-of-truth endpoint for customer identity data (email,
    phone, name).  The Customers page overlays this onto transaction-derived
    data so canonical contact details always win.
    """
    from email_parser.database import get_all_customers
    return jsonify(get_all_customers())


@app.route("/api/customers/update", methods=["POST"])
@require_role("manager")
def api_update_customer():
    """Update personal info fields across all items for a customer.

    Body: { customer_name: str, customer_id?: int,
            fields: {customer_email, customer_phone, chapter, ...} }
    Updates every item row matching this customer name. customer_id, when
    given, is used directly for the customers-table syncs (status, venmo,
    chapter, etc.) instead of re-deriving it by name — see
    update_customer_info() for why that re-derivation can miss.
    """
    data = request.get_json(force=True)
    customer_name = (data.get("customer_name") or "").strip()
    try:
        customer_id = int(data.get("customer_id")) if data.get("customer_id") else None
    except (TypeError, ValueError):
        customer_id = None
    fields = data.get("fields") or {}
    if not customer_name:
        return jsonify({"error": "customer_name is required"}), 400
    if not fields:
        return jsonify({"error": "fields object is required"}), 400

    # Only allow personal-info columns, not transaction data
    allowed = {"customer_email", "customer_phone", "chapter", "handicap",
               "date_of_birth", "shirt_size", "customer",
               "first_name", "last_name", "middle_name", "suffix",
               "address", "address2", "city", "state", "zip",
               "archived", "venmo_username", "current_player_status",
               "payment_method", "payment_handle"}
    safe = {k: v for k, v in fields.items() if k in allowed}
    if not safe:
        return jsonify({"error": "No valid fields to update"}), 400

    try:
        updated = update_customer_info(customer_name, safe, customer_id=customer_id)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    return jsonify({"status": "ok", "items_updated": updated})


@app.route("/api/customers/venmo-handles")
@require_role("view-only")
def api_customer_venmo_handles():
    """Return all customers with Venmo handles set."""
    from email_parser.database import get_customer_venmo_handles
    return jsonify(get_customer_venmo_handles())


# ---------------------------------------------------------------------------
# Membership renewal endpoints
# ---------------------------------------------------------------------------

@app.route("/api/customers/<int:customer_id>/memberships")
@require_role("view-only")
def api_get_memberships(customer_id):
    """Return the membership term history for a customer (newest first)."""
    from email_parser.memberships import get_memberships_for_customer
    return jsonify(get_memberships_for_customer(customer_id))


@app.route("/api/memberships/current")
@require_role("view-only")
def api_current_memberships():
    """Return {customer_id: {expires_at, started_at, ...}} for the latest term per customer.

    Used by the Customers list page to render the Renewal column in one fetch
    instead of N round-trips.
    """
    from email_parser.memberships import get_current_term_map
    return jsonify(get_current_term_map())


@app.route("/api/customers/<int:customer_id>/memberships", methods=["POST"])
@require_role("admin")
def api_add_membership(customer_id):
    """Add a manual membership term (e.g. backfill a legacy renewal)."""
    from email_parser.memberships import add_manual_term
    data = request.get_json(force=True) or {}
    started_at = (data.get("started_at") or "").strip()
    expires_at = (data.get("expires_at") or "").strip() or None
    notes = data.get("notes")
    if not started_at:
        return jsonify({"error": "started_at is required (YYYY-MM-DD)"}), 400
    try:
        result = add_manual_term(customer_id, started_at, expires_at, notes)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    return jsonify({"status": "ok", "term": result})


@app.route("/api/memberships/<int:term_id>", methods=["PATCH"])
@require_role("admin")
def api_update_membership(term_id):
    """Update started_at / expires_at / notes on a membership term."""
    from email_parser.memberships import update_term
    data = request.get_json(force=True) or {}
    updated = update_term(term_id, data)
    return jsonify({"status": "ok", "updated": updated})


@app.route("/api/memberships/<int:term_id>", methods=["DELETE"])
@require_role("admin")
def api_delete_membership(term_id):
    """Delete a membership term (admin cleanup of mistaken manual entries)."""
    from email_parser.memberships import delete_term
    deleted = delete_term(term_id)
    return jsonify({"status": "ok", "deleted": deleted})


@app.route("/api/admin/run-membership-reminders", methods=["POST"])
@require_role("admin")
def api_run_membership_reminders():
    """Manually trigger the daily membership reminders job. Returns counts."""
    counts = run_membership_reminders()
    if counts is None:
        return jsonify({"error": "run failed — check server logs"}), 500
    return jsonify({"status": "ok", "counts": counts})


@app.route("/api/memberships/<int:term_id>/preview-notice")
@require_role("admin")
def api_preview_membership_notice(term_id):
    """Render a notice email for a specific term without sending it.

    Query: ?window=30d|7d|dayof|lapsed|confirmation
           ?with_roster_buttons=0|1 (optional — overrides the per-window default;
              omit to use the default: lapsed includes buttons, others don't)
    Returns: {to, subject, html, term, customer, can_send, reason}
    """
    window = (request.args.get("window") or "").strip()
    raw_roster = request.args.get("with_roster_buttons")
    if raw_roster is None or raw_roster == "":
        with_roster = None  # apply per-window default
    else:
        with_roster = raw_roster in ("1", "true", "yes")
    from email_parser.memberships import preview_notice
    try:
        preview = preview_notice(term_id, window, with_roster_buttons=with_roster)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    return jsonify(preview)


@app.route("/api/memberships/<int:term_id>/send-notice", methods=["POST"])
@require_role("admin")
def api_send_membership_notice(term_id):
    """Send a specific notice for a term right now and stamp the matching column.

    Body: {window: "30d" | "7d" | "dayof" | "lapsed" | "confirmation",
           subject?: "optional admin-edited subject",
           with_roster_buttons?: bool — overrides per-window default;
              omit to use the default (lapsed includes buttons, others don't)}
    """
    data = request.get_json(force=True) or {}
    window = (data.get("window") or "").strip()
    subject_override = (data.get("subject") or "").strip() or None
    if "with_roster_buttons" in data:
        with_roster = bool(data.get("with_roster_buttons"))
    else:
        with_roster = None  # apply per-window default
    from email_parser.memberships import send_notice_now
    result = send_notice_now(
        term_id, window, _membership_send_email,
        subject_override=subject_override,
        with_roster_buttons=with_roster,
    )
    if not result.get("ok"):
        return jsonify(result), 400
    return jsonify(result)


@app.route("/m/roster/<token>")
def public_roster_choice(token):
    """One-click roster keep/remove from the lapsed-notice email.

    No login required — the token is HMAC-signed and expires after 30 days.
    Renders a simple HTML confirmation page and notifies admin via email.
    """
    from email_parser.memberships import apply_roster_choice
    result = apply_roster_choice(token, send_email=_membership_send_email)
    if not result.get("ok"):
        body = f"""<h2 style="color:#b91c1c;">{result.get('error', 'Invalid link')}</h2>
<p>If you'd like to update your roster preference manually, please reply to your most recent membership email or write to <a href="mailto:admin@thegolffellowship.com">admin@thegolffellowship.com</a>.</p>"""
        return _public_page("Membership link invalid", body), 400

    action = result["action"]
    if action == "keep":
        headline = "Got it — we'll keep you on the rosters."
        sub = ("You'll continue to receive Golf Genius weekly event invitations. "
               "If you'd still like to renew your membership, the link is below.")
        bg = "#16a34a"
    else:
        headline = "Got it — we'll remove you from the rosters."
        sub = ("You'll stop receiving Golf Genius weekly event invitations within "
               "the next few days. You're always welcome back — renewing your "
               "membership at the link below puts you straight back on the list.")
        bg = "#dc2626"

    body = f"""<div style="background:{bg}; color:#fff; padding:1rem 1.25rem; border-radius:8px;">
  <h2 style="margin:0 0 0.5rem; font-size:1.25rem;">{headline}</h2>
  <p style="margin:0; opacity:0.92;">{sub}</p>
</div>
<p style="margin-top:1.5rem;">A confirmation has been sent to our admin team.</p>
<p style="margin-top:1.5rem;">
  <a href="https://thegolffellowship.com/shop/ols/products/tgf-membership"
     style="display:inline-block; background:#2563eb; color:#fff; padding:0.7rem 1.4rem; border-radius:6px; text-decoration:none; font-weight:600;">
    Renew Membership — $75
  </a>
</p>"""
    return _public_page("Roster preference saved", body)


def _public_page(title: str, body_html: str) -> str:
    """Tiny standalone HTML shell for the public roster-choice confirmation page."""
    return f"""<!doctype html>
<html><head><meta charset="utf-8"><title>{title} · The Golf Fellowship</title>
<meta name="viewport" content="width=device-width, initial-scale=1">
<style>
  body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
         color:#111827; background:#f9fafb; margin:0; padding:2rem 1rem; }}
  .card {{ max-width:560px; margin:0 auto; background:#fff; padding:1.75rem;
          border-radius:10px; box-shadow:0 1px 3px rgba(0,0,0,.06); }}
  h1 {{ font-size:1.05rem; color:#6b7280; margin:0 0 1.25rem; font-weight:600;
        text-transform:uppercase; letter-spacing:0.05em; }}
</style>
</head><body><div class="card">
<h1>The Golf Fellowship</h1>
{body_html}
</div></body></html>"""


@app.route("/api/customers/sync-roles", methods=["POST"])
@require_role("admin")
def api_sync_customer_roles():
    """Replace all roles for a customer.

    Body: { customer_name: str, customer_id?: int, roles: [...] }
    customer_id, when given, is used directly instead of re-deriving it via
    an exact first+last name match against the customers table — that match
    misses for customers whose name doesn't split cleanly into two tokens
    (middle names, suffixes) or who aren't in the customers table under that
    exact spelling, silently 404ing a save the frontend already believed had
    succeeded.
    """
    from email_parser.database import _connect
    data = request.get_json(force=True)
    customer_name = (data.get("customer_name") or "").strip()
    try:
        customer_id = int(data.get("customer_id")) if data.get("customer_id") else None
    except (TypeError, ValueError):
        customer_id = None
    roles = data.get("roles", [])
    if not customer_name:
        return jsonify({"error": "customer_name required"}), 400
    valid_roles = {"golfer", "manager", "admin", "owner", "course_contact", "sponsor", "vendor"}
    roles = [r for r in roles if r in valid_roles]
    with _connect() as conn:
        cid = None
        if customer_id:
            row = conn.execute(
                "SELECT customer_id FROM customers WHERE customer_id = ?", (customer_id,)
            ).fetchone()
            cid = row["customer_id"] if row else None
        if not cid:
            row = conn.execute(
                "SELECT customer_id FROM customers WHERE LOWER(first_name || ' ' || last_name) = LOWER(?)",
                (customer_name,)
            ).fetchone()
            if not row:
                return jsonify({"error": "Customer not found"}), 404
            cid = row["customer_id"]
        conn.execute("DELETE FROM customer_roles WHERE customer_id = ?", (cid,))
        for r in roles:
            conn.execute("INSERT OR IGNORE INTO customer_roles (customer_id, role_type) VALUES (?,?)", (cid, r))
        conn.commit()
    return jsonify({"status": "ok", "roles": roles})


@app.route("/api/customer-roles")
@require_role("view-only")
def api_customer_roles():
    """Return a map of customer_id → {roles: [...], first_timer_ever: bool}.

    Used by frontend pages (Customers, Events, Transactions) to display
    role badges and flag STATUS mismatches between self-selected
    items.user_status and authoritative customer_roles.
    """
    from email_parser.database import _connect
    result = {}
    with _connect() as conn:
        # Build role lists per customer
        role_rows = conn.execute(
            "SELECT customer_id, role_type FROM customer_roles ORDER BY customer_id, role_type"
        ).fetchall()
        for r in role_rows:
            cid = str(r["customer_id"])
            if cid not in result:
                result[cid] = {"roles": [], "first_timer_ever": True}
            result[cid]["roles"].append(r["role_type"])

        # Add first_timer_ever, current status, and chapter for every customer.
        # Status is read from customer_statuses (canonical) with fallback to
        # current_player_status for customers not yet in the new table.
        customer_rows = conn.execute(
            """SELECT c.customer_id, c.first_name, c.last_name,
                      c.first_timer_ever, c.current_player_status, c.chapter,
                      s.status_name, s.display_name
               FROM customers c
               LEFT JOIN customer_statuses cs_latest
                      ON cs_latest.id = (
                          SELECT id FROM customer_statuses
                          WHERE customer_id = c.customer_id
                          ORDER BY set_at DESC LIMIT 1
                      )
               LEFT JOIN statuses s ON s.status_id = cs_latest.status_id"""
        ).fetchall()
        name_to_id = {}
        for c in customer_rows:
            cid = str(c["customer_id"])
            if cid not in result:
                result[cid] = {"roles": [], "first_timer_ever": bool(c["first_timer_ever"])}
            else:
                result[cid]["first_timer_ever"] = bool(c["first_timer_ever"])
            # Prefer status_name from customer_statuses; fall back to current_player_status
            result[cid]["status_name"] = c["status_name"] or None
            result[cid]["status_display_name"] = c["display_name"] or None
            result[cid]["current_player_status"] = c["current_player_status"]
            result[cid]["chapter"] = c["chapter"]
            # Build name→id map for frontend fallback when items.customer_id is null
            name_key = f"{(c['first_name'] or '')} {(c['last_name'] or '')}".strip().lower()
            if name_key:
                name_to_id[name_key] = cid
        result["_by_name"] = name_to_id

        # Count 1ST TIMER registrations per customer for second-use detection
        timer_rows = conn.execute(
            "SELECT customer_id, COUNT(*) as cnt FROM items "
            "WHERE customer_id IS NOT NULL "
            "AND UPPER(user_status) LIKE '%TIMER%' "
            "AND COALESCE(transaction_status, 'active') IN ('active', 'transferred', 'wd') "
            "GROUP BY customer_id"
        ).fetchall()
        for t in timer_rows:
            cid = str(t["customer_id"])
            if cid in result:
                result[cid]["first_timer_used"] = t["cnt"]

    return jsonify(result)


@app.route("/api/customers/<int:customer_id>/status", methods=["POST"])
@require_role("manager")
def api_set_customer_status(customer_id):
    """Insert a new row into customer_statuses (status history).

    Body: { status_name: str, notes?: str }
    Valid status_name values: member, member_plus, guest, 1st_timer, former
    """
    from email_parser.database import set_customer_status, _connect
    data = request.get_json(silent=True) or {}
    status_name = (data.get("status_name") or "").strip().lower()
    valid = {"member", "member_plus", "guest", "1st_timer", "former"}
    if status_name not in valid:
        return jsonify({"error": f"Invalid status_name. Must be one of: {', '.join(sorted(valid))}"}), 400
    notes = data.get("notes") or "manual status change"
    try:
        new_id = set_customer_status(customer_id, status_name, db_path=None, notes=notes)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    return jsonify({"status": "ok", "id": new_id, "status_name": status_name})


@app.route("/api/customers/<int:customer_id>/pace", methods=["POST"])
@require_role("manager")
def api_set_customer_pace(customer_id):
    """One-tap pace rating (task #23, Kerry-ratified scale 1 slow → 3 fast).

    Body: { rating: 1|2|3 }. Always writes an explicit value with
    source='manager' — never NULL, because the boot seed is
    fill-only-if-NULL and clearing a seeded player would resurrect the
    seed value on the next deploy. Manager edits win forever.
    """
    from email_parser.database import set_customer_pace_rating
    data = request.get_json(silent=True) or {}
    try:
        rating = int(data.get("rating"))
    except (TypeError, ValueError):
        return jsonify({"error": "rating must be 1, 2, or 3"}), 400
    if rating not in (1, 2, 3):
        return jsonify({"error": "rating must be 1, 2, or 3"}), 400
    try:
        result = set_customer_pace_rating(customer_id, rating)
    except ValueError as e:
        return jsonify({"error": str(e)}), 404
    return jsonify({"status": "ok", **result})


@app.route("/api/customers/<int:customer_id>/status-history")
@require_role("view-only")
def api_customer_status_history(customer_id):
    """Return full status history for a customer (newest first)."""
    from email_parser.database import _connect
    with _connect() as conn:
        rows = conn.execute(
            """SELECT cs.id, cs.set_at, s.status_name, s.display_name, cs.notes
               FROM customer_statuses cs
               JOIN statuses s ON s.status_id = cs.status_id
               WHERE cs.customer_id = ?
               ORDER BY cs.set_at DESC""",
            (customer_id,),
        ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/customers/create", methods=["POST"])
@require_role("manager")
def api_create_customer():
    """Create a new standalone customer."""
    data = request.get_json(force=True)
    first_name = (data.get("first_name") or "").strip()
    last_name = (data.get("last_name") or "").strip()
    name = (data.get("name") or "").strip()
    # Build name from parts if not given directly
    if not name and (first_name or last_name):
        name = " ".join(filter(None, [first_name, last_name]))
    if not name:
        return jsonify({"error": "name is required"}), 400
    result = create_customer(
        name,
        email=data.get("email", ""),
        phone=data.get("phone", ""),
        chapter=data.get("chapter", ""),
        first_name=first_name,
        last_name=last_name,
        middle_name=data.get("middle_name", ""),
        suffix=data.get("suffix", ""),
    )
    if result is None:
        return jsonify({"error": "Customer already exists"}), 409
    return jsonify({"status": "ok", "item": result})


@app.route("/api/customers/parse-roster", methods=["POST"])
@require_role("manager")
def api_parse_roster():
    """Parse an uploaded Excel file and return headers + preview rows."""
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    file = request.files["file"]
    if not file.filename:
        return jsonify({"error": "No file selected"}), 400

    import io
    from openpyxl import load_workbook
    try:
        wb = load_workbook(io.BytesIO(file.read()), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)

        # Read up to first 5 rows to find the actual header row.
        # Spreadsheets often have a title/metadata row before real headers.
        candidate_rows = []
        for raw_row in rows_iter:
            candidate_rows.append(raw_row)
            if len(candidate_rows) >= 5:
                break
        if not candidate_rows:
            return jsonify({"error": "Empty spreadsheet"}), 400

        # Heuristic: the header row is the first row where >40% of cells
        # are non-empty.  Title rows typically have only 1-2 filled cells.
        header_idx = 0
        total_cols = len(candidate_rows[0])
        for i, row in enumerate(candidate_rows):
            non_empty = sum(1 for c in row if c is not None and str(c).strip())
            if non_empty >= max(2, total_cols * 0.4):
                header_idx = i
                break

        header_row = candidate_rows[header_idx]
        headers = [str(h).strip() if h else f"Column {i+1}"
                   for i, h in enumerate(header_row)]

        # Data rows = remaining candidate rows after header + rest of sheet
        data_candidate = candidate_rows[header_idx + 1:]
        preview = []
        for row in data_candidate:
            preview.append([str(c).strip() if c is not None else "" for c in row])
        for row in rows_iter:
            if len(preview) >= 100:
                break
            preview.append([str(c).strip() if c is not None else "" for c in row])

        wb.close()
        return jsonify({"headers": headers, "preview": preview,
                        "total_rows": len(preview)})
    except Exception as e:
        return jsonify({"error": f"Failed to parse file: {str(e)}"}), 400


@app.route("/api/customers/import-roster", methods=["POST"])
@require_role("manager")
def api_import_roster():
    """Import roster data with column mapping.

    Body: { mapping: {db_field: excel_col_index, ...}, data: [[...], ...],
            new_fields: [{name: "field_name", col_index: N}, ...] }
    """
    data = request.get_json(force=True)
    mapping = data.get("mapping") or {}
    rows_data = data.get("data") or []
    new_fields = data.get("new_fields") or []
    if not mapping or not rows_data:
        return jsonify({"error": "mapping and data are required"}), 400

    # Create any new custom fields first
    fields_created = []
    for nf in new_fields:
        field_name = (nf.get("name") or "").strip().lower().replace(" ", "_")
        col_idx = nf.get("col_index")
        if field_name and col_idx is not None:
            try:
                created = add_custom_field(field_name)
                if created:
                    fields_created.append(field_name)
                # Add to the mapping
                mapping[field_name] = col_idx
            except ValueError as e:
                return jsonify({"error": str(e)}), 400

    # mapping is like {"customer": 0, "customer_email": 2, ...}
    # rows_data is the array of arrays from the preview
    import_rows = []
    for row in rows_data:
        mapped = {}
        for db_field, col_idx in mapping.items():
            if col_idx is not None and 0 <= col_idx < len(row):
                val = str(row[col_idx]).strip() if row[col_idx] else ""
                if val:
                    mapped[db_field] = val
        # Support name from first+last OR full name
        if not mapped.get("customer"):
            first = mapped.get("first_name", "")
            last = mapped.get("last_name", "")
            if first or last:
                mapped["customer"] = " ".join(filter(None, [first, last]))
        if mapped.get("customer"):
            import_rows.append(mapped)

    if not import_rows:
        return jsonify({"error": "No valid rows to import (customer name required)"}), 400

    try:
        result = import_roster(import_rows)
    except Exception as e:
        logger.exception("Roster import failed")
        return jsonify({"error": f"Import failed: {str(e)}"}), 500
    result["fields_created"] = fields_created
    return jsonify(result)


@app.route("/api/customers/preview-roster", methods=["POST"])
@require_role("manager")
def api_preview_roster():
    """Preview a roster import with AI name parsing and duplicate detection.

    Body: { mapping: {db_field: excel_col_index, ...}, data: [[...], ...] }
    Returns enriched row data with parsed names, match status, and validation warnings.
    """
    data = request.get_json(force=True)
    mapping = data.get("mapping") or {}
    rows_data = data.get("data") or []
    if not mapping or not rows_data:
        return jsonify({"error": "mapping and data are required"}), 400

    # Build mapped rows
    preview_rows = []
    for row in rows_data:
        mapped = {}
        for db_field, col_idx in mapping.items():
            if col_idx is not None and 0 <= col_idx < len(row):
                val = str(row[col_idx]).strip() if row[col_idx] else ""
                if val:
                    mapped[db_field] = val
        # Support name from first+last OR full name
        if not mapped.get("customer"):
            first = mapped.get("first_name", "")
            last = mapped.get("last_name", "")
            if first or last:
                mapped["customer"] = " ".join(filter(None, [first, last]))
        if mapped.get("customer"):
            preview_rows.append(mapped)

    if not preview_rows:
        return jsonify({"error": "No valid rows"}), 400

    try:
        result = preview_roster_import(preview_rows)
    except Exception as e:
        logger.exception("Roster preview failed")
        return jsonify({"error": f"Preview analysis failed: {str(e)}"}), 500
    return jsonify(result)


@app.route("/api/customers/merge", methods=["POST"])
@require_role("admin")
def api_merge_customers():
    """Merge one customer into another.

    Body: { source, target, source_customer_id?, target_customer_id? }
    The customer_ids, when given, are used directly instead of re-deriving
    them by name — see merge_customers() for why that re-derivation could
    previously produce a silently-incomplete "split identity" merge.
    """
    data = request.get_json(force=True)
    source = (data.get("source") or "").strip()
    target = (data.get("target") or "").strip()
    if not source or not target:
        return jsonify({"error": "source and target customer names required"}), 400
    if source == target:
        return jsonify({"error": "source and target cannot be the same"}), 400
    try:
        source_customer_id = int(data.get("source_customer_id")) if data.get("source_customer_id") else None
        target_customer_id = int(data.get("target_customer_id")) if data.get("target_customer_id") else None
    except (TypeError, ValueError):
        return jsonify({"error": "source_customer_id/target_customer_id must be integers"}), 400
    try:
        result = merge_customers(
            source, target,
            source_customer_id=source_customer_id,
            target_customer_id=target_customer_id,
        )
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    return jsonify(result)


@app.route("/api/customers/aliases", methods=["GET"])
@require_role("manager")
def api_get_aliases():
    """Get aliases for a customer. Query: ?customer_name=..."""
    customer_name = request.args.get("customer_name", "").strip()
    if not customer_name:
        return jsonify({"error": "customer_name is required"}), 400
    aliases = get_customer_aliases(customer_name)
    return jsonify({"aliases": aliases})


@app.route("/api/customers/aliases", methods=["POST"])
@require_role("manager")
def api_add_alias():
    """Add an alias for a customer."""
    data = request.get_json(force=True)
    customer_name = (data.get("customer_name") or "").strip()
    alias_type = (data.get("alias_type") or "").strip()
    alias_value = (data.get("alias_value") or "").strip()
    if not customer_name or not alias_type or not alias_value:
        return jsonify({"error": "customer_name, alias_type, and alias_value are required"}), 400
    try:
        result = add_customer_alias(customer_name, alias_type, alias_value)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    return jsonify(result)


@app.route("/api/customers/aliases/<int:alias_id>", methods=["DELETE"])
@require_role("manager")
def api_delete_alias(alias_id):
    """Delete an alias by ID."""
    deleted = delete_customer_alias(alias_id)
    if not deleted:
        return jsonify({"error": "Alias not found"}), 404
    return jsonify({"status": "ok"})


@app.route("/api/customers/winnings")
@require_role("view-only")
def api_customer_winnings():
    """Get payout/winnings history for a customer.

    Query: ?customer_name=...&customer_id=... — customer_id, when given, is
    used directly instead of an unqualified name match (see
    get_customer_winnings() for the same-name collision risk that avoids).
    """
    from email_parser.database import get_customer_winnings
    name = request.args.get("customer_name", "").strip()
    if not name:
        return jsonify({"error": "customer_name required"}), 400
    try:
        customer_id = int(request.args.get("customer_id")) if request.args.get("customer_id") else None
    except (TypeError, ValueError):
        customer_id = None
    return jsonify(get_customer_winnings(name, customer_id=customer_id))


@app.route("/api/customers/<int:customer_id>/gg-cards")
@require_role("manager")
def api_customer_gg_cards(customer_id):
    """Races this customer appears in (persisted GG standings rows).

    Powers the Customers Points tab: race_key + member_card_id feed the
    existing points-race detail endpoint; rank/points make the summary
    strip. Manager role matches the Contests endpoints the tab chains to.
    """
    from email_parser.database import get_customer_gg_cards
    return jsonify({"cards": get_customer_gg_cards(customer_id)})


@app.route("/api/customers/from-rsvp", methods=["POST"])
@require_role("manager")
def api_create_customer_from_rsvp():
    """Create a customer from an unmatched RSVP and link them."""
    data = request.get_json(force=True)
    name = (data.get("name") or "").strip()
    email = (data.get("email") or "").strip()
    if not name:
        return jsonify({"error": "name is required"}), 400
    try:
        result = create_customer_from_rsvp(name, email)
        return jsonify(result)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        logger.exception("Error creating customer from RSVP: %s", e)
        return jsonify({"error": f"Server error: {e}"}), 500


@app.route("/api/customers/link-rsvp", methods=["POST"])
@require_role("manager")
def api_link_rsvp_to_customer():
    """Link an unmatched RSVP email to an existing customer."""
    data = request.get_json(force=True)
    rsvp_email = (data.get("rsvp_email") or "").strip()
    target_name = (data.get("target_customer") or "").strip()
    rsvp_player_name = (data.get("rsvp_player_name") or "").strip()
    if not rsvp_email or not target_name:
        return jsonify({"error": "rsvp_email and target_customer are required"}), 400
    try:
        result = link_rsvp_to_customer(rsvp_email, target_name, rsvp_player_name)
        return jsonify(result)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        logger.exception("Error linking RSVP to customer: %s", e)
        return jsonify({"error": f"Server error: {e}"}), 500


@app.route("/api/chapters")
@require_role("view-only")
def api_chapters():
    """Return all chapters with their IDs."""
    from email_parser.database import _connect
    with _connect() as conn:
        rows = conn.execute(
            "SELECT chapter_id, name, short_code, timezone, status FROM chapters ORDER BY name"
        ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/courses")
@require_role("view-only")
def api_courses():
    """Return all courses with chapter linkage and aliases."""
    from email_parser.database import _connect
    with _connect() as conn:
        rows = conn.execute(
            """SELECT c.course_id, c.name, c.chapter_id, c.city, c.state, c.status,
                      ch.name as chapter_name
               FROM courses c
               LEFT JOIN chapters ch ON ch.chapter_id = c.chapter_id
               ORDER BY c.name"""
        ).fetchall()
        courses = [dict(r) for r in rows]

        # Attach aliases
        alias_map: dict[int, list[str]] = {}
        for a in conn.execute("SELECT course_id, alias_name FROM course_aliases").fetchall():
            alias_map.setdefault(a["course_id"], []).append(a["alias_name"])
        for c in courses:
            c["aliases"] = alias_map.get(c["course_id"], [])

    return jsonify(courses)


@app.route("/api/courses", methods=["POST"])
@require_role("manager")
def api_create_course():
    """Create a new course. Body: {name, chapter_id?, city?, state?}."""
    from email_parser.database import _connect
    d = request.json or {}
    name = (d.get("name") or "").strip()
    if not name:
        return jsonify({"error": "Course name required"}), 400
    with _connect() as conn:
        existing = conn.execute(
            "SELECT course_id FROM courses WHERE LOWER(name) = LOWER(?)", (name,)
        ).fetchone()
        if existing:
            return jsonify({"error": f"Course '{name}' already exists", "course_id": existing["course_id"]}), 409
        cur = conn.execute(
            """INSERT INTO courses (name, chapter_id, city, state)
               VALUES (?, ?, ?, ?)""",
            (name, d.get("chapter_id"), d.get("city"), d.get("state")),
        )
        conn.commit()
        return jsonify({"course_id": cur.lastrowid, "name": name})


@app.route("/api/events")
@require_role("view-only")
def api_events():
    """Return all events with registration counts and aliases."""
    return jsonify(get_all_events())


@app.route("/api/events/aliases")
@require_role("view-only")
def api_event_aliases():
    """Return alias_name → canonical_event_name map."""
    return jsonify(get_all_event_aliases())


@app.route("/api/events/<int:event_id>/aliases", methods=["POST"])
@require_role("manager")
def api_add_event_alias_to_event(event_id):
    """Add an alias pointing to this event's canonical name."""
    data = request.get_json(silent=True) or {}
    alias_name = (data.get("alias_name") or "").strip()
    if not alias_name:
        return jsonify({"error": "alias_name required"}), 400
    evs = get_all_events()
    ev = next((e for e in evs if e["id"] == event_id), None)
    if not ev:
        return jsonify({"error": "Event not found"}), 404
    inserted = add_event_alias(alias_name, ev["item_name"])
    return jsonify({"status": "ok", "inserted": inserted, "canonical": ev["item_name"]})


@app.route("/api/events/<int:event_id>/aliases", methods=["DELETE"])
@require_role("manager")
def api_delete_event_alias_from_event(event_id):
    """Remove an alias from this event."""
    data = request.get_json(silent=True) or {}
    alias_name = (data.get("alias_name") or "").strip()
    if not alias_name:
        return jsonify({"error": "alias_name required"}), 400
    deleted = delete_event_alias(alias_name)
    return jsonify({"status": "ok", "deleted": deleted})


@app.route("/events/<int:event_id>/starter-sheet")
@require_role("manager")
def starter_sheet_page(event_id):
    """Print-optimized Starter Sheet for an event (B5) — tee times / groups /
    players / cart split, rendered from the saved pairings."""
    from email_parser.database import get_event_print_pack
    pack = get_event_print_pack(event_id)
    if not pack:
        return "Event not found", 404
    return render_template("starter_sheet.html", pack=pack)


@app.route("/events/<int:event_id>/cart-signs")
@require_role("manager")
def cart_signs_page(event_id):
    """Print-optimized Cart Signs for an event (B5) — one card per cart
    (seats 1&2 = Cart A, 3&4 = Cart B), rendered from the saved pairings."""
    from email_parser.database import get_event_print_pack
    pack = get_event_print_pack(event_id)
    if not pack:
        return "Event not found", 404
    return render_template("cart_signs.html", pack=pack)


@app.route("/api/events/<int:event_id>/pairings", methods=["GET"])
@require_role("view-only")
def api_get_pairings(event_id):
    """Return saved pairings for an event plus the computed tee-time slots."""
    try:
        from email_parser.database import _connect
        with _connect() as conn:
            ev = conn.execute("SELECT * FROM events WHERE id = ?", (event_id,)).fetchone()
        if not ev:
            return jsonify({"error": "Event not found"}), 404
        ev = dict(ev)
        pairings = get_event_pairings(event_id)
        slots_9 = _pairing_time_slots(ev, "9")
        slots_18 = _pairing_time_slots(ev, "18")
        # Current active player list — used by UI to detect unassigned players
        INACTIVE = ("credited", "refunded", "transferred", "wd")
        ph = ",".join("?" * len(INACTIVE))
        _pconn = get_connection()
        try:
            player_rows = _pconn.execute(f"""
                SELECT DISTINCT i.customer AS name, i.holes, i.tee_choice,
                                c.pace_rating, c.current_player_status,
                                i.user_status, i.customer_id
                FROM events e
                LEFT JOIN event_aliases ea ON ea.canonical_event_name = e.item_name
                JOIN items i ON (
                    i.item_name = e.item_name COLLATE NOCASE
                    OR i.item_name = ea.alias_name COLLATE NOCASE
                    OR i.event_id = e.id
                )
                LEFT JOIN customers c ON c.customer_id = i.customer_id
                WHERE e.id = ?
                  AND COALESCE(i.transaction_status,'active') NOT IN ({ph})
                  AND i.parent_item_id IS NULL
                ORDER BY i.customer COLLATE NOCASE
            """, (event_id, *INACTIVE)).fetchall()
        finally:
            _pconn.close()
        # Player TIER for the pairing-card colour bands: 'member' |
        # 'alumni' | 'guest', from derive_member_financial_status_bulk —
        # the same D1 truth Player Rankings chips with, so a FORMER member
        # reads as ALUMNI instead of being lumped in with guests (Kerry
        # 2026-07-31, "Why is Wade Amen shown Pink?"). The roster decides
        # the tier; the order label only splits GUEST from 1ST TIMER.
        from email_parser.database import (
            _ls_is_member as _isM, derive_member_financial_status_bulk)
        event_players = [dict(r) for r in player_rows]
        _cids = [d["customer_id"] for d in event_players if d.get("customer_id")]
        _tier = {}
        if _cids:
            _tc = get_connection()
            try:
                _tier = derive_member_financial_status_bulk(_tc, _cids)
            except Exception:
                logger.exception("Member-tier derivation failed (non-fatal)")
            finally:
                _tc.close()
        for d in event_players:
            cps = d.pop("current_player_status", None)
            ms = _tier.get(d.get("customer_id"))
            if not ms:
                # No membership rows to derive from — fall back to the
                # roster status, which still separates member from guest.
                ms = "member" if _isM(cps, d.get("user_status")) else "guest"
            d["member_status"] = ms
            d["is_member"] = ms == "member"
        # Match Play matches still pending among this roster — lets the
        # PAIRINGS tab badge opponents on SAVED pairings too (rule 8's
        # visual denotation), not just on a fresh generate.
        mp_matches = []
        try:
            from email_parser.database import detect_match_play_pairings
            mp_matches = detect_match_play_pairings(event_id).get("matches", [])
        except Exception:
            logger.exception("MP detection failed for event %d (non-fatal)", event_id)
        # Partner-request list (who asked for whom + suppression state)
        # so the PAIRINGS tab can show its Requests chip without a
        # second round trip.
        partner_requests = []
        try:
            from email_parser.database import get_event_partner_requests
            partner_requests = get_event_partner_requests(event_id).get("requests", [])
        except Exception:
            logger.exception("Partner-request list failed for event %d (non-fatal)", event_id)
        # Season points for the STANDINGS view's points column, so a SAVED
        # sheet shows them too and not only a freshly generated one
        # (Kerry 2026-07-31). Read-only: a big max_age never triggers a GG
        # fetch on a plain GET — the point-of-use refresh belongs to
        # Generate, and this endpoint is hit on every panel open.
        standings_points, standings_enrolled = {}, {}
        try:
            from email_parser.database import _standings_rank_map
            _rk = (request.args.get("race_key") or "").strip() or None
            _, _, _, _meta = _standings_rank_map(
                ev.get("chapter"), _rk, max_age_hours=10 ** 6,
                event_name=ev.get("item_name"), _with_meta=True)
            # Keyed by customer_id AND name — see the note in
            # generate_event_pairings. Name-only made GG's spelling of a
            # player ("MURPHY, Mike") fail to match our roster's
            # ("Michael Murphy") even when the id had resolved.
            standings_points = {str(k): v for k, v in _meta["points"].items()}
            standings_enrolled = {str(k): v for k, v in _meta["enrolled"].items()}
        except Exception:
            logger.exception("Standings points lookup failed for event %d "
                             "(non-fatal)", event_id)
        return jsonify({
            "pairings": pairings,
            "slots_9": slots_9,
            "slots_18": slots_18,
            "event_players": event_players,
            "mp_matches": mp_matches,
            "partner_requests": partner_requests,
            "standings_points": standings_points,
            "standings_enrolled": standings_enrolled,
            "event": {
                "format": ev.get("format"),
                "start_type": ev.get("start_type"),
                "start_type_18": ev.get("start_type_18"),
                "tee_time_count": ev.get("tee_time_count"),
                "tee_time_count_18": ev.get("tee_time_count_18"),
                "nine_side": ev.get("nine_side") or "Front",
                # The saved pairing METHOD travels with the panel's own
                # fetch, not the events-list payload — that list can be
                # stale or predate the column, which is exactly why the
                # STANDINGS button kept coming back unselected (Kerry
                # 2026-07-31, third report).
                "pairing_mode": ev.get("pairing_mode") or None,
                "pairing_race_key": ev.get("pairing_race_key") or None,
                "allow_fivesomes": 1 if ev.get("allow_fivesomes") else 0,
            },
        })
    except Exception as e:
        logger.exception("Failed to get pairings for event %d", event_id)
        return jsonify({"error": str(e)}), 500


@app.route("/api/events/<int:event_id>/pairings/matchplay", methods=["GET"])
@require_role("manager")
def api_pairings_matchplay(event_id):
    """Potential Match Play matches implied by the season state among this
    event's roster (rule 8 amendment: the manager confirms/declines each
    one before Generate runs — a decline drops that constraint)."""
    try:
        from email_parser.database import detect_match_play_pairings
        return jsonify(detect_match_play_pairings(event_id))
    except ValueError as e:
        return jsonify({"error": str(e)}), 404
    except Exception as e:
        logger.exception("Match Play detection failed for event %d", event_id)
        return jsonify({"error": str(e)}), 500


@app.route("/api/events/<int:event_id>/pairings/generate", methods=["POST"])
@require_role("manager")
def api_generate_pairings(event_id):
    """Run the pairing generator and return proposed groups (not saved)."""
    data = request.get_json(silent=True) or {}
    mode = data.get("mode", "random")
    protect = bool(data.get("protect_partner_requests", True))
    seeds = data.get("seeds", [])
    mp_pairs = data.get("mp_pairs", [])
    if mode not in ("random", "abcd", "standings"):
        return jsonify({"error": "mode must be 'random', 'abcd' or "
                                 "'standings'"}), 400
    if not isinstance(mp_pairs, list):
        return jsonify({"error": "mp_pairs must be a list of [name, name] pairs"}), 400
    try:
        result = generate_event_pairings(
            event_id,
            mode=mode,
            protect_partner_requests=protect,
            seeds=seeds,
            mp_pairs=mp_pairs,
            race_key=(data.get("race_key") or "").strip() or None,
        )
        return jsonify(result)
    except ValueError as e:
        return jsonify({"error": str(e)}), 404
    except Exception as e:
        logger.exception("Pairing generation failed for event %d", event_id)
        return jsonify({"error": str(e)}), 500


@app.route("/api/events/<int:event_id>/pairings/save", methods=["POST"])
@require_role("manager")
def api_save_pairings(event_id):
    """Save pairings for an event (replaces any previous save)."""
    data = request.get_json(silent=True) or {}
    groups_by_holes = data.get("groups_by_holes")
    if not groups_by_holes or not isinstance(groups_by_holes, dict):
        return jsonify({"error": "groups_by_holes required"}), 400
    try:
        save_event_pairings(event_id, groups_by_holes)
        return jsonify({"status": "ok"})
    except Exception as e:
        logger.exception("Failed to save pairings for event %d", event_id)
        return jsonify({"error": str(e)}), 500


@app.route("/api/events/<int:event_id>/pairings/gg-rounds")
@require_role("manager")
def api_pairings_gg_rounds(event_id):
    """Rounds on this event's chapter GG tee sheet, ranked by how well
    each matches the event. Backs the PAIRINGS tab's GG SHEET button when
    the auto-match is not confident enough to act on alone."""
    from email_parser.database import gg_teesheet_round_options
    try:
        res = gg_teesheet_round_options(event_id)
        if res.get("error"):
            return jsonify(res), 404
        return jsonify(res)
    except Exception as e:
        logger.exception("GG round listing failed for event %d", event_id)
        return jsonify({"error": str(e)}), 502


@app.route("/api/events/<int:event_id>/pairings/import-gg", methods=["POST"])
@require_role("manager")
def api_pairings_import_gg(event_id):
    """Pull this event's pairings from Golf Genius (Kerry 2026-07-31) —
    for events whose sheet was built in GG instead of by our generator.
    REPLACES the event's saved pairings for the chosen leg."""
    from email_parser.database import import_gg_pairings_for_event
    data = request.get_json(silent=True) or {}
    holes = str(data.get("holes") or "").strip() or None
    if holes not in (None, "9", "18"):
        return jsonify({"error": "holes must be 9 or 18"}), 400
    try:
        res = import_gg_pairings_for_event(
            event_id,
            round_id=data.get("round_id"),
            holes=holes,
            apply=bool(data.get("apply", True)))
        if res.get("error"):
            return jsonify(res), 409
        return jsonify(res)
    except Exception as e:
        logger.exception("GG pairings import failed for event %d", event_id)
        return jsonify({"error": str(e)}), 502


@app.route("/api/events/<int:event_id>/pairings/race-options")
@require_role("manager")
def api_pairing_race_options(event_id):
    """Season-contest standings this event's pairings can be ordered by.

    The chapter's own City NET race is flagged as the default; a TGF-wide
    event defaults to THE FELLOWSHIP CUP (Kerry 2026-07-30). Offered as a
    choice rather than inferred, because the right race is not always the
    chapter's own.
    """
    from email_parser.database import get_connection, pairing_race_options
    conn = get_connection()
    try:
        row = conn.execute("SELECT chapter, item_name FROM events WHERE id = ?",
                           (event_id,)).fetchone()
        if not row:
            return jsonify({"error": "Event not found."}), 404
        return jsonify({"chapter": row["chapter"],
                        "options": pairing_race_options(row["chapter"],
                                                        row["item_name"])})
    finally:
        conn.close()


@app.route("/api/events/<int:event_id>/pairings/requests", methods=["GET"])
@require_role("manager")
def api_pairings_requests(event_id):
    """Partner requests on this event's roster: who asked for whom,
    whether the text matched a rostered player, and suppression state."""
    try:
        from email_parser.database import get_event_partner_requests
        return jsonify(get_event_partner_requests(event_id))
    except Exception as e:
        logger.exception("Partner-request list failed for event %d", event_id)
        return jsonify({"error": str(e)}), 500


@app.route("/api/events/<int:event_id>/pairings/requests/suppress", methods=["POST"])
@require_role("manager")
def api_pairings_request_suppress(event_id):
    """Suppress or restore one player's partner request for this event.
    Suppressed requests stay listed (badged SUPPRESSED) but the
    generator ignores them on the next run."""
    data = request.get_json(silent=True) or {}
    requester = (data.get("requester") or "").strip()
    if not requester:
        return jsonify({"error": "requester required"}), 400
    try:
        from email_parser.database import set_partner_request_suppression
        return jsonify(set_partner_request_suppression(
            event_id, requester, bool(data.get("suppressed", True))))
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        logger.exception("Request suppression failed for event %d", event_id)
        return jsonify({"error": str(e)}), 500


@app.route("/api/events/<int:event_id>/pairings/requests/match", methods=["POST"])
@require_role("manager")
def api_pairings_request_match(event_id):
    """Manually bind a partner request to a rostered player (signup
    text that didn't auto-resolve, e.g. a nickname). partner = null
    clears the manual match."""
    data = request.get_json(silent=True) or {}
    requester = (data.get("requester") or "").strip()
    if not requester:
        return jsonify({"error": "requester required"}), 400
    try:
        from email_parser.database import set_partner_request_match
        return jsonify(set_partner_request_match(
            event_id, requester, data.get("partner")))
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        logger.exception("Request match failed for event %d", event_id)
        return jsonify({"error": str(e)}), 500


@app.route("/api/events/<int:event_id>/pairings/requests/approve", methods=["POST"])
@require_role("manager")
def api_pairings_request_approve(event_id):
    """Approve an OUTRANKED partner request, or revoke that approval
    (Kerry 2026-07-31). First-come is the default; the manager is the
    override. An approved request JOINS the group that already claimed
    its partner rather than displacing anyone — a foursome remains the
    ceiling, and the request stays outranked if honoring it would make
    five."""
    data = request.get_json(silent=True) or {}
    requester = (data.get("requester") or "").strip()
    if not requester:
        return jsonify({"error": "requester required"}), 400
    try:
        from email_parser.database import set_partner_request_approval
        return jsonify(set_partner_request_approval(
            event_id, requester, bool(data.get("approved", True)),
            approved_by=session.get("role")))
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        logger.exception("Request approval failed for event %d", event_id)
        return jsonify({"error": str(e)}), 500


@app.route("/api/events/<int:event_id>/pairings/switch-side", methods=["POST"])
@require_role("manager")
def api_pairings_switch_side(event_id):
    """Flip the event's 9-hole side (Front ↔ Back nine). Shotgun events
    also get their SAVED 9-hole slot labels shifted (1A ↔ 10A, …) so
    the sheet and printables follow without a regenerate."""
    try:
        from email_parser.database import switch_event_pairings_side
        return jsonify(switch_event_pairings_side(event_id))
    except ValueError as e:
        return jsonify({"error": str(e)}), 404
    except Exception as e:
        logger.exception("Side switch failed for event %d", event_id)
        return jsonify({"error": str(e)}), 500


@app.route("/api/admin/gg-points-identity-audit", methods=["GET", "POST"])
@require_role("admin")
def api_gg_points_identity_audit():
    """Standings rows that never resolved to a customer_id — report + repair.

    An unresolved row still shows on the Contests board but is invisible
    to anything that joins on identity (pairings order, the points column,
    flighting, payouts). Re-runs the resolver, links what the nickname
    rung can now match, captures the alias, and returns whoever is left.
    """
    try:
        from email_parser.database import audit_gg_points_identities
        return jsonify(audit_gg_points_identities())
    except Exception as e:
        logger.exception("GG points identity audit failed")
        return jsonify({"error": str(e)}), 500


@app.route("/api/events/<int:event_id>/pairings", methods=["DELETE"])
@require_role("manager")
def api_delete_pairings(event_id):
    """Clear saved pairings (and their history) for an event."""
    try:
        delete_event_pairings(event_id)
        return jsonify({"status": "ok"})
    except Exception as e:
        logger.exception("Failed to delete pairings for event %d", event_id)
        return jsonify({"error": str(e)}), 500


@app.route("/api/events/sync", methods=["POST"])
@require_role("manager")
def api_sync_events():
    """Scan items and auto-create event entries for event-type items."""
    try:
        result = sync_events_from_items()
        return jsonify({"status": "ok", **result})
    except Exception as e:
        logger.exception("Event sync failed")
        return jsonify({"error": str(e)}), 500


@app.route("/api/events/<int:event_id>", methods=["PATCH"])
@require_role("manager")
def api_update_event(event_id):
    """Update fields on an event."""
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "Request body must be JSON."}), 400
    err = _validate_update_fields(data)
    if err:
        return jsonify({"error": err}), 400
    if update_event(event_id, data):
        return jsonify({"status": "ok"})
    return jsonify({"error": "not found or no valid fields"}), 404


@app.route("/api/events/<int:event_id>", methods=["DELETE"])
@require_role("admin")
def api_delete_event(event_id):
    """Delete an event. Admin only."""
    if delete_event(event_id):
        return jsonify({"status": "ok"})
    return jsonify({"error": "not found"}), 404


@app.route("/api/events/<int:event_id>/cancel", methods=["POST"])
@require_role("admin")
def api_cancel_event(event_id):
    """Cancel or postpone an event, silently removing comps and RSVP-only players."""
    data = request.get_json(silent=True) or {}
    status = data.get("status", "cancelled")
    reason = data.get("reason", "").strip()
    if status not in ("cancelled", "postponed"):
        return jsonify({"error": "status must be 'cancelled' or 'postponed'"}), 400
    if not reason:
        return jsonify({"error": "reason is required"}), 400

    # Set the event status
    if not set_event_status(event_id, status, reason):
        return jsonify({"error": "Event not found"}), 404

    # Silently credit comps and RSVP-only players
    players = get_cancellation_players(event_id)
    silent_note = f"Event {status} — {reason}"
    silent_count = 0
    for item in players.get("silent", []):
        try:
            credit_item(item["id"], note=silent_note)
            silent_count += 1
        except Exception:
            pass

    return jsonify({
        "status": "ok",
        "silent_removed": silent_count,
        "paid_players": len(players.get("paid", [])),
    })


@app.route("/api/events/<int:event_id>/restore", methods=["POST"])
@require_role("admin")
def api_restore_event(event_id):
    """Restore a cancelled/postponed event to active. Only allowed if no player actions taken."""
    if not can_restore_event(event_id):
        return jsonify({"error": "Cannot restore: player actions have already been applied."}), 400
    if set_event_status(event_id, "active", ""):
        return jsonify({"status": "ok"})
    return jsonify({"error": "Event not found"}), 404


@app.route("/api/events/<int:event_id>/cancellation-players", methods=["GET"])
@require_role("admin")
def api_get_cancellation_players(event_id):
    """Return paid players who need a credit/refund action after cancellation."""
    players = get_cancellation_players(event_id)
    return jsonify(players.get("paid", []))


@app.route("/api/events/<int:event_id>/cancel-bulk", methods=["POST"])
@require_role("admin")
def api_cancel_bulk(event_id):
    """Apply credit or refund to all eligible paid players for a cancelled event."""
    data = request.get_json(silent=True) or {}
    action = data.get("action")  # 'credit' or 'refund'
    note = data.get("note", "").strip() or "Event cancelled"
    if action not in ("credit", "refund"):
        return jsonify({"error": "action must be 'credit' or 'refund'"}), 400

    players = get_cancellation_players(event_id)
    results = {"ok": [], "failed": []}
    for item in players.get("paid", []):
        item_id = item["id"]
        try:
            if action == "credit":
                success = credit_item(item_id, note=note)
            else:
                method = item.get("auto_refund_method", "")
                success = refund_item(item_id, method=method, note=note)
            if success:
                results["ok"].append(item_id)
            else:
                results["failed"].append({"id": item_id, "reason": "already actioned"})
        except Exception as e:
            results["failed"].append({"id": item_id, "reason": str(e)})

    return jsonify({"status": "ok", "results": results})


def _cancel_recipient_list(plan, outcome_by_item):
    """Per-recipient send list + personalized vars for cancellation
    notices. SHARED by cancel-preview (predicted outcomes) and
    cancel-execute (actual outcomes) so what the manager previews is
    what gets sent. Returns (recipients, players_without_email)."""
    recipients, no_email = [], []
    for rec in plan["recipients"]:
        if not rec["email"]:
            no_email.append(rec["customer"])
            continue
        outs = {outcome_by_item[iid][0] for iid in rec["item_ids"]
                if iid in outcome_by_item}
        amount_str = f"${rec['amount']:,.2f}"
        if rec["kind"] != "paid" or rec["amount"] <= 0:
            amount_str = "$0.00"
            line = ("No payment was on file for this event, so there is "
                    "nothing to settle.")
            outcome = "none"
        elif "credited" in outs:
            line = (f"Your {amount_str} entry has been converted to a full "
                    "credit on your account — it will be applied "
                    "automatically when you register for a future event.")
            outcome = "credited"
        elif "refunded" in outs:
            methods = {outcome_by_item[iid][1] for iid in rec["item_ids"]
                       if outcome_by_item.get(iid, ("", ""))[0] == "refunded"}
            mtxt = next((m for m in methods if m), "your original payment method")
            line = f"Your {amount_str} is being refunded via {mtxt}."
            outcome = "refunded"
        else:
            line = "We'll follow up with you separately about your entry."
            outcome = "skipped"
        recipients.append({
            "player_name": rec["customer"],
            "email": rec["email"],
            "outcome": outcome,
            "vars": {"credit_amount": amount_str, "credit_line": line},
        })
    return recipients, no_email


def _cancel_event_vars(event_id, reason, status):
    ev = next((e for e in get_all_events() if e.get("id") == event_id), {}) or {}
    return {
        "event_name": ev.get("item_name") or "",
        "event_date": ev.get("event_date") or "",
        "course": ev.get("course") or "",
        "chapter": ev.get("chapter") or "",
        "reason": reason,
        "status_label": status,  # 'cancelled' / 'postponed' reads well in copy
    }


def _cancel_predicted_outcomes(paid_by_id, mode, actions):
    """What WOULD happen to each paid item — the preview's stand-in for
    execute's real results (skip/credit/refund per mode)."""
    custom_map = {a["item_id"]: a for a in (actions or [])
                  if a.get("item_id") is not None}
    outcomes = {}
    for iid, item in paid_by_id.items():
        if mode in ("credit", "refund"):
            act = mode
            method = item.get("auto_refund_method", "")
        else:
            entry = custom_map.get(iid, {})
            act = entry.get("action") or "skip"
            method = entry.get("method") or item.get("auto_refund_method", "")
        outcomes[iid] = (("credited" if act == "credit"
                          else "refunded" if act == "refund"
                          else "skipped"), method if act == "refund" else
                         (method if act == "credit" else ""))
    return outcomes


@app.route("/api/events/<int:event_id>/cancel-preview", methods=["POST"])
@require_role("admin")
def api_cancel_preview(event_id):
    """Dry-run of the one-tap cancellation emails (Kerry 2026-07-14):
    each player's FULLY RENDERED subject + body with their exact amount,
    using the same recipient builder and template renderer as the send —
    zero writes. Body: {status, reason, mode, actions?, subject, html_body}.
    """
    from email_parser.database import plan_event_cancellation_notice
    data = request.get_json(silent=True) or {}
    status = data.get("status", "cancelled")
    reason = (data.get("reason") or "").strip()
    mode = data.get("mode", "credit")
    subject_tpl = (data.get("subject") or "").strip()
    body_tpl = (data.get("html_body") or "").strip()
    if status not in ("cancelled", "postponed"):
        return jsonify({"error": "status must be 'cancelled' or 'postponed'"}), 400
    if mode not in ("credit", "refund", "custom"):
        return jsonify({"error": "mode must be 'credit', 'refund', or 'custom'"}), 400
    if not subject_tpl or not body_tpl:
        return jsonify({"error": "subject and html_body are required"}), 400

    try:
        plan = plan_event_cancellation_notice(event_id)
    except Exception as e:
        logger.exception("Cancel preview failed for event %d", event_id)
        return jsonify({"error": str(e)}), 500
    players = get_cancellation_players(event_id)
    paid_by_id = {p["id"]: p for p in players.get("paid", [])}
    outcome_by_item = _cancel_predicted_outcomes(paid_by_id, mode,
                                                 data.get("actions"))
    recipients, no_email = _cancel_recipient_list(plan, outcome_by_item)
    event_vars = _cancel_event_vars(event_id, reason or "—", status)

    previews = []
    for r in recipients:
        variables = {**event_vars, **r.get("vars", {}),
                     "player_name": r.get("player_name") or "Player"}
        previews.append({
            "player_name": r["player_name"],
            "email": r["email"],
            "outcome": r["outcome"],
            "credit_amount": r["vars"]["credit_amount"],
            "subject": render_msg_template(subject_tpl, variables),
            "html": render_msg_template(body_tpl, variables),
        })
    return jsonify({
        "recipients": previews,
        "players_without_email": no_email,
        "n_recipients": len(previews),
    })


@app.route("/api/events/<int:event_id>/cancel-execute", methods=["POST"])
@require_role("admin")
def api_cancel_execute(event_id):
    """One-tap cancellation (Kerry-ratified 2026-07-14): set status, settle
    every player, and send notification emails carrying each player's
    EXACT credit amount — one execute, correct by construction (the plan
    with amounts + emails is captured BEFORE anything is credited, so the
    old credited-players-drop-off-the-email-audience trap can't occur).

    Body: {
        status: 'cancelled'|'postponed',
        reason: str (required),
        mode: 'credit'|'refund'|'custom' (default credit),
        actions: [{item_id, action: credit|refund|skip, method?}]
                 (custom mode only; unlisted items are skipped),
        send_email: bool (default true),
        subject: str, html_body: str (required when send_email —
                 vars: {player_name} {event_name} {event_date} {course}
                 {chapter} {reason} {status_label} {credit_amount}
                 {credit_line})
    }
    """
    from email_parser.database import plan_event_cancellation_notice
    data = request.get_json(silent=True) or {}
    status = data.get("status", "cancelled")
    reason = (data.get("reason") or "").strip()
    mode = data.get("mode", "credit")
    badge = (data.get("badge") or "").strip()[:32]
    clear_rsvps = bool(data.get("clear_rsvps", True))
    send_email = bool(data.get("send_email", True))
    subject_tpl = (data.get("subject") or "").strip()
    body_tpl = (data.get("html_body") or "").strip()
    if status not in ("cancelled", "postponed"):
        return jsonify({"error": "status must be 'cancelled' or 'postponed'"}), 400
    if not reason:
        return jsonify({"error": "reason is required"}), 400
    if mode not in ("credit", "refund", "custom"):
        return jsonify({"error": "mode must be 'credit', 'refund', or 'custom'"}), 400
    if send_email and (not subject_tpl or not body_tpl):
        return jsonify({"error": "subject and html_body are required when send_email is true"}), 400

    # Plan FIRST — amounts and emails come from still-active items
    plan = plan_event_cancellation_notice(event_id)
    players = get_cancellation_players(event_id)
    paid_by_id = {p["id"]: p for p in players.get("paid", [])}

    if not set_event_status(event_id, status, reason, badge=badge or None):
        return jsonify({"error": "Event not found"}), 404

    note = f"Event {status} — {reason}"

    # Silent removals (comps / RSVP-only — nothing owed)
    silent_removed = 0
    for iid in plan["silent_items"]:
        try:
            if credit_item(iid, note=note):
                silent_removed += 1
        except Exception:
            logger.exception("Silent credit failed for item %s", iid)

    # Paid players
    custom_map = {}
    if mode == "custom":
        for a in (data.get("actions") or []):
            if a.get("item_id") is not None:
                custom_map[a["item_id"]] = a
    outcome_by_item = {}
    ok = failed = skipped = 0
    for iid, item in paid_by_id.items():
        if mode in ("credit", "refund"):
            act = mode
            method = item.get("auto_refund_method", "")
        else:
            entry = custom_map.get(iid, {})
            act = entry.get("action") or "skip"
            method = entry.get("method") or item.get("auto_refund_method", "")
        if act == "skip":
            outcome_by_item[iid] = ("skipped", "")
            skipped += 1
            continue
        try:
            if act == "credit":
                success = credit_item(iid, note=note)
            else:
                success = refund_item(iid, method=method, note=note)
            outcome_by_item[iid] = (("credited" if act == "credit" else "refunded")
                                    if success else "failed", method)
            if success:
                ok += 1
            else:
                failed += 1
        except Exception:
            logger.exception("Cancel action failed for item %s", iid)
            outcome_by_item[iid] = ("failed", method)
            failed += 1

    # Clear the RSVP roster (Kerry 2026-07-14: the credit pass rightly
    # skips never-paid rsvp_only/gg_rsvp rows, but they kept the roster
    # populated after cancellation). They were already captured as
    # notification recipients by the plan above, so they still get the
    # email.
    rsvps_cleared = 0
    if clear_rsvps:
        try:
            from email_parser.database import clear_event_rsvp_items
            rsvps_cleared = clear_event_rsvp_items(
                event_id, note=f"{note} — RSVP removed")
        except Exception:
            logger.exception("RSVP clear failed for event %d", event_id)

    # Notification emails — every player (settled, skipped, RSVP-only).
    # Same builder the preview endpoint uses, fed with ACTUAL outcomes.
    email_result = {"sent": 0, "failed": 0, "errors": []}
    no_email = []
    if send_email:
        event_vars = _cancel_event_vars(event_id, reason, status)
        recipients, no_email = _cancel_recipient_list(plan, outcome_by_item)
        if recipients:
            email_result = send_bulk_emails(
                recipients=recipients,
                subject_template=subject_tpl,
                body_template=body_tpl,
                event_vars=event_vars,
            )
            role = session.get("role", "unknown")
            error_emails = [e["recipient"] for e in email_result.get("errors", [])]
            for r in recipients:
                try:
                    log_message({
                        "event_name": event_vars["event_name"],
                        "template_id": None,
                        "channel": "email",
                        "recipient_name": r.get("player_name"),
                        "recipient_address": r["email"],
                        "subject": render_msg_template(
                            subject_tpl, {**event_vars, **r.get("vars", {}),
                                          "player_name": r.get("player_name") or "Player"}),
                        "body_preview": render_msg_template(
                            body_tpl, {**event_vars, **r.get("vars", {}),
                                       "player_name": r.get("player_name") or "Player"})[:200],
                        "audience": "cancellation",
                        "status": "sent" if r["email"] not in error_emails else "failed",
                        "sent_by": role,
                    })
                except Exception:
                    logger.exception("Failed to log cancellation email to %s", r["email"])

    return jsonify({
        "status": "ok",
        "event_status": status,
        "badge": badge or None,
        "silent_removed": silent_removed,
        "rsvps_cleared": rsvps_cleared,
        "actioned": ok,
        "failed": failed,
        "skipped": skipped,
        "emails_sent": email_result.get("sent", 0),
        "emails_failed": email_result.get("failed", 0),
        "email_errors": email_result.get("errors", []),
        "players_without_email": no_email,
    })


@app.route("/api/events/<int:event_id>/cancel-apply", methods=["POST"])
@require_role("admin")
def api_cancel_apply(event_id):
    """Apply per-player actions from the one-by-one staging list.

    Body: { "actions": [{"item_id": 123, "action": "credit"|"refund"|"skip", "note": "..."}] }
    """
    data = request.get_json(silent=True) or {}
    actions = data.get("actions", [])
    if not actions:
        return jsonify({"error": "No actions provided"}), 400

    results = {"ok": [], "skipped": [], "failed": []}
    for entry in actions:
        item_id = entry.get("item_id")
        action = entry.get("action")
        note = (entry.get("note") or "Event cancelled").strip()
        method = (entry.get("method") or "").strip()

        if action == "skip":
            results["skipped"].append(item_id)
            continue
        try:
            if action == "credit":
                success = credit_item(item_id, note=note)
            elif action == "refund":
                success = refund_item(item_id, method=method, note=note)
            else:
                results["skipped"].append(item_id)
                continue
            if success:
                results["ok"].append(item_id)
            else:
                results["failed"].append({"id": item_id, "reason": "already actioned or not found"})
        except Exception as e:
            results["failed"].append({"id": item_id, "reason": str(e)})

    return jsonify({"status": "ok", "results": results})


@app.route("/api/players/<path:customer_name>/credits", methods=["GET"])
@require_role("manager")
def api_player_credits(customer_name):
    """Return all credited items for a player."""
    credits = get_player_credits(customer_name)
    total = sum(c["credit_amount"] for c in credits)
    return jsonify({"credits": credits, "total_credit": total})


@app.route("/api/events/<path:event_name>/rsvp-credits", methods=["GET"])
@require_role("manager")
def api_event_rsvp_credits(event_name):
    """Return credit info for all RSVP-only players in an event."""
    credit_map = get_event_rsvp_credit_map(event_name)
    return jsonify(credit_map)


@app.route("/api/events/<path:event_name>/balance-due-sends", methods=["GET"])
@require_role("manager")
def api_event_balance_due_sends(event_name):
    """Return {item_id: {sent_at, status, count}} map of balance-due email sends
    for this event. Frontend uses this to switch the button label from
    'Send Venmo Email' to 'Remind' on items that have already been emailed."""
    from email_parser.database import _connect as _db_connect
    out: dict[str, dict] = {}
    with _db_connect() as conn:
        rows = conn.execute(
            """SELECT body_preview, status, sent_at FROM message_log
               WHERE event_name = ?
                 AND body_preview LIKE 'Balance due%(item:%'
               ORDER BY sent_at ASC""",
            (event_name,),
        ).fetchall()
    for r in rows:
        bp = r["body_preview"] or ""
        # Extract "(item:N)" tag
        marker = bp.rfind("(item:")
        if marker < 0:
            continue
        end = bp.find(")", marker)
        if end < 0:
            continue
        item_id = bp[marker + 6:end].strip()
        if not item_id:
            continue
        cur = out.get(item_id)
        if cur is None:
            out[item_id] = {"sent_at": r["sent_at"], "status": r["status"], "count": 1}
        else:
            cur["count"] += 1
            cur["sent_at"] = r["sent_at"]  # keep latest
            cur["status"] = r["status"]
    return jsonify(out)


@app.route("/api/rsvps/<int:item_id>/credit-info", methods=["GET"])
@require_role("manager")
def api_rsvp_credit_info_by_item(item_id):
    """Return full credit analysis for an RSVP-only item (by items.id)."""
    from email_parser.database import _connect, _calc_event_pricing_breakdown
    with _connect() as conn:
        item = conn.execute("SELECT * FROM items WHERE id = ?", (item_id,)).fetchone()
        if not item:
            return jsonify({"error": "Item not found"}), 404
        item = dict(item)

        credits = get_player_credits(item["customer"],
                                     customer_id=item.get("customer_id"))
        if not credits:
            return jsonify({"error": "No credits on file for this player"}), 404

        total_credit = sum(c["credit_amount"] for c in credits)

        event_row = conn.execute(
            "SELECT * FROM events WHERE item_name = ? COLLATE NOCASE",
            (item["item_name"],),
        ).fetchone()
        event = dict(event_row) if event_row else {}

        most_recent = credits[0]
        # For non-combo events the holes value is locked to the event format.
        # Inheriting from the credited source item (e.g. a 9-hole credit applied
        # at an 18-hole event) was producing the wrong subtotal because the
        # PER_GAME_ADDON differs ($16 vs $30) — see _calc_event_pricing_breakdown.
        _evt_fmt = (event.get("format") or "")
        if _evt_fmt == "18 Holes":
            _holes = "18"
        elif _evt_fmt == "9 Holes":
            _holes = "9"
        else:
            _holes = most_recent.get("holes") or "9"
        prev = {
            "user_status": most_recent.get("user_status") or "MEMBER",
            "holes": _holes,
            "side_games": most_recent.get("side_games") or "NONE",
            "tee_choice": most_recent.get("tee_choice") or "",
        }

        breakdown = _calc_event_pricing_breakdown(
            event, prev["user_status"], prev["holes"], prev["side_games"]
        )
        event_price = breakdown["total"] if breakdown else None
        event_subtotal = breakdown["subtotal"] if breakdown else None
        # Balance due is computed against subtotal — difference is paid via Venmo (no tx fee)
        amount_owed = round((event_subtotal or 0.0) - total_credit, 2) if event_subtotal is not None else None

        venmo_username = None
        cust_id = item.get("customer_id")
        if cust_id:
            row = conn.execute(
                "SELECT venmo_username FROM customers WHERE customer_id = ?", (cust_id,)
            ).fetchone()
            if row:
                venmo_username = row["venmo_username"] or None
        if not venmo_username:
            row = conn.execute(
                """SELECT venmo_username FROM customers
                   WHERE TRIM(first_name||' '||last_name) = ? COLLATE NOCASE
                   AND venmo_username IS NOT NULL AND venmo_username != ''
                   LIMIT 1""",
                (item.get("customer", ""),),
            ).fetchone()
            if row:
                venmo_username = row["venmo_username"]

    return jsonify({
        "item_id": item_id,
        "customer": item["customer"],
        "credits": [
            {
                "id": c["id"],
                "item_name": c.get("item_name") or "",
                "event_name": c.get("item_name") or "",
                "origin_event": c.get("origin_event") or c.get("item_name") or "",
                "item_price": f"${c.get('credit_amount', 0):.2f}",
                "credit_amount": round(c.get("credit_amount") or 0, 2),
                "order_date": c.get("order_date") or "",
            }
            for c in credits
        ],
        "total_credit": total_credit,
        "event_price": event_price,
        "event_subtotal": event_subtotal,
        "amount_owed": amount_owed,
        "previous_selections": prev,
        "venmo_username": venmo_username,
    })


def _arm_excess_venmo_watch(result, data):
    """When Apply Credit & Register refunds the excess via Venmo, arm a refund
    watch on the new excess-credit item so the provider receipt auto-records
    the refund (Kerry 2026-07-16: the button should open Venmo and self-verify
    — no separate link click, no manual record). Fire-and-forget; the 2-min
    expense cycle backs up the ~75s/180s quick sweeps scheduled here."""
    if (data or {}).get("excess_action") != "venmo":
        return
    ecid = result.get("excess_credit_id")
    amount = result.get("excess")
    if not ecid or not amount or amount <= 0:
        return
    from email_parser.database import create_refund_watch
    ev = (data or {}).get("excess_venmo") or {}
    handle = (ev.get("handle") or "").strip().lstrip("@")
    memo = (ev.get("memo") or "").strip()
    try:
        create_refund_watch(ecid, method="Venmo", amount=amount,
                            memo=memo, handle=handle)
    except Exception:
        logger.warning("apply-credit excess Venmo watch failed for item %s",
                       ecid, exc_info=True)
        return
    result["excess_venmo_watch_armed"] = True
    if getattr(scheduler, "running", False):
        now = datetime.now()
        for jid, secs in (("venmo_quick_check_a", 75), ("venmo_quick_check_b", 180)):
            try:
                scheduler.add_job(_quick_expense_check, "date",
                                  run_date=now + timedelta(seconds=secs),
                                  id=jid, replace_existing=True, coalesce=True,
                                  misfire_grace_time=120)
            except Exception:
                logger.exception("Failed to schedule quick receipt check %s", jid)


@app.route("/api/rsvps/<int:item_id>/apply-credit", methods=["POST"])
@require_role("manager")
def api_apply_credit_to_rsvp(item_id):
    """Apply a player's credits to their RSVP-only registration."""
    data = request.get_json(silent=True) or {}
    credited_item_ids = data.get("credited_item_ids", [])
    excess_action = data.get("excess_action", "keep")
    holes = data.get("holes", "")
    side_games = data.get("side_games", "")
    tee_choice = data.get("tee_choice", "")
    user_status = data.get("user_status", "")

    if not credited_item_ids:
        return jsonify({"error": "credited_item_ids required"}), 400
    if excess_action not in ("keep", "note", "venmo"):
        return jsonify({"error": "excess_action must be 'keep', 'note', or 'venmo'"}), 400

    result = apply_credit_to_rsvp(
        rsvp_item_id=item_id,
        credited_item_ids=credited_item_ids,
        excess_action=excess_action,
        holes=holes,
        side_games=side_games,
        tee_choice=tee_choice,
        user_status=user_status,
    )
    if not result.get("ok"):
        return jsonify({"error": result.get("error", "Failed")}), 400
    _arm_excess_venmo_watch(result, data)
    _send_credit_entry_confirmation(item_id, result)
    result["balance_email"] = _maybe_auto_send_balance_email(item_id, result, data)
    return jsonify(result)


@app.route("/api/rsvps/trigger-credit-alerts", methods=["POST"])
@require_role("admin")
def api_trigger_credit_alerts():
    """Manually trigger credit alert email scan (for testing)."""
    try:
        _send_rsvp_credit_alerts()
        return jsonify({"status": "ok"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/rsvps/gg/<int:rsvp_id>/credit-info", methods=["GET"])
@require_role("manager")
def api_gg_rsvp_credit_info(rsvp_id):
    """Full credit analysis for a GG RSVP (by rsvps.id, not items.id).

    Delegates to get_rsvp_credit_info — the SAME code path that renders
    the roster CREDIT badge and sends the credit-alert emails — instead
    of a hand-rolled duplicate of its resolution/pricing logic. The
    duplicate drifted twice in one day (email-only resolution, then a
    partial customer_id fix) while the DB function worked the whole time
    (Kerry 2026-07-20: Anthis's badge + alert email fine, modal 404).
    """
    from email_parser.database import _connect, get_rsvp_credit_info
    info = get_rsvp_credit_info(rsvp_id)
    if not info:
        return jsonify({"error": "No credits on file for this player "
                                 "(or the RSVP isn't matched to an event)"}), 404

    # Venmo handle for the excess-refund deep link
    venmo_username = None
    with _connect() as conn:
        row = conn.execute(
            """SELECT venmo_username FROM customers
               WHERE TRIM(first_name||' '||last_name) = ? COLLATE NOCASE
               AND venmo_username IS NOT NULL AND venmo_username != ''
               LIMIT 1""",
            (info.get("player_name") or "",),
        ).fetchone()
        if row:
            venmo_username = row["venmo_username"]
        if not venmo_username and info.get("player_email"):
            row = conn.execute(
                """SELECT c.venmo_username FROM customers c
                   JOIN customer_emails ce ON ce.customer_id = c.customer_id
                   WHERE LOWER(ce.email) = ?
                   AND c.venmo_username IS NOT NULL AND c.venmo_username != ''
                   LIMIT 1""",
                ((info.get("player_email") or "").strip().lower(),),
            ).fetchone()
            if row:
                venmo_username = row["venmo_username"]

    return jsonify({
        "rsvp_id": rsvp_id,
        "customer": info.get("player_name"),
        "credits": [
            {**c, "item_price": f"${(c.get('credit_amount') or 0):.2f}"}
            for c in info.get("credits", [])
        ],
        "total_credit": info.get("total_credit"),
        "event_price": info.get("new_event_price"),
        "event_subtotal": info.get("new_event_subtotal"),
        "amount_owed": info.get("amount_owed"),
        "previous_selections": info.get("selections"),
        "venmo_username": venmo_username,
    })


@app.route("/api/rsvps/gg/<int:rsvp_id>/apply-credit", methods=["POST"])
@require_role("manager")
def api_gg_rsvp_apply_credit(rsvp_id):
    """Apply credits for a GG RSVP (synthetic row): creates rsvp_only item then applies credit."""
    from email_parser.database import _connect
    data = request.get_json(silent=True) or {}
    credited_item_ids = data.get("credited_item_ids", [])
    excess_action = data.get("excess_action", "keep")
    holes = data.get("holes", "")
    side_games = data.get("side_games", "")
    tee_choice = data.get("tee_choice", "")
    user_status = data.get("user_status", "")

    if not credited_item_ids:
        return jsonify({"error": "credited_item_ids required"}), 400

    with _connect() as conn:
        rsvp = conn.execute("SELECT * FROM rsvps WHERE id = ?", (rsvp_id,)).fetchone()
        if not rsvp:
            return jsonify({"error": "RSVP not found"}), 404
        rsvp = dict(rsvp)

    event_name = rsvp.get("matched_event") or ""
    player_email = (rsvp.get("player_email") or "").strip().lower()
    player_name = rsvp.get("player_name") or ""

    # Resolve canonical customer name via email so the created item uses the proper name
    canonical_name = player_name
    if player_email:
        with _connect() as conn:
            card = conn.execute(
                """SELECT customer FROM items WHERE LOWER(customer_email) = ?
                   AND customer IS NOT NULL AND customer != ''
                   ORDER BY order_date DESC LIMIT 1""",
                (player_email,),
            ).fetchone()
            if card:
                canonical_name = card["customer"]

    # Create the rsvp_only item (idempotent)
    new_item_id = create_rsvp_only_item(
        event_name=event_name,
        player_name=canonical_name,
        player_email=player_email,
        rsvp_id=rsvp_id,
    )

    result = apply_credit_to_rsvp(
        rsvp_item_id=new_item_id,
        credited_item_ids=credited_item_ids,
        excess_action=excess_action,
        holes=holes,
        side_games=side_games,
        tee_choice=tee_choice,
        user_status=user_status,
    )
    if not result.get("ok"):
        return jsonify({"error": result.get("error", "Failed")}), 400
    _arm_excess_venmo_watch(result, data)
    _send_credit_entry_confirmation(new_item_id, result)
    result["balance_email"] = _maybe_auto_send_balance_email(new_item_id, result, data)
    return jsonify({**result, "item_id": new_item_id})


@app.route("/api/items/<int:item_id>/reverse-credit-application", methods=["POST"])
@require_role("manager")
def api_reverse_credit_application(item_id):
    """Undo a credit application: restore source credits, revert registration to RSVP."""
    result = reverse_credit_application(item_id)
    if not result.get("ok"):
        return jsonify({"error": result.get("error", "Failed")}), 400
    return jsonify(result)


# ─────────────────────────────────────────────────────────────────────────────
# Balance-due email (Venmo prefilled link)
# ─────────────────────────────────────────────────────────────────────────────

VENMO_RECIPIENT_USERNAME = "tgf-payments"


@app.route("/pay/venmo")
def pay_venmo_bounce():
    """Deep-link bounce for Venmo links embedded in EMAILS.

    Kerry 2026-07-15 (Richard Palacios's balance-due payment): the
    https venmo.com universal link renders a literal '+' for every memo
    space no matter how it's encoded — the same quirk that made the
    in-app pay buttons use the native venmo:// scheme (2026-07-08).
    Emails can't link venmo:// directly (Gmail strips app-scheme
    hrefs), so they link HERE: this page fires the app scheme (where
    %20 decodes to real spaces) and keeps the web link + a copyable
    memo as fallback. No auth — it lives in members' inboxes and serves
    nothing beyond what the link itself carries.
    """
    import html as _html
    from urllib.parse import quote as _q
    to = re.sub(r"[^A-Za-z0-9_.@-]", "", (request.args.get("to") or ""))[:60].lstrip("@")
    note = (request.args.get("note") or "")[:200]
    try:
        amount = round(float(request.args.get("amount") or 0), 2)
    except (TypeError, ValueError):
        amount = 0
    if not to or amount <= 0:
        return Response("Invalid payment link.", status=400)
    amt = f"{amount:.2f}"
    app_link = (f"venmo://paycharge?txn=pay&recipients={_q(to)}"
                f"&amount={amt}&note={_q(note)}")
    web_link = (f"https://venmo.com/{_q(to)}?txn=pay&amount={amt}"
                f"&note={_q(note)}")
    e = _html.escape
    page = f"""<!DOCTYPE html><html><head>
<meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">
<title>Opening Venmo…</title></head>
<body style="font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif;
             max-width:420px;margin:3rem auto;padding:0 1rem;color:#0f172a;text-align:center;">
<h2 style="margin-bottom:0.25rem;">Opening Venmo…</h2>
<p style="color:#475569;">Pay <strong>${e(amt)}</strong> to <strong>@{e(to)}</strong></p>
<p style="background:#f1f5f9;border-radius:8px;padding:0.6rem 0.8rem;font-size:0.9rem;">Memo: {e(note)}</p>
<p><a href="{e(app_link)}" style="display:inline-block;background:#3D95CE;color:#fff;text-decoration:none;
    padding:0.75rem 1.5rem;border-radius:8px;font-weight:600;">Open the Venmo app</a></p>
<p style="font-size:0.85rem;color:#64748b;">App didn't open?
  <a href="{e(web_link)}">Pay on venmo.com</a> (the memo may lose its spaces there — copy it from above).</p>
<script>setTimeout(function() {{ window.location.href = {json.dumps(app_link)}; }}, 400);</script>
</body></html>"""
    return Response(page, mimetype="text/html")


def _build_balance_due_email(item_id: int) -> dict | None:
    """Assemble the balance-due email payload for a credit-transfer item.

    Returns a dict with subject, html_body, recipient (intended), override_recipient
    (testing destination), venmo_link, amount_owed, player_name, event_name.
    Returns None if the item is not a credit-transfer with a positive balance_due.
    """
    from urllib.parse import quote
    from email_parser.database import _connect as _db_connect

    with _db_connect() as conn:
        item = conn.execute("SELECT * FROM items WHERE id = ?", (item_id,)).fetchone()
        if not item:
            return {"error": "Item not found"}
        item = dict(item)

        if (item.get("merchant") or "") != "Paid Separately (Credit Transfer)":
            return {"error": "Item is not a credit transfer"}
        cnote = item.get("credit_note") or ""
        if not cnote.startswith("balance_due:"):
            return {"error": "No balance due on this item"}
        try:
            amount_owed = float(cnote.split(":", 1)[1])
        except (ValueError, IndexError):
            return {"error": "Could not parse balance amount"}
        if amount_owed <= 0:
            return {"error": "Balance is zero — nothing owed"}

        event_name = item.get("item_name") or ""
        ev_row = conn.execute(
            "SELECT * FROM events WHERE item_name = ? COLLATE NOCASE", (event_name,)
        ).fetchone()
        event = dict(ev_row) if ev_row else {}

    player_name = (item.get("customer") or "").strip()
    # Customer Info primary email is canonical; only fall back to the
    # historical items.customer_email if no primary is on file.
    try:
        player_email = _resolve_player_email(item)
    except Exception:
        logger.warning("_build_balance_due_email: email resolution failed for %s", player_name, exc_info=True)
        player_email = (item.get("customer_email") or "").strip()

    first_name = player_name.split(" ", 1)[0] if player_name else "there"
    event_date = event.get("event_date") or ""
    course = event.get("course") or ""

    memo = f"{player_name} - Balance due for {event_name}"
    # https venmo.com links render the note's encoded spaces as literal '+'
    # chars in the prefilled memo no matter how they're encoded (Kerry,
    # 2026-07-15). Route through the /pay/venmo bounce page, which fires the
    # native venmo:// scheme (decodes %20 correctly) with a venmo.com
    # fallback for desktop. Gmail strips venmo:// hrefs, so the email link
    # must be https — the bounce page is the bridge.
    base_url = os.getenv("APP_BASE_URL", "https://tgf-tracker.up.railway.app").rstrip("/")
    venmo_link = (
        f"{base_url}/pay/venmo?to={quote(VENMO_RECIPIENT_USERNAME)}"
        f"&amount={amount_owed:.2f}&note={quote(memo)}"
    )

    subject = f"Balance due for {event_name} — ${amount_owed:.2f}"

    when_line = ""
    if event_date or course:
        bits = [b for b in [event_date, course] if b]
        when_line = f"<p style='color:#475569; margin:0.25rem 0;'>{' &middot; '.join(bits)}</p>"

    html_body = f"""
<div style="font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif;
            max-width:560px; color:#0f172a; line-height:1.5;">
  <p>Hi {first_name},</p>
  <p>Thanks for RSVPing to <strong>{event_name}</strong>!</p>
  {when_line}
  <p>We applied your existing credit toward this event. The balance remaining is
  <strong style="color:#b45309;">${amount_owed:.2f}</strong>.</p>

  <p>Please send <strong>${amount_owed:.2f}</strong> via Venmo to
  <strong>@{VENMO_RECIPIENT_USERNAME}</strong>. Tap the button below to open Venmo
  with the amount and memo prefilled:</p>

  <p style="margin:1.25rem 0;">
    <a href="{venmo_link}"
       style="display:inline-block; background:#3D95CE; color:#fff; text-decoration:none;
              padding:0.75rem 1.5rem; border-radius:6px; font-weight:600;
              font-size:1rem;">
      Pay ${amount_owed:.2f} on Venmo
    </a>
  </p>

  <p style="font-size:0.85rem; color:#64748b;">
    If the button doesn&apos;t open Venmo, send the payment manually to
    <strong>@{VENMO_RECIPIENT_USERNAME}</strong> with the memo:<br>
    <code style="background:#f1f5f9; padding:2px 6px; border-radius:4px;">{memo}</code>
  </p>

  <p>See you on the course!</p>
  <p style="color:#64748b;">— The Golf Fellowship</p>
</div>"""

    intended_recipient = player_email
    # Set BALANCE_DUE_EMAIL_OVERRIDE in Railway env to redirect all balance-due
    # emails to a testing address. Unset = live (sends to the player).
    override_recipient = (os.getenv("BALANCE_DUE_EMAIL_OVERRIDE") or "").strip()
    override_active = bool(override_recipient)

    return {
        "item_id": item_id,
        "player_name": player_name,
        "player_email": player_email,
        "event_name": event_name,
        "event_date": event_date,
        "course": course,
        "amount_owed": amount_owed,
        "venmo_link": venmo_link,
        "memo": memo,
        "subject": subject,
        "html_body": html_body,
        "intended_recipient": intended_recipient,
        "override_recipient": override_recipient,
        "override_active": override_active,
    }


@app.route("/api/items/<int:item_id>/balance-due-email/preview", methods=["GET"])
@require_role("manager")
def api_balance_due_email_preview(item_id):
    """Preview the balance-due Venmo email for a credit-transfer item."""
    result = _build_balance_due_email(item_id)
    if not result:
        return jsonify({"error": "Failed to build email"}), 400
    if "error" in result:
        return jsonify(result), 400
    return jsonify(result)


def _send_credit_entry_confirmation(item_id, result):
    """Auto-send the entry-confirmation email when the applied credit covered
    the WHOLE fee (amount_owed <= 0). Delegates to the shared db helper. Never
    raises — a mail failure must not break the apply-credit response. Kill
    switch: AUTO_CREDIT_ENTRY_EMAIL=0.

    RESTORED v2.139.4: v2.129.25's edit dropped this def while keeping both
    call sites, so every Apply Credit since 2026-07-20 hit a NameError AFTER
    the DB write — the credit applied but the modal showed 'Internal server
    error' and the balance-due auto-email never ran (the Fieber case)."""
    try:
        if (os.getenv("AUTO_CREDIT_ENTRY_EMAIL", "1") or "1").strip().lower() \
                not in ("1", "true", "yes", "on"):
            return
        from email_parser.database import send_entry_confirmation_email
        send_entry_confirmation_email(item_id, result)
    except Exception:
        logger.warning("credit-entry confirmation email failed for item %s",
                       item_id, exc_info=True)


def _maybe_auto_send_balance_email(item_id: int, result: dict,
                                   data: dict) -> dict | None:
    """Auto-send the standard balance-due Venmo email after Apply Credit
    when the member still owes money (Kerry 2026-07-20: 'auto-send the
    standard email we created with the prepared Venmo link, with an
    option to uncheck'). The modal passes auto_email; default ON."""
    try:
        if not data.get("auto_email", True):
            return {"skipped": "unchecked"}
        owed = result.get("remaining_owed")
        if owed is None:
            owed = result.get("amount_owed") or 0
        if float(owed) <= 0:
            return None
        return _send_balance_due_email_now(item_id)
    except Exception:
        logger.warning("auto balance-due email failed for item %s",
                       item_id, exc_info=True)
        return {"error": "auto-send failed"}


def _send_balance_due_email_now(item_id: int) -> dict:
    """Build + send the standard balance-due Venmo email for an item.
    Shared by the manual send route and the Apply Credit auto-send
    (Kerry 2026-07-20: when the member owes more, auto-send the
    standard email with the prepared Venmo link)."""
    payload = _build_balance_due_email(item_id)
    if not payload:
        return {"error": "Failed to build email"}
    if "error" in payload:
        return payload

    tenant_id = os.getenv("AZURE_TENANT_ID")
    client_id = os.getenv("AZURE_CLIENT_ID")
    client_secret = os.getenv("AZURE_CLIENT_SECRET")
    from_address = os.getenv("EMAIL_ADDRESS")
    if not all([tenant_id, client_id, client_secret, from_address]):
        return {"error": "Email credentials not configured on server"}

    # During testing, route to admin override; remove env var to send to player.
    to_address = payload["override_recipient"] if payload["override_active"] else payload["intended_recipient"]
    if not to_address:
        return {"error": "No recipient address available"}

    ok = send_mail_graph(
        tenant_id=tenant_id,
        client_id=client_id,
        client_secret=client_secret,
        from_address=from_address,
        to_address=to_address,
        subject=payload["subject"],
        html_body=payload["html_body"],
    )

    status = "sent" if ok else "failed"
    try:
        from email_parser.database import log_message
        log_message({
            "event_name": payload["event_name"],
            "channel": "email",
            "recipient_name": payload["player_name"],
            "recipient_address": to_address,
            "subject": payload["subject"],
            # item_id tag enables per-item lookup in get_balance_due_sends_for_event
            "body_preview": f"Balance due ${payload['amount_owed']:.2f} via Venmo (item:{item_id})",
            "status": status,
            "sent_by": session.get("role", "unknown"),
        })
    except Exception:
        logger.warning("Failed to log balance-due email", exc_info=True)

    return {"status": status, "to": to_address,
            "amount_owed": payload["amount_owed"]}


@app.route("/api/items/<int:item_id>/balance-due-email/send", methods=["POST"])
@require_role("manager")
def api_balance_due_email_send(item_id):
    """Send the balance-due Venmo email. During testing, sends to override recipient."""
    res = _send_balance_due_email_now(item_id)
    if res.get("error"):
        code = 500 if "credentials" in res["error"] else 400
        return jsonify(res), code
    return jsonify(res)


@app.route("/api/items/<int:item_id>/entry-confirmation/send", methods=["POST"])
@require_role("manager")
def api_send_entry_confirmation(item_id):
    """Manually (re)send the entry-confirmation email for a registered item —
    the retroactive/resend path for registrations made before the auto-email,
    or a re-send on request (Kerry 2026-07-16). force=True skips the balance
    guard."""
    from email_parser.database import send_entry_confirmation_email
    res = send_entry_confirmation_email(item_id, force=True)
    if not res.get("ok"):
        return jsonify({"error": res.get("error", "Failed to send")}), 400
    return jsonify({"status": "ok", **res})


@app.route("/api/items/<int:credit_item_id>/apply-credit-info")
@require_role("manager")
def api_credit_item_apply_info(credit_item_id):
    """Return price info for applying a credit item to a selected event."""
    from email_parser.database import _connect, _calc_event_pricing_breakdown, _parse_dollar
    event_name = request.args.get("event_name", "").strip()
    with _connect() as conn:
        item = conn.execute("SELECT * FROM items WHERE id = ?", (credit_item_id,)).fetchone()
        if not item:
            return jsonify({"error": "Item not found"}), 404
        item = dict(item)
        if item.get("transaction_status") != "credited":
            return jsonify({"error": "Item is not a credit"}), 400
        credit_amount = _parse_dollar(item.get("item_price")) + _parse_dollar(item.get("transaction_fees") or "0")
        event_price = None
        event_subtotal = None
        amount_owed = None
        if event_name:
            ev_row = conn.execute(
                "SELECT * FROM events WHERE item_name = ? COLLATE NOCASE", (event_name,)
            ).fetchone()
            event = dict(ev_row) if ev_row else {}
            breakdown = _calc_event_pricing_breakdown(
                event,
                item.get("user_status") or "MEMBER",
                item.get("holes") or "9",
                item.get("side_games") or "NONE",
            )
            if breakdown is not None:
                event_price = breakdown["total"]
                event_subtotal = breakdown["subtotal"]
                # Balance due is computed against subtotal — paid via Venmo (no tx fee)
                amount_owed = round(event_subtotal - credit_amount, 2)
        # Look up the customer's Venmo handle for excess-refund link generation
        venmo_username = None
        cust_id = item.get("customer_id")
        if cust_id:
            row = conn.execute(
                "SELECT venmo_username FROM customers WHERE customer_id = ?", (cust_id,)
            ).fetchone()
            if row:
                venmo_username = row["venmo_username"] or None
    return jsonify({
        "credit_item_id": credit_item_id,
        "customer": item.get("customer"),
        "source_event": item.get("item_name"),
        "credit_amount": credit_amount,
        "event_price": event_price,
        "event_subtotal": event_subtotal,
        "amount_owed": amount_owed,
        "venmo_username": venmo_username,
        "previous_selections": {
            "user_status": item.get("user_status") or "MEMBER",
            "holes": item.get("holes") or "9",
            "side_games": item.get("side_games") or "NONE",
            "tee_choice": item.get("tee_choice") or "",
        },
    })


@app.route("/api/items/<int:credit_item_id>/apply-to-event", methods=["POST"])
@require_role("manager")
def api_apply_credit_item_to_event(credit_item_id):
    """Apply an existing credit item to register a player in a different event."""
    from email_parser.database import _connect
    data = request.get_json(silent=True) or {}
    event_name = (data.get("event_name") or "").strip()
    if not event_name:
        return jsonify({"error": "event_name required"}), 400
    excess_action = data.get("excess_action", "keep")
    holes = data.get("holes", "")
    side_games = data.get("side_games", "")
    tee_choice = data.get("tee_choice", "")
    user_status = data.get("user_status", "")

    with _connect() as conn:
        item = conn.execute("SELECT * FROM items WHERE id = ?", (credit_item_id,)).fetchone()
        if not item:
            return jsonify({"error": "Item not found"}), 404
        item = dict(item)
        if item.get("transaction_status") != "credited":
            return jsonify({"error": "Item is not a credit"}), 400

    player_name = item.get("customer") or ""
    player_email = item.get("customer_email") or ""
    uid = f"manual-credit-{credit_item_id}"

    from email_parser.database import _resolve_or_create_customer
    with _connect() as conn:
        existing = conn.execute("SELECT id FROM items WHERE email_uid = ?", (uid,)).fetchone()
        if existing:
            new_item_id = existing["id"]
        else:
            ev_row = conn.execute(
                "SELECT chapter FROM events WHERE item_name = ? COLLATE NOCASE", (event_name,)
            ).fetchone()
            chapter = (ev_row["chapter"] if ev_row else "") or ""
            # This synthetic rsvp_only row has no transaction to trigger the
            # normal _resolve_or_create_customer() call, so resolve/create the
            # canonical customers row now — otherwise customer_id stays NULL
            # forever (Membership Terms, roles, and status edits all require
            # it), and apply_credit_to_rsvp() runs against this row immediately.
            cid = _resolve_or_create_customer(conn, player_name, player_email or None)
            conn.execute(
                """INSERT INTO items (email_uid, merchant, customer, customer_email, item_name,
                   item_price, transaction_status, order_date, chapter, customer_id)
                   VALUES (?, 'Golf Genius RSVP', ?, ?, ?, '', 'rsvp_only', date('now'), ?, ?)""",
                (uid, player_name, player_email, event_name, chapter, cid),
            )
            conn.commit()
            new_item_id = conn.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]

    result = apply_credit_to_rsvp(
        rsvp_item_id=new_item_id,
        credited_item_ids=[credit_item_id],
        excess_action=excess_action,
        holes=holes or item.get("holes") or "",
        side_games=side_games or item.get("side_games") or "",
        tee_choice=tee_choice or item.get("tee_choice") or "",
        user_status=user_status or item.get("user_status") or "",
    )
    if not result.get("ok"):
        with _connect() as conn:
            conn.execute(
                "DELETE FROM items WHERE email_uid = ? AND transaction_status = 'rsvp_only'", (uid,)
            )
            conn.commit()
        return jsonify({"error": result.get("error", "Failed")}), 400
    return jsonify({**result, "item_id": new_item_id})


@app.route("/api/events", methods=["POST"])
@require_role("manager")
def api_create_event():
    """Manually create a new event."""
    data = request.get_json(silent=True)
    if not data or not data.get("item_name"):
        return jsonify({"error": "item_name is required."}), 400
    event = create_event(
        item_name=data["item_name"],
        event_date=data.get("event_date"),
        course=data.get("course"),
        chapter=data.get("chapter"),
        format=data.get("format"),
        start_type=data.get("start_type"),
        start_time=data.get("start_time"),
        tee_time_count=data.get("tee_time_count"),
        tee_time_interval=data.get("tee_time_interval"),
        start_time_18=data.get("start_time_18"),
        start_type_18=data.get("start_type_18"),
        tee_time_count_18=data.get("tee_time_count_18"),
        tee_direction=data.get("tee_direction"),
        tee_direction_18=data.get("tee_direction_18"),
        nine_side=data.get("nine_side"),
        course_cost=data.get("course_cost"),
        tgf_markup=data.get("tgf_markup"),
        side_game_fee=data.get("side_game_fee"),
        transaction_fee_pct=data.get("transaction_fee_pct"),
        course_cost_9=data.get("course_cost_9"),
        course_cost_18=data.get("course_cost_18"),
        tgf_markup_9=data.get("tgf_markup_9"),
        tgf_markup_18=data.get("tgf_markup_18"),
        side_game_fee_9=data.get("side_game_fee_9"),
        side_game_fee_18=data.get("side_game_fee_18"),
        tgf_markup_final=data.get("tgf_markup_final"),
        tgf_markup_final_9=data.get("tgf_markup_final_9"),
        tgf_markup_final_18=data.get("tgf_markup_final_18"),
        course_cost_breakdown=data.get("course_cost_breakdown"),
        course_cost_breakdown_9=data.get("course_cost_breakdown_9"),
        course_cost_breakdown_18=data.get("course_cost_breakdown_18"),
        per_game_addon=data.get("per_game_addon"),
    )
    if event:
        # allow_fivesomes rides on update_event rather than widening
        # create_event's 30-argument signature for one boolean.
        if data.get("allow_fivesomes") and event.get("id"):
            try:
                update_event(event["id"],
                             {"allow_fivesomes": 1 if data["allow_fivesomes"] else 0})
                event["allow_fivesomes"] = 1
            except Exception:
                logger.exception("Could not set allow_fivesomes on new event")
        return jsonify({"status": "ok", "event": event}), 201
    return jsonify({"error": "Event already exists with that name."}), 409


@app.route("/api/events/merge", methods=["POST"])
@require_role("admin")
def api_merge_events():
    """Merge source event into target event. Admin only.

    All items, RSVPs, and overrides from the source event are reassigned
    to the target event, then the source event is deleted.
    """
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "JSON body required."}), 400
    source_id = data.get("source_id")
    target_id = data.get("target_id")
    if not source_id or not target_id:
        return jsonify({"error": "source_id and target_id are required."}), 400
    if source_id == target_id:
        return jsonify({"error": "Cannot merge an event into itself."}), 400
    result = merge_events(source_id, target_id)
    if result:
        return jsonify({"status": "ok", **result})
    return jsonify({"error": "Source or target event not found."}), 404


@app.route("/api/events/mvp-unlinks")
@require_role("manager")
def api_mvp_unlinks():
    """Return list of event names explicitly unlinked from same-day TGF MVP combining."""
    return jsonify(get_mvp_unlinked_events())


@app.route("/api/events/mvp-unlink", methods=["POST"])
@require_role("admin")
def api_mvp_unlink():
    """Unlink an event from same-day TGF MVP combining."""
    data = request.get_json(silent=True)
    if not data or not data.get("event_name"):
        return jsonify({"error": "event_name required"}), 400
    set_mvp_unlink(data["event_name"], unlink=True)
    return jsonify({"status": "ok"})


@app.route("/api/events/mvp-relink", methods=["POST"])
@require_role("admin")
def api_mvp_relink():
    """Re-link a previously unlinked event for same-day TGF MVP combining."""
    data = request.get_json(silent=True)
    if not data or not data.get("event_name"):
        return jsonify({"error": "event_name required"}), 400
    set_mvp_unlink(data["event_name"], unlink=False)
    return jsonify({"status": "ok"})


@app.route("/api/events/tgf-mvp")
@require_role("manager")
def api_tgf_mvp_determination():
    """Compute City MVP per linked same-day event (highest net Stableford
    points among NET buyers, ratified tiebreakers) and the TGF MVP."""
    event = (request.args.get("event") or "").strip()
    if not event:
        return jsonify({"error": "event parameter required"}), 400
    try:
        from email_parser.database import determine_tgf_mvp
        return jsonify(determine_tgf_mvp(event))
    except Exception as e:
        logger.exception("TGF MVP determination failed")
        return jsonify({"error": str(e)}), 500


@app.route("/api/events/game-results")
@require_role("manager")
def api_event_game_results():
    """Shadow-computed side-game winners from our imported scorecards
    (individual_net / individual_gross / skins). `flights` is the prize
    matrix's flight count for this game at the event's buyer count —
    the Games tab (which owns the matrix amounts) passes it in."""
    event = (request.args.get("event") or "").strip()
    game = (request.args.get("game") or "").strip()
    flights = request.args.get("flights", default=1, type=int)
    if not event or not game:
        return jsonify({"error": "event and game parameters required"}), 400
    try:
        from email_parser.database import determine_event_game_results
        return jsonify(determine_event_game_results(event, game,
                                                    flights=max(1, flights)))
    except Exception as e:
        logger.exception("Game results determination failed")
        return jsonify({"error": str(e)}), 500


@app.route("/api/events/record-game-payouts", methods=["POST"])
@require_role("admin")
def api_record_game_payouts():
    """Record the Games tab's determined winners into the PAYOUTS tab
    (tgf_events/tgf_payouts + ledger reconciliation). Every row ties to
    a customer via the payout name-resolution cascade; team payouts
    arrive pre-split per member."""
    d = request.get_json(silent=True) or {}
    event = (d.get("event_name") or "").strip()
    payouts = d.get("payouts") or []
    if not event:
        return jsonify({"error": "event_name required"}), 400
    for p in payouts:
        if not p.get("golferName") or not isinstance(p.get("amount"), (int, float)):
            return jsonify({"error": "each payout needs golferName and numeric amount"}), 400
    try:
        from email_parser.database import (assemble_event_game_payouts,
                                           record_event_game_payouts)
        if not payouts:
            # assemble server-side (the normal path — single source of truth)
            asm = assemble_event_game_payouts(event)
            if asm.get("error"):
                return jsonify(asm), 400
            payouts = asm["rows"]
            if not payouts:
                return jsonify({"error": "nothing determined to record",
                                "notes": asm.get("notes")}), 400
        result = record_event_game_payouts(event, payouts,
                                           force=bool(d.get("force")))
    except Exception as e:
        logger.exception("Record game payouts failed")
        return jsonify({"error": str(e)}), 500
    status = 200
    if result.get("error"):
        status = 409 if result.get("needs_force") else 400
    return jsonify(result), status


@app.route("/api/events/game-payouts-preview")
@require_role("admin")
def api_game_payouts_preview():
    """Server-assembled payout rows for one event (no write) — feeds the
    Record Payouts confirm dialog."""
    event = (request.args.get("event") or "").strip()
    if not event:
        return jsonify({"error": "event parameter required"}), 400
    try:
        from email_parser.database import assemble_event_game_payouts
        return jsonify(assemble_event_game_payouts(event))
    except Exception as e:
        logger.exception("Payout preview failed")
        return jsonify({"error": str(e)}), 500


@app.route("/api/events/gg-game-results")
@require_role("manager")
def api_gg_game_results():
    """GG-RECORDED winners for one event (CTP / Longest Putt / HIO /
    TEAM Net — manually entered into GG post-round; the portal is the
    source of record, pulled by import_gg_game_results)."""
    event = (request.args.get("event") or "").strip()
    if not event:
        return jsonify({"error": "event parameter required"}), 400
    try:
        from email_parser.database import get_gg_game_results
        return jsonify(get_gg_game_results(event))
    except Exception as e:
        logger.exception("GG game results read failed")
        return jsonify({"error": str(e)}), 500


@app.route("/api/events/orphaned-items")
@require_role("manager")
def api_orphaned_items():
    """Return items whose item_name doesn't match any event."""
    return jsonify(get_orphaned_items())


@app.route("/api/sunset")
@require_role("view-only")
def api_sunset():
    """Return sunset and civil twilight times for a chapter + date, in Central Time."""
    import pytz
    import requests as _requests
    from datetime import datetime as _dt

    # TODO: pull chapter coordinates from a chapters table when full platform is built
    CHAPTER_COORDS = {
        "San Antonio": (29.4241, -98.4936),
        "Austin": (30.2672, -97.7431),
    }

    date_str = request.args.get("date")
    chapter = request.args.get("chapter")
    if not date_str or not chapter:
        return jsonify({"error": "date and chapter are required"}), 400
    coords = CHAPTER_COORDS.get(chapter)
    if not coords:
        return jsonify({"error": f"Unknown chapter: {chapter}"}), 400

    try:
        resp = _requests.get(
            "https://api.sunrise-sunset.org/json",
            params={"lat": coords[0], "lng": coords[1], "date": date_str, "formatted": 0},
            timeout=10,
        )
        data = resp.json()
        if data.get("status") != "OK":
            return jsonify({"error": "Sunrise-Sunset API error"}), 502

        results = data["results"]
        central = pytz.timezone("America/Chicago")

        def to_central_12h(iso_str):
            utc_dt = _dt.fromisoformat(iso_str.replace("Z", "+00:00"))
            local_dt = utc_dt.astimezone(central)
            return local_dt.strftime("%-I:%M %p"), local_dt.strftime("%H:%M")

        sunset_12h, sunset_24h = to_central_12h(results["sunset"])
        twilight_12h, twilight_24h = to_central_12h(results["civil_twilight_end"])

        return jsonify({
            "sunset": sunset_12h,
            "sunset_24h": sunset_24h,
            "civil_twilight_end": twilight_12h,
            "civil_twilight_end_24h": twilight_24h,
        })
    except _requests.RequestException:
        return jsonify({"error": "Failed to reach Sunrise-Sunset API"}), 502
    except Exception as exc:
        logger.exception("Sunset API error")
        return jsonify({"error": str(exc)}), 500


@app.route("/api/events/resolve-orphan", methods=["POST"])
@require_role("admin")
def api_resolve_orphan():
    """Reassign orphaned items to an existing event. Admin only."""
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "JSON body required."}), 400
    old_name = data.get("old_item_name")
    target = data.get("target_event")
    if not old_name or not target:
        return jsonify({"error": "old_item_name and target_event are required."}), 400
    result = resolve_orphaned_items(old_name, target)
    return jsonify({"status": "ok", **result})


@app.route("/api/parse-warnings")
@require_role("manager")
def api_parse_warnings():
    """Return open parse warnings (items flagged during parsing)."""
    status = request.args.get("status", "open")
    return jsonify(get_parse_warnings(status))


@app.route("/api/action-items")
@require_role("manager")
def api_notification_action_items():
    """Return pending action items for admin/manager review.

    Aggregates parse warnings and GUEST registrations needing guest names.
    """
    items = []

    # 1. Parse warnings
    for w in get_parse_warnings("open"):
        items.append({
            "type": "parse_warning",
            "id": f"pw-{w['id']}",
            "pw_id": w["id"],
            "title": w.get("item_name") or "Unknown item",
            "message": w.get("message", ""),
            "customer": w.get("customer"),
            "order_id": w.get("order_id"),
            "item_id": w.get("item_id"),
            "code": w.get("warning_code"),
            "created_at": w.get("created_at"),
        })

    # 2. GUEST registrations needing guest name assignment
    #    Only flag when: same buyer has another item in the same order (multi-item purchase)
    #    AND no guest_name or partner_request is available to identify the guest.
    conn = get_connection()
    try:
        guests = conn.execute(
            """SELECT i.id, i.customer, i.item_name, i.order_date, i.user_status, i.notes
               FROM items i
               WHERE i.user_status LIKE '%GUEST%'
                 AND COALESCE(i.transaction_status, 'active') = 'active'
                 AND (i.notes IS NULL OR i.notes NOT LIKE '%Purchased by%')
                 AND i.email_uid NOT LIKE 'manual-%'
                 AND EXISTS (
                     SELECT 1 FROM items peer
                     WHERE peer.email_uid = i.email_uid
                       AND peer.id != i.id
                       AND peer.customer = i.customer COLLATE NOCASE
                 )
                 AND COALESCE(i.guest_name, '') = ''
                 AND COALESCE(i.partner_request, '') = ''
               ORDER BY i.order_date DESC"""
        ).fetchall()
        for g in guests:
            g = dict(g)
            items.append({
                "type": "guest_name_needed",
                "id": f"guest-{g['id']}",
                "item_id": g["id"],
                "title": g.get("item_name") or "Unknown event",
                "message": f"GUEST registration under \"{g['customer']}\" — confirm or enter the actual guest player's name.",
                "customer": g.get("customer"),
                "created_at": g.get("order_date"),
            })
    finally:
        conn.close()

    return jsonify(items)


@app.route("/api/parse-warnings/<int:warning_id>/dismiss", methods=["POST"])
@require_role("manager")
def api_dismiss_parse_warning(warning_id):
    """Dismiss a parse warning."""
    if dismiss_parse_warning(warning_id):
        return jsonify({"status": "ok"})
    return jsonify({"error": "Warning not found."}), 404


@app.route("/api/parse-warnings/<int:warning_id>/resolve", methods=["POST"])
@require_role("manager")
def api_resolve_parse_warning(warning_id):
    """Mark a parse warning as resolved."""
    if resolve_parse_warning(warning_id):
        return jsonify({"status": "ok"})
    return jsonify({"error": "Warning not found."}), 404


# ---------------------------------------------------------------------------
# Routes — Credit / Transfer
# ---------------------------------------------------------------------------
@app.route("/api/items/<int:item_id>/credit", methods=["POST"])
@require_role("manager")
def api_credit_item(item_id):
    """Mark an item as credited (money held for future event)."""
    data = request.get_json(silent=True) or {}
    if credit_item(item_id, note=data.get("note", "")):
        return jsonify({"status": "ok"})
    return jsonify({"error": "Item not found or already credited/transferred."}), 400


@app.route("/api/items/<int:item_id>/wd", methods=["POST"])
@require_role("manager")
def api_wd_item(item_id):
    """Mark an item as WD (withdrawn) with optional partial credit."""
    data = request.get_json(silent=True) or {}
    note = data.get("note", "")
    credits = data.get("credits")  # dict like {"included_games": 14, ...}
    credit_amount = data.get("credit_amount", "")
    if wd_item(item_id, note=note, credits=credits, credit_amount=credit_amount):
        return jsonify({"status": "ok"})
    return jsonify({"error": "Item not found or already credited/transferred/WD."}), 400


# MANAGER-ACCESSIBLE money actions (Kerry 2026-07-30). Chapter managers run
# the events and must be able to put money back on the spot: Robert could not
# credit Carlos Zapata when he dropped out of side games he had already bought
# into. /credit was already manager-level, but the SAME credit modal's Refund
# and Partial Refund buttons were admin-only, so the flow 403'd halfway
# through. refund / partial-refund / payout-credit / refund-watch are now
# manager, matching /credit and /wd.
#
# NOT chapter-scoped: any manager can act on any event. Chapter managers
# already carry session["chapter"], so scoping is a small follow-up — flagged
# here rather than silently assumed.
@app.route("/api/items/<int:item_id>/refund", methods=["POST"])
@require_role("manager")
def api_refund_item(item_id):
    """Mark an item as refunded via GoDaddy or Venmo."""
    data = request.get_json(silent=True) or {}
    method = data.get("method", "")
    if method and method not in ("GoDaddy", "Venmo", "Zelle", "PayPal", "Cash App"):
        return jsonify({"error": "Invalid refund method. Must be GoDaddy, Venmo, Zelle, PayPal, or Cash App."}), 400
    if refund_item(item_id, method=method, note=data.get("note", "")):
        return jsonify({"status": "ok"})
    return jsonify({"error": "Item not found or already credited/transferred."}), 400


@app.route("/api/items/<int:item_id>/payout-credit", methods=["POST"])
@app.route("/api/items/<int:item_id>/payout-wd-credit", methods=["POST"])  # legacy alias
@require_role("manager")
def api_payout_credit(item_id):
    """Record a cash payout of a player credit (WD or standalone credited row).

    Body: {"method": "Venmo|Zelle|Check|GoDaddy|PayPal", "date": "YYYY-MM-DD", "note": "..."}
    """
    data = request.get_json(silent=True) or {}
    method = (data.get("method") or "").strip()
    if method and method not in ("GoDaddy", "Venmo", "Zelle", "Check", "PayPal", "Cash App"):
        return jsonify({"error": "Invalid method. Must be GoDaddy, Venmo, Zelle, Check, PayPal, or Cash App."}), 400
    refund_date = (data.get("date") or "").strip()
    if refund_date:
        try:
            datetime.strptime(refund_date, "%Y-%m-%d")
        except ValueError:
            return jsonify({"error": "date must be YYYY-MM-DD"}), 400
    result = payout_credit(
        item_id, method=method, note=data.get("note", ""), refund_date=refund_date,
    )
    if not result.get("ok"):
        return jsonify({"error": result.get("error", "Payout failed")}), 400
    return jsonify({"status": "ok", "amount": result.get("amount"), "date": result.get("date")})


@app.route("/api/items/<int:item_id>/refund-watch", methods=["POST"])
@require_role("manager")
def api_create_refund_watch(item_id):
    """Register a refund-payment watch when the admin taps a P2P pay link
    (Kerry 2026-07-15). Body: {method, amount, memo, handle?}. Also
    schedules the ~75s/~180s quick inbox sweeps so the provider's receipt
    verifies the refund within a couple of minutes."""
    from email_parser.database import create_refund_watch
    data = request.get_json(silent=True) or {}
    method = (data.get("method") or "").strip()
    if method not in ("Venmo", "PayPal", "Cash App", "Zelle"):
        return jsonify({"error": "method must be Venmo, PayPal, Cash App, or Zelle"}), 400
    result = create_refund_watch(
        item_id, method=method, amount=data.get("amount"),
        memo=(data.get("memo") or "").strip(),
        handle=(data.get("handle") or "").strip(),
    )
    if not result.get("ok"):
        return jsonify({"error": result.get("error", "watch failed")}), 400
    # Quick receipt sweeps (same coalescing jobs as the payouts Pay flow)
    if getattr(scheduler, "running", False):
        now = datetime.now()
        for jid, secs in (("venmo_quick_check_a", 75), ("venmo_quick_check_b", 180)):
            try:
                scheduler.add_job(_quick_expense_check, "date",
                                  run_date=now + timedelta(seconds=secs),
                                  id=jid, replace_existing=True, coalesce=True,
                                  misfire_grace_time=120)
            except Exception:
                logger.exception("Failed to schedule quick receipt check %s", jid)
    return jsonify({"status": "ok", **result})


@app.route("/api/items/<int:item_id>/refund-watch", methods=["GET"])
@require_role("manager")
def api_get_refund_watch(item_id):
    """Latest refund watch for a credit item — the modal polls this to
    flip to 'verified' when the receipt lands."""
    from email_parser.database import get_refund_watch
    w = get_refund_watch(item_id)
    return jsonify(w or {})


@app.route("/api/refund-watches", methods=["GET"])
@require_role("manager")
def api_list_refund_watches():
    """Open (unverified) refund watches — pending-refund indicators."""
    from email_parser.database import get_open_refund_watches
    return jsonify(get_open_refund_watches())


@app.route("/api/hio-pot", methods=["GET"])
@require_role("manager")
def api_hio_pot():
    """Running Hole-In-One pot: per-event matrix contributions accrued
    across all past (non-rained-out) events, minus recorded HIO payouts
    (Kerry 2026-07-20)."""
    from email_parser.database import get_hio_pot
    return jsonify(get_hio_pot())


@app.route("/api/hio-pot/carry-in", methods=["POST"])
@require_role("admin")
def api_set_hio_carry_in():
    """Set the Hole-In-One pot CARRY-IN — the balance brought forward from
    before the Tracker started accruing (Kerry 2026-07-20: ratified $1,822).

    `get_hio_pot` has always READ `hio_pot_carry_in` from app_settings, but
    nothing in the app ever wrote it — no route, no UI, no MCP tool — so
    the figure could not be set or corrected and the running pot showed
    only what the Tracker itself had accrued (Kerry 2026-07-31: "Hole In
    One Pot is not persisting"). app_settings lives in the DB, so on a
    Railway volume this survives redeploys.
    """
    from email_parser.database import set_app_setting, get_hio_pot
    data = request.get_json(silent=True) or {}
    raw = data.get("carry_in")
    if raw is None or str(raw).strip() == "":
        return jsonify({"error": "carry_in is required."}), 400
    try:
        carry_in = round(float(str(raw).replace("$", "").replace(",", "").strip()), 2)
    except (TypeError, ValueError):
        return jsonify({"error": "carry_in must be a number."}), 400
    if carry_in < 0:
        return jsonify({"error": "carry_in cannot be negative."}), 400
    set_app_setting("hio_pot_carry_in", str(carry_in))
    note = (data.get("note") or "").strip()
    if note:
        set_app_setting("hio_pot_carry_in_note", note)
    return jsonify({"status": "ok", **get_hio_pot()})


@app.route("/api/refunds/overview", methods=["GET"])
@require_role("admin")
def api_refunds_overview():
    """Consolidated REFUNDS console: OUTSTANDING (held credits) / IN FLIGHT
    (open watches) / COMPLETED (recorded payouts). Kerry 2026-07-15 — one
    place so nothing falls through the cracks."""
    from email_parser.database import get_refunds_overview
    try:
        days = int(request.args.get("completed_days", 120))
    except (TypeError, ValueError):
        days = 120
    days = max(7, min(days, 365))
    return jsonify(get_refunds_overview(completed_days=days))


@app.route("/api/items/<int:item_id>/partial-refund", methods=["POST"])
@require_role("manager")
def api_partial_refund_item(item_id):
    """Partially refund specific components (e.g., one side game) while keeping player active.

    Creates a -PAY child row showing the refund, and updates the parent's side_games.
    """
    data = request.get_json(silent=True) or {}
    method = data.get("method", "")
    # "Credit" (Kerry 2026-07-14) keeps the money in the house: the child
    # row is a CREDITED item (picked up by get_player_credits → Apply
    # Credit / balance emails) instead of an outbound refund.
    if method and method not in ("Credit", "GoDaddy", "Venmo", "Zelle", "PayPal", "Cash App"):
        return jsonify({"error": "Invalid refund method."}), 400
    as_credit = (method == "Credit")
    refunded_components = data.get("components", {})  # e.g. {"gross_games": 30}
    new_side_games = data.get("new_side_games")  # e.g. "NET" (after removing GROSS)
    # Event Downgrade (Kerry 2026-07-14): refunding the 18-vs-9 price
    # difference on a 9/18 Combo event also flips the registration to 9.
    new_holes = data.get("new_holes")
    if new_holes is not None and str(new_holes) not in ("9", "18"):
        return jsonify({"error": "new_holes must be 9 or 18."}), 400
    note = data.get("note", "")
    total = sum(refunded_components.values())

    # Build description
    comp_labels = ", ".join(f"{k.replace('_', ' ').title()}" for k in refunded_components.keys())
    if as_credit:
        refund_desc = f"Partial credit: {comp_labels} (held for a future event)"
    else:
        refund_desc = f"Refund {comp_labels} via {method}" if method else f"Refund {comp_labels}"

    import time as _time
    from email_parser.database import _connect
    uid = f"manual-refund-{int(_time.time() * 1000)}"
    with _connect() as conn:
        parent = conn.execute("SELECT * FROM items WHERE id = ?", (item_id,)).fetchone()
        if not parent:
            return jsonify({"error": "Item not found."}), 404
        parent = dict(parent)

        # Snapshot parent's mutable fields BEFORE modifying
        parent_snap = {}
        for fld in ("side_games", "holes", "tee_choice", "user_status"):
            if parent.get(fld) is not None:
                parent_snap[fld] = parent[fld]

        # Compute new side_games from current DB value based on refunded components
        current_sg = (parent.get("side_games") or "NONE").strip().upper()
        refunding_net = "net_games" in refunded_components
        refunding_gross = "gross_games" in refunded_components
        computed_new_sg = current_sg
        if current_sg == "BOTH":
            if refunding_net and refunding_gross:
                computed_new_sg = "NONE"
            elif refunding_net:
                computed_new_sg = "GROSS"
            elif refunding_gross:
                computed_new_sg = "NET"
        elif current_sg == "NET" and refunding_net:
            computed_new_sg = "NONE"
        elif current_sg == "GROSS" and refunding_gross:
            computed_new_sg = "NONE"

        # Update parent side_games if changed
        if computed_new_sg != current_sg:
            conn.execute("UPDATE items SET side_games = ? WHERE id = ?",
                         (computed_new_sg, item_id))

        # Event Downgrade: flip the registration's holes (the previous
        # value is already preserved in parent_snapshot for reversal)
        if new_holes is not None and str(parent.get("holes") or "") != str(new_holes):
            conn.execute("UPDATE items SET holes = ? WHERE id = ?",
                         (str(new_holes), item_id))

        # Create the child row with parent snapshot (customer_id copied from
        # parent, same as transfer_item() — already resolved, no new lookup).
        # Refund → -PAY child (money out). Credit → a CREDITED child with a
        # POSITIVE price: get_player_credits() surfaces any credited row
        # (parent or child), so it flows into Apply Credit / balance emails.
        cur = conn.execute(
            """INSERT INTO items (email_uid, merchant, customer, item_name, item_price,
               side_games, notes, parent_item_id, parent_snapshot, transaction_status, order_date,
               customer_id)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (uid,
             "Partial Credit" if as_credit else (f"Refund ({method})" if method else "Partial Refund"),
             parent["customer"], parent["item_name"],
             f"${total:.2f}" if as_credit else f"-${total:.2f}",
             None,
             refund_desc + (f" — {note}" if note else ""),
             item_id,
             json.dumps(parent_snap) if parent_snap else None,
             "credited" if as_credit else "active",
             today_central_str(),
             parent.get("customer_id")),
        )
        new_child_id = cur.lastrowid

        # ── Accounting: flat entry for partial refund ──
        # A CREDIT writes NO acct entry — it's an internal ledger move
        # (unified-financial-model rule: only real outbound payments hit
        # acct_transactions); the money leaves when the credit is later
        # applied or refunded through those flows.
        if not as_credit:
            try:
                from email_parser.database import _write_acct_entry
                refund_source = method.lower().replace(" ", "_") if method else "manual"
                _m = (method or "").lower()
                refund_account = "Venmo" if "venmo" in _m else ("PayPal" if "paypal" in _m else "TGF Checking")
                _write_acct_entry(
                    conn,
                    item_id=new_child_id,
                    event_name=parent["item_name"],
                    customer=parent["customer"],
                    order_id=parent.get("order_id", ""),
                    entry_type="expense",
                    category="refund",
                    source=refund_source,
                    amount=float(total),
                    description=f"Partial refund ({method}): {parent['customer']} — {parent['item_name']}",
                    account=refund_account,
                    source_ref=f"partial-refund-{new_child_id}",
                    date=today_central_str(),
                )
            except Exception:
                logger.warning("Failed to create accounting entry for partial refund %d", item_id, exc_info=True)

        conn.commit()

    return jsonify({"status": "ok", "refunded": total,
                    "new_side_games": computed_new_sg,
                    "new_holes": new_holes})


@app.route("/api/items/<int:item_id>/transfer", methods=["POST"])
@require_role("manager")
def api_transfer_item(item_id):
    """Transfer an item to a different event."""
    data = request.get_json(silent=True)
    if not data or not data.get("target_event"):
        return jsonify({"error": "target_event is required."}), 400
    new_item = transfer_item(item_id, data["target_event"], note=data.get("note", ""))
    if new_item:
        return jsonify({"status": "ok", "new_item": new_item})
    return jsonify({"error": "Item not found or already credited/transferred."}), 400


@app.route("/api/items/<int:item_id>/reverse-credit", methods=["POST"])
@require_role("manager")
def api_reverse_credit(item_id):
    """Reverse a credit or transfer, restoring the original item to active."""
    if reverse_credit(item_id):
        return jsonify({"status": "ok"})
    return jsonify({"error": "Item not found or not in credited/transferred state."}), 400


@app.route("/api/events/add-player", methods=["POST"])
@require_role("manager")
def api_add_player():
    """Add a player to an event (comp, RSVP only, or paid separately)."""
    data = request.get_json(silent=True)
    if not data or not data.get("event_name") or not data.get("customer"):
        return jsonify({"error": "event_name and customer are required."}), 400
    err = validate_json_fields(data)
    if err:
        return jsonify({"error": err}), 400
    mode = data.get("mode", "comp")
    if mode not in ("comp", "rsvp", "paid_separately"):
        return jsonify({"error": "Invalid mode."}), 400
    try:
        item = add_player_to_event(
            event_name=data["event_name"],
            customer=data["customer"],
            mode=mode,
            side_games=data.get("side_games", ""),
            tee_choice=data.get("tee_choice", ""),
            handicap=data.get("handicap", ""),
            user_status=data.get("user_status", data.get("member_status", "")),
            payment_amount=data.get("payment_amount", ""),
            payment_source=data.get("payment_source", ""),
            customer_email=data.get("customer_email", ""),
            customer_phone=data.get("customer_phone", ""),
            holes=data.get("holes", ""),
            order_date=data.get("order_date", ""),
        )
        if item:
            return jsonify({"status": "ok", "item": item}), 201
        return jsonify({"error": "Failed to add player."}), 500
    except Exception as e:
        logger.exception("Error adding player: %s", e)
        return jsonify({"error": f"Server error: {e}"}), 500


@app.route("/api/events/add-payment", methods=["POST"])
@require_role("manager")
def api_add_payment():
    """Add an additional payment record for an existing event player."""
    data = request.get_json(silent=True)
    if not data or not data.get("event_name") or not data.get("customer"):
        return jsonify({"error": "event_name and customer are required."}), 400
    if not data.get("payment_amount") or not data.get("payment_source"):
        return jsonify({"error": "payment_amount and payment_source are required."}), 400
    err = validate_json_fields(data)
    if err:
        return jsonify({"error": err}), 400
    try:
        item = add_payment_to_event(
            event_name=data["event_name"],
            customer=data["customer"],
            payment_item=data.get("payment_item", ""),
            payment_amount=data.get("payment_amount", ""),
            payment_source=data.get("payment_source", ""),
            note=data.get("note", ""),
            order_date=data.get("order_date", ""),
        )
        if isinstance(item, dict) and item.get("error"):
            return jsonify({"error": item["error"]}), 400
        if item:
            return jsonify({"status": "ok", "item": item}), 201
        return jsonify({"error": "Failed to add payment."}), 500
    except Exception as e:
        logger.exception("Error adding payment: %s", e)
        return jsonify({"error": f"Server error: {e}"}), 500


@app.route("/api/events/add-payment/quote", methods=["GET"])
@require_role("manager")
def api_add_payment_quote():
    """Suggested amounts (from the event's pricing setup) + the player's
    available credit for the Add Payment modal (Kerry 2026-07-29)."""
    event_name = (request.args.get("event_name") or "").strip()
    customer = (request.args.get("customer") or "").strip()
    if not event_name or not customer:
        return jsonify({"error": "event_name and customer are required."}), 400
    try:
        return jsonify(get_add_payment_quote(event_name, customer))
    except Exception as e:
        logger.exception("Error building add-payment quote: %s", e)
        return jsonify({"error": f"Server error: {e}"}), 500


@app.route("/api/events/upgrade-rsvp", methods=["POST"])
@require_role("manager")
def api_upgrade_rsvp():
    """Upgrade an RSVP-only placeholder to a full paid registration."""
    data = request.get_json(silent=True)
    if not data or not data.get("item_id"):
        return jsonify({"error": "item_id is required."}), 400
    item = upgrade_rsvp_to_paid(
        item_id=data["item_id"],
        payment_amount=data.get("payment_amount", ""),
        payment_source=data.get("payment_source", ""),
        side_games=data.get("side_games", ""),
        tee_choice=data.get("tee_choice", ""),
        handicap=data.get("handicap", ""),
        user_status=data.get("user_status", data.get("member_status", "")),
    )
    if item:
        return jsonify({"status": "ok", "item": item})
    return jsonify({"error": "Item not found or not in RSVP-only state."}), 400


@app.route("/api/events/send-reminder", methods=["POST"])
@require_role("manager")
def api_send_reminder():
    """Send a payment reminder email to an RSVP-only player."""
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "JSON body required"}), 400
    to_email = (data.get("to_email") or "").strip()
    player_name = data.get("player_name", "Player")
    event_name = data.get("event_name", "the upcoming event")
    if not to_email:
        return jsonify({"error": "to_email is required"}), 400

    tenant_id = os.getenv("AZURE_TENANT_ID")
    client_id = os.getenv("AZURE_CLIENT_ID")
    client_secret = os.getenv("AZURE_CLIENT_SECRET")
    from_address = os.getenv("EMAIL_ADDRESS")
    if not all([tenant_id, client_id, client_secret, from_address]):
        return jsonify({"error": "Email credentials not configured"}), 500

    subject = f"Payment Reminder — {event_name}"
    html_body = (
        f"<p>Hi {player_name},</p>"
        f"<p>This is a friendly reminder that we have you down for "
        f"<strong>{event_name}</strong>, but we haven't received your payment yet.</p>"
        f"<p>Please complete your registration at your earliest convenience.</p>"
        f"<p>Thanks,<br>The Golf Fellowship</p>"
    )

    ok = send_mail_graph(
        tenant_id=tenant_id,
        client_id=client_id,
        client_secret=client_secret,
        from_address=from_address,
        to_address=to_email,
        subject=subject,
        html_body=html_body,
    )
    if ok:
        return jsonify({"status": "ok", "message": f"Reminder sent to {to_email}"})
    return jsonify({"error": "Failed to send reminder email"}), 500


@app.route("/api/events/send-reminder-all", methods=["POST"])
@require_role("manager")
def api_send_reminder_all():
    """Send payment reminder emails to ALL RSVP-only players for an event."""
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "JSON body required"}), 400
    event_name = (data.get("event_name") or "").strip()
    if not event_name:
        return jsonify({"error": "event_name is required"}), 400

    tenant_id = os.getenv("AZURE_TENANT_ID")
    client_id = os.getenv("AZURE_CLIENT_ID")
    client_secret = os.getenv("AZURE_CLIENT_SECRET")
    from_address = os.getenv("EMAIL_ADDRESS")
    if not all([tenant_id, client_id, client_secret, from_address]):
        return jsonify({"error": "Email credentials not configured"}), 500

    # Find all RSVP-only players for this event with email addresses (case-insensitive)
    items = get_all_items()
    rsvp_players = [
        i for i in items
        if (i.get("item_name") or "").lower() == event_name.lower()
        and (i.get("transaction_status") or "active") == "rsvp_only"
    ]

    if not rsvp_players:
        return jsonify({"error": "No RSVP-only players found for this event"}), 404

    sent = 0
    failed = 0
    skipped_no_email = 0
    for player in rsvp_players:
        to_email = _resolve_player_email(player)
        if not to_email:
            skipped_no_email += 1
            continue
        player_name = player.get("customer") or "Player"
        subject = f"Payment Reminder — {event_name}"
        html_body = (
            f"<p>Hi {player_name},</p>"
            f"<p>This is a friendly reminder that we have you down for "
            f"<strong>{event_name}</strong>, but we haven't received your payment yet.</p>"
            f"<p>Please complete your registration at your earliest convenience.</p>"
            f"<p>Thanks,<br>The Golf Fellowship</p>"
        )
        ok = send_mail_graph(
            tenant_id=tenant_id, client_id=client_id,
            client_secret=client_secret, from_address=from_address,
            to_address=to_email, subject=subject, html_body=html_body,
        )
        if ok:
            sent += 1
        else:
            failed += 1

    if sent == 0:
        return jsonify({"error": "All reminder emails failed to send", "sent": sent, "failed": failed, "total": len(rsvp_players)}), 500
    status = "ok" if failed == 0 else "partial"
    return jsonify({"status": status, "sent": sent, "failed": failed, "total": len(rsvp_players)})


# ---------------------------------------------------------------------------
# Routes — Messaging (Bulk Email Communications)
# ---------------------------------------------------------------------------

@app.route("/api/messages/templates", methods=["GET"])
@require_role("manager")
def api_get_templates():
    """Return all message templates."""
    return jsonify(get_message_templates())


@app.route("/api/messages/templates", methods=["POST"])
@require_role("manager")
def api_create_template():
    """Create a new message template."""
    data = request.get_json(silent=True)
    if not data or not data.get("name"):
        return jsonify({"error": "name is required"}), 400
    template = create_message_template(data)
    return jsonify(template), 201


@app.route("/api/messages/templates/<int:template_id>", methods=["PATCH"])
@require_role("manager")
def api_update_template(template_id):
    """Update a message template."""
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "JSON body required"}), 400
    result = update_message_template(template_id, data)
    if result is None:
        return jsonify({"error": "Template not found"}), 404
    return jsonify(result)


@app.route("/api/messages/templates/<int:template_id>", methods=["DELETE"])
@require_role("admin")
def api_delete_template(template_id):
    """Delete a non-system message template."""
    if delete_message_template(template_id):
        return jsonify({"status": "ok"})
    return jsonify({"error": "Template not found or is a system template"}), 400


@app.route("/api/messages/send", methods=["POST"])
@require_role("manager")
def api_send_messages():
    """Send a message to a filtered audience for an event.

    Body: {
        event_name: str,
        template_id: int (optional — use template subject/body),
        subject: str (overrides template subject if provided),
        html_body: str (overrides template body if provided),
        audience: str (all|playing|rsvp_only|net|gross|both|not_playing|custom),
        custom_emails: [str] (required when audience=custom — specific email addresses),
        exclude_ids: [int] (optional — item IDs to exclude)
    }
    """
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "JSON body required"}), 400

    event_name = (data.get("event_name") or "").strip()
    if not event_name:
        return jsonify({"error": "event_name is required"}), 400

    # Resolve subject + body from template or direct input
    subject_tpl = data.get("subject") or ""
    body_tpl = data.get("html_body") or ""
    template_id = data.get("template_id")

    if template_id:
        tpl = get_message_template(template_id)
        if not tpl:
            return jsonify({"error": "Template not found"}), 404
        if not subject_tpl:
            subject_tpl = tpl.get("subject") or ""
        if not body_tpl:
            body_tpl = tpl.get("html_body") or ""

    if not subject_tpl or not body_tpl:
        return jsonify({"error": "subject and html_body are required (directly or via template)"}), 400

    # Build event variables for template rendering
    all_events = get_all_events()
    event_info = next((e for e in all_events if (e["item_name"] or "").lower() == event_name.lower()), {})
    event_vars = {
        "event_name": event_name,
        "event_date": event_info.get("event_date") or "",
        "course": event_info.get("course") or "",
        "chapter": event_info.get("chapter") or "",
    }
    event_status = (event_info.get("status") or "active")

    # Filter audience
    audience = (data.get("audience") or "all").lower()
    custom_emails = set()
    if audience == "custom":
        raw = data.get("custom_emails") or []
        custom_emails = {e.strip().lower() for e in raw if isinstance(e, str) and e.strip()}
        if not custom_emails:
            return jsonify({"error": "custom_emails list is required for custom audience"}), 400
    exclude_ids = set(data.get("exclude_ids") or [])
    items = get_all_items()

    # Get event aliases for matching
    all_aliases = get_all_event_aliases()
    alias_to_canonical = {}
    for alias_name, canonical in all_aliases.items():
        alias_to_canonical[alias_name] = canonical

    def matches_event(item):
        iname = item.get("item_name") or ""
        if iname == event_name:
            return True
        if alias_to_canonical.get(iname) == event_name:
            return True
        return False

    registrants = [i for i in items if matches_event(i)]

    # Use the canonical resolver — customer_emails.is_primary first,
    # items.customer_email only as a fallback.
    def resolve_email(r):
        return _resolve_player_email(r)

    # Get RSVP override data for playing/not_playing filtering
    rsvp_overrides = {}
    try:
        overrides = get_rsvp_overrides(event_name)
        for ov in overrides:
            rsvp_overrides[ov["item_id"]] = ov["status"]
    except Exception:
        logger.warning("Failed to load RSVP overrides for %s", event_name, exc_info=True)

    rsvps_for_event = {}
    rsvp_list = []
    try:
        rsvp_list = get_rsvps_for_event(event_name)
        for rv in rsvp_list:
            if rv.get("matched_item_id"):
                rsvps_for_event[rv["matched_item_id"]] = rv["response"]
    except Exception:
        logger.warning("Failed to load RSVPs for event %s", event_name, exc_info=True)

    # Build GG RSVP synthetic rows (unmatched RSVPs with player_email)
    email_overrides = {}
    try:
        email_overrides = get_rsvp_email_overrides(event_name)
    except Exception:
        logger.warning("Failed to load RSVP email overrides for %s", event_name, exc_info=True)

    reg_emails = {(r.get("customer_email") or "").strip().lower() for r in registrants if r.get("customer_email")}
    reg_names = {(r.get("customer") or "").strip().lower() for r in registrants if r.get("customer")}
    gg_rsvp_rows = []
    for rv in rsvp_list:
        if rv.get("response") != "PLAYING":
            continue
        if rv.get("matched_item_id"):
            continue
        email = (rv.get("player_email") or "").strip().lower()
        if email and email_overrides.get(email) == "not_playing":
            continue
        if email and email in reg_emails:
            continue
        resolved = (rv.get("resolved_name") or "").strip().lower()
        first_name = (rv.get("player_name") or "").strip().lower()
        if resolved and resolved in reg_names:
            continue
        if first_name and any(n.startswith(first_name) for n in reg_names):
            continue
        if not (rv.get("player_email") or "").strip():
            continue  # No email — can't message them
        gg_rsvp_rows.append({
            "id": f"gg-rsvp-{len(gg_rsvp_rows)}",
            "customer": rv.get("resolved_name") or rv.get("player_name") or "Unknown",
            "customer_email": (rv.get("player_email") or "").strip(),
            "item_name": event_name,
            "transaction_status": "gg_rsvp",
            "side_games": "",
        })

    all_registrants = registrants + gg_rsvp_rows

    def get_rsvp_status(item):
        item_id = item["id"]
        if isinstance(item_id, str) and item_id.startswith("gg-rsvp"):
            return "playing"  # GG RSVP players are playing by definition
        override = rsvp_overrides.get(item_id)
        if override and override != "none":
            return override  # playing, not_playing, manual_green
        rsvp_resp = rsvps_for_event.get(item_id)
        if rsvp_resp == "PLAYING":
            return "playing"
        if rsvp_resp == "NOT PLAYING":
            return "not_playing"
        return "unknown"

    def classify_side_games(sg):
        sg = (sg or "").strip().upper()
        if sg in ("NET", "GROSS", "BOTH", "NONE"):
            return sg
        return "NONE"

    filtered = []
    for r in all_registrants:
        rid = r["id"]
        if not (isinstance(rid, str) and rid.startswith("gg-rsvp")) and rid in exclude_ids:
            continue
        status = (r.get("transaction_status") or "active")
        # Skip transferred always. Skip credited only while the event is
        # still ACTIVE — on a cancelled/postponed event the credited
        # players ARE the audience (they were credited BY the
        # cancellation, and excluding them made the post-cancel email
        # silently reach nobody — Kerry, rained-out s9.18, 2026-07-14).
        if status == "transferred":
            continue
        if status == "credited" and event_status == "active":
            continue
        email = resolve_email(r)
        if not email:
            continue

        sg = classify_side_games(r.get("side_games"))
        rsvp = get_rsvp_status(r)

        if audience == "all":
            filtered.append(r)
        elif audience == "playing":
            if rsvp in ("playing", "manual_green"):
                filtered.append(r)
        elif audience == "rsvp_only":
            if status in ("rsvp_only", "gg_rsvp"):
                filtered.append(r)
        elif audience == "net":
            if sg in ("NET", "BOTH"):
                filtered.append(r)
        elif audience == "gross":
            if sg in ("GROSS", "BOTH"):
                filtered.append(r)
        elif audience == "both":
            if sg == "BOTH":
                filtered.append(r)
        elif audience == "not_playing":
            if rsvp == "not_playing":
                filtered.append(r)
        elif audience == "custom":
            if email.lower() in custom_emails:
                filtered.append(r)
        else:
            filtered.append(r)

    # Build recipient list from filtered registrants
    recipients = [
        {"player_name": r.get("customer") or "Player", "email": resolve_email(r)}
        for r in filtered
    ]

    # Add extra recipients (manually entered emails not on the player list)
    extra_recipients = data.get("extra_recipients") or []
    seen_emails = {r["email"].lower() for r in recipients}
    for er in extra_recipients:
        email = (er.get("email") or "").strip()
        if email and email.lower() not in seen_emails:
            recipients.append({"player_name": er.get("name") or email.split("@")[0], "email": email})
            seen_emails.add(email.lower())

    if not recipients:
        return jsonify({"error": "No recipients found matching the audience filter", "sent": 0, "failed": 0}), 404

    # Send with throttle
    result = send_bulk_emails(
        recipients=recipients,
        subject_template=subject_tpl,
        body_template=body_tpl,
        event_vars=event_vars,
    )

    # Log each send
    role = session.get("role", "unknown")
    body_preview = render_msg_template(body_tpl, {**event_vars, "player_name": "..."})[:200]
    error_emails = [e["recipient"] for e in result.get("errors", [])]
    for r in recipients:
        email = r["email"]
        was_sent = email not in error_emails
        log_message({
            "event_name": event_name,
            "template_id": template_id,
            "channel": "email",
            "recipient_name": r.get("player_name"),
            "recipient_address": email,
            "subject": render_msg_template(subject_tpl, {**event_vars, "player_name": r.get("player_name") or "Player"}),
            "body_preview": body_preview,
            "status": "sent" if was_sent else "failed",
            "error_message": None if was_sent else "Send failed",
            "sent_by": role,
        })

    status = "ok" if result["failed"] == 0 else "partial"
    return jsonify({
        "status": status,
        "sent": result["sent"],
        "failed": result["failed"],
        "total": len(recipients),
        "errors": result["errors"],
    })


@app.route("/api/messages/preview", methods=["POST"])
@require_role("manager")
def api_preview_message():
    """Render a message template with sample data. Returns rendered subject + body."""
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "JSON body required"}), 400

    subject_tpl = data.get("subject") or ""
    body_tpl = data.get("html_body") or ""
    template_id = data.get("template_id")

    if template_id:
        tpl = get_message_template(template_id)
        if tpl:
            if not subject_tpl:
                subject_tpl = tpl.get("subject") or ""
            if not body_tpl:
                body_tpl = tpl.get("html_body") or ""

    variables = {
        "player_name": data.get("player_name", "John Doe"),
        "event_name": data.get("event_name", "Sample Event"),
        "event_date": data.get("event_date", "2026-03-15"),
        "course": data.get("course", "Sample Course"),
        "chapter": data.get("chapter", "San Antonio"),
    }

    return jsonify({
        "subject": render_msg_template(subject_tpl, variables),
        "html_body": render_msg_template(body_tpl, variables),
    })


@app.route("/api/messages/log", methods=["GET"])
@require_role("manager")
def api_message_log():
    """Return message send history, optionally filtered by event."""
    event_name = request.args.get("event_name")
    limit = min(int(request.args.get("limit", 200)), 1000)
    return jsonify(get_message_log(event_name=event_name, limit=limit))


@app.route("/api/messages/log/<path:event_name>", methods=["GET"])
@require_role("manager")
def api_message_log_event(event_name):
    """Return message log for a specific event."""
    return jsonify(get_message_log(event_name=event_name))


@app.route("/api/events/seed", methods=["POST"])
@require_role("admin")
def api_seed_events():
    """Batch-create events from a JSON list. Admin only."""
    data = request.get_json(silent=True)
    if not data or not isinstance(data.get("events"), list):
        return jsonify({"error": "Body must be JSON with 'events' array."}), 400
    result = seed_events(data["events"])
    return jsonify({"status": "ok", **result})


# ---------------------------------------------------------------------------
# Routes — RSVP
# ---------------------------------------------------------------------------
@app.route("/rsvps")
def rsvps_page():
    if session.get("role") == "view-only":
        return redirect("/events")
    return render_template("rsvps.html")


@app.route("/api/rsvps")
@require_role("view-only")
def api_rsvps():
    """Return RSVPs, optionally filtered by event or response."""
    event = request.args.get("event", "")
    response = request.args.get("response", "")
    return jsonify(get_all_rsvps(event_name=event, response=response))


# event_name → monotonic timestamp of the last on-open auto-audit.
# In-memory is fine: single Railway instance, and the ingest + nightly
# audits back it up if the process restarts.
_rsvp_onopen_audit_last: dict = {}
_RSVP_ONOPEN_AUDIT_SECONDS = 900  # at most once per event per 15 min


@app.route("/api/rsvps/event/<path:event_name>")
@require_role("view-only")
def api_rsvps_for_event(event_name):
    """Return the latest RSVP per player for a specific event.

    Opening an event self-heals its RSVP matches (Kerry 2026-07-15):
    manager/admin sessions trigger the clear-mismatches + rematch audit
    inline (throttled per event) BEFORE the read, so what the manager
    sees is already corrected — no manual Audit RSVPs button required.
    """
    if session.get("role") in ("manager", "admin"):
        last = _rsvp_onopen_audit_last.get(event_name)
        if last is None or time.monotonic() - last > _RSVP_ONOPEN_AUDIT_SECONDS:
            _rsvp_onopen_audit_last[event_name] = time.monotonic()
            try:
                res = audit_event_rsvps(event_name)
                if res.get("cleared") or res.get("rematched"):
                    logger.info("On-open RSVP audit for %s: %s", event_name, res)
            except Exception:
                logger.exception("On-open RSVP audit failed for %s", event_name)
    return jsonify(get_rsvps_for_event(event_name))


@app.route("/api/rsvps/bulk")
@require_role("view-only")
def api_rsvps_bulk():
    """Return all RSVPs, overrides, and email overrides grouped by event.

    Used by the events page to show accurate player counts on collapsed cards
    without requiring per-event fetches.
    """
    return jsonify(get_all_rsvps_bulk())


@app.route("/api/rsvps/stats")
@require_role("view-only")
def api_rsvp_stats():
    """Return RSVP summary statistics."""
    return jsonify(get_rsvp_stats())


@app.route("/api/rsvps/check-now", methods=["POST"])
@require_role("manager")
def api_rsvp_check_now():
    """Manually trigger an RSVP inbox check."""
    rsvp_address = os.getenv("RSVP_EMAIL_ADDRESS")
    if not rsvp_address:
        return jsonify({"error": "RSVP_EMAIL_ADDRESS not configured."}), 400

    try:
        check_rsvp_inbox()
        stats = get_rsvp_stats()
        return jsonify({"status": "ok", "stats": stats})
    except Exception as e:
        logger.exception("Manual RSVP check failed")
        return jsonify({"error": str(e)}), 500


@app.route("/api/rsvps/rematch", methods=["POST"])
@require_role("manager")
def api_rsvp_rematch():
    """Re-run matching logic on unmatched RSVPs."""
    try:
        result = rematch_rsvps()
        return jsonify({"status": "ok", **result})
    except Exception as e:
        logger.exception("RSVP rematch failed")
        return jsonify({"error": str(e)}), 500


@app.route("/api/rsvps/audit-event/<path:event_name>", methods=["POST"])
@require_role("manager")
def api_audit_event_rsvps(event_name):
    """Audit and fix RSVP matches for a specific event.

    Clears bad matches (email mismatch) and re-attempts matching.
    """
    try:
        result = audit_event_rsvps(event_name)
        return jsonify({"status": "ok", **result})
    except Exception as e:
        logger.exception("RSVP audit failed for event: %s", event_name)
        return jsonify({"error": str(e)}), 500


@app.route("/api/rsvps/<int:rsvp_id>/match", methods=["POST"])
@require_role("admin")
def api_manual_match_rsvp(rsvp_id):
    """Manually assign an RSVP to an event. Admin only."""
    data = request.get_json(silent=True)
    if not data or not data.get("event_name"):
        return jsonify({"error": "event_name is required."}), 400
    if manual_match_rsvp(rsvp_id, data["event_name"]):
        return jsonify({"status": "ok"})
    return jsonify({"error": "RSVP not found."}), 404


@app.route("/api/rsvps/<int:rsvp_id>/unmatch", methods=["POST"])
@require_role("admin")
def api_unmatch_rsvp(rsvp_id):
    """Clear the match for an RSVP. Admin only."""
    if unmatch_rsvp(rsvp_id):
        return jsonify({"status": "ok"})
    return jsonify({"error": "RSVP not found."}), 404


@app.route("/api/rsvps/overrides/<path:event_name>")
@require_role("view-only")
def api_rsvp_overrides(event_name):
    """Return manual RSVP overrides for an event.

    Returns {"by_item": {item_id: status}, "by_email": {email: status}}.
    """
    return jsonify({
        "by_item": get_rsvp_overrides(event_name),
        "by_email": get_rsvp_email_overrides(event_name),
    })


@app.route("/api/rsvps/overrides", methods=["POST"])
@require_role("manager")
def api_set_rsvp_override():
    """Set a manual RSVP override for a registrant (by item_id or player_email)."""
    data = request.get_json(force=True)
    item_id = data.get("item_id")
    player_email = data.get("player_email")
    event_name = data.get("event_name")
    status = data.get("status", "none")
    if not event_name:
        return jsonify({"error": "event_name required"}), 400
    if not item_id and not player_email:
        return jsonify({"error": "item_id or player_email required"}), 400
    if status not in ("none", "playing", "not_playing", "manual_green"):
        return jsonify({"error": "status must be none, playing, not_playing, or manual_green"}), 400
    if player_email:
        set_rsvp_email_override(player_email, event_name, status)
        return jsonify({"status": "ok", "player_email": player_email, "event_name": event_name, "rsvp_status": status})
    set_rsvp_override(int(item_id), event_name, status)
    return jsonify({"status": "ok", "item_id": item_id, "event_name": event_name, "rsvp_status": status})


@app.route("/api/rsvps/config-status")
@require_role("view-only")
def api_rsvp_config_status():
    """Check whether RSVP email credentials are configured."""
    rsvp_ok = bool(os.getenv("RSVP_EMAIL_ADDRESS"))
    tenant_ok = bool(
        os.getenv("RSVP_AZURE_TENANT_ID") or os.getenv("AZURE_TENANT_ID")
    )
    return jsonify({
        "configured": rsvp_ok and tenant_ok,
        "rsvp_email": rsvp_ok,
        "azure_credentials": tenant_ok,
    })


@app.route("/api/report/send-now", methods=["POST"])
@require_role("manager")
def api_send_report_now():
    """Manually trigger the daily report."""
    if not os.getenv("DAILY_REPORT_TO"):
        return jsonify({"error": "DAILY_REPORT_TO not configured in .env"}), 400
    try:
        send_daily_report()
        return jsonify({"status": "ok", "sent_to": os.getenv("DAILY_REPORT_TO")})
    except Exception as e:
        logger.exception("Manual report send failed")
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# Routes — AI Support Chat & Feedback
# ---------------------------------------------------------------------------


def _send_feedback_notification(feedback: dict):
    """Send an instant email notification when a new bug/feature is submitted."""
    notify_to = os.getenv("FEEDBACK_NOTIFY_TO") or os.getenv("DAILY_REPORT_TO")
    if not notify_to:
        return

    tenant_id = os.getenv("AZURE_TENANT_ID")
    client_id = os.getenv("AZURE_CLIENT_ID")
    client_secret = os.getenv("AZURE_CLIENT_SECRET")
    from_addr = os.getenv("EMAIL_ADDRESS")
    if not all([tenant_id, client_id, client_secret, from_addr]):
        return

    fb_type = feedback.get("type", "feedback").capitalize()
    label = "Bug Report" if feedback.get("type") == "bug" else "Feature Request"
    color = "#dc2626" if feedback.get("type") == "bug" else "#2563eb"
    page = feedback.get("page") or "Unknown"
    role = feedback.get("role") or "Unknown"
    created = feedback.get("created_at") or "—"
    message = feedback.get("message") or ""

    html = f"""\
<html><body style="font-family: Arial, sans-serif; color: #333; max-width: 600px;">
<h2 style="color: {color};">New {label} Submitted</h2>
<table style="font-size: 14px; margin-bottom: 16px;">
  <tr><td style="padding:4px 12px 4px 0; font-weight:600;">Type:</td>
      <td><span style="background:{color}; color:#fff; padding:2px 10px; border-radius:10px; font-size:12px;">{fb_type}</span></td></tr>
  <tr><td style="padding:4px 12px 4px 0; font-weight:600;">Page:</td><td>{page}</td></tr>
  <tr><td style="padding:4px 12px 4px 0; font-weight:600;">Submitted by:</td><td>{role}</td></tr>
  <tr><td style="padding:4px 12px 4px 0; font-weight:600;">Time:</td><td>{created}</td></tr>
</table>
<div style="background:#f9fafb; border-left:4px solid {color}; padding:12px 16px; margin-bottom:16px; white-space:pre-wrap;">{message}</div>
<p style="font-size:12px; color:#999;">This is an automated notification from TGF Transaction Tracker.</p>
</body></html>"""

    subject = f"[TGF {label}] New submission from {page} page"

    try:
        ok = send_mail_graph(
            tenant_id=tenant_id,
            client_id=client_id,
            client_secret=client_secret,
            from_address=from_addr,
            to_address=notify_to,
            subject=subject,
            html_body=html,
        )
        if not ok:
            logger.warning("Feedback notification email failed to send to %s", notify_to)
    except Exception:
        logger.exception("Failed to send feedback notification email")

_TGF_SYSTEM_PROMPT = """You are the TGF Assistant, an AI helper built into The Golf Fellowship's Transaction Tracker.
You help managers and admins understand and use the platform.

Key facts about the platform:
- Pages: Transactions (main item list), Events (event roster + RSVP circles), Customers (player directory with merge), RSVP Log, Matrix (admin pairings), Audit (email verification).
- Transaction items are parsed from emails via AI. Each row = one line item from a purchase.
- Events are auto-created from transaction item names when they match golf event patterns.
- Event aliases link variant item names to a canonical event (e.g. "San Antonio Kickoff NORTHERN HILLS" → "San Antonio Kickoff CEDAR CREEK").
- RSVP circles show player status: green = paid, yellow/dotted = RSVP only (no payment), red = not playing, gray = no response. Real Golf Genius RSVPs override manual green.
- Orphan banner appears when transaction items don't match any event — admins can create the event or add an alias.
- Customer merge combines two player records (e.g. "Jdub Wade" + "John Wade").
- Credits: items can be credited (money on account), transferred to another event, or reversed.
- Auth: PIN-based, two tiers — Admin (full access) and Manager (no audit/matrix).
- Bulk "Remind All" sends payment reminder emails to RSVP-only players on an event.
- Database backup is available at /admin/backup (admin only).

When answering:
- Be concise and helpful. Use specific page names and button labels.
- If you don't know something specific about TGF data, say so — don't guess at numbers.
- For bugs or feature requests, encourage the user to use the Report a Bug or Request a Feature buttons.
- You can explain any feature, workflow, or concept in the platform.
"""


@app.route("/api/support/chat", methods=["POST"])
def api_support_chat():
    """Streaming AI chat endpoint for the support widget."""
    user_role = session.get("role")
    if not user_role:
        return jsonify({"error": "Not authenticated."}), 401

    data = request.get_json(silent=True)
    if not data or not data.get("message"):
        return jsonify({"error": "Message is required."}), 400

    api_key = os.getenv("ANTHROPIC_API_KEY")
    if not api_key:
        return jsonify({"error": "AI not configured (missing API key)."}), 503

    messages = data.get("history", [])
    messages.append({"role": "user", "content": data.get("message", "")})

    page = data.get("page", "")
    role_context = f"\nThe user is a {user_role} currently on the {page} page." if page else f"\nThe user is a {user_role}."

    def generate():
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)
        try:
            with client.messages.stream(
                model="claude-sonnet-4-5-20250929",
                max_tokens=1024,
                system=_TGF_SYSTEM_PROMPT + role_context,
                messages=messages,
            ) as stream:
                for text in stream.text_stream:
                    yield f"data: {json.dumps({'text': text})}\n\n"
            yield "data: {\"done\": true}\n\n"
        except Exception as e:
            logger.exception("Support chat error")
            yield f"data: {json.dumps({'error': str(e)})}\n\n"

    return Response(generate(), mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


@app.route("/api/support/feedback", methods=["POST"])
def api_support_feedback_post():
    """Log a bug report or feature request."""
    user_role = session.get("role")
    if not user_role:
        return jsonify({"error": "Not authenticated."}), 401

    data = request.get_json(silent=True)
    if not data or not data.get("message"):
        return jsonify({"error": "Message is required."}), 400
    err = validate_json_fields(data)
    if err:
        return jsonify({"error": err}), 400

    fb_type = data.get("type", "bug")
    if fb_type not in ("bug", "feature"):
        return jsonify({"error": "Type must be 'bug' or 'feature'."}), 400

    result = save_feedback(
        feedback_type=fb_type,
        message=data["message"],
        page=data.get("page", ""),
        role=user_role,
    )

    # Send instant email notification for new feedback
    _send_feedback_notification(result)

    return jsonify({"status": "ok", "feedback": result})


@app.route("/api/support/feedback", methods=["GET"])
@require_role("admin")
def api_support_feedback_get():
    """Return all feedback (admin only)."""
    rows = get_all_feedback()
    return jsonify({"feedback": rows})


@app.route("/api/support/feedback/<int:feedback_id>", methods=["PATCH"])
@require_role("admin")
def api_support_feedback_update(feedback_id):
    """Update feedback status (admin only)."""
    data = request.get_json(silent=True)
    if not data or not data.get("status"):
        return jsonify({"error": "Status is required."}), 400
    new_status = data.get("status", "")
    if new_status not in ("open", "resolved", "dismissed"):
        return jsonify({"error": "Status must be 'open', 'resolved', or 'dismissed'."}), 400
    ok = update_feedback_status(feedback_id, new_status)
    if not ok:
        return jsonify({"error": "Feedback not found."}), 404
    return jsonify({"status": "ok"})


@app.route("/api/support/test-digest", methods=["POST"])
@require_role("admin")
def api_test_digest():
    """Send the daily digest email right now (admin only)."""
    try:
        send_daily_report()
        return jsonify({"status": "ok", "message": "Daily digest sent."})
    except Exception as e:
        logger.exception("Test digest failed")
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# Routes — Handicap Calculator
# ---------------------------------------------------------------------------
@app.route("/handicaps")
def page_handicaps():
    return render_template("handicaps.html")


@app.route("/api/handicaps/players")
@require_role("member")
def api_handicap_players():
    """Return all players with their current handicap index.

    Also runs a quick auto-link pass for any unlinked players so that
    newly added customers are matched to their handicap records.
    """
    try:
        relink_all_unlinked_players()
    except Exception:
        logger.debug("Auto-link pass failed (non-critical)", exc_info=True)
    players = get_all_handicap_players()
    return jsonify(players)


@app.route("/api/handicaps/rounds")
@require_role("member")
def api_handicap_rounds():
    """Return rounds for a single player (?player=Name) or all rounds."""
    player_name = request.args.get("player")
    rounds = get_handicap_rounds(player_name=player_name)
    return jsonify(rounds)


@app.route("/api/handicaps/for-customer")
@require_role("view-only")
def api_handicap_for_customer():
    """Return handicap data for a customer by looking up their linked player name.

    Query: ?customer_name=John+Smith
    Returns: {player_name, handicap_index, rounds: [...], settings: {...}}
    or {error: "not linked"} if no handicap player is linked to this customer.
    """
    customer_name = request.args.get("customer_name", "").strip()
    if not customer_name:
        return jsonify({"error": "customer_name required"}), 400

    conn = get_connection()
    try:
        # Find linked player name for this customer
        link = conn.execute(
            "SELECT player_name FROM handicap_player_links "
            "WHERE LOWER(customer_name) = LOWER(?)",
            (customer_name,),
        ).fetchone()
        if not link:
            return jsonify({"error": "not linked", "customer_name": customer_name})

        player_name = link["player_name"]
    finally:
        conn.close()

    # Get their handicap index from the players list
    all_players = get_all_handicap_players()
    player_info = next((p for p in all_players if p["player_name"] == player_name), None)

    # Get their rounds
    rounds = get_handicap_rounds(player_name=player_name)

    # Get settings for the frontend calc
    cfg = get_handicap_settings()

    return jsonify({
        "player_name": player_name,
        "handicap_index": player_info["handicap_index"] if player_info else None,
        "active_rounds": player_info["active_rounds"] if player_info else 0,
        "total_rounds": player_info["total_rounds"] if player_info else 0,
        "rounds": rounds,
        "settings": cfg,
    })


@app.route("/api/handicaps/index-map")
@require_role("member")
def api_handicap_index_map():
    """Return a map of customer_name (lowercase) → handicap_index for all linked players.

    Lightweight endpoint used by the events page to display live HCP values.
    """
    players = get_all_handicap_players()
    index_map = {}
    for p in players:
        cname = p.get("customer_name")
        if cname and p.get("handicap_index") is not None:
            idx9 = p["handicap_index"]
            # A STARTING handicap is TYPED as an 18-hole number, and the
            # 9-hole half is derived from it. Re-deriving 18 from that
            # rounded half loses a tenth on every odd input — 12.5 came
            # back as 12.4 on the roster (Kerry 2026-07-31). Report the
            # number that was actually entered.
            src = p.get("handicap_source") or "computed"
            idx18 = (p.get("starting_handicap_18")
                     if src == "starting" and p.get("starting_handicap_18") is not None
                     else round(idx9 * 2, 1))
            entry = {
                "index_9": idx9,
                "index_18": idx18,
                # 'computed' from real rounds vs 'starting' placeholder.
                # The roster must never show a stand-in as if it were an
                # established index — Golf Genius flights people regardless
                # and that is exactly the behaviour we are not copying.
                "source": p.get("handicap_source") or "computed",
                "rounds": p.get("active_rounds") or 0,
                "customer_id": p.get("customer_id"),
            }
            index_map[cname.lower()] = entry
            # ALSO key by customer_id. A name key cannot survive the same
            # person being spelled differently on an order row than on the
            # canonical record, and that is precisely how a handicap set on
            # one screen goes missing on the other (Kerry 2026-07-31). Same
            # fix as the standings points map — guiding principle 6.
            if p.get("customer_id"):
                index_map[f"cid:{p['customer_id']}"] = entry
    return jsonify(index_map)


@app.route("/api/handicaps/rounds/<int:round_id>", methods=["DELETE"])
@require_role("manager")
def api_delete_handicap_round(round_id):
    """Delete a single round by id. Manager or admin."""
    if delete_handicap_round(round_id):
        return jsonify({"status": "ok"})
    return jsonify({"error": "not found"}), 404


@app.route("/api/handicaps/purge-invalid", methods=["POST"])
@require_role("admin")
def api_purge_invalid_rounds():
    """Delete all rounds with 18-hole ratings (rating > 50). Admin only."""
    conn = get_connection()
    try:
        rows = conn.execute(
            "SELECT id, player_name, round_date, course_name, rating "
            "FROM handicap_rounds WHERE rating > 50"
        ).fetchall()
        for row in rows:
            conn.execute("DELETE FROM handicap_rounds WHERE id = ?", (row["id"],))
        conn.commit()
        return jsonify({"status": "ok", "deleted": len(rows),
                        "rounds": [dict(r) for r in rows]})
    finally:
        conn.close()


@app.route("/api/handicaps/players/<path:player_name>", methods=["DELETE"])
@require_role("admin")
def api_delete_handicap_player(player_name):
    """Delete all rounds for a player. Admin only."""
    count = delete_all_handicap_rounds_for_player(player_name)
    return jsonify({"status": "ok", "deleted": count})


@app.route("/api/handicaps/settings", methods=["GET"])
@require_role("member")
def api_get_handicap_settings():
    """Return current handicap calculation settings."""
    return jsonify(get_handicap_settings())


@app.route("/api/handicaps/settings", methods=["PATCH"])
@require_role("admin")
def api_update_handicap_settings():
    """Update handicap calculation settings. Admin only."""
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "Request body must be JSON."}), 400
    allowed = {"lookback_months", "min_rounds", "multiplier"}
    filtered = {k: v for k, v in data.items() if k in allowed}
    if not filtered:
        return jsonify({"error": f"No valid settings keys. Allowed: {', '.join(allowed)}"}), 400
    # Validate types
    try:
        if "lookback_months" in filtered:
            v = int(filtered["lookback_months"])
            if v < 1 or v > 120:
                return jsonify({"error": "lookback_months must be 1–120"}), 400
        if "min_rounds" in filtered:
            v = int(filtered["min_rounds"])
            if v < 1 or v > 20:
                return jsonify({"error": "min_rounds must be 1–20"}), 400
        if "multiplier" in filtered:
            v = float(filtered["multiplier"])
            if v <= 0 or v > 2:
                return jsonify({"error": "multiplier must be between 0 and 2"}), 400
    except (ValueError, TypeError) as e:
        return jsonify({"error": f"Invalid value: {e}"}), 400
    update_handicap_settings(filtered)
    return jsonify({"status": "ok", "settings": get_handicap_settings()})


@app.route("/api/handicaps/import-preview", methods=["POST"])
@require_role("manager")
def api_handicap_import_preview():
    """Parse uploaded Excel and return headers + first 10 data rows for mapping."""
    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify({"error": "No file uploaded"}), 400

    import io
    from openpyxl import load_workbook
    try:
        wb = load_workbook(io.BytesIO(file.read()), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)

        # Read first few rows to detect header
        candidate_rows = []
        for raw_row in rows_iter:
            candidate_rows.append(raw_row)
            if len(candidate_rows) >= 5:
                break
        if not candidate_rows:
            return jsonify({"error": "Empty spreadsheet"}), 400

        # Find header row: first row where >40% of cells are non-empty
        header_idx = 0
        total_cols = len(candidate_rows[0])
        for i, row in enumerate(candidate_rows):
            non_empty = sum(1 for c in row if c is not None and str(c).strip())
            if non_empty >= max(2, total_cols * 0.4):
                header_idx = i
                break

        header_row = candidate_rows[header_idx]
        headers = [str(h).strip() if h else f"Column {i+1}"
                   for i, h in enumerate(header_row)]

        # Collect preview rows (up to 10)
        preview = []
        for row in candidate_rows[header_idx + 1:]:
            preview.append([str(c).strip() if c is not None else "" for c in row])
        for row in rows_iter:
            if len(preview) >= 10:
                break
            preview.append([str(c).strip() if c is not None else "" for c in row])

        # Count total data rows (re-open for accurate count)
        wb.close()

        # Auto-detect column mapping from header names
        def _find_col(candidates):
            for cand in candidates:
                for idx, h in enumerate(headers):
                    if h.lower() == cand.lower():
                        return idx
            return None

        auto_mapping = {
            "player_name":     _find_col(["name", "player", "player_name", "player name"]),
            "player_email":    _find_col(["email", "player_email", "player email", "e-mail"]),
            "round_date":      _find_col(["play at", "date", "round_date", "played"]),
            "round_id":        _find_col(["round id", "round_id", "roundid"]),
            "course_name":     _find_col(["course name", "course", "course_name"]),
            "tee_name":        _find_col(["tee name", "tee", "tee_name", "tees"]),
            "adjusted_score":  _find_col(["adjusted score", "adj score", "score", "adjusted_score"]),
            "rating":          _find_col(["rating", "course rating"]),
            "slope":           _find_col(["slope", "slope rating"]),
            "differential":    _find_col(["differential", "diff"]),
        }

        return jsonify({
            "headers": headers,
            "preview": preview,
            "auto_mapping": auto_mapping,
        })
    except Exception as e:
        logger.exception("Handicap import preview failed")
        return jsonify({"error": f"Failed to parse file: {str(e)}"}), 400


@app.route("/api/handicaps/import", methods=["POST"])
@require_role("manager")
def api_handicap_import():
    """Import handicap rounds from uploaded Excel with column mapping.

    Accepts multipart/form-data with:
      - file: the Excel file
      - mapping: JSON object {field_name: col_index, ...}
    """
    file = request.files.get("file")
    mapping_json = request.form.get("mapping", "{}")
    if not file or not file.filename:
        return jsonify({"error": "No file uploaded"}), 400

    try:
        mapping = json.loads(mapping_json)
    except json.JSONDecodeError:
        return jsonify({"error": "Invalid mapping JSON"}), 400

    required = {"player_name", "round_date", "adjusted_score", "rating", "slope"}
    missing = required - set(k for k, v in mapping.items() if v is not None)
    if missing:
        return jsonify({"error": f"Required column mapping missing: {', '.join(missing)}"}), 400

    import io
    from openpyxl import load_workbook
    try:
        wb = load_workbook(io.BytesIO(file.read()), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)

        # Skip to header row (same detection as preview)
        candidate_rows = []
        for raw_row in rows_iter:
            candidate_rows.append(raw_row)
            if len(candidate_rows) >= 5:
                break
        header_idx = 0
        if candidate_rows:
            total_cols = len(candidate_rows[0])
            for i, row in enumerate(candidate_rows):
                non_empty = sum(1 for c in row if c is not None and str(c).strip())
                if non_empty >= max(2, total_cols * 0.4):
                    header_idx = i
                    break

        # Build list of data rows
        all_data_rows = list(candidate_rows[header_idx + 1:]) + list(rows_iter)
        wb.close()

        # Parse date helper — handles datetime objects, MM/DD/YYYY, YYYY-MM-DD,
        # and sparse "D-Mon" / "D-Mon-YY" strings (e.g. "7-Feb", "7-Feb-25")
        def _parse_date(val):
            if val is None:
                return ""
            s = str(val).strip()
            # openpyxl may return a datetime object directly
            if hasattr(val, "strftime"):
                return val.strftime("%Y-%m-%d")
            # MM/DD/YYYY
            if "/" in s:
                parts = s.split("/")
                if len(parts) == 3:
                    m, d, y = parts
                    return f"{y.zfill(4)}-{m.zfill(2)}-{d.zfill(2)}"
            # D-Mon-YY or D-Mon (e.g. "7-Feb-25" or "7-Feb")
            if "-" in s:
                from datetime import datetime as _dt
                for fmt in ("%d-%b-%y", "%d-%b-%Y", "%d-%b"):
                    try:
                        parsed = _dt.strptime(s, fmt)
                        if fmt == "%d-%b":
                            # No year supplied — use current year
                            parsed = parsed.replace(year=_dt.now().year)
                        return parsed.strftime("%Y-%m-%d")
                    except ValueError:
                        continue
            return s  # already YYYY-MM-DD or unknown format

        rounds = []
        last_player_name = None  # support fill-down name format
        last_player_email = None  # support fill-down email format
        for row in all_data_rows:
            def _get(field):
                idx = mapping.get(field)
                if idx is None or idx >= len(row):
                    return None
                val = row[idx]
                return str(val).strip() if val is not None else None

            player_name = _get("player_name")
            if player_name:
                last_player_name = player_name
            elif last_player_name:
                player_name = last_player_name

            if not player_name:
                continue

            # Email: fill-down like player_name (email only appears on first row per player)
            player_email = _get("player_email")
            if player_email:
                last_player_email = player_email
            elif not player_email and player_name == last_player_name:
                player_email = last_player_email

            rounds.append({
                "player_name": player_name,
                "player_email": player_email,
                "round_date":  _parse_date(row[mapping["round_date"]] if mapping.get("round_date") is not None and mapping["round_date"] < len(row) else None),
                "round_id":    _get("round_id"),
                "course_name": _get("course_name"),
                "tee_name":    _get("tee_name"),
                "adjusted_score": _get("adjusted_score"),
                "rating":      _get("rating"),
                "slope":       _get("slope"),
                "differential": _get("differential"),
            })

        if not rounds:
            return jsonify({"error": "No data rows found in the file"}), 400

        result = import_handicap_rounds(rounds)
        return jsonify(result)

    except Exception as e:
        logger.exception("Handicap import failed")
        return jsonify({"error": f"Import failed: {str(e)}"}), 500


# ---------------------------------------------------------------------------
# Routes — Golf Genius Sync
# ---------------------------------------------------------------------------

@app.route("/api/handicaps/auto-link", methods=["POST"])
@require_role("admin")
def api_handicap_auto_link():
    """Re-attempt matching all unlinked handicap players to customer records."""
    result = relink_all_unlinked_players()
    return jsonify(result)


@app.route("/api/handicaps/link-player", methods=["POST"])
@require_role("manager")
def api_handicap_link_player():
    """Link a single handicap player to a customer name."""
    data = request.get_json(force=True)
    player_name = (data.get("player_name") or "").strip()
    customer_name = (data.get("customer_name") or "").strip()
    if not player_name:
        return jsonify({"error": "player_name required"}), 400
    if not customer_name:
        return jsonify({"error": "customer_name required"}), 400

    conn = get_connection()
    try:
        conn.execute(
            "INSERT INTO handicap_player_links (player_name, customer_name) VALUES (?, ?) "
            "ON CONFLICT(player_name) DO UPDATE SET customer_name = excluded.customer_name, "
            "linked_at = datetime('now')",
            (player_name, customer_name),
        )
        conn.commit()
    finally:
        conn.close()
    return jsonify({"status": "ok", "player_name": player_name, "customer_name": customer_name})


@app.route("/api/handicaps/repair-swapped-links", methods=["POST"])
@require_role("admin")
def api_repair_swapped_links():
    """Fix links where player_name and customer_name were swapped due to a bug.

    Detection: if the stored player_name exists as a customer in items but NOT
    in handicap_rounds, and the stored customer_name exists in handicap_rounds
    but NOT as a customer in items — the link is backwards and needs swapping.
    """
    conn = get_connection()
    try:
        links = conn.execute(
            "SELECT player_name, customer_name FROM handicap_player_links "
            "WHERE customer_name IS NOT NULL"
        ).fetchall()

        # Build lookup sets
        hcp_players = {r["player_name"].lower() for r in conn.execute(
            "SELECT DISTINCT player_name FROM handicap_rounds"
        ).fetchall()}
        item_customers = {r["customer"].lower() for r in conn.execute(
            "SELECT DISTINCT customer FROM items WHERE customer IS NOT NULL"
        ).fetchall()}

        swapped = []
        for lnk in links:
            pn = lnk["player_name"]
            cn = lnk["customer_name"]
            pn_l = pn.lower()
            cn_l = cn.lower()

            # If stored player_name looks like a customer (in items) but not a
            # handicap player, AND stored customer_name looks like a handicap
            # player but not a customer — they're swapped.
            if (pn_l in item_customers and pn_l not in hcp_players and
                    cn_l in hcp_players and cn_l not in item_customers):
                swapped.append((pn, cn))

        # Fix the swapped links
        for old_pn, old_cn in swapped:
            # Delete the wrong row, insert corrected one
            conn.execute(
                "DELETE FROM handicap_player_links WHERE player_name = ?",
                (old_pn,),
            )
            conn.execute(
                "INSERT INTO handicap_player_links (player_name, customer_name) "
                "VALUES (?, ?) ON CONFLICT(player_name) DO UPDATE SET "
                "customer_name = excluded.customer_name, linked_at = datetime('now')",
                (old_cn, old_pn),
            )

        conn.commit()
    finally:
        conn.close()

    return jsonify({
        "status": "ok",
        "repaired": len(swapped),
        "details": [f"{pn} ↔ {cn}" for pn, cn in swapped],
    })


@app.route("/api/handicaps/unlink-player", methods=["POST"])
@require_role("manager")
def api_handicap_unlink_player():
    """Unlink a handicap player from their customer."""
    data = request.get_json(force=True)
    player_name = (data.get("player_name") or "").strip()
    if not player_name:
        return jsonify({"error": "player_name required"}), 400

    conn = get_connection()
    try:
        conn.execute(
            "UPDATE handicap_player_links SET customer_name = NULL WHERE player_name = ?",
            (player_name,),
        )
        conn.commit()
    finally:
        conn.close()
    return jsonify({"status": "ok", "player_name": player_name})


@app.route("/api/customers/names")
@require_role("view-only")
def api_customer_names():
    """Return a sorted list of unique customer names for autocomplete/linking."""
    conn = get_connection()
    try:
        rows = conn.execute(
            "SELECT DISTINCT customer FROM items WHERE customer IS NOT NULL AND TRIM(customer) != '' ORDER BY customer COLLATE NOCASE"
        ).fetchall()
    finally:
        conn.close()
    return jsonify([r["customer"] for r in rows])


@app.route("/api/handicaps/link-debug")
@require_role("admin")
def api_handicap_link_debug():
    """Full export diagnostic: show every handicap player's link status, email, chapter, and index."""
    from email_parser.database import _connect, get_all_handicap_players
    all_players = get_all_handicap_players()
    player_map = {p["player_name"]: p for p in all_players}

    with _connect() as conn:
        links = conn.execute(
            "SELECT player_name, customer_name FROM handicap_player_links ORDER BY player_name"
        ).fetchall()
        link_map = {r["player_name"]: r["customer_name"] for r in links}

        details = []
        for pname in sorted(player_map.keys(), key=str.lower):
            p = player_map[pname]
            cname = link_map.get(pname)
            email_items = None
            email_alias = None
            chapter = None
            if cname:
                row = conn.execute(
                    "SELECT customer_email FROM items WHERE LOWER(customer)=LOWER(?) "
                    "AND customer_email IS NOT NULL AND TRIM(customer_email) != '' "
                    "ORDER BY id DESC LIMIT 1", (cname,)
                ).fetchone()
                email_items = row["customer_email"].strip().lower() if row else None
                row2 = conn.execute(
                    "SELECT alias_value FROM customer_aliases "
                    "WHERE LOWER(customer_name)=LOWER(?) AND alias_type='email' LIMIT 1",
                    (cname,)
                ).fetchone()
                email_alias = row2["alias_value"].strip().lower() if row2 else None
                row3 = conn.execute(
                    "SELECT chapter FROM items WHERE LOWER(customer)=LOWER(?) "
                    "AND chapter IS NOT NULL AND TRIM(chapter) != '' "
                    "ORDER BY id DESC LIMIT 1", (cname,)
                ).fetchone()
                chapter = row3["chapter"] if row3 else None

            email = email_items or email_alias or None
            idx = p["handicap_index"]
            would_export = bool(cname and email and idx is not None)

            details.append({
                "player_name": pname,
                "customer_name": cname,
                "linked": bool(cname),
                "email_from_items": email_items,
                "email_from_aliases": email_alias,
                "email": email,
                "chapter": chapter,
                "handicap_index_9": idx,
                "handicap_index_18": round(idx * 2, 1) if idx is not None else None,
                "would_export": would_export,
                "missing": (
                    "not linked" if not cname else
                    "no email" if not email else
                    "no index" if idx is None else
                    None
                ),
            })

        summary = {
            "total_players": len(details),
            "linked": sum(1 for d in details if d["linked"]),
            "unlinked": sum(1 for d in details if not d["linked"]),
            "have_email": sum(1 for d in details if d["email"]),
            "have_index": sum(1 for d in details if d["handicap_index_9"] is not None),
            "would_export": sum(1 for d in details if d["would_export"]),
            "missing_email": [d["player_name"] for d in details if d["linked"] and not d["email"]],
            "missing_index": [d["player_name"] for d in details if d["linked"] and d["email"] and d["handicap_index_9"] is None],
        }

    return jsonify({"summary": summary, "players": details})


@app.route("/api/handicaps/unlinked-players")
@require_role("admin")
def api_handicap_unlinked_players():
    """Return handicap players with no linked customer record."""
    conn = get_connection()
    try:
        rows = conn.execute(
            """SELECT l.player_name, l.customer_name
               FROM handicap_player_links l
               WHERE l.customer_name IS NULL"""
        ).fetchall()
    finally:
        conn.close()
    return jsonify([{"player_name": r["player_name"]} for r in rows])


@app.route("/api/handicaps/create-customers-for-unlinked", methods=["POST"])
@require_role("admin")
def api_create_customers_for_unlinked():
    """Auto-create archived customer records for all unlinked handicap players."""
    conn = get_connection()
    try:
        unlinked = conn.execute(
            "SELECT player_name FROM handicap_player_links WHERE customer_name IS NULL"
        ).fetchall()
    finally:
        conn.close()

    created = 0
    linked = 0
    skipped = 0
    for row in unlinked:
        player_name = row["player_name"]
        # Try to find an existing customer with this name
        conn2 = get_connection()
        try:
            existing = conn2.execute(
                "SELECT customer FROM items WHERE customer = ? COLLATE NOCASE LIMIT 1",
                (player_name,)
            ).fetchone()
        finally:
            conn2.close()

        if existing:
            # Link to existing customer, update link
            conn3 = get_connection()
            try:
                conn3.execute(
                    "UPDATE handicap_player_links SET customer_name = ? WHERE player_name = ?",
                    (existing["customer"], player_name)
                )
                conn3.commit()
            finally:
                conn3.close()
            linked += 1
        else:
            # Create a new archived customer record
            parts = player_name.split(None, 1)
            first_name = parts[0] if parts else player_name
            last_name = parts[1] if len(parts) > 1 else ""
            today = today_central_str()

            from email_parser.database import _resolve_or_create_customer
            conn3 = get_connection()
            try:
                # This bulk tool has no transaction to trigger the normal
                # _resolve_or_create_customer() call, so resolve/create the
                # canonical customers row now — otherwise customer_id stays
                # NULL forever (Membership Terms, roles, and status edits
                # all require it).
                cid = _resolve_or_create_customer(
                    conn3, player_name, None,
                    first_name=first_name, last_name=last_name or None,
                )
                conn3.execute(
                    """INSERT INTO items (email_uid, item_index, merchant, customer, first_name,
                       last_name, order_date, item_name, archived, customer_id)
                       VALUES (?, 0, 'Handicap Import', ?, ?, ?, ?, 'Handicap Import', 1, ?)""",
                    (f"handicap_import_{player_name}_{today}", player_name,
                     first_name, last_name, today, cid)
                )
                # Now link them
                conn3.execute(
                    "UPDATE handicap_player_links SET customer_name = ? WHERE player_name = ?",
                    (player_name, player_name)
                )
                conn3.commit()
            finally:
                conn3.close()
            created += 1

    return jsonify({"created": created, "linked": linked, "total": len(unlinked)})


@app.route("/api/handicaps/export-preview")
@require_role("manager")
def api_handicap_export_preview():
    """JSON preview of what the CSV export would contain, with diagnostics."""
    chapter = request.args.get("chapter", "").strip()
    data = get_handicap_export_data(chapter=chapter if chapter else None)
    return jsonify(data)


@app.route("/api/handicaps/export-csv")
@require_role("manager")
def api_handicap_export_csv():
    """Download a Golf Genius-ready CSV for the given chapter.

    Query params:
        chapter: "San Antonio" | "Austin" | (omit for all)
    """
    chapter = request.args.get("chapter", "").strip()
    data = get_handicap_export_data(chapter=chapter if chapter else None)

    import io as _io, csv as _csv
    buf = _io.StringIO()
    writer = _csv.writer(buf)
    writer.writerow(["Last", "First", "Suffix", "Chapter", "Handicap Index", "Email"])
    for row in data["rows"]:
        writer.writerow([row["last_name"], row["first_name"], row["suffix"],
                         row["chapter"], row["handicap_index"], row["email"]])

    chapter_slug = chapter.lower().replace(" ", "_") if chapter else "all"
    filename = f"tgf_handicaps_{chapter_slug}_{datetime.now().strftime('%Y%m%d')}.csv"
    return Response(
        buf.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# Track running sync jobs (chapter key → {"status", "message", "timestamp"})
_gg_sync_jobs: dict[str, dict] = {}


@app.route("/api/handicaps/sync-golf-genius", methods=["POST"])
@require_role("admin")
def api_sync_golf_genius():
    """Trigger an on-demand Golf Genius handicap sync for a chapter.

    Body JSON:
        {"chapter": "San Antonio" | "Austin" | "all",
         "test_player_email": "email@example.com"}   # optional: limit to 1 player for testing
    """
    from golf_genius_sync import sync_handicaps_to_league
    import threading

    body = request.get_json(silent=True) or {}
    chapter = body.get("chapter", "").strip()
    test_player_email = (body.get("test_player_email") or "").strip().lower() or None

    gg_email = os.getenv("GOLF_GENIUS_EMAIL", "").strip()
    gg_password = os.getenv("GOLF_GENIUS_PASSWORD", "").strip()
    sa_league_id = os.getenv("GOLF_GENIUS_SA_LEAGUE_ID", "514047").strip()
    austin_league_id = os.getenv("GOLF_GENIUS_AUSTIN_LEAGUE_ID", "514705").strip()

    if not gg_email or not gg_password:
        return jsonify({
            "status": "error",
            "message": "GOLF_GENIUS_EMAIL and GOLF_GENIUS_PASSWORD environment variables are not set",
        }), 400

    chapters_to_sync = []
    if chapter.lower() in ("san antonio", "sa", ""):
        chapters_to_sync.append(("San Antonio", sa_league_id, "san_antonio"))
    if chapter.lower() in ("austin", "atx", ""):
        chapters_to_sync.append(("Austin", austin_league_id, "austin"))

    if not chapters_to_sync:
        return jsonify({"status": "error", "message": f"Unknown chapter: {chapter}"}), 400

    # Mark jobs as running
    for _, _, key in chapters_to_sync:
        _gg_sync_jobs[key] = {
            "status": "running",
            "message": "Sync in progress…",
            "timestamp": datetime.utcnow().isoformat(),
            "rows_submitted": 0,
        }

    def _run_sync():
        for chap, league_id, key in chapters_to_sync:
            try:
                export = get_handicap_export_data(
                    chapter=chap,
                    test_player_email=test_player_email,
                )
                rows = export["rows"]
                if not rows:
                    msg = (
                        f"No player found with email '{test_player_email}' in {chap}"
                        if test_player_email
                        else f"No players with email + handicap index for {chap}"
                    )
                    _gg_sync_jobs[key] = {
                        "status": "skipped",
                        "message": msg,
                        "rows_submitted": 0,
                        "timestamp": datetime.utcnow().isoformat(),
                    }
                    continue

                result = sync_handicaps_to_league(
                    rows=rows,
                    league_id=league_id,
                    email=gg_email,
                    password=gg_password,
                )
                _gg_sync_jobs[key] = result
            except Exception as exc:
                logger.exception("GG sync error for %s", chap)
                _gg_sync_jobs[key] = {
                    "status": "error",
                    "message": str(exc),
                    "rows_submitted": 0,
                    "timestamp": datetime.utcnow().isoformat(),
                }

        # Persist results
        try:
            update_handicap_settings({"last_gg_sync": json.dumps(_gg_sync_jobs)})
        except Exception:
            logger.warning("Failed to persist GG sync results", exc_info=True)

    threading.Thread(target=_run_sync, daemon=True).start()
    return jsonify({"status": "started", "chapters": [c[0] for c in chapters_to_sync]})


@app.route("/api/handicaps/sync-status")
@require_role("manager")
def api_handicap_sync_status():
    """Return the current/last Golf Genius sync status."""
    # Merge in-memory jobs with persisted last result
    persisted = {}
    try:
        settings = get_handicap_settings()
        raw = settings.get("last_gg_sync")
        if raw:
            persisted = json.loads(raw)
    except Exception:
        logger.debug("Failed to load persisted GG sync results", exc_info=True)

    merged = {**persisted, **_gg_sync_jobs}
    return jsonify(merged)


# ---------------------------------------------------------------------------
# Routes — Handicap Email Cards
# ---------------------------------------------------------------------------

@app.route("/api/handicaps/preview-email", methods=["POST"])
@require_role("manager")
def api_handicap_preview_email():
    """Preview a handicap card email for a player."""
    data = request.get_json(silent=True) or {}
    player_name = (data.get("player_name") or "").strip()
    if not player_name:
        return jsonify({"error": "player_name is required"}), 400

    card_data = build_handicap_card_data(player_name)
    html = build_handicap_card_html(card_data)

    first = card_data.get("first_name") or ""
    last = card_data.get("last_name") or ""
    display = f"{first} {last}".strip() or player_name
    subject = f"TGF Handicap Update \u2014 {display}"

    return jsonify({
        "html": html,
        "subject": subject,
        "email": card_data.get("email") or "",
        "has_email": bool(card_data.get("email")),
        "has_index": card_data.get("handicap_index_9") is not None,
        "player_name": player_name,
    })


@app.route("/api/handicaps/send-email", methods=["POST"])
@require_role("manager")
def api_handicap_send_email():
    """Send a handicap card email to a single player."""
    data = request.get_json(silent=True) or {}
    player_name = (data.get("player_name") or "").strip()
    if not player_name:
        return jsonify({"error": "player_name is required"}), 400

    card_data = build_handicap_card_data(player_name)

    email = card_data.get("email") or ""
    if not email:
        return jsonify({"error": f"No email address found for {player_name}"}), 400

    if card_data.get("handicap_index_9") is None:
        return jsonify({"error": f"{player_name} does not have a handicap index yet"}), 400

    html = build_handicap_card_html(card_data)

    first = card_data.get("first_name") or ""
    last = card_data.get("last_name") or ""
    display = f"{first} {last}".strip() or player_name
    subject = f"TGF Handicap Update \u2014 {display}"

    tenant_id = os.getenv("AZURE_TENANT_ID")
    client_id = os.getenv("AZURE_CLIENT_ID")
    client_secret = os.getenv("AZURE_CLIENT_SECRET")
    from_address = os.getenv("EMAIL_ADDRESS")

    if not all([tenant_id, client_id, client_secret, from_address]):
        return jsonify({"error": "Email credentials not configured on server"}), 500

    ok = send_mail_graph(
        tenant_id=tenant_id,
        client_id=client_id,
        client_secret=client_secret,
        from_address=from_address,
        to_address=email,
        subject=subject,
        html_body=html,
    )

    status = "sent" if ok else "failed"
    try:
        log_message({
            "event_name": "handicap-card",
            "channel": "email",
            "recipient_name": player_name,
            "recipient_address": email,
            "subject": subject,
            "body_preview": f"Handicap card: {card_data.get('handicap_index_9')}N",
            "status": status,
            "sent_by": session.get("role", "unknown"),
        })
    except Exception:
        logger.warning("Failed to log handicap card email", exc_info=True)

    if ok:
        return jsonify({"status": "ok", "email": email})
    return jsonify({"error": "Failed to send email — check server logs"}), 500


@app.route("/api/handicaps/send-bulk-email", methods=["POST"])
@require_role("manager")
def api_handicap_send_bulk_email():
    """Send handicap card emails to eligible players.

    Filters: chapter (optional), event_name (optional).
    When event_name is given, only players registered for that event
    who also have an established TGF handicap will receive cards.
    """
    import time as _time

    data = request.get_json(silent=True) or {}
    chapter = (data.get("chapter") or "").strip() or None
    event_name = (data.get("event_name") or "").strip() or None

    tenant_id = os.getenv("AZURE_TENANT_ID")
    client_id = os.getenv("AZURE_CLIENT_ID")
    client_secret = os.getenv("AZURE_CLIENT_SECRET")
    from_address = os.getenv("EMAIL_ADDRESS")

    if not all([tenant_id, client_id, client_secret, from_address]):
        return jsonify({"error": "Email credentials not configured on server"}), 500

    export = get_handicap_export_data(chapter=chapter)
    eligible_rows = export.get("rows") or []

    # If filtering by event, restrict to players registered for that event
    # and compute event-specific skip counts
    skipped_no_email = 0
    skipped_no_index = 0
    if event_name:
        all_items = get_all_items()
        aliases = get_all_event_aliases()
        # Collect customer names registered for this event (active only)
        event_customers = set()
        for item in all_items:
            iname = item.get("item_name") or ""
            if iname.lower() == event_name.lower() or (aliases.get(iname) or "").lower() == event_name.lower():
                if item.get("transaction_status") in (None, "active", "rsvp_only", "gg_rsvp"):
                    cname = (item.get("customer") or "").strip().lower()
                    if cname:
                        event_customers.add(cname)

        # Build player_name → customer_name map from handicap links
        conn = get_connection()
        try:
            links = conn.execute(
                "SELECT player_name, customer_name FROM handicap_player_links "
                "WHERE customer_name IS NOT NULL"
            ).fetchall()
        finally:
            conn.close()
        player_to_customer = {r["player_name"]: r["customer_name"] for r in links}

        # Filter eligible rows to only those whose linked customer is in the event
        eligible_rows = [
            r for r in eligible_rows
            if player_to_customer.get(r["player_name"], "").strip().lower() in event_customers
        ]

        # Count event-specific skips: event registrants not in eligible list
        eligible_customers = {
            player_to_customer.get(r["player_name"], "").strip().lower()
            for r in eligible_rows
        }
        for cname_l in event_customers:
            if cname_l not in eligible_customers:
                skipped_no_index += 1  # no handicap, no link, or no email
    else:
        skipped_no_email = len(export.get("no_email") or [])
        skipped_no_index = len(export.get("no_index") or [])

    sent = 0
    failed = 0
    errors = []
    role = session.get("role", "unknown")

    for i, row in enumerate(eligible_rows):
        pname = row["player_name"]
        email = row.get("email") or ""
        if not email:
            continue

        try:
            card_data = build_handicap_card_data(pname)
            if card_data.get("handicap_index_9") is None:
                continue

            html = build_handicap_card_html(card_data)
            first = card_data.get("first_name") or ""
            last = card_data.get("last_name") or ""
            display = f"{first} {last}".strip() or pname
            subject = f"TGF Handicap Update \u2014 {display}"

            ok = send_mail_graph(
                tenant_id=tenant_id,
                client_id=client_id,
                client_secret=client_secret,
                from_address=from_address,
                to_address=email,
                subject=subject,
                html_body=html,
            )

            status = "sent" if ok else "failed"
            if ok:
                sent += 1
            else:
                failed += 1
                errors.append({"player": pname, "email": email, "error": "send_mail_graph returned False"})

            try:
                log_message({
                    "event_name": "handicap-card",
                    "channel": "email",
                    "recipient_name": pname,
                    "recipient_address": email,
                    "subject": subject,
                    "body_preview": f"Handicap card: {card_data.get('handicap_index_9')}N",
                    "status": status,
                    "sent_by": role,
                })
            except Exception:
                logger.warning("Failed to log handicap card email for %s", pname, exc_info=True)

        except Exception as exc:
            failed += 1
            errors.append({"player": pname, "email": email, "error": str(exc)})

        # Throttle to avoid rate limiting (300ms between sends)
        if i < len(eligible_rows) - 1:
            _time.sleep(0.3)

    return jsonify({
        "status": "ok" if failed == 0 else "partial",
        "sent": sent,
        "failed": failed,
        "skipped_no_email": skipped_no_email,
        "skipped_no_index": skipped_no_index,
        "total_eligible": len(eligible_rows),
        "errors": errors[:20],  # limit error details
    })


# ---------------------------------------------------------------------------
# Routes — Participation Analysis
# ---------------------------------------------------------------------------
# Identifies players by last-event date, 12-month frequency, and trend vs the
# prior 12 months. Powers the /participation page used to spot dormant players
# and send re-engagement emails. "Event" = a non-membership, non-season-contest
# items row with transaction_status active or rsvp_only and parent_item_id NULL
# (skips child payment rows). Audience: every customer whose canonical status
# is not 'former' (MEMBER / MEMBER+ / GUEST / 1st TIMER), plus customers with
# no status row at all.

# Default re-engagement email template. Edited per-send in the composer.
PARTICIPATION_DEFAULT_SUBJECT = "We miss you on the tee, {first_name}"
PARTICIPATION_DEFAULT_BODY_HTML = """<p>Hi {first_name},</p>

<p>It's been about <strong>{days_since} days</strong> since your last
TGF round{last_event_phrase}, and we wanted to check in.</p>

<p>The {chapter} chapter has events booked over the next several weeks
and we'd love to get you back out there. Tee times, side games, and
the usual good company — same as you remember.</p>

<p>Take a look at what's on the schedule and grab a spot:</p>

<p style="margin:1.25rem 0;">
  <a href="https://thegolffellowship.com/events"
     style="display:inline-block;background:#16a34a;color:#fff;padding:0.7rem 1.4rem;border-radius:6px;text-decoration:none;font-weight:600;">
    See upcoming events
  </a>
</p>

<p>If something's keeping you off the course — schedule, handicap,
travel, anything — just reply to this email and let us know. We're
happy to help.</p>

<p>See you soon,<br>
The Golf Fellowship</p>"""


def _participation_event_filter_sql(alias: str = "i") -> str:
    """SQL fragment selecting event-participation items only.

    Excludes membership renewals, season contest enrollments, and child
    payment rows. Both paid (active) and RSVP-only rows count as
    "played" for the purposes of last-event / frequency.
    """
    return f"""
        {alias}.customer_id IS NOT NULL
        AND COALESCE({alias}.transaction_status, 'active') IN ('active', 'rsvp_only')
        AND UPPER(COALESCE({alias}.item_name, '')) NOT LIKE '%MEMBERSHIP%'
        AND UPPER(COALESCE({alias}.item_name, '')) NOT LIKE '%SEASON CONTEST%'
        AND {alias}.parent_item_id IS NULL
    """


def _get_participation_rows(conn: sqlite3.Connection) -> list[dict]:
    """Return one row per active customer with last-event + frequency stats.

    "Played" requires an actual event date — items.item_name must join to a
    row in the events table that has a non-null event_date <= today. Items
    with no matching events row are silently dropped from the play counts
    (rather than falling back to order_date, which makes purchase timing
    masquerade as play timing). The price of that strictness is some legacy
    items don't count; the alternative was showing purchase dates as if
    they were play dates and letting future registrations look like recent
    plays — which they aren't.

    next_event is the soonest upcoming registration (events.event_date >
    today) per customer, surfaced as its own column so a player who's
    re-engaged after a long dormancy is visibly distinct from one who hasn't.
    """
    today = today_central_str()
    ev = _participation_event_filter_sql("i")

    rows = conn.execute(
        f"""
        WITH latest_status AS (
            SELECT cs.customer_id, s.status_name
              FROM customer_statuses cs
              JOIN statuses s ON s.status_id = cs.status_id
              JOIN (
                  SELECT customer_id, MAX(id) AS max_id
                    FROM customer_statuses
                   GROUP BY customer_id
              ) latest ON latest.customer_id = cs.customer_id
                      AND latest.max_id = cs.id
        ),
        primary_email AS (
            SELECT ce.customer_id, MIN(ce.email) AS email
              FROM customer_emails ce
             WHERE ce.is_primary = 1
             GROUP BY ce.customer_id
        ),
        played_items AS (
            -- Every items row that joins to an events row with a known,
            -- non-future event_date. This is the strict definition of
            -- "played" — items without a matching events row are dropped.
            -- items.event_id is the authoritative link (set at insert /
            -- backfill); the TRIM + COLLATE NOCASE name join is only the
            -- fallback for legacy rows without one. Name-only joins missed
            -- real plays whenever the item snapshot and the events row
            -- drifted apart (e.g. "s9.16 TPC OAKS" vs the renamed
            -- "s9.16 TPC San Antonio | Oaks" — the Arias case, 2026-07-08).
            SELECT i.customer_id,
                   e.event_date AS played_date,
                   e.item_name  AS item_name,   -- canonical name from events
                   i.id         AS item_id
              FROM items i
              JOIN events e
                ON e.id = i.event_id
                OR (i.event_id IS NULL
                    AND TRIM(e.item_name) = TRIM(i.item_name) COLLATE NOCASE)
             WHERE {ev}
               AND e.event_date IS NOT NULL
               AND e.event_date <= DATE(?)
        ),
        upcoming_items AS (
            -- Every registration whose event_date is strictly in the future.
            SELECT i.customer_id,
                   e.event_date AS upcoming_date,
                   e.item_name  AS item_name,
                   i.id         AS item_id
              FROM items i
              JOIN events e
                ON e.id = i.event_id
                OR (i.event_id IS NULL
                    AND TRIM(e.item_name) = TRIM(i.item_name) COLLATE NOCASE)
             WHERE {ev}
               AND e.event_date IS NOT NULL
               AND e.event_date > DATE(?)
        ),
        last_event AS (
            SELECT pi.customer_id,
                   MAX(pi.played_date) AS last_event_date,
                   COUNT(*)            AS plays_lifetime,
                   -- Pick the item_name from the row that owns MAX(played_date);
                   -- tie-break by items.id DESC so multi-event days resolve
                   -- to the most recently inserted registration.
                   (SELECT pi2.item_name FROM played_items pi2
                     WHERE pi2.customer_id = pi.customer_id
                     ORDER BY pi2.played_date DESC, pi2.item_id DESC
                     LIMIT 1) AS last_event_name
              FROM played_items pi
             GROUP BY pi.customer_id
        ),
        next_event AS (
            SELECT ui.customer_id,
                   MIN(ui.upcoming_date) AS next_event_date,
                   (SELECT ui2.item_name FROM upcoming_items ui2
                     WHERE ui2.customer_id = ui.customer_id
                     ORDER BY ui2.upcoming_date ASC, ui2.item_id ASC
                     LIMIT 1) AS next_event_name
              FROM upcoming_items ui
             GROUP BY ui.customer_id
        ),
        plays_12 AS (
            SELECT pi.customer_id, COUNT(*) AS n
              FROM played_items pi
             WHERE pi.played_date >= DATE(?, '-12 months')
             GROUP BY pi.customer_id
        ),
        plays_prior_12 AS (
            SELECT pi.customer_id, COUNT(*) AS n
              FROM played_items pi
             WHERE pi.played_date >= DATE(?, '-24 months')
               AND pi.played_date <  DATE(?, '-12 months')
             GROUP BY pi.customer_id
        )
        SELECT
            c.customer_id,
            TRIM(COALESCE(c.first_name, '') || ' ' || COALESCE(c.last_name, '')) AS name,
            c.first_name, c.last_name,
            c.chapter,
            COALESCE(ls.status_name, c.current_player_status) AS status_raw,
            pe.email AS email,
            le.last_event_date,
            le.last_event_name,
            ne.next_event_date,
            ne.next_event_name,
            COALESCE(le.plays_lifetime, 0) AS plays_lifetime,
            COALESCE(p12.n, 0)             AS plays_12mo,
            COALESCE(pp12.n, 0)            AS plays_prior_12mo
          FROM customers c
          LEFT JOIN latest_status   ls   ON ls.customer_id   = c.customer_id
          LEFT JOIN primary_email   pe   ON pe.customer_id   = c.customer_id
          LEFT JOIN last_event      le   ON le.customer_id   = c.customer_id
          LEFT JOIN next_event      ne   ON ne.customer_id   = c.customer_id
          LEFT JOIN plays_12        p12  ON p12.customer_id  = c.customer_id
          LEFT JOIN plays_prior_12  pp12 ON pp12.customer_id = c.customer_id
         WHERE COALESCE(c.account_status, 'active') = 'active'
           AND COALESCE(ls.status_name, c.current_player_status, '') NOT IN
               ('former', 'expired_member', 'inactive')
         ORDER BY le.last_event_date IS NULL, le.last_event_date DESC, c.last_name COLLATE NOCASE
        """,
        (today, today, today, today, today),
    ).fetchall()

    # Map current_player_status / status_name → user-facing label.
    label_map = {
        "member":         "MEMBER",
        "member_plus":    "MEMBER+",
        "guest":          "GUEST",
        "1st_timer":      "1st TIMER",
        "active_member":  "MEMBER",
        "active_guest":   "GUEST",
        "first_timer":    "1st TIMER",
    }

    out = []
    for r in rows:
        d = dict(r)
        d["status"] = label_map.get((d.get("status_raw") or "").lower(),
                                    (d.get("status_raw") or "").upper() or "—")
        # days_since: integer days from today (Central) to last_event_date.
        last = d.get("last_event_date")
        if last:
            try:
                ld = datetime.strptime(last[:10], "%Y-%m-%d").date()
                td = datetime.strptime(today, "%Y-%m-%d").date()
                d["days_since"] = max(0, (td - ld).days)
            except Exception:
                d["days_since"] = None
        else:
            d["days_since"] = None

        # trend: simple delta of plays_12mo vs plays_prior_12mo. Values are
        # 'up', 'flat', 'down', or 'new' (no prior-period plays at all).
        p12 = d["plays_12mo"]
        pp12 = d["plays_prior_12mo"]
        if pp12 == 0 and p12 > 0:
            d["trend"] = "new"
        elif p12 > pp12:
            d["trend"] = "up"
        elif p12 < pp12:
            d["trend"] = "down"
        else:
            d["trend"] = "flat"
        d["trend_delta"] = p12 - pp12

        out.append(d)
    return out


def _render_participation_email(row: dict, subject_tpl: str, body_tpl: str) -> dict:
    """Render the merge variables into the subject + body for one player."""
    first = (row.get("first_name") or "").strip() or "there"
    last_event = (row.get("last_event_date") or "")[:10]
    days_since = row.get("days_since")
    chapter = (row.get("chapter") or "").strip() or "TGF"
    plays_12 = row.get("plays_12mo") or 0

    last_event_phrase = ""
    if last_event:
        last_event_phrase = f" (on {last_event})"

    vars_ = {
        "first_name": first,
        "last_name": (row.get("last_name") or "").strip(),
        "days_since": days_since if days_since is not None else "—",
        "last_event": last_event or "—",
        "last_event_phrase": last_event_phrase,
        "chapter": chapter,
        "plays_12mo": plays_12,
    }
    try:
        subject = subject_tpl.format(**vars_)
    except (KeyError, IndexError):
        subject = subject_tpl
    try:
        body = body_tpl.format(**vars_)
    except (KeyError, IndexError):
        body = body_tpl

    # Wrap the body in the same minimal HTML shell the membership emails use.
    html = (
        "<!doctype html><html><body style=\"font-family:-apple-system,"
        "BlinkMacSystemFont,'Segoe UI',sans-serif;color:#111827;max-width:600px;"
        "margin:0 auto;padding:1.5rem;\">"
        + body +
        "<hr style=\"border:none;border-top:1px solid #e5e7eb;margin:1.5rem 0;\">"
        "<p style=\"font-size:0.78rem;color:#6b7280;\">"
        "The Golf Fellowship &middot; "
        "<a href=\"https://thegolffellowship.com\" style=\"color:#2563eb;\">"
        "thegolffellowship.com</a></p>"
        "</body></html>"
    )
    return {"subject": subject, "html": html}


@app.route("/participation")
def page_participation():
    if session.get("role") == "view-only":
        return redirect("/events")
    return render_template("participation.html")


@app.route("/api/participation/players")
@require_role("manager")
def api_participation_players():
    """Return per-customer participation summary for the /participation page."""
    conn = get_connection()
    try:
        rows = _get_participation_rows(conn)
    finally:
        conn.close()
    return jsonify({
        "as_of": today_central_str(),
        "default_subject": PARTICIPATION_DEFAULT_SUBJECT,
        "default_body_html": PARTICIPATION_DEFAULT_BODY_HTML,
        "rows": rows,
    })


@app.route("/api/participation/preview-email", methods=["POST"])
@require_role("manager")
def api_participation_preview_email():
    """Render the re-engagement email for ONE customer (merge-vars filled)."""
    data = request.get_json(silent=True) or {}
    customer_id = data.get("customer_id")
    subject_tpl = data.get("subject") or PARTICIPATION_DEFAULT_SUBJECT
    body_tpl = data.get("body_html") or PARTICIPATION_DEFAULT_BODY_HTML
    if not customer_id:
        return jsonify({"error": "customer_id required"}), 400

    conn = get_connection()
    try:
        rows = _get_participation_rows(conn)
    finally:
        conn.close()
    row = next((r for r in rows if r["customer_id"] == customer_id), None)
    if not row:
        return jsonify({"error": "customer not found or not in audience"}), 404

    rendered = _render_participation_email(row, subject_tpl, body_tpl)
    return jsonify({
        "customer_id": customer_id,
        "name": row.get("name"),
        "email": row.get("email") or "",
        "has_email": bool(row.get("email")),
        "days_since": row.get("days_since"),
        "last_event_date": row.get("last_event_date"),
        "subject": rendered["subject"],
        "html": rendered["html"],
    })


@app.route("/api/participation/send-email", methods=["POST"])
@require_role("manager")
def api_participation_send_email():
    """Send the re-engagement email to one or more customers."""
    data = request.get_json(silent=True) or {}
    ids = data.get("customer_ids") or []
    if not isinstance(ids, list) or not ids:
        return jsonify({"error": "customer_ids (non-empty list) required"}), 400
    subject_tpl = data.get("subject") or PARTICIPATION_DEFAULT_SUBJECT
    body_tpl = data.get("body_html") or PARTICIPATION_DEFAULT_BODY_HTML

    tenant_id = os.getenv("AZURE_TENANT_ID")
    client_id = os.getenv("AZURE_CLIENT_ID")
    client_secret = os.getenv("AZURE_CLIENT_SECRET")
    from_address = os.getenv("EMAIL_ADDRESS")
    if not all([tenant_id, client_id, client_secret, from_address]):
        return jsonify({"error": "Email credentials not configured on server"}), 500

    conn = get_connection()
    try:
        rows = _get_participation_rows(conn)
    finally:
        conn.close()
    by_id = {r["customer_id"]: r for r in rows}

    results = []
    sent = 0
    skipped = 0
    failed = 0
    for cid in ids:
        row = by_id.get(cid)
        if not row:
            results.append({"customer_id": cid, "status": "skipped",
                            "reason": "customer not in audience"})
            skipped += 1
            continue
        email = (row.get("email") or "").strip()
        if not email:
            results.append({"customer_id": cid, "name": row.get("name"),
                            "status": "skipped", "reason": "no primary email"})
            skipped += 1
            continue
        # #127 guardrail 2 (Kerry-ratified): gg_roster historical profiles
        # are excluded from ALL marketing flows — contact data on the
        # profile is fine, sending to it from a marketing composer is not.
        gconn = get_connection()
        try:
            src = gconn.execute(
                "SELECT acquisition_source FROM customers WHERE customer_id=?",
                (cid,)).fetchone()
        finally:
            gconn.close()
        if src and (src["acquisition_source"] or "") == "gg_roster":
            results.append({"customer_id": cid, "name": row.get("name"),
                            "status": "skipped",
                            "reason": "historical profile (gg_roster) — "
                                      "excluded from marketing flows"})
            skipped += 1
            continue
        rendered = _render_participation_email(row, subject_tpl, body_tpl)
        try:
            ok = send_mail_graph(
                tenant_id=tenant_id,
                client_id=client_id,
                client_secret=client_secret,
                from_address=from_address,
                to_address=email,
                subject=rendered["subject"],
                html_body=rendered["html"],
            )
        except Exception as exc:
            logger.exception("participation send failed for %s", cid)
            results.append({"customer_id": cid, "name": row.get("name"),
                            "email": email, "status": "error", "reason": str(exc)})
            failed += 1
            continue

        status = "sent" if ok else "failed"
        try:
            log_message({
                "event_name": "participation-reengagement",
                "channel": "email",
                "recipient_name": row.get("name") or "",
                "recipient_address": email,
                "subject": rendered["subject"],
                "body_preview": (f"Re-engagement; last event "
                                 f"{row.get('last_event_date') or 'never'} "
                                 f"({row.get('days_since')} days)"),
                "status": status,
                "sent_by": session.get("role", "unknown"),
            })
        except Exception:
            logger.warning("Failed to log participation email", exc_info=True)

        if ok:
            sent += 1
            results.append({"customer_id": cid, "name": row.get("name"),
                            "email": email, "status": "sent"})
        else:
            failed += 1
            results.append({"customer_id": cid, "name": row.get("name"),
                            "email": email, "status": "failed",
                            "reason": "send_mail_graph returned false"})

    return jsonify({
        "requested": len(ids),
        "sent": sent,
        "skipped": skipped,
        "failed": failed,
        "results": results,
    })


# ---------------------------------------------------------------------------
# Routes — Season Contests
# ---------------------------------------------------------------------------

@app.route("/api/season-contests")
@require_role("member")
def api_season_contests():
    """List season contest enrollments with optional filters."""
    from email_parser.database import get_season_contest_enrollments
    contest_type = request.args.get("contest_type")
    chapter = request.args.get("chapter")
    season = request.args.get("season")
    enrollments = get_season_contest_enrollments(contest_type, chapter, season)
    return jsonify(enrollments)


@app.route("/api/season-contests", methods=["POST"])
@require_role("manager")
def api_enroll_season_contest():
    """Manually enroll a customer in a season contest by customer_id.

    Body: { customer_id, contest_type, chapter, season }
    Looks up the canonical customer_name from the customers table so the
    enrollment is always tied to a real customer record (FK enforced in app layer).
    """
    from email_parser.database import enroll_season_contest, _connect
    data = request.get_json(silent=True) or {}
    customer_id = data.get("customer_id")
    contest_type = (data.get("contest_type") or "City Match Play").strip()
    chapter = (data.get("chapter") or "").strip()
    season = (data.get("season") or "").strip()
    if not customer_id:
        return jsonify({"error": "customer_id required"}), 400
    # Resolve canonical name — refuse to enroll if customer doesn't exist
    with _connect() as conn:
        row = conn.execute(
            """SELECT TRIM(COALESCE(NULLIF(company_name,''),
                   NULLIF(TRIM(first_name || ' ' || last_name), ''))) AS customer_name
               FROM customers WHERE customer_id = ?""",
            (int(customer_id),),
        ).fetchone()
    if not row or not row["customer_name"]:
        return jsonify({"error": "Customer not found"}), 404
    customer_name = row["customer_name"]
    enrollment = enroll_season_contest(customer_name, contest_type, chapter, season, manually_enrolled=True)
    return jsonify(enrollment), 201


@app.route("/api/season-contests/sync", methods=["POST"])
@require_role("manager")
def api_sync_season_contests():
    """Scan all items and enroll customers in season contests."""
    from email_parser.database import sync_season_contests_from_items
    result = sync_season_contests_from_items()
    return jsonify(result)


@app.route("/api/season-contests/customer/<path:customer_name>")
@require_role("manager")
def api_customer_season_contests(customer_name):
    """Get season contest enrollments for a specific customer."""
    from email_parser.database import get_customer_season_contests
    enrollments = get_customer_season_contests(customer_name)
    return jsonify(enrollments)


@app.route("/api/season-contests/removals")
@require_role("view-only")
def api_season_contest_removals():
    """List removal records (the Enrollment tab's recordation list)."""
    from email_parser.database import get_season_contest_removals
    removals = get_season_contest_removals(
        request.args.get("contest_type"),
        request.args.get("chapter"),
        request.args.get("season"),
    )
    return jsonify(removals)


@app.route("/api/season-contests/points-race")
@require_role("member")
def api_season_contest_points_race():
    """Persisted GG points-race standings joined with live buy-in status.

    Serves the gg_points_standings snapshot (instant, survives restarts);
    auto-refreshes from the Golf Genius portal when the snapshot is empty
    or >12h old, and ?force=1 refreshes on demand. If GG is unreachable
    the stale snapshot is served with gg_error set.
    """
    from email_parser.database import get_points_race_standings
    race = request.args.get("race", "san_antonio_net")
    force = request.args.get("force") == "1"
    try:
        data = get_points_race_standings(race, force_refresh=force)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        logger.exception("Points race load failed")
        return jsonify({"error": f"Points race load failed: {e}"}), 500
    return jsonify(data)


@app.route("/api/season-contests/points-race/live")
@require_role("member")
def api_points_race_live():
    """Season standings + TODAY'S championship points, added together
    (Kerry 2026-07-31: "the Championship is in addition to the regular
    season total ... everything earned tomorrow adds on").

    Member tier, and PII-free like every other points read: names, points
    and thru only. Polled ~1/min by the member scoreboard; the GG walk is
    server-side cached so many viewers collapse to one fetch.
    """
    from email_parser.database import get_points_race_live
    race = request.args.get("race", "san_antonio_net")
    # force=1 (admin Refresh) forces the SEASON snapshot re-walk and still
    # returns the merged live view — the old refresh button bypassed the
    # overlay entirely and painted GG's portal raw, which after close-out
    # (but before GG posts championship points to the portal) read as if
    # the event never happened (Kerry, 2026-08-01 evening).
    force = request.args.get("force") == "1"
    try:
        return jsonify(get_points_race_live(race, force_refresh=force))
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        logger.exception("Live points race load failed")
        return jsonify({"error": f"Live points race load failed: {e}"}), 500


@app.route("/api/season-contests/points-race/champ-boards", methods=["GET", "POST"])
@require_role("manager")
def api_champ_points_boards():
    """Read (or, as admin, repoint) the championship POINTS boards.

    A DIAL rather than code so next season's Golf Genius tournament ids
    are a settings change instead of a deploy.
    """
    from email_parser.database import champ_points_boards, set_app_setting
    if request.method == "POST":
        if session.get("role") != "admin":
            return jsonify({"error": "Admin only."}), 403
        boards = (request.get_json(silent=True) or {}).get("boards")
        if not isinstance(boards, dict) or not boards:
            return jsonify({"error": "boards must be a non-empty object"}), 400
        for key, b in boards.items():
            if not isinstance(b, dict) or not str(b.get("url", "")).startswith("https://"):
                return jsonify({"error": f"{key}: needs an https url"}), 400
            if "golfgenius.com" not in b["url"]:
                return jsonify({"error": f"{key}: must be a golfgenius.com url"}), 400
        set_app_setting("gg_champ_points_boards", json.dumps(boards))
    return jsonify({"boards": champ_points_boards()})


@app.route("/api/season-contests/points-race/fellowship-cup")
@require_role("member")
def api_fellowship_cup_projection():
    """Combined NET-race reset projection (THE FELLOWSHIP CUP tab)."""
    from email_parser.database import get_fellowship_cup_projection
    force = request.args.get("force") == "1"
    try:
        return jsonify(get_fellowship_cup_projection(force_refresh=force))
    except Exception as e:
        logger.exception("Fellowship Cup projection failed")
        return jsonify({"error": f"Projection failed: {e}"}), 500


@app.route("/api/season-contests/lone-star-cup")
@require_role("member")
def api_lone_star_cup():
    """LONE STAR CUP projected rosters + alternates pool (member LSC tab)."""
    from email_parser.database import get_lone_star_cup_projection
    try:
        return jsonify(get_lone_star_cup_projection())
    except Exception as e:
        logger.exception("Lone Star Cup projection failed")
        return jsonify({"error": f"Projection failed: {e}"}), 500


@app.route("/api/scoring/import", methods=["POST"])
@require_role("admin")
def api_import_scorecards():
    """Import scorecards from a GG tournament page (body: tournament_url, event_code)."""
    from email_parser.database import import_gg_scorecards
    body = request.get_json(silent=True) or {}
    url = (body.get("tournament_url") or "").strip()
    if not url:
        return jsonify({"error": "tournament_url required"}), 400
    try:
        return jsonify(import_gg_scorecards(url, event_code=(body.get("event_code") or "").strip() or None))
    except Exception as e:
        logger.exception("Scorecard import failed")
        return jsonify({"error": f"Import failed: {e}"}), 502


@app.route("/api/scoring/rounds")
@require_role("member")
def api_scoring_rounds():
    from email_parser.database import get_scoring_rounds_list
    rows = get_scoring_rounds_list(
        request.args.get("player"), request.args.get("event"),
        int(request.args.get("customer_id") or 0) or None,
        int(request.args.get("limit") or 100))
    # 2026-only ruling (Kerry, 2026-07-12): the pinless/member tier never
    # sees archive-era rounds — logged-in staff sessions see everything.
    if session.get("role") not in ("view-only", "manager", "admin"):
        from email_parser.timezone_utils import today_central
        season = f"{today_central().year}-01-01"
        rows = [r for r in rows
                if (r.get("round_date") or "") >= season
                and not str(r.get("source") or "").startswith("gg_history")]
    return jsonify(rows)


@app.route("/api/scoring/scorecard/<int:scoring_round_id>")
@require_role("member")
def api_scorecard(scoring_round_id):
    from email_parser.database import get_scorecard
    card = get_scorecard(scoring_round_id)
    return (jsonify(card), 200) if card else (jsonify({"error": "not found"}), 404)


# ═══════════════════════════════════════════════════════════════════════
# LIVE SCORING TEST CENTER (admin sandbox)
#
# Stage 1 of the untether-from-GG plan (docs/claude/game-engine.md): stand
# our own leaderboard next to GG's and diff until we reproduce it exactly.
# Every route here is admin-only and every write lands in ls_test_* tables —
# nothing on this surface can touch a production scoring row.
# ═══════════════════════════════════════════════════════════════════════

@app.route("/admin/test-center")
@require_role("admin")
def test_center_page():
    return render_template("test_center.html", SHELL_TITLE="Test Center",
                           SHELL_ACTIVE="admin")


@app.route("/api/test-center/sessions")
@require_role("admin")
def api_ls_sessions():
    from email_parser.database import ls_list_sessions
    return jsonify(ls_list_sessions())


@app.route("/api/test-center/sessions", methods=["POST"])
@require_role("admin")
def api_ls_create_session():
    """Create a sandbox session — synthetic, or seeded from a real event."""
    from email_parser.database import (ls_create_session,
                                       ls_seed_session_from_event)
    body = request.get_json(silent=True) or {}
    who = session.get("role")
    event_name = (body.get("seed_from_event") or "").strip()
    try:
        if event_name:
            res = ls_seed_session_from_event(
                event_name, name=(body.get("name") or "").strip() or None,
                created_by=who)
            return (jsonify(res), 400) if "error" in res else (jsonify(res), 201)
        res = ls_create_session(
            (body.get("name") or "").strip() or "Untitled test",
            holes=int(body.get("holes") or 9),
            course_id=int(body.get("course_id") or 0) or None,
            tee_id=int(body.get("tee_id") or 0) or None,
            championship=bool(body.get("championship")),
            notes=(body.get("notes") or "").strip() or None,
            created_by=who)
        return jsonify(res), 201
    except Exception as e:
        logger.exception("Test Center session create failed")
        return jsonify({"error": f"Create failed: {e}"}), 500


@app.route("/api/test-center/sessions/<int:session_id>")
@require_role("admin")
def api_ls_get_session(session_id):
    from email_parser.database import ls_get_session
    data = ls_get_session(session_id)
    return (jsonify(data), 200) if data else (jsonify({"error": "not found"}), 404)


@app.route("/api/test-center/sessions/<int:session_id>", methods=["PATCH"])
@require_role("admin")
def api_ls_update_session(session_id):
    from email_parser.database import ls_update_session
    return jsonify(ls_update_session(session_id, request.get_json(silent=True) or {}))


@app.route("/api/test-center/sessions/<int:session_id>", methods=["DELETE"])
@require_role("admin")
def api_ls_delete_session(session_id):
    from email_parser.database import ls_delete_session
    return jsonify(ls_delete_session(session_id))


@app.route("/api/test-center/sessions/<int:session_id>/players",
           methods=["POST"])
@require_role("admin")
def api_ls_add_player(session_id):
    from email_parser.database import ls_add_player
    body = request.get_json(silent=True) or {}
    ph = body.get("playing_handicap")
    try:
        return jsonify(ls_add_player(
            session_id, (body.get("player_name") or "").strip(),
            customer_id=int(body.get("customer_id") or 0) or None,
            playing_handicap=float(ph) if ph not in (None, "") else None,
            flight=(body.get("flight") or "").strip() or None,
            team_num=int(body.get("team_num") or 0) or None,
            buys_net=body.get("buys_net", True),
            buys_gross=body.get("buys_gross", True),
            is_member=body.get("is_member", True))), 201
    except (TypeError, ValueError) as e:
        return jsonify({"error": f"Bad player payload: {e}"}), 400


@app.route("/api/test-center/players/<int:player_id>", methods=["PATCH"])
@require_role("admin")
def api_ls_update_player(player_id):
    from email_parser.database import ls_update_player
    return jsonify(ls_update_player(player_id, request.get_json(silent=True) or {}))


@app.route("/api/test-center/players/<int:player_id>", methods=["DELETE"])
@require_role("admin")
def api_ls_delete_player(player_id):
    from email_parser.database import ls_delete_player
    return jsonify(ls_delete_player(player_id))


@app.route("/api/test-center/sessions/<int:session_id>/score",
           methods=["POST"])
@require_role("admin")
def api_ls_set_score(session_id):
    """One hole, one player — the Stage-2 write path in miniature."""
    from email_parser.database import ls_set_score
    body = request.get_json(silent=True) or {}
    try:
        player_id = int(body["player_id"])
        hole = int(body["hole_number"])
    except (KeyError, TypeError, ValueError):
        return jsonify({"error": "player_id and hole_number required"}), 400
    strokes = body.get("strokes")
    strokes = int(strokes) if strokes not in (None, "") else None
    if strokes is not None and not (1 <= strokes <= 20):
        return jsonify({"error": "strokes must be between 1 and 20"}), 400
    sr = body.get("strokes_received")
    return jsonify(ls_set_score(session_id, player_id, hole, strokes,
                                int(sr) if sr not in (None, "") else None))


@app.route("/api/test-center/sessions/<int:session_id>/hole",
           methods=["POST"])
@require_role("admin")
def api_ls_set_course_hole(session_id):
    from email_parser.database import ls_set_course_hole
    body = request.get_json(silent=True) or {}

    def _int(key):
        v = body.get(key)
        return int(v) if v not in (None, "") else None
    try:
        hole = int(body["hole_number"])
    except (KeyError, TypeError, ValueError):
        return jsonify({"error": "hole_number required"}), 400
    return jsonify(ls_set_course_hole(session_id, hole, _int("par"),
                                      _int("yardage"), _int("stroke_index")))


@app.route("/api/test-center/sessions/<int:session_id>/autoplay",
           methods=["POST"])
@require_role("admin")
def api_ls_autoplay(session_id):
    from email_parser.database import ls_autoplay
    body = request.get_json(silent=True) or {}
    return jsonify(ls_autoplay(
        session_id,
        through_hole=int(body.get("through_hole") or 0) or None,
        seed=int(body["seed"]) if str(body.get("seed") or "").strip() else None,
        overwrite=bool(body.get("overwrite"))))


@app.route("/api/test-center/sessions/<int:session_id>/clear-scores",
           methods=["POST"])
@require_role("admin")
def api_ls_clear_scores(session_id):
    from email_parser.database import ls_clear_scores
    return jsonify(ls_clear_scores(session_id))


@app.route("/api/test-center/sessions/<int:session_id>/contest",
           methods=["POST"])
@require_role("admin")
def api_ls_contest(session_id):
    """Record a CTP / Longest Putt / HIO winner — measured, never derived."""
    from email_parser.database import ls_record_contest
    body = request.get_json(silent=True) or {}
    pid = body.get("player_id")
    res = ls_record_contest(session_id, body.get("kind") or "",
                            int(body.get("hole_number") or 0) or None,
                            int(pid) if pid not in (None, "") else None,
                            (body.get("note") or "").strip() or None)
    return (jsonify(res), 400) if "error" in res else jsonify(res)


@app.route("/api/test-center/sessions/<int:session_id>/refresh",
           methods=["POST"])
@require_role("admin")
def api_ls_refresh(session_id):
    """Re-pull a seeded session's scores from GG in place — the live path.

    Body may carry `tournament_url` to fetch fresh cards from Golf Genius
    first; without it the session re-syncs from whatever `scoring_rounds`
    already holds.
    """
    from email_parser.database import ls_refresh_session_from_gg
    body = request.get_json(silent=True) or {}
    res = ls_refresh_session_from_gg(
        session_id, (body.get("tournament_url") or "").strip() or None)
    return (jsonify(res), 400) if "error" in res else jsonify(res)


@app.route("/api/test-center/sessions/<int:session_id>/leaderboard")
@require_role("admin")
def api_ls_leaderboard(session_id):
    from email_parser.database import ls_leaderboard
    res = ls_leaderboard(session_id)
    return (jsonify(res), 404) if "error" in res else jsonify(res)


@app.route("/api/test-center/sessions/<int:session_id>/parity")
@require_role("admin")
def api_ls_parity(session_id):
    """The Stage-1 confidence gate: our engine vs GG, player by player."""
    from email_parser.database import ls_parity
    res = ls_parity(session_id)
    return (jsonify(res), 400) if "error" in res else jsonify(res)


@app.route("/api/test-center/flight-lab")
@require_role("admin")
def api_ls_flight_lab():
    """Run one event's real field through BOTH flighting modes, side by side.

    ?event=&game=&min_flight_size=&index_scale=&tie_direction=
    Read-only. Where GG's own per-game flights were captured, both modes are
    graded against them — that is how the rule gets derived from history
    rather than recollection.
    """
    from email_parser.database import ls_flight_lab
    event = (request.args.get("event") or "").strip()
    if not event:
        return jsonify({"error": "event required"}), 400
    over = {}
    if request.args.get("min_flight_size"):
        over["min_flight_size"] = int(request.args["min_flight_size"])
    for k in ("index_scale", "tie_direction"):
        if request.args.get(k):
            over[k] = request.args[k].strip()
    res = ls_flight_lab(event, (request.args.get("game")
                                or "individual_net").strip(), over)
    return (jsonify(res), 404) if "error" in res else jsonify(res)


@app.route("/api/test-center/flightable-events")
@require_role("admin")
def api_ls_flightable_events():
    """Events worth running through the lab: anything with registrations."""
    from email_parser.database import (get_connection,
                                       _ensure_gg_game_flights_tables)
    conn = get_connection()
    try:
        # Lazily created elsewhere — a DB that has never run a flight import
        # would otherwise 500 on the subquery below.
        _ensure_gg_game_flights_tables(conn)
        rows = conn.execute(
            """SELECT e.item_name, e.event_date, e.course,
                      COUNT(i.id) AS n_items,
                      (SELECT COUNT(*) FROM gg_game_flights g
                        WHERE g.event_id = e.id) AS gg_flights
               FROM events e
               JOIN items i ON LOWER(i.item_name) = LOWER(e.item_name)
                           AND COALESCE(i.transaction_status,'active') = 'active'
               GROUP BY e.id HAVING n_items > 0
               ORDER BY e.event_date DESC LIMIT 120""").fetchall()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        logger.exception("flightable-events failed")
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/test-center/scorable-events")
@require_role("admin")
def api_ls_scorable_events():
    """Events a session can be seeded from.

    Includes UPCOMING events that have registrations but no cards yet — a
    session for those seeds from the field so the pre-flight can happen
    days before the round, with scores arriving later via Pull from GG.
    Restricting this to events with scorecards made the pre-flight
    impossible: the only events you could shadow were ones already played.
    """
    from email_parser.database import get_connection
    conn = get_connection()
    try:
        rows = conn.execute(
            """SELECT e.item_name, e.event_date, e.course, e.chapter,
                      (SELECT COUNT(*) FROM scoring_rounds sr
                        WHERE sr.event_id = e.id) AS n_rounds,
                      (SELECT COUNT(*) FROM items i
                        WHERE LOWER(i.item_name) = LOWER(e.item_name)
                          AND COALESCE(i.transaction_status,'active') NOT IN
                              ('credited','refunded','transferred','rsvp_only')
                      ) AS n_registered
               FROM events e
               ORDER BY e.event_date DESC, e.item_name""").fetchall()
        out = [dict(r) for r in rows
               if (r["n_rounds"] or 0) > 0 or (r["n_registered"] or 0) > 0]
        return jsonify(out[:250])
    finally:
        conn.close()


@app.route("/api/courses/tees")
@require_role("manager")
def api_course_tees():
    """Courses enriched with imported tee data (slope/rating/yardage/holes)."""
    from email_parser.database import list_courses
    return jsonify(list_courses())


# ── Member portal (M1) — magic-link access, own data only ──────────────
# No PIN/role gate: the signed token IS the credential, and customer_id is
# derived exclusively from it (docs/claude/member-portal.md). Bumping
# customers.portal_token_version revokes a member's outstanding links.

def _portal_cid_or_none():
    from email_parser.database import verify_portal_token
    return verify_portal_token((request.args.get("t") or "").strip())


@app.route("/me")
def member_portal_page():
    # The page itself carries no data — its JS presents the token to the
    # /api/me endpoints. Render even with a bad token so the error is shown.
    return render_template("me.html")


@app.route("/api/me/summary")
def api_me_summary():
    cid = _portal_cid_or_none()
    if not cid:
        return jsonify({"error": "invalid or revoked link"}), 401
    from email_parser.database import get_member_summary
    summary = get_member_summary(cid)
    return (jsonify(summary), 200) if summary else (jsonify({"error": "not found"}), 404)


@app.route("/api/me/scorecards")
def api_me_scorecards():
    cid = _portal_cid_or_none()
    if not cid:
        return jsonify({"error": "invalid or revoked link"}), 401
    from email_parser.database import get_scoring_rounds_list
    return jsonify(get_scoring_rounds_list(None, None, cid, 100))


@app.route("/api/me/scorecard/<int:scoring_round_id>")
def api_me_scorecard(scoring_round_id):
    cid = _portal_cid_or_none()
    if not cid:
        return jsonify({"error": "invalid or revoked link"}), 401
    from email_parser.database import get_scorecard
    card = get_scorecard(scoring_round_id)
    if not card or card["round"].get("customer_id") != cid:
        return jsonify({"error": "not found"}), 404
    return jsonify(card)


@app.route("/api/customers/<int:customer_id>/portal-link")
@require_role("manager")
def api_customer_portal_link(customer_id):
    """Generate (or re-fetch) a member's magic link for sharing."""
    from email_parser.database import make_portal_token
    tok = make_portal_token(customer_id)
    if not tok:
        return jsonify({"error": "unknown customer"}), 404
    return jsonify({"customer_id": customer_id,
                    "url": request.url_root.rstrip("/") + "/me?t=" + tok})


_points_detail_cache: dict = {}
_POINTS_DETAIL_CACHE_TTL = 600


@app.route("/api/season-contests/points-race/detail")
@require_role("member")
def api_season_contest_points_race_detail():
    """One player's per-round points breakdown (GG row expansion), live.

    Proxies GG's season_points_v2/individual_info XHR for the given
    member_card_id and returns the parsed tables. Cached 10 minutes per
    (race, card) — expansion detail is browse-heavy but changes per round.
    """
    from email_parser.database import _GG_POINTS_RACES
    from golf_genius_sync import fetch_points_race_member_detail
    race_key = request.args.get("race", "san_antonio_net")
    card = (request.args.get("card") or "").strip()
    race = _GG_POINTS_RACES.get(race_key)
    if not race:
        return jsonify({"error": f"Unknown race {race_key!r}"}), 400
    if not card.isdigit():
        return jsonify({"error": "card must be a numeric member_card_id"}), 400

    now = time.time()
    cached = _points_detail_cache.get((race_key, card))
    if cached and now - cached[0] < _POINTS_DETAIL_CACHE_TTL:
        return jsonify(cached[1])

    try:
        data = fetch_points_race_member_detail(
            page_id=race["page_id"], member_card_id=card,
            league_id=race["league_id"], host=race["host"])
        from email_parser.database import substitute_gg_tournament_names
        data["tables"] = substitute_gg_tournament_names(data["tables"])
    except Exception as e:
        logger.exception("Points race detail fetch failed")
        return jsonify({"error": f"Golf Genius fetch failed: {e}"}), 502

    _points_detail_cache[(race_key, card)] = (now, data)
    return jsonify(data)


@app.route("/api/season-contests/points-race/champ-card")
@require_role("member")
def api_season_contest_champ_card():
    """One player's LIVE championship hole-by-hole card (championship day).

    PII-free: the payload is the same name + scores Golf Genius already
    shows on its public board, plus OUR computed net/points per hole.
    Caching (45s per player) lives in the database layer so many viewers
    polling collapse to one GG walk.
    """
    from email_parser.database import fetch_champ_player_card
    race_key = request.args.get("race", "")
    cid = (request.args.get("cid") or "").strip()
    if not cid.isdigit():
        return jsonify({"error": "cid must be a numeric customer id"}), 400
    data = fetch_champ_player_card(race_key, int(cid))
    if data.get("error") and not data.get("holes"):
        code = 404 if "not on the championship" in data["error"] else 502
        return jsonify(data), code
    return jsonify(data)


@app.route("/api/season-contests/monthly-points")
@require_role("member")
def api_season_contest_monthly_points():
    """Combined monthly points races (both chapters) with winner + purse.

    Served from the persisted DB snapshot (gg_data_snapshots) so opening
    the MONTHLY tab never waits on Golf Genius. ?force=1 (the Refresh
    button) live-refetches and updates the snapshot; a daily scheduler
    job does the same, bounding staleness at ~24h. Purse = $1 per active
    TGF member at the close of the month; ties split it.
    """
    from email_parser.database import load_gg_snapshot, refresh_monthly_points_snapshot
    force = request.args.get("force") == "1"
    snapshot = load_gg_snapshot("monthly_points")
    if snapshot and not force:
        return jsonify(snapshot)
    try:
        data = refresh_monthly_points_snapshot()
    except Exception as e:
        logger.exception("Monthly points fetch failed")
        if snapshot:
            stale = dict(snapshot)
            stale["gg_error"] = str(e)
            return jsonify(stale)
        return jsonify({"error": f"Golf Genius fetch failed: {e}"}), 502
    return jsonify(data)


@app.route("/api/season-contests/<int:enrollment_id>", methods=["DELETE"])
@require_role("manager")
def api_delete_season_contest(enrollment_id):
    """Remove a season contest enrollment and record the removal.

    Optional JSON body from the refund modal:
      { reason, refund_amount, refund_method, note }

    remove_season_contest_enrollment() snapshots the enrollment into
    season_contest_removals (permanent recordation shown at the bottom of
    the Enrollment tab), clears the matching contest flag on the source
    purchase item (without which the next sync silently re-enrolled the
    player), and deletes the enrollment — all in one transaction.
    """
    from email_parser.database import remove_season_contest_enrollment
    data = request.get_json(silent=True) or {}
    refund_amount = data.get("refund_amount")
    try:
        refund_amount = float(refund_amount) if refund_amount not in (None, "") else None
    except (TypeError, ValueError):
        return jsonify({"error": "refund_amount must be a number"}), 400
    removal = remove_season_contest_enrollment(
        enrollment_id,
        reason=data.get("reason"),
        refund_amount=refund_amount,
        refund_method=data.get("refund_method"),
        note=data.get("note"),
    )
    if removal is None:
        return jsonify({"error": "Not found"}), 404
    return jsonify({"ok": True, "removal": removal})


# ---------------------------------------------------------------------------
# Routes — Contests page
# ---------------------------------------------------------------------------

def _matchplay_v2_flag():
    """SHELL_V2-style kill switch for the Match Play visual rollout (#228).
    Default ON as of the live go (Kerry, 2026-07-18) — members/managers now
    see the new Match Play tab everywhere. Kill switch preserved: set env
    MATCHPLAY_V2=0 on Railway to instantly revert everyone to the legacy tab.
    The admin preview route forces it on regardless."""
    return os.environ.get("MATCHPLAY_V2", "1") == "1"


@app.route("/contests")
def contests_page():
    return render_template("contests.html", MATCHPLAY_V2=_matchplay_v2_flag())


@app.route("/admin/matchplay-preview")
@require_role("admin")
def matchplay_v2_preview():
    """Admin-only SANDBOX: the NEW (platform-claude-approved, #228) Match
    Play tab on LIVE data — same /api/cmp/* endpoints, all current
    pools/brackets/standings/payouts — so Kerry can play with it before it
    goes live to members. Forces MATCHPLAY_V2 on for this route only;
    /contests + /member/contests stay on the current tab until the env flag
    is flipped. Read-through preview; changes no data.

    ?view=member renders the MEMBER presentation (read-only scoreboards);
    ?view=admin (default) renders the manager/admin editing view. The sandbox
    toggle bar (SANDBOX_PREVIEW flag) flips between them in-page."""
    view = (request.args.get("view") or "admin").strip().lower()
    return render_template("contests.html", MATCHPLAY_V2=True,
                           SANDBOX_PREVIEW=True,
                           member_mode=(view == "member"))


# ── Member view (v2.53.0, Kerry): pinless read-only pages members can
# reach from a plain shared URL. Season Contests + Handicaps only; the
# APIs they call are the @require_role("member") public read tier.
@app.route("/member")
def member_home():
    # Spotlight is the member landing (Kerry 2026-07-14: make /member about
    # the individual player first).
    return redirect("/member/spotlight")


@app.route("/member/contests")
def member_contests():
    return render_template("contests.html", member_mode=True,
                           MATCHPLAY_V2=_matchplay_v2_flag())


@app.route("/member/handicaps")
def member_handicaps():
    return render_template("handicaps.html", member_mode=True)


@app.route("/member/spotlight")
def member_spotlight():
    """Player Spotlight, member view (Kerry GO 2026-07-12: show
    winnings, 2026-era data only — satisfied by construction, the
    payload reads live races/payouts/handicaps, never archive rows)."""
    return render_template("spotlight.html", member_mode=True)


# ---------------------------------------------------------------------------
# Routes — City Match Play (CMP)
# ---------------------------------------------------------------------------

@app.route("/api/cmp/pools")
@require_role("member")
def api_cmp_pools():
    season = request.args.get("season", "")
    chapter = request.args.get("chapter", "")
    if not season or not chapter:
        return jsonify({"error": "season and chapter required"}), 400
    from email_parser.database import cmp_get_pools
    return jsonify(cmp_get_pools(season, chapter))


@app.route("/api/cmp/pools", methods=["POST"])
@require_role("manager")
def api_cmp_create_pool():
    data = request.get_json(silent=True) or {}
    season = (data.get("season") or "").strip()
    chapter = (data.get("chapter") or "").strip()
    pool_name = (data.get("pool_name") or "").strip()
    if not season or not chapter or not pool_name:
        return jsonify({"error": "season, chapter, and pool_name required"}), 400
    from email_parser.database import cmp_create_pool
    pool = cmp_create_pool(season, chapter, pool_name)
    return jsonify(pool), 201


@app.route("/api/cmp/pools/<int:pool_id>", methods=["DELETE"])
@require_role("manager")
def api_cmp_delete_pool(pool_id):
    from email_parser.database import cmp_delete_pool
    cmp_delete_pool(pool_id)
    return jsonify({"ok": True})


@app.route("/api/cmp/pools/<int:pool_id>/members", methods=["POST"])
@require_role("manager")
def api_cmp_add_member(pool_id):
    data = request.get_json(silent=True) or {}
    customer_name = (data.get("customer_name") or "").strip()
    customer_id = data.get("customer_id")
    if not customer_name:
        return jsonify({"error": "customer_name required"}), 400
    from email_parser.database import cmp_add_member
    member = cmp_add_member(pool_id, customer_name, customer_id)
    return jsonify(member), 201


@app.route("/api/cmp/pools/<int:pool_id>/members/<path:customer_name>", methods=["DELETE"])
@require_role("manager")
def api_cmp_remove_member(pool_id, customer_name):
    from email_parser.database import cmp_remove_member
    cmp_remove_member(pool_id, customer_name)
    return jsonify({"ok": True})


@app.route("/api/cmp/pools/<int:pool_id>/members/<path:customer_name>/withdrawn", methods=["POST"])
@require_role("manager")
def api_cmp_set_member_withdrawn(pool_id, customer_name):
    data = request.get_json(silent=True) or {}
    from email_parser.database import cmp_set_member_withdrawn
    return jsonify(cmp_set_member_withdrawn(
        pool_id, customer_name, bool(data.get("withdrawn", True)),
        reason=data.get("reason")))


@app.route("/api/cmp/matches", methods=["POST"])
@require_role("manager")
def api_cmp_save_match():
    data = request.get_json(silent=True) or {}
    pool_id = data.get("pool_id")
    player1 = (data.get("player1_name") or "").strip()
    player2 = (data.get("player2_name") or "").strip()
    if not pool_id or not player1 or not player2:
        return jsonify({"error": "pool_id, player1_name, player2_name required"}), 400
    p1_stab = data.get("player1_stableford")
    p2_stab = data.get("player2_stableford")
    if p1_stab is not None:
        p1_stab = float(p1_stab)
    if p2_stab is not None:
        p2_stab = float(p2_stab)
    winner = (data.get("winner_name") or "").strip() or None
    margin = (data.get("margin") or "").strip() or None
    event_id = data.get("event_id")
    if event_id is not None:
        try:
            event_id = int(event_id)
        except (TypeError, ValueError):
            event_id = None
    from email_parser.database import cmp_save_match
    match = cmp_save_match(
        pool_id, player1, player2,
        winner_name=winner, margin=margin,
        p1_stableford=p1_stab, p2_stableford=p2_stab,
        match_date=data.get("match_date"), notes=data.get("notes"),
        event_id=event_id,
    )
    if match.get("blocked") == "result_locked":
        return jsonify({"error": "This result is locked (verified against "
                        "Golf Genius) and can't be changed from here.",
                        **match}), 409
    return jsonify(match)


@app.route("/api/cmp/matches", methods=["DELETE"])
@require_role("manager")
def api_cmp_clear_match():
    data = request.get_json(silent=True) or {}
    pool_id = data.get("pool_id")
    player1 = (data.get("player1_name") or "").strip()
    player2 = (data.get("player2_name") or "").strip()
    if not pool_id or not player1 or not player2:
        return jsonify({"error": "pool_id, player1_name, player2_name required"}), 400
    from email_parser.database import cmp_clear_match
    try:
        cmp_clear_match(pool_id, player1, player2)
    except ValueError as e:
        return jsonify({"error": str(e)}), 409
    return jsonify({"ok": True})


@app.route("/api/cmp/matches")
@require_role("member")
def api_cmp_get_matches():
    pool_id = request.args.get("pool_id", type=int)
    if not pool_id:
        return jsonify({"error": "pool_id required"}), 400
    from email_parser.database import cmp_get_matches
    return jsonify(cmp_get_matches(pool_id))


@app.route("/api/cmp/live-match")
@require_role("member")
def api_cmp_live_match():
    """Live GG match-play detail for an in-progress match (polled ~1/min by the
    scoreboard). Read-only; server-side cached so concurrent viewers collapse
    to one GG walk."""
    chapter = request.args.get("chapter", "")
    a = request.args.get("a", "")
    b = request.args.get("b", "")
    if not chapter or not a or not b:
        return jsonify({"error": "chapter, a, b required"}), 400
    from email_parser.database import cmp_fetch_live_match
    return jsonify(cmp_fetch_live_match(chapter, a, b))


@app.route("/api/cmp/standings")
@require_role("member")
def api_cmp_standings():
    season = request.args.get("season", "")
    chapter = request.args.get("chapter", "")
    if not season or not chapter:
        return jsonify({"error": "season and chapter required"}), 400
    from email_parser.database import cmp_get_standings, sct_get_active_config
    active = sct_get_active_config("match_play", season, chapter)
    advance = 2
    if active:
        advance = int(active["config"].get("advance_per_pool", 2))
    return jsonify(cmp_get_standings(season, chapter, advance_per_pool=advance))


@app.route("/api/cmp/bracket")
@require_role("member")
def api_cmp_get_bracket():
    season = request.args.get("season", "")
    chapter = request.args.get("chapter", "")
    if not season or not chapter:
        return jsonify({"error": "season and chapter required"}), 400
    from email_parser.database import cmp_get_bracket
    return jsonify(cmp_get_bracket(season, chapter))


@app.route("/api/cmp/bracket", methods=["POST"])
@require_role("manager")
def api_cmp_save_bracket():
    data = request.get_json(silent=True) or {}
    season = (data.get("season") or "").strip()
    chapter = (data.get("chapter") or "").strip()
    round_ = (data.get("round") or "").strip()
    slot = data.get("slot")
    if not season or not chapter or not round_ or slot is None:
        return jsonify({"error": "season, chapter, round, slot required"}), 400
    p_stab = float(data["player_stableford"]) if data.get("player_stableford") is not None else None
    o_stab = float(data["opponent_stableford"]) if data.get("opponent_stableford") is not None else None
    winner = (data.get("winner_name") or "").strip() or None
    margin = (data.get("margin") or "").strip() or None
    event_id = int(data["event_id"]) if data.get("event_id") else None
    from email_parser.database import cmp_save_bracket_slot
    row = cmp_save_bracket_slot(
        season, chapter, round_, int(slot),
        data.get("player_name"), p_stab,
        data.get("opponent_name"), o_stab,
        winner, margin, event_id,
    )
    return jsonify(row)


@app.route("/api/cmp/bracket", methods=["DELETE"])
@require_role("manager")
def api_cmp_clear_bracket():
    season = (request.args.get("season") or "").strip()
    chapter = (request.args.get("chapter") or "").strip()
    if not season or not chapter:
        return jsonify({"error": "season and chapter required"}), 400
    from email_parser.database import cmp_clear_bracket
    deleted = cmp_clear_bracket(season, chapter)
    return jsonify({"ok": True, "deleted": deleted})


@app.route("/api/cmp/consolation", methods=["POST"])
@require_role("manager")
def api_cmp_consolation():
    """D-MP-08: record or clear the 3rd-place consolation match between the
    two semifinal losers. Body: season, chapter, loser_a, loser_b,
    winner_name (omit/empty to clear → reverts to the fallback split)."""
    data = request.get_json(silent=True) or {}
    season = (data.get("season") or "").strip()
    chapter = (data.get("chapter") or "").strip()
    loser_a = (data.get("loser_a") or "").strip()
    loser_b = (data.get("loser_b") or "").strip()
    if not season or not chapter or not loser_a or not loser_b:
        return jsonify({"error": "season, chapter, loser_a, loser_b required"}), 400
    from email_parser.database import cmp_record_consolation
    try:
        row = cmp_record_consolation(
            season, chapter, loser_a, loser_b,
            winner_name=(data.get("winner_name") or "").strip() or None,
            margin=(data.get("margin") or "").strip() or None,
            by=session.get("role"))
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    return jsonify(row)


# ── Match Play versioned config + config-driven operations (v2.34.0) ──

@app.route("/api/cmp/config")
@require_role("member")
def api_cmp_get_config():
    """Active config for a season/chapter (snapshot-resolved) + structure
    for the currently enrolled N when season+chapter are given."""
    season = (request.args.get("season") or "").strip()
    chapter = (request.args.get("chapter") or "").strip()
    from email_parser.database import cmp_enrolled_entrants, sct_get_active_config
    active = sct_get_active_config("match_play", season or None, chapter or None)
    if not active:
        return jsonify({"error": "No Match Play config template found"}), 404
    out = dict(active)
    if season and chapter:
        from email_parser.match_play import structure_for_n
        entrants = cmp_enrolled_entrants(season, chapter)
        out["n_enrolled"] = len(entrants)
        out["entrants"] = entrants
        try:
            out["structure"] = structure_for_n(active["config"], len(entrants))
        except ValueError as e:
            out["structure_error"] = str(e)
    return jsonify(out)


@app.route("/api/cmp/config/versions")
@require_role("view-only")
def api_cmp_config_versions():
    from email_parser.database import sct_list_versions
    return jsonify(sct_list_versions("match_play"))


@app.route("/api/cmp/config/versions/<int:version_id>")
@require_role("view-only")
def api_cmp_config_version_detail(version_id):
    from email_parser.database import sct_get_version
    v = sct_get_version(version_id)
    if not v:
        return jsonify({"error": "Version not found"}), 404
    return jsonify(v)


@app.route("/api/cmp/config/versions", methods=["POST"])
@require_role("admin")
def api_cmp_save_config_version():
    data = request.get_json(silent=True) or {}
    config = data.get("config")
    if not isinstance(config, dict):
        return jsonify({"error": "config (object) required"}), 400
    from email_parser.database import sct_save_version
    try:
        result = sct_save_version("match_play", config,
                                  saved_by=session.get("role", "admin"),
                                  notes=(data.get("notes") or "").strip() or None)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    return jsonify(result), 201


@app.route("/api/cmp/config/snapshot", methods=["POST"])
@require_role("admin")
def api_cmp_pin_config_snapshot():
    """Re-pin a season/chapter to a specific (or the current) config version."""
    data = request.get_json(silent=True) or {}
    season = (data.get("season") or "").strip()
    chapter = (data.get("chapter") or "").strip()
    if not season or not chapter:
        return jsonify({"error": "season and chapter required"}), 400
    from email_parser.database import sct_pin_snapshot
    try:
        snap = sct_pin_snapshot("match_play", season, chapter,
                                version_id=data.get("version_id"),
                                by=session.get("role", "admin"))
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    return jsonify(snap)


@app.route("/api/cmp/structure")
@require_role("member")
def api_cmp_structure():
    """Computed matrix row(s). ?n=12 for one field size, else all N.
    ?version_id previews a specific config version instead of the active one."""
    season = (request.args.get("season") or "").strip() or None
    chapter = (request.args.get("chapter") or "").strip() or None
    from email_parser.database import sct_get_active_config, sct_get_version
    from email_parser.match_play import structure_for_n
    version_id = request.args.get("version_id", type=int)
    if version_id:
        v = sct_get_version(version_id)
        if not v:
            return jsonify({"error": "Version not found"}), 404
        active = {"version_no": v["version_no"], "config": v["config"]}
    else:
        active = sct_get_active_config("match_play", season, chapter)
    if not active:
        return jsonify({"error": "No Match Play config template found"}), 404
    cfg = active["config"]
    n_arg = request.args.get("n", type=int)
    if n_arg:
        try:
            return jsonify(structure_for_n(cfg, n_arg))
        except ValueError as e:
            return jsonify({"error": str(e)}), 400
    rows = []
    for n in range(int(cfg.get("n_min", 4)), int(cfg.get("n_max", 32)) + 1):
        try:
            rows.append(structure_for_n(cfg, n))
        except ValueError:
            continue
    return jsonify({"config_version": active["version_no"], "rows": rows})


@app.route("/api/cmp/pools/auto-assign", methods=["POST"])
@require_role("manager")
def api_cmp_auto_assign_pools():
    data = request.get_json(silent=True) or {}
    season = (data.get("season") or "").strip()
    chapter = (data.get("chapter") or "").strip()
    if not season or not chapter:
        return jsonify({"error": "season and chapter required"}), 400
    from email_parser.database import cmp_auto_assign_pools
    try:
        result = cmp_auto_assign_pools(season, chapter,
                                       created_by=session.get("role"),
                                       force=bool(data.get("force")))
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    return jsonify(result)


@app.route("/api/cmp/bracket/seed", methods=["POST"])
@require_role("manager")
def api_cmp_seed_knockout():
    data = request.get_json(silent=True) or {}
    season = (data.get("season") or "").strip()
    chapter = (data.get("chapter") or "").strip()
    if not season or not chapter:
        return jsonify({"error": "season and chapter required"}), 400
    from email_parser.database import cmp_seed_knockout
    try:
        result = cmp_seed_knockout(season, chapter,
                                   created_by=session.get("role"),
                                   force=bool(data.get("force")))
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    return jsonify(result)


@app.route("/api/cmp/payouts")
@require_role("member")
def api_cmp_payouts():
    season = (request.args.get("season") or "").strip()
    chapter = (request.args.get("chapter") or "").strip()
    if not season or not chapter:
        return jsonify({"error": "season and chapter required"}), 400
    from email_parser.database import cmp_get_payout_sheet
    return jsonify(cmp_get_payout_sheet(season, chapter))


# ---------------------------------------------------------------------------
# Routes — Authentication
# ---------------------------------------------------------------------------
@app.route("/api/auth/login", methods=["POST"])
def api_auth_login():
    """Authenticate with a PIN and set the session role."""
    if not _check_login_rate_limit():
        return jsonify({"error": "Too many login attempts. Please try again in 15 minutes."}), 429

    # Re-read .env so PIN changes take effect without a server restart
    load_dotenv(override=True)

    data = request.get_json(silent=True)
    if not data or not data.get("pin"):
        return jsonify({"error": "PIN is required."}), 400

    pin = str(data.get("pin", "")).strip()
    admin_pin = os.getenv("ADMIN_PIN", "")
    austin_pin = os.getenv("AUSTIN_MANAGER_PIN", "")
    sa_pin = os.getenv("SA_MANAGER_PIN", "")
    viewonly_pin = os.getenv("VIEWONLY_PIN", "")
    manager_pin = os.getenv("MANAGER_PIN", "")

    def _ok(role, chapter=None):
        session["role"] = role
        if chapter:
            session["chapter"] = chapter
        else:
            session.pop("chapter", None)
        return jsonify({"status": "ok", "role": role, "chapter": chapter})

    # Chapter-manager PINs (Kerry, 2026-07-08): each chapter manager gets
    # their own PIN whose session carries the chapter, so every tab lands
    # pre-filtered to their chapter. The LEGACY shared MANAGER_PIN is
    # demoted to view-only (Kerry: "keep the old pin as a VIEWONLY_PIN").
    if admin_pin and secrets.compare_digest(pin, admin_pin):
        return _ok("admin")
    elif austin_pin and secrets.compare_digest(pin, austin_pin):
        return _ok("manager", "Austin")
    elif sa_pin and secrets.compare_digest(pin, sa_pin):
        return _ok("manager", "San Antonio")
    elif viewonly_pin and secrets.compare_digest(pin, viewonly_pin):
        return _ok("view-only")
    elif manager_pin and secrets.compare_digest(pin, manager_pin):
        return _ok("view-only")
    else:
        return jsonify({"error": "Invalid PIN."}), 401


@app.route("/api/auth/role")
def api_auth_role():
    """Return the current session role + chapter (or null if not logged in)."""
    return jsonify({"role": session.get("role"),
                    "chapter": session.get("chapter")})


@app.route("/api/auth/logout", methods=["POST"])
def api_auth_logout():
    """Clear the session role."""
    session.pop("role", None)
    session.pop("chapter", None)
    return jsonify({"status": "ok"})


# ═══════════════════════════════════════════════════════════════════════════
# Accounting Module — Multi-Entity Bookkeeping
# ═══════════════════════════════════════════════════════════════════════════

@app.route("/accounting")
def accounting_page():
    """Multi-entity accounting dashboard (admin only)."""
    if session.get("role") != "admin":
        return redirect("/events")
    return render_template("accounting.html")


@app.route("/api/member-metric", methods=["POST"])
@require_role("member")
def api_member_metric():
    """Anonymous member-page beacon (v2.62.0, Kerry): page opens + clicks.

    PII-free by design: whitelisted event names, truncated path/label,
    nothing identifying stored. Bodies over 1KB are ignored."""
    from email_parser.database import log_member_event
    raw = request.get_data(cache=False, as_text=False)
    if not raw or len(raw) > 1024:
        return jsonify({"ok": False}), 400
    import json as _json
    try:
        data = _json.loads(raw)
    except Exception:
        return jsonify({"ok": False}), 400
    event = str(data.get("event") or "")
    if event not in ("open", "click", "nudge"):
        return jsonify({"ok": False}), 400
    path = str(data.get("path") or "")
    if not path.startswith("/member"):
        return jsonify({"ok": False}), 400
    try:
        log_member_event(event, path, str(data.get("detail") or ""))
    except Exception:
        logger.warning("member metric write failed", exc_info=True)
    return jsonify({"ok": True})


@app.route("/api/member-traffic")
@require_role("admin")
def api_member_traffic():
    """Aggregated member-side traffic for the admin Traffic view."""
    from email_parser.database import get_member_traffic_summary
    return jsonify(get_member_traffic_summary())


@app.route("/traffic")
def traffic_page():
    """Member-traffic dashboard (admin only) — v2.62.0, Kerry."""
    if session.get("role") != "admin":
        return redirect("/events")
    return render_template("traffic.html")


# ── GG History review (admin) — v2.74.x ──
# Kerry's identity review queue + archive coverage browser for the GG
# history initiative (docs/claude/gg-history.md). Decisions here are
# rule-3b territory only when they touch member-facing surfaces — the
# queue itself is admin-internal.

@app.route("/admin/gg-history")
def gg_history_page():
    if session.get("role") != "admin":
        return redirect("/events")
    return render_template("gg_history.html")


@app.route("/api/gg-history/overview")
@require_role("admin")
def api_gg_history_overview():
    from email_parser import gg_history as ggh
    return jsonify(ggh.portal_overview())


@app.route("/api/gg-history/pending-names")
@require_role("admin")
def api_gg_history_pending_names():
    from email_parser import gg_history as ggh
    return jsonify(ggh.review_queue())


@app.route("/api/gg-history/standings")
@require_role("admin")
def api_gg_history_standings():
    from email_parser import gg_history as ggh
    return jsonify(ggh.standings_browser(
        portal=request.args.get("portal") or None,
        q=request.args.get("q") or None,
        contest=request.args.get("contest") or None,
        limit=min(int(request.args.get("limit", 200)), 1000)))


@app.route("/api/gg-history/resolve", methods=["POST"])
@require_role("admin")
def api_gg_history_resolve():
    from email_parser import gg_history as ggh
    data = request.get_json(silent=True) or {}
    try:
        link_id = int(data.get("id"))
    except (TypeError, ValueError):
        return jsonify({"error": "id required"}), 400
    cid = data.get("customer_id")
    res = ggh.resolve_name_link(
        link_id, str(data.get("action") or ""),
        customer_id=int(cid) if cid else None)
    return (jsonify(res), 400) if res.get("error") else jsonify(res)


@app.route("/api/gg-history/customer-search")
@require_role("admin")
def api_gg_history_customer_search():
    from email_parser import gg_history as ggh
    return jsonify(ggh.search_customers_for_link(
        request.args.get("q", ""), limit=10))


# ── Player Spotlight (ADMIN PREVIEW v1 — Kerry, 2026-07-10) ──
# Destined for the pinless member view after CA/CD iteration: the payloads
# are PII-free by design, so opening it up later is a role change only.
# Until Kerry ratifies, ALL THREE routes stay admin.

@app.route("/spotlight")
def spotlight_page():
    """Player Spotlight (admin preview) — search a player, see their story."""
    if session.get("role") != "admin":
        return redirect("/events")
    return render_template("spotlight.html")


@app.route("/api/spotlight/search")
@require_role("member")
def api_spotlight_search():
    """Name typeahead. PII-free payload (member-tier-ready)."""
    from email_parser.database import search_spotlight_players
    return jsonify({"players": search_spotlight_players(
        request.args.get("q", ""))})


@app.route("/api/spotlight/player")
@require_role("member")
def api_spotlight_player():
    """One player's spotlight payload. PII-free (member-tier-ready)."""
    from email_parser.database import get_player_spotlight
    try:
        cid = int(request.args.get("cid", ""))
    except ValueError:
        return jsonify({"error": "cid must be an integer"}), 400
    try:
        return jsonify(get_player_spotlight(cid))
    except Exception as e:
        logger.exception("Spotlight payload failed")
        return jsonify({"error": f"Spotlight failed: {e}"}), 500


@app.route("/courses")
def courses_page():
    """Course database editor (admin only) — v2.57.0, Kerry."""
    if session.get("role") != "admin":
        return redirect("/events")
    return render_template("course_db.html")


@app.route("/api/course-db")
@require_role("admin")
def api_course_db_list():
    from email_parser.database import list_courses, list_chapter_names
    return jsonify({"courses": list_courses(), "chapters": list_chapter_names()})


@app.route("/api/course-db/<int:course_id>", methods=["PATCH"])
@require_role("admin")
def api_course_db_update(course_id):
    from email_parser.database import update_course
    try:
        return jsonify(update_course(course_id, request.get_json(force=True) or {}))
    except ValueError as e:
        return jsonify({"error": str(e)}), 400


# ── Entities ──────────────────────────────────────────────────────────────

@app.route("/api/accounting/entities")
@require_role("admin")
def api_acct_entities():
    return jsonify(get_all_acct_entities())


@app.route("/api/accounting/entities", methods=["POST"])
@require_role("admin")
def api_acct_create_entity():
    d = request.json or {}
    if not d.get("name") or not d.get("short_name"):
        return jsonify({"error": "name and short_name required"}), 400
    try:
        return jsonify(create_acct_entity(d["name"], d["short_name"], d.get("color", "#2563eb")))
    except Exception as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/accounting/entities/<int:eid>", methods=["PATCH"])
@require_role("admin")
def api_acct_update_entity(eid):
    d = request.json or {}
    return jsonify(update_acct_entity(eid, **d))


# ── Categories ────────────────────────────────────────────────────────────

@app.route("/api/accounting/categories")
@require_role("admin")
def api_acct_categories():
    entity_id = request.args.get("entity_id", type=int)
    cat_type = request.args.get("type")
    return jsonify(get_acct_categories(entity_id=entity_id, cat_type=cat_type))


@app.route("/api/accounting/categories", methods=["POST"])
@require_role("admin")
def api_acct_create_category():
    d = request.json or {}
    if not d.get("name") or not d.get("type"):
        return jsonify({"error": "name and type required"}), 400
    return jsonify(create_acct_category(
        d["name"], d["type"], d.get("entity_id"), d.get("parent_id"), d.get("icon"),
    ))


@app.route("/api/accounting/categories/<int:cid>", methods=["PATCH"])
@require_role("admin")
def api_acct_update_category(cid):
    d = request.json or {}
    return jsonify(update_acct_category(cid, **d))


@app.route("/api/accounting/categories/<int:cid>", methods=["DELETE"])
@require_role("admin")
def api_acct_delete_category(cid):
    delete_acct_category(cid)
    return jsonify({"status": "ok"})


# ── Accounts ──────────────────────────────────────────────────────────────

@app.route("/api/accounting/accounts")
@require_role("admin")
def api_acct_accounts():
    entity_id = request.args.get("entity_id", type=int)
    return jsonify(get_acct_accounts(entity_id=entity_id))


@app.route("/api/accounting/accounts", methods=["POST"])
@require_role("admin")
def api_acct_create_account():
    d = request.json or {}
    if not d.get("name") or not d.get("account_type"):
        return jsonify({"error": "name and account_type required"}), 400
    return jsonify(create_acct_account(
        d["name"], d["account_type"], d.get("entity_id"),
        d.get("institution"), d.get("last_four"), d.get("opening_balance", 0),
    ))


@app.route("/api/accounting/accounts/<int:aid>", methods=["PATCH"])
@require_role("admin")
def api_acct_update_account(aid):
    d = request.json or {}
    return jsonify(update_acct_account(aid, **d))


@app.route("/api/accounting/accounts/balances")
@require_role("admin")
def api_acct_account_balances():
    return jsonify(get_acct_account_balances())


# ── Transactions ──────────────────────────────────────────────────────────

@app.route("/api/accounting/vendors")
@require_role("admin")
def api_acct_list_vendors():
    """List vendor-role customers with their ledger activity.

    Vendors live in the customers table (the expense ledger references
    payees by customer_id) but are tagged with the 'vendor' role and
    excluded from every people-facing surface — this list on the
    Accounting page is their home.
    """
    from email_parser.database import _connect
    with _connect() as conn:
        rows = conn.execute(
            """SELECT c.customer_id, c.company_name, c.first_name, c.last_name,
                      c.phone,
                      (SELECT COUNT(*) FROM acct_transactions a
                       WHERE a.customer_id = c.customer_id
                         AND COALESCE(a.status, 'active') != 'merged') AS n_ledger_rows
               FROM customers c
               WHERE EXISTS (SELECT 1 FROM customer_roles r
                             WHERE r.customer_id = c.customer_id
                               AND r.role_type = 'vendor')
               ORDER BY COALESCE(NULLIF(c.company_name, ''), c.last_name) COLLATE NOCASE"""
        ).fetchall()
    out = []
    for r in rows:
        d = dict(r)
        d["display_name"] = d.get("company_name") or \
            f"{d.get('first_name', '')} {d.get('last_name', '')}".strip()
        out.append(d)
    return jsonify(out)


@app.route("/api/accounting/vendors", methods=["POST"])
@require_role("admin")
def api_acct_create_vendor():
    """Create a new vendor (customer with vendor role) for transaction linking."""
    from email_parser.database import _connect
    d = request.json or {}
    name = (d.get("name") or "").strip()
    phone = (d.get("phone") or "").strip() or None
    if not name:
        return jsonify({"error": "name is required"}), 400

    with _connect() as conn:
        # Check if a vendor with this company name already exists
        existing = conn.execute(
            "SELECT customer_id FROM customers WHERE LOWER(COALESCE(company_name,'')) = LOWER(?) LIMIT 1",
            (name,),
        ).fetchone()
        if not existing:
            # Also check personal first+last name — a vendor whose display
            # name happens to exactly match an existing personal customer
            # (e.g. a sole-proprietor course pro billed under their own
            # name) would otherwise get a second, disconnected customers
            # row instead of reusing the real one.
            parts = name.split(None, 1)
            if len(parts) == 2:
                existing = conn.execute(
                    """SELECT customer_id FROM customers
                       WHERE LOWER(first_name) = LOWER(?) AND LOWER(last_name) = LOWER(?)
                       LIMIT 1""",
                    (parts[0], parts[1]),
                ).fetchone()

        if existing:
            cid = existing["customer_id"]
        else:
            cursor = conn.execute(
                """INSERT INTO customers
                       (first_name, last_name, company_name, phone, acquisition_source, account_status)
                   VALUES ('', ?, ?, ?, 'vendor', 'active')""",
                (name, name, phone),
            )
            cid = cursor.lastrowid

        conn.execute(
            "INSERT OR IGNORE INTO customer_roles (customer_id, role_type) VALUES (?, 'vendor')",
            (cid,),
        )
        conn.commit()

        row = conn.execute(
            """SELECT c.customer_id, c.first_name, c.last_name, c.company_name,
                      c.current_player_status, c.chapter, 1 as is_vendor
               FROM customers c WHERE c.customer_id = ?""",
            (cid,),
        ).fetchone()

    data = dict(row)
    data["display_name"] = data.get("company_name") or f"{data.get('first_name','')} {data.get('last_name','')}".strip()
    return jsonify(data), 201


@app.route("/api/accounting/customers")
@require_role("admin")
def api_acct_customers():
    """Return all customers and vendors for transaction linking."""
    from email_parser.database import _connect
    with _connect() as conn:
        rows = conn.execute(
            """SELECT c.customer_id, c.first_name, c.last_name,
                      c.company_name, c.current_player_status, c.chapter,
                      EXISTS(
                          SELECT 1 FROM customer_roles r
                          WHERE r.customer_id = c.customer_id AND r.role_type = 'vendor'
                      ) as is_vendor
               FROM customers c
               WHERE c.account_status = 'active'
               ORDER BY c.last_name COLLATE NOCASE, c.first_name COLLATE NOCASE"""
        ).fetchall()

    result = []
    for r in rows:
        d = dict(r)
        d["display_name"] = d.get("company_name") or f"{d.get('first_name','')} {d.get('last_name','')}".strip()
        result.append(d)
    return jsonify(result)


@app.route("/api/accounting/smart-fill", methods=["POST"])
@require_role("admin")
def api_acct_smart_fill():
    """Auto-assign accounts and default splits to unsplit transactions."""
    from email_parser.database import _connect
    d = request.json or {}
    dry_run = d.get("dry_run", True)

    with _connect() as conn:
        # Find transactions with no splits
        unsplit = conn.execute(
            """SELECT t.id, t.type, t.source, t.total_amount, t.account_id,
                      t.event_name, t.description
               FROM acct_transactions t
               WHERE t.total_amount > 0
                 AND t.type IN ('income', 'expense')
                 AND COALESCE(t.status, 'active') = 'active'
                 AND NOT EXISTS (SELECT 1 FROM acct_splits s WHERE s.transaction_id = t.id)
               ORDER BY t.date DESC"""
        ).fetchall()

        # Look up TGF Checking account id
        tgf_checking = conn.execute(
            "SELECT id FROM acct_accounts WHERE LOWER(name) LIKE '%tgf checking%' LIMIT 1"
        ).fetchone()
        tgf_checking_id = tgf_checking["id"] if tgf_checking else None

        # Look up Venmo account
        venmo_acct = conn.execute(
            "SELECT id FROM acct_accounts WHERE LOWER(name) LIKE '%venmo%' OR account_type='venmo' LIMIT 1"
        ).fetchone()
        venmo_id = venmo_acct["id"] if venmo_acct else None

        # Look up TGF entity
        tgf_entity = conn.execute(
            "SELECT id FROM acct_entities WHERE LOWER(short_name) = 'tgf' LIMIT 1"
        ).fetchone()
        tgf_entity_id = tgf_entity["id"] if tgf_entity else None

        default_entity = conn.execute("SELECT id FROM acct_entities LIMIT 1").fetchone()
        default_entity_id = default_entity["id"] if default_entity else None

        # Look up "Event Revenue" income category
        event_rev_cat = conn.execute(
            "SELECT id FROM acct_categories WHERE type='income' AND LOWER(name) LIKE '%event revenue%' LIMIT 1"
        ).fetchone()
        event_rev_cat_id = event_rev_cat["id"] if event_rev_cat else None

        applied = []
        for t in unsplit:
            source = (t["source"] or "").lower()
            desc = (t["description"] or "").lower()

            # Determine account to assign
            new_account_id = t["account_id"]
            if not new_account_id:
                if source == "godaddy" or desc.startswith("godaddy order"):
                    new_account_id = tgf_checking_id
                elif source == "venmo":
                    new_account_id = venmo_id

            # Determine split params
            entity_id = tgf_entity_id if t["type"] == "income" else default_entity_id
            category_id = event_rev_cat_id if t["type"] == "income" else None

            info = {
                "id": t["id"],
                "description": t["description"],
                "amount": t["total_amount"],
                "type": t["type"],
                "source": t["source"],
                "new_account_id": new_account_id,
                "entity_id": entity_id,
                "category_id": category_id,
            }
            applied.append(info)

            if not dry_run:
                # Update account if changed
                if new_account_id and new_account_id != t["account_id"]:
                    conn.execute("UPDATE acct_transactions SET account_id=? WHERE id=?",
                                 (new_account_id, t["id"]))
                # Create a single split for the full amount
                conn.execute(
                    """INSERT INTO acct_splits (transaction_id, entity_id, category_id, amount, memo)
                       VALUES (?, ?, ?, ?, '')""",
                    (t["id"], entity_id, category_id, t["total_amount"])
                )

        if not dry_run:
            conn.commit()

    return jsonify({"count": len(applied), "transactions": applied if dry_run else []})


@app.route("/api/accounting/transactions")
@require_role("admin")
def api_acct_transactions():
    return jsonify(get_acct_transactions(
        entity_id=request.args.get("entity_id", type=int),
        account_id=request.args.get("account_id", type=int),
        category_id=request.args.get("category_id", type=int),
        start_date=request.args.get("start_date"),
        end_date=request.args.get("end_date"),
        search=request.args.get("search"),
        txn_type=request.args.get("type"),
        limit=request.args.get("limit", 200, type=int),
        offset=request.args.get("offset", 0, type=int),
    ))


@app.route("/api/accounting/transactions/unified")
@require_role("admin")
def api_acct_unified_transactions():
    return jsonify(get_unified_transactions(
        entity_id=request.args.get("entity_id", type=int),
        account_id=request.args.get("account_id", type=int),
        category_id=request.args.get("category_id", type=int),
        start_date=request.args.get("start_date"),
        end_date=request.args.get("end_date"),
        search=request.args.get("search"),
        txn_type=request.args.get("type"),
        source=request.args.get("source"),
        review_status=request.args.get("review_status"),
        ledger_status=request.args.get("ledger_status"),
        limit=request.args.get("limit", 200, type=int),
        offset=request.args.get("offset", 0, type=int),
    ))


@app.route("/api/accounting/transactions/<int:tid>")
@require_role("admin")
def api_acct_transaction(tid):
    txn = get_acct_transaction(tid)
    if not txn:
        return jsonify({"error": "not found"}), 404
    return jsonify(txn)


@app.route("/api/accounting/transactions", methods=["POST"])
@require_role("admin")
def api_acct_create_transaction():
    d = request.json or {}
    required = ["date", "description", "total_amount", "type"]
    for f in required:
        if f not in d:
            return jsonify({"error": f"{f} is required"}), 400
    splits = d.get("splits", [])
    if not splits:
        return jsonify({"error": "At least one split is required"}), 400
    # Validate split total matches transaction total
    split_total = sum(s.get("amount", 0) for s in splits)
    if abs(split_total - float(d["total_amount"])) > 0.01:
        return jsonify({"error": f"Split total ({split_total:.2f}) doesn't match transaction amount ({d['total_amount']})"}), 400
    try:
        txn = _create_acct_ledger_entry(
            date=d["date"], description=d["description"],
            total_amount=float(d["total_amount"]), txn_type=d["type"],
            account_id=d.get("account_id"), transfer_to_account_id=d.get("transfer_to_account_id"),
            notes=d.get("notes"), receipt_path=d.get("receipt_path"),
            source=d.get("source", "manual"), source_ref=d.get("source_ref"),
            splits=splits, tag_ids=d.get("tag_ids"),
            customer_id=d.get("customer_id"),
        )
        return jsonify(txn), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/accounting/transactions/<int:tid>", methods=["PUT"])
@require_role("admin")
def api_acct_update_transaction(tid):
    d = request.json or {}
    splits = d.get("splits")
    if splits is not None and "total_amount" in d:
        split_total = sum(s.get("amount", 0) for s in splits)
        if abs(split_total - float(d["total_amount"])) > 0.01:
            return jsonify({"error": f"Split total ({split_total:.2f}) doesn't match transaction amount ({d['total_amount']})"}), 400
    try:
        txn = update_acct_transaction(tid, **d)
        return jsonify(txn)
    except Exception as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/accounting/transactions/<int:tid>", methods=["DELETE"])
@require_role("admin")
def api_acct_delete_transaction(tid):
    delete_acct_transaction(tid)
    return jsonify({"status": "ok"})


@app.route("/api/accounting/transactions/<int:tid>/reconcile", methods=["POST"])
@require_role("admin")
def api_acct_reconcile(tid):
    d = request.json or {}
    return jsonify(reconcile_acct_transaction(tid, d.get("reconciled", True)))


# ── Tags ──────────────────────────────────────────────────────────────────

@app.route("/api/accounting/tags")
@require_role("admin")
def api_acct_tags():
    return jsonify(get_acct_tags())


@app.route("/api/accounting/tags", methods=["POST"])
@require_role("admin")
def api_acct_create_tag():
    d = request.json or {}
    if not d.get("name"):
        return jsonify({"error": "name required"}), 400
    try:
        return jsonify(create_acct_tag(d["name"], d.get("color", "#6b7280")))
    except Exception as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/accounting/tags/<int:tid>", methods=["DELETE"])
@require_role("admin")
def api_acct_delete_tag(tid):
    delete_acct_tag(tid)
    return jsonify({"status": "ok"})


# ── Reports ───────────────────────────────────────────────────────────────

@app.route("/api/accounting/reports/summary")
@require_role("admin")
def api_acct_report_summary():
    return jsonify(get_acct_summary(
        entity_id=request.args.get("entity_id", type=int),
        start_date=request.args.get("start_date"),
        end_date=request.args.get("end_date"),
    ))


@app.route("/api/accounting/reports/monthly")
@require_role("admin")
def api_acct_report_monthly():
    return jsonify(get_acct_monthly_totals(
        entity_id=request.args.get("entity_id", type=int),
        months=request.args.get("months", 12, type=int),
    ))


@app.route("/api/accounting/reports/categories")
@require_role("admin")
def api_acct_report_categories():
    return jsonify(get_acct_category_breakdown(
        entity_id=request.args.get("entity_id", type=int),
        txn_type=request.args.get("type", "expense"),
        start_date=request.args.get("start_date"),
        end_date=request.args.get("end_date"),
    ))


# ── CSV Import ────────────────────────────────────────────────────────────

@app.route("/api/accounting/import/preview", methods=["POST"])
@require_role("admin")
def api_acct_import_preview():
    if "file" in request.files:
        csv_text = request.files["file"].read().decode("utf-8", errors="replace")
    elif request.json and "csv_text" in request.json:
        csv_text = request.json["csv_text"]
    else:
        return jsonify({"error": "No CSV data provided"}), 400
    # Auto-detect columns from headers; caller can override with explicit indices
    d = request.form if request.files else (request.json or {})
    overrides = {}
    for key in ("date_col", "description_col", "amount_col", "category_col", "memo_col"):
        val = d.get(key)
        if val is not None and val != "":
            overrides[key] = int(val)
    result = preview_acct_csv(csv_text, **overrides)
    return jsonify(result)


@app.route("/api/accounting/import/commit", methods=["POST"])
@require_role("admin")
def api_acct_import_commit():
    d = request.json or {}
    if not d.get("rows") or not d.get("account_id") or not d.get("entity_id"):
        return jsonify({"error": "rows, account_id, and entity_id required"}), 400
    result = import_acct_csv(
        d["rows"], d["account_id"], d["entity_id"],
        transfer_account_id=d.get("transfer_account_id"),
    )
    return jsonify(result)


# ── Recurring ─────────────────────────────────────────────────────────────

@app.route("/api/accounting/recurring")
@require_role("admin")
def api_acct_recurring():
    return jsonify(get_acct_recurring())


@app.route("/api/accounting/recurring", methods=["POST"])
@require_role("admin")
def api_acct_create_recurring():
    d = request.json or {}
    required = ["description", "amount", "type", "entity_id", "frequency", "next_date"]
    for f in required:
        if f not in d:
            return jsonify({"error": f"{f} required"}), 400
    return jsonify(create_acct_recurring(
        d["description"], float(d["amount"]), d["type"], d["entity_id"],
        d["frequency"], d["next_date"], d.get("category_id"), d.get("account_id"),
    ))


@app.route("/api/accounting/recurring/<int:rid>", methods=["DELETE"])
@require_role("admin")
def api_acct_delete_recurring(rid):
    delete_acct_recurring(rid)
    return jsonify({"status": "ok"})


# ── Receipt Upload ────────────────────────────────────────────────────────

@app.route("/api/accounting/upload-receipt", methods=["POST"])
@require_role("admin")
def api_acct_upload_receipt():
    """Upload a receipt image/PDF and return the file path."""
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400
    f = request.files["file"]
    if not f.filename:
        return jsonify({"error": "No filename"}), 400
    # Save to receipts directory
    receipts_dir = os.path.join(os.path.dirname(DB_PATH), "receipts")
    os.makedirs(receipts_dir, exist_ok=True)
    safe_name = f"{secrets.token_hex(8)}_{f.filename}"
    path = os.path.join(receipts_dir, safe_name)
    f.save(path)
    return jsonify({"path": path, "filename": safe_name})


# ── AI Bookkeeper ─────────────────────────────────────────────────────────

@app.route("/api/accounting/ai/categorize", methods=["POST"])
@require_role("admin")
def api_acct_ai_categorize():
    """Auto-categorize transactions using learned rules + AI."""
    d = request.json or {}
    descriptions = d.get("descriptions", [])
    txn_types = d.get("types", [])
    if not descriptions:
        return jsonify({"error": "descriptions required"}), 400
    results = auto_categorize_transactions(descriptions, txn_types)
    return jsonify(results)


@app.route("/api/accounting/ai/review-queue")
@require_role("admin")
def api_acct_ai_review_queue():
    """Return transactions needing categorization."""
    return jsonify(get_acct_review_queue())


@app.route("/api/accounting/ai/stats")
@require_role("admin")
def api_acct_ai_stats():
    """Return categorization coverage stats."""
    return jsonify(get_acct_categorization_stats())


@app.route("/api/accounting/ai/batch")
@require_role("admin")
def api_acct_ai_batch():
    """Return a batch of pending expense_transactions with AI suggestions pre-populated."""
    limit = int(request.args.get("limit", 20))
    offset = int(request.args.get("offset", 0))
    return jsonify(get_expense_batch_preview(limit=limit, offset=offset))


@app.route("/api/accounting/ai/batch-approve", methods=["POST"])
@require_role("admin")
def api_acct_ai_batch_approve():
    """Approve and promote selected expense_transactions into the ledger."""
    d = request.json or {}
    items = d.get("items", [])
    if not items:
        return jsonify({"error": "items required"}), 400
    result = batch_approve_expenses(items)
    return jsonify(result)


@app.route("/api/accounting/liabilities")
@require_role("admin")
def api_accounting_liabilities():
    """Return all liability buckets for the Liabilities Dashboard."""
    return jsonify(get_accounting_liabilities())


@app.route("/api/accounting/month-close")
@require_role("admin")
def api_accounting_month_close():
    """Return month-close checklist status and financial position."""
    return jsonify(get_month_close_status())


@app.route("/api/accounting/liabilities/update", methods=["POST"])
@require_role("admin")
def api_accounting_liabilities_update():
    """Update a manual liability value."""
    d = request.json or {}
    key = d.get("key", "").strip()
    value = d.get("value")
    allowed_keys = {
        "hio_pot", "season_contests_total", "lone_star_cup_shirts",
        "chapter_manager_payouts", "grandparent_loan", "member_credits_2025",
        "irs_balance", "chase_biz_7680", "chase_sapphire_6159",
    }
    if not key or key not in allowed_keys:
        return jsonify({"error": "invalid key"}), 400
    try:
        value = float(value)
    except (TypeError, ValueError):
        return jsonify({"error": "value must be a number"}), 400
    set_coo_manual_value(key, value)
    return jsonify({"ok": True, "key": key, "value": value})


# ── Contractor Payouts ────────────────────────────────────────────────────────

@app.route("/api/accounting/contractors")
@require_role("admin")
def api_contractors_list():
    return jsonify(get_contractor_payouts())


@app.route("/api/accounting/contractors/managers")
@require_role("admin")
def api_contractors_managers():
    return jsonify(get_contractor_managers())


@app.route("/api/accounting/contractors", methods=["POST"])
@require_role("admin")
def api_contractors_add():
    d = request.json or {}
    mgr_id = d.get("manager_customer_id")
    if not mgr_id:
        return jsonify({"error": "manager_customer_id required"}), 400
    try:
        amount = float(d.get("amount_owed", 0))
    except (TypeError, ValueError):
        return jsonify({"error": "amount_owed must be a number"}), 400
    new_id = add_contractor_payout(
        manager_customer_id=int(mgr_id),
        event_name=d.get("event_name") or None,
        event_date=d.get("event_date") or None,
        amount_owed=amount,
        chapter_id=d.get("chapter_id") or None,
        notes=d.get("notes") or None,
    )
    return jsonify({"ok": True, "id": new_id})


@app.route("/api/accounting/contractors/<int:payout_id>", methods=["PATCH"])
@require_role("admin")
def api_contractors_update(payout_id):
    d = request.json or {}
    amount_paid = d.get("amount_paid")
    if amount_paid is not None:
        try:
            amount_paid = float(amount_paid)
        except (TypeError, ValueError):
            return jsonify({"error": "amount_paid must be a number"}), 400
    ok = update_contractor_payout(
        payout_id=payout_id,
        amount_paid=amount_paid,
        status=d.get("status") or None,
        payment_method=d.get("payment_method") or None,
        notes=d.get("notes"),
    )
    return jsonify({"ok": ok})


@app.route("/api/accounting/contractors/<int:payout_id>", methods=["DELETE"])
@require_role("admin")
def api_contractors_delete(payout_id):
    ok = delete_contractor_payout(payout_id)
    return jsonify({"ok": ok})


@app.route("/api/accounting/ai/bulk-categorize", methods=["POST"])
@require_role("admin")
def api_acct_ai_bulk_categorize():
    """AI-categorize all uncategorized transactions in one shot."""
    queue = get_acct_review_queue()
    if not queue:
        return jsonify({"updated": 0, "message": "All transactions are categorized"})

    descriptions = [t["description"] for t in queue]
    types = [t["type"] for t in queue]
    suggestions = auto_categorize_transactions(descriptions, types)

    updated = 0
    for txn, suggestion in zip(queue, suggestions):
        if not suggestion or suggestion["confidence"] == "none":
            continue
        cat_id = suggestion.get("category_id")
        ent_id = suggestion.get("entity_id")
        if not cat_id:
            continue

        # Update the first split with the suggested category + entity
        from email_parser.database import _connect
        with _connect() as conn:
            split = conn.execute(
                "SELECT id, entity_id FROM acct_splits WHERE transaction_id = ? LIMIT 1",
                (txn["id"],),
            ).fetchone()
            if split:
                updates = {"category_id": cat_id}
                if ent_id:
                    updates["entity_id"] = ent_id
                evt_id = suggestion.get("event_id")
                if evt_id:
                    updates["event_id"] = evt_id
                set_clause = ", ".join(f"{k} = ?" for k in updates)
                conn.execute(
                    f"UPDATE acct_splits SET {set_clause} WHERE id = ?",
                    (*updates.values(), split["id"]),
                )
                conn.commit()
                updated += 1

    return jsonify({"updated": updated, "total": len(queue)})


# ── Reset & Account Rules ─────────────────────────────────────────────────

@app.route("/api/accounting/reset", methods=["POST"])
@require_role("admin")
def api_acct_reset():
    """Wipe all accounting data and re-seed entities + categories."""
    result = reset_acct_data()
    return jsonify(result)


@app.route("/api/accounting/accounts/<int:aid>/rules")
@require_role("admin")
def api_acct_get_rules(aid):
    return jsonify(get_acct_account_rules(aid))


@app.route("/api/accounting/accounts/<int:aid>/rules", methods=["POST"])
@require_role("admin")
def api_acct_set_rule(aid):
    d = request.json or {}
    if not d.get("rule_type") or "rule_value" not in d:
        return jsonify({"error": "rule_type and rule_value required"}), 400
    set_acct_account_rule(aid, d["rule_type"], d["rule_value"])
    return jsonify({"status": "ok"})


@app.route("/api/accounting/keyword-rules")
@require_role("admin")
def api_acct_keyword_rules():
    return jsonify(get_acct_keyword_rules())


@app.route("/api/accounting/keyword-rules", methods=["POST"])
@require_role("admin")
def api_acct_create_keyword_rule():
    d = request.json or {}
    if not d.get("keyword"):
        return jsonify({"error": "keyword is required"}), 400
    return jsonify(create_acct_keyword_rule(
        keyword=d["keyword"],
        match_type=d.get("match_type", "contains"),
        category_id=d.get("category_id"),
        entity_id=d.get("entity_id"),
    ))


@app.route("/api/accounting/keyword-rules/<int:rule_id>", methods=["PATCH"])
@require_role("admin")
def api_acct_update_keyword_rule(rule_id):
    return jsonify(update_acct_keyword_rule(rule_id, request.json or {}))


@app.route("/api/accounting/keyword-rules/<int:rule_id>", methods=["DELETE"])
@require_role("admin")
def api_acct_delete_keyword_rule(rule_id):
    return jsonify(delete_acct_keyword_rule(rule_id))


@app.route("/api/accounting/events-list")
@require_role("admin")
def api_acct_events_list():
    """Return events from the events directory for linking to accounting transactions."""
    events = get_all_events()
    return jsonify([{
        "id": e["id"], "item_name": e["item_name"],
        "event_date": e.get("event_date"), "course": e.get("course"),
        "chapter": e.get("chapter"),
    } for e in events])


# ── Allocations ───────────────────────────────────────────────────────────

@app.route("/api/accounting/allocations")
@require_role("admin")
def api_acct_allocations():
    """Return allocation records with totals grouped by bucket."""
    return jsonify(get_acct_allocations(
        month=request.args.get("month"),
        event=request.args.get("event"),
        chapter=request.args.get("chapter"),
    ))


@app.route("/api/accounting/allocations/calculate", methods=["POST"])
@require_role("admin")
def api_acct_calculate_allocation():
    """Calculate allocation for a specific order."""
    d = request.json or {}
    order_id = d.get("order_id")
    if not order_id:
        return jsonify({"error": "order_id required"}), 400
    try:
        result = calculate_order_allocation(order_id)
        if not result:
            return jsonify({"error": f"No active items found for order {order_id}"}), 404
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/accounting/allocations/calculate-all", methods=["POST"])
@require_role("admin")
def api_acct_calculate_all_allocations():
    """Calculate allocations for all orders that don't have allocations yet."""
    from email_parser.database import _connect
    with _connect() as conn:
        order_ids = conn.execute(
            """SELECT DISTINCT order_id FROM items
               WHERE order_id IS NOT NULL AND order_id != ''
                 AND COALESCE(transaction_status, 'active') = 'active'
                 AND order_id NOT IN (SELECT DISTINCT order_id FROM acct_allocations)
               ORDER BY order_date DESC"""
        ).fetchall()
    calculated = 0
    errors = 0
    for row in order_ids:
        try:
            calculate_order_allocation(row["order_id"])
            calculated += 1
        except Exception:
            errors += 1
    return jsonify({"calculated": calculated, "errors": errors, "total_orders": len(order_ids)})


# ── Event Financial Summary (Unified Financial Model, Issue #242) ─────────

@app.route("/api/events/<event_name>/financial-summary")
@require_role("manager")
def api_event_financial_summary(event_name):
    """Return unified financial summary for an event from the accounting system."""
    try:
        return jsonify(get_event_financial_summary(event_name))
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/accounting/backfill", methods=["POST"])
@require_role("admin")
def api_backfill_financials():
    """Backfill accounting entries for existing items missing them (Issue #242)."""
    try:
        result = backfill_financial_entries()
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/accounting/backfill-acct-transactions", methods=["POST"])
@require_role("admin")
def api_backfill_acct_transactions():
    """Backfill flat acct_transactions entries for all 2026 items."""
    try:
        result = backfill_acct_transactions()
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/accounting/verify-event/<event_name>")
@require_role("manager")
def api_verify_event_accounting(event_name):
    """Verify acct_transactions totals for an event."""
    from email_parser.database import _connect
    try:
        with _connect() as conn:
            rows = conn.execute(
                """SELECT entry_type, category, source,
                          COUNT(*) as count,
                          COALESCE(SUM(amount), 0) as total
                   FROM acct_transactions
                   WHERE event_name = ? COLLATE NOCASE
                   AND COALESCE(status, 'active') = 'active'
                   AND entry_type IS NOT NULL
                   GROUP BY entry_type, category, source
                   ORDER BY entry_type, category""",
                (event_name,),
            ).fetchall()
            breakdown = [dict(r) for r in rows]

            income = sum(r["total"] for r in breakdown if r["entry_type"] == "income")
            fees = sum(r["total"] for r in breakdown if r["entry_type"] == "expense" and r["category"] == "processing_fee")
            refunds = sum(r["total"] for r in breakdown if r["entry_type"] == "expense" and r["category"] == "refund")
            contra = sum(r["total"] for r in breakdown if r["entry_type"] == "contra")
            net = round(income - fees - refunds - contra, 2)

            summary = get_event_financial_summary(event_name)

            return jsonify({
                "event_name": event_name,
                "acct_transactions_breakdown": breakdown,
                "totals": {
                    "income": round(income, 2),
                    "processing_fees": round(fees, 2),
                    "refunds": round(refunds, 2),
                    "contra": round(contra, 2),
                    "net": net,
                },
                "financial_summary": summary,
                "accounting_verified": summary.get("accounting_verified", False),
            })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/audit/scan-price-mismatches", methods=["POST"])
@require_role("admin")
def api_scan_price_mismatches():
    """Scan all items for side_games / item_price mismatches and create parse warnings."""
    try:
        result = scan_price_games_mismatches()
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/admin/reconcile-orphan-venmo", methods=["POST"])
@require_role("admin")
def api_reconcile_orphan_venmo():
    """Sweep credit-transfer items with balance_due and net them against
    prior orphan Venmo / manual +PAY items by the same customer (14-day window).

    Body params (optional JSON): {"days": 14, "dry_run": false}
    """
    try:
        data = request.get_json(silent=True) or {}
        days = int(data.get("days", 14))
        dry_run = bool(data.get("dry_run", False))
        result = reconcile_orphan_venmo_payments(max_days_back=days, dry_run=dry_run)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Expense Transactions & Action Items ───────────────────────────────────

@app.route("/api/accounting/expense-transactions")
@require_role("admin")
def api_expense_transactions():
    return jsonify(get_expense_transactions(
        date_from=request.args.get("date_from"),
        date_to=request.args.get("date_to"),
        source_type=request.args.get("source_type"),
        review_status=request.args.get("review_status"),
        event_name=request.args.get("event_name"),
        limit=request.args.get("limit", 100, type=int),
    ))


@app.route("/api/accounting/expense-transactions/<int:tid>")
@require_role("admin")
def api_get_expense_transaction(tid):
    from email_parser.database import _connect, get_expense_suggestions, suggest_for_merchant
    with _connect() as conn:
        row = conn.execute("SELECT * FROM expense_transactions WHERE id = ?", (tid,)).fetchone()
        if not row:
            return jsonify({"error": "not found"}), 404
        result = dict(row)
        # Add AI suggestion for pending expenses
        if result.get("review_status") == "pending" and result.get("merchant"):
            suggestion_data = get_expense_suggestions(conn)
            result["suggestion"] = suggest_for_merchant(result["merchant"], suggestion_data)
    return jsonify(result)


@app.route("/api/accounting/expense-transactions/<int:tid>", methods=["PATCH"])
@require_role("admin")
def api_update_expense_transaction(tid):
    d = request.json or {}
    result = update_expense_transaction(tid, d)
    # If this update flipped a Venmo IN expense to approved, try the balance-due matcher
    # AND stamp the player's Venmo @handle on their customer record (best-effort).
    if (result
            and result.get("source_type") == "venmo"
            and result.get("transaction_type") == "received"
            and result.get("review_status") in ("approved", "corrected")):
        try:
            from email_parser.database import (
                auto_match_venmo_inbound_to_balance_due,
                capture_venmo_handle_for_customer,
            )
            capture_venmo_handle_for_customer(tid)
            match_result = auto_match_venmo_inbound_to_balance_due([tid])
            if match_result.get("matched"):
                result["balance_due_match"] = match_result
        except Exception:
            logger.warning("venmo balance-due auto-match failed for exp %s", tid, exc_info=True)
    if (result
            and result.get("source_type") == "venmo"
            and result.get("transaction_type") == "payout"
            and result.get("review_status") in ("approved", "corrected")):
        try:
            from email_parser.database import auto_match_venmo_payouts_to_tgf
            match_result = auto_match_venmo_payouts_to_tgf([tid])
            if match_result.get("matched"):
                result["payout_match"] = match_result
        except Exception:
            logger.warning("venmo payout auto-match failed for exp %s", tid, exc_info=True)
    return jsonify(result)


@app.route("/api/accounting/auto-match-venmo-balance-due", methods=["POST"])
@require_role("admin")
def api_auto_match_venmo_balance_due():
    """Manually trigger the Venmo IN → balance-due matcher across all approved
    Venmo IN expense_transactions. Useful for backfilling after enabling the feature
    or after manually approving a batch."""
    from email_parser.database import auto_match_venmo_inbound_to_balance_due
    return jsonify(auto_match_venmo_inbound_to_balance_due())


@app.route("/api/tgf/auto-match-venmo-payouts", methods=["POST"])
@require_role("admin")
def api_auto_match_venmo_payouts():
    """Sweep ALL outbound Venmo payout receipts against pending tgf_payouts
    and mark matches PAID (Kerry, 2026-07-08). New receipts also match
    automatically as their emails arrive; this endpoint is the backfill /
    catch-up trigger."""
    from email_parser.database import auto_match_venmo_payouts_to_tgf
    return jsonify(auto_match_venmo_payouts_to_tgf())


def _quick_expense_check():
    """One-shot Venmo-receipt sweep, run shortly after an admin taps Pay.
    check_expense_inbox() parses any new Venmo receipt AND calls the
    payout matcher inline, so this both ingests the receipt and flips the
    tgf_payout to PAID. Scheduler jobs must never raise."""
    try:
        check_expense_inbox(days_back=1)
    except Exception:
        logger.exception("Quick Venmo receipt check failed")


@app.route("/api/tgf/schedule-venmo-check", methods=["POST"])
@require_role("admin")
def api_schedule_venmo_check():
    """After an admin taps a Pay deep link (Venmo/PayPal/Cash App),
    schedule two one-shot inbox sweeps (~75s and ~180s later) so the
    receipt is caught and the payout flips to PAID within a couple
    minutes instead of waiting for the 5-minute scheduler cycle
    (Kerry, 2026-07-13). Repeated taps within the window COALESCE onto
    the same two job ids (replace_existing), so paying a batch triggers
    one sweep shortly after the LAST tap rather than a pile-up. The
    normal 5-minute cycle stays the backstop; cost is unchanged — the
    expense dedup bills each email at most once regardless of how often
    the inbox is swept."""
    if not getattr(scheduler, "running", False):
        return jsonify({"scheduled": False, "reason": "scheduler not running "
                        "in this process — 5-minute cycle will catch it"})
    now = datetime.now()
    scheduled = []
    for jid, secs in (("venmo_quick_check_a", 75), ("venmo_quick_check_b", 180)):
        try:
            scheduler.add_job(_quick_expense_check, "date",
                              run_date=now + timedelta(seconds=secs),
                              id=jid, replace_existing=True, coalesce=True,
                              misfire_grace_time=120)
            scheduled.append(secs)
        except Exception:
            logger.exception("Failed to schedule quick Venmo check %s", jid)
    return jsonify({"scheduled": True, "in_seconds": scheduled})


@app.route("/api/admin/venmo-debug")
@require_role("admin")
def api_admin_venmo_debug():
    """Diagnostic: search expense_transactions by payer name + show match data.

    Query: ?payer=<name fragment> (case-insensitive LIKE on merchant)
    Returns: expense rows + the customer_aliases / customers.venmo_username lookups
    that the matcher would attempt.
    """
    from email_parser.database import _connect
    payer = (request.args.get("payer") or "").strip()
    if not payer:
        return jsonify({"error": "payer query param required"}), 400
    with _connect() as conn:
        expenses = [
            dict(r) for r in conn.execute(
                """SELECT id, source_type, transaction_type, merchant, amount,
                          transaction_date, review_status, matched_item_id,
                          other_party_handle, customer_id
                   FROM expense_transactions
                   WHERE merchant LIKE ? COLLATE NOCASE
                   ORDER BY id DESC LIMIT 20""",
                (f"%{payer}%",),
            ).fetchall()
        ]
        aliases = [
            dict(r) for r in conn.execute(
                """SELECT customer_name, alias_type, alias_value
                   FROM customer_aliases
                   WHERE alias_value LIKE ? COLLATE NOCASE
                   ORDER BY customer_name""",
                (f"%{payer}%",),
            ).fetchall()
        ]
        venmo_handles = [
            dict(r) for r in conn.execute(
                """SELECT customer_id, first_name, last_name, venmo_username
                   FROM customers
                   WHERE (LOWER(first_name || ' ' || COALESCE(last_name,'')) LIKE ?
                       OR LOWER(COALESCE(venmo_username,'')) LIKE ?)
                   LIMIT 20""",
                (f"%{payer.lower()}%", f"%{payer.lower()}%"),
            ).fetchall()
        ]
        balance_due_items = [
            dict(r) for r in conn.execute(
                """SELECT id, customer, customer_id, item_name, credit_note,
                          item_price, transaction_status
                   FROM items
                   WHERE merchant = 'Paid Separately (Credit Transfer)'
                     AND credit_note LIKE 'balance_due:%'
                     AND (customer LIKE ? COLLATE NOCASE)
                   LIMIT 20""",
                (f"%{payer}%",),
            ).fetchall()
        ]
    return jsonify({
        "expense_transactions": expenses,
        "customer_aliases": aliases,
        "customers_with_matching_name_or_handle": venmo_handles,
        "balance_due_items": balance_due_items,
    })


@app.route("/api/accounting/expense-transactions/<int:tid>", methods=["DELETE"])
@require_role("admin")
def api_delete_expense_transaction(tid):
    from email_parser.database import _connect
    with _connect() as conn:
        conn.execute("DELETE FROM expense_transactions WHERE id = ?", (tid,))
        conn.commit()
    return jsonify({"deleted": True})


@app.route("/api/accounting/block-merchant", methods=["POST"])
@require_role("admin")
def api_block_merchant():
    d = request.json or {}
    merchant = (d.get("merchant") or "").strip()
    if not merchant:
        return jsonify({"error": "merchant required"}), 400
    blocked = block_merchant(merchant)
    return jsonify({"blocked": blocked})


@app.route("/api/accounting/action-items")
@require_role("admin")
def api_action_items():
    return jsonify(get_action_items(
        status=request.args.get("status"),
        category=request.args.get("category"),
        limit=request.args.get("limit", 100, type=int),
    ))


@app.route("/api/accounting/action-items/<int:aid>", methods=["PATCH"])
@require_role("admin")
def api_update_action_item(aid):
    d = request.json or {}
    return jsonify(update_action_item(aid, d))


@app.route("/api/accounting/pending-review")
@require_role("admin")
def api_pending_review():
    return jsonify(get_pending_review_count())


@app.route("/api/accounting/classify-email", methods=["POST"])
@require_role("admin")
def api_classify_email():
    """Classify an email (for testing)."""
    d = request.json or {}
    result = classify_email(
        d.get("subject", ""), d.get("from_addr", ""), d.get("body_text", ""),
    )
    return jsonify(result)


@app.route("/api/accounting/check-expense-inbox", methods=["POST"])
@require_role("admin")
def api_check_expense_inbox():
    """Manually trigger expense email processing.

    JSON body options:
        force: true — reprocess all emails (skip dedup for expenses/actions)
        days_back: 7 — how many days to look back (default 14)
    """
    data = request.get_json(silent=True) or {}
    force = data.get("force", False)
    days_back = data.get("days_back", 14)
    try:
        result = check_expense_inbox(force=force, days_back=days_back)
        return jsonify({"status": "ok", "result": result})
    except Exception as e:
        logger.exception("Manual expense inbox check failed")
        return jsonify({"error": str(e)}), 500


@app.route("/api/accounting/mail-folders", methods=["GET"])
@require_role("admin")
def api_list_mail_folders():
    """Debug: list all mail folders visible to Graph API."""
    from email_parser.fetcher import _get_graph_token, _request_with_retry, GRAPH_BASE
    tenant_id = os.getenv("AZURE_TENANT_ID")
    client_id = os.getenv("AZURE_CLIENT_ID")
    client_secret = os.getenv("AZURE_CLIENT_SECRET")
    address = os.getenv("EXPENSE_EMAIL_ADDRESS") or os.getenv("RSVP_EMAIL_ADDRESS") or os.getenv("EMAIL_ADDRESS")
    token = _get_graph_token(tenant_id, client_id, client_secret)
    if not token:
        return jsonify({"error": "Could not get token"}), 500
    headers = {"Authorization": f"Bearer {token}"}

    def _list(parent_id=None, depth=0):
        url = (f"{GRAPH_BASE}/users/{address}/mailFolders/{parent_id}/childFolders"
               if parent_id else f"{GRAPH_BASE}/users/{address}/mailFolders")
        try:
            resp = _request_with_retry("get", url, headers=headers, params={"$top": "100"}, timeout=15)
            if resp.status_code != 200:
                return []
            results = []
            for f in resp.json().get("value", []):
                entry = {"name": f["displayName"], "id": f["id"],
                         "total": f.get("totalItemCount", 0), "depth": depth}
                results.append(entry)
                if f.get("childFolderCount", 0) > 0:
                    results.extend(_list(f["id"], depth + 1))
            return results
        except Exception as e:
            return [{"error": str(e)}]

    folders = _list()
    return jsonify({"folders": folders})


@app.route("/api/accounting/expense-inbox-audit", methods=["POST"])
@require_role("admin")
def api_expense_inbox_audit():
    """Preview what's in the inbox without processing — classify only.

    Returns a list of emails with their classification and whether they were already processed.
    JSON body: { days_back: 7 }
    """
    data = request.get_json(silent=True) or {}
    days_back = data.get("days_back", 7)

    tenant_id = os.getenv("AZURE_TENANT_ID")
    client_id = os.getenv("AZURE_CLIENT_ID")
    client_secret = os.getenv("AZURE_CLIENT_SECRET")
    address = os.getenv("EXPENSE_EMAIL_ADDRESS") or os.getenv("RSVP_EMAIL_ADDRESS") or os.getenv("EMAIL_ADDRESS")
    if not all([tenant_id, client_id, client_secret, address]):
        return jsonify({"error": "Azure AD credentials not configured"}), 400

    try:
        emails = fetch_all_emails(
            tenant_id=tenant_id, client_id=client_id, client_secret=client_secret,
            email_address=address, since_date=datetime.now() - timedelta(days=days_back),
            max_emails=300,
            include_subfolders=["2025 Chase", "2025 Venmo", "Payouts", "Invoices"],
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    known_uids = get_known_email_uids()
    conn = get_connection()
    try:
        expense_uids = {r["email_uid"] for r in conn.execute(
            "SELECT email_uid FROM expense_transactions WHERE email_uid IS NOT NULL"
        ).fetchall()}
        action_uids = {r["email_uid"] for r in conn.execute(
            "SELECT email_uid FROM action_items WHERE email_uid IS NOT NULL"
        ).fetchall()}
    finally:
        conn.close()

    results = []
    for e in emails:
        uid = e.get("uid", "")
        status = "new"
        if uid in known_uids:
            status = "order"
        elif uid in expense_uids:
            status = "expense_saved"
        elif uid in action_uids:
            status = "action_saved"
        results.append({
            "subject": e.get("subject", ""),
            "from": e.get("from", ""),
            "date": (e.get("date") or "")[:10],
            "status": status,
        })

    counts = {"order": 0, "expense_saved": 0, "action_saved": 0, "new": 0}
    for r in results:
        counts[r["status"]] = counts.get(r["status"], 0) + 1

    return jsonify({"total": len(results), "counts": counts, "emails": results})


# ═══════════════════════════════════════════════════════════════════════════
# COO Dashboard
# ═══════════════════════════════════════════════════════════════════════════

@app.route("/coo")
@require_role("admin")
def coo_page():
    return render_template("coo.html")


@app.route("/api/coo/action-items")
@require_role("admin")
def api_coo_action_items():
    return jsonify(get_action_items(
        status=request.args.get("status"),
        category=request.args.get("category"),
    ))


@app.route("/api/coo/action-items/<int:aid>", methods=["PATCH"])
@require_role("admin")
def api_coo_update_action_item(aid):
    d = request.json or {}
    return jsonify(update_action_item(aid, d))


@app.route("/api/coo/action-items/batch-dismiss", methods=["POST"])
@require_role("admin")
def api_coo_batch_dismiss():
    d = request.json or {}
    return jsonify(batch_dismiss_action_items(
        item_ids=d.get("item_ids"),
        category=d.get("category"),
        status_filter=d.get("status_filter", "open"),
    ))


@app.route("/api/coo/action-items/consolidate", methods=["POST"])
@require_role("admin")
def api_coo_consolidate():
    return jsonify(consolidate_action_items())


@app.route("/api/coo/financial-snapshot")
@require_role("admin")
def api_coo_financial_snapshot():
    return jsonify(get_coo_financial_snapshot())


@app.route("/api/coo/manual-values", methods=["POST"])
@require_role("admin")
def api_coo_manual_values():
    d = request.json or {}
    if "key" not in d or "value" not in d:
        return jsonify({"error": "key and value required"}), 400
    set_coo_manual_value(d["key"], float(d["value"]))
    return jsonify({"status": "ok"})


@app.route("/api/coo/review-queue")
@require_role("admin")
def api_coo_review_queue():
    return jsonify(get_coo_review_queue())


@app.route("/api/coo/chat", methods=["POST"])
@require_role("admin")
def api_coo_chat():
    """COO Chat — routes to specialist agent, responds as Chief of Staff.
    Now persists all messages to coo_chat_sessions/coo_chat_messages so the AI
    retains full conversation context across page reloads and sessions."""
    from email_parser.database import route_to_agent, get_coo_agent, log_agent_action
    d = request.json or {}
    user_message = d.get("message", "")
    session_id = d.get("session_id")
    context = d.get("context", {})

    # Legacy support: if caller sends "messages" array instead of "message" string
    if not user_message and d.get("messages"):
        msgs = d["messages"]
        user_message = msgs[-1].get("content", "") if msgs else ""

    if not user_message:
        return jsonify({"error": "message required"}), 400

    api_key = os.getenv("ANTHROPIC_API_KEY")
    if not api_key:
        return jsonify({"error": "ANTHROPIC_API_KEY not configured"}), 500

    # Create or load session
    if session_id:
        session = get_chat_session(session_id)
        if not session:
            session = create_chat_session()
    else:
        session = create_chat_session()
    session_id = session["id"]

    # Save the user message
    add_chat_message(session_id, "user", user_message)

    # Auto-title the session from the first user message
    if session.get("title") == "New Chat" and user_message:
        short_title = user_message[:60] + ("..." if len(user_message) > 60 else "")
        update_chat_session_title(session_id, short_title)

    # Build message history from the DB (full session context for the AI)
    session = get_chat_session(session_id)
    messages = [{"role": m["role"], "content": m["content"]} for m in session.get("messages", [])]

    try:
        # Build full business context from all tracker modules
        try:
            full_context = build_coo_full_context()
        except Exception:
            full_context = "(Business intelligence temporarily unavailable)"

        # Build master context — summaries of ALL past sessions
        try:
            master_context = get_chat_master_context(exclude_session_id=session_id)
        except Exception:
            master_context = ""

        # Route the latest user message to a specialist agent
        routed_agent = route_to_agent(user_message)

        # Get the specialist's system prompt
        agent = get_coo_agent(routed_agent)
        specialist_prompt = agent["system_prompt"] if agent else ""

        # Always respond as Chief of Staff, with specialist context
        cos_agent = get_coo_agent("Chief of Staff")
        cos_prompt = cos_agent["system_prompt"] if cos_agent else ""

        system_prompt = f"""{cos_prompt}

--- SPECIALIST CONTEXT ---
For this question, the {routed_agent} provided analysis context:
{specialist_prompt}

--- FULL BUSINESS INTELLIGENCE ---
Live data from the TGF Transaction Tracker as of {datetime.now().strftime('%Y-%m-%d %H:%M')}:

{full_context}"""

        if master_context:
            system_prompt += f"""

--- PERSISTENT MEMORY ---
{master_context}"""

        client = _anthropic.Anthropic(api_key=api_key)
        resp = client.messages.create(
            model="claude-sonnet-4-5-20250929",
            max_tokens=1500,
            system=system_prompt,
            messages=messages,
        )

        assistant_content = resp.content[0].text

        # Save the assistant response
        add_chat_message(session_id, "assistant", assistant_content, routed_to=routed_agent)

        # Auto-summarize the session after each exchange (lightweight — no extra AI call)
        # Build summary from key topics discussed: all user messages, condensed
        updated_session = get_chat_session(session_id)
        user_msgs = [m["content"] for m in updated_session.get("messages", []) if m["role"] == "user"]
        asst_msgs = [m["content"] for m in updated_session.get("messages", []) if m["role"] == "assistant"]
        # Summary = first 3 user questions + last assistant key points
        topics = "; ".join(msg[:80] for msg in user_msgs[:5])
        last_answer = asst_msgs[-1][:150] if asst_msgs else ""
        auto_summary = f"Topics: {topics}. Last response: {last_answer}"
        update_chat_session_summary(session_id, auto_summary[:500])

        # Log the routing decision
        log_agent_action(routed_agent, "chat_routing",
                         f"Routed question to {routed_agent}: {user_message[:100]}",
                         outcome="response_generated")

        return jsonify({
            "role": "assistant",
            "content": assistant_content,
            "routed_to": routed_agent,
            "session_id": session_id,
        })
    except Exception as e:
        return jsonify({"error": str(e), "session_id": session_id}), 500


# ── COO Chat Session Management ────────────────────────────

@app.route("/api/coo/chat-sessions", methods=["GET"])
@require_role("admin")
def api_coo_chat_sessions():
    """List recent chat sessions."""
    return jsonify(get_chat_sessions(limit=30))


@app.route("/api/coo/chat-sessions", methods=["POST"])
@require_role("admin")
def api_coo_create_chat_session():
    """Create a new chat session."""
    d = request.json or {}
    return jsonify(create_chat_session(title=d.get("title", "New Chat")))


@app.route("/api/coo/chat-sessions/<int:sid>", methods=["GET"])
@require_role("admin")
def api_coo_get_chat_session(sid):
    """Get a chat session with all messages."""
    sess = get_chat_session(sid)
    if not sess:
        return jsonify({"error": "Session not found"}), 404
    return jsonify(sess)


@app.route("/api/coo/chat-sessions/<int:sid>", methods=["DELETE"])
@require_role("admin")
def api_coo_delete_chat_session(sid):
    """Delete a chat session."""
    return jsonify(delete_chat_session(sid))


@app.route("/api/coo/chat-sessions/<int:sid>", methods=["PATCH"])
@require_role("admin")
def api_coo_rename_chat_session(sid):
    """Rename a chat session."""
    d = request.json or {}
    title = d.get("title", "").strip()
    if not title:
        return jsonify({"error": "title required"}), 400
    update_chat_session_title(sid, title)
    return jsonify({"status": "ok", "title": title})


@app.route("/api/coo/send-daily-email", methods=["POST"])
@require_role("admin")
def api_coo_send_daily_email():
    """Manually trigger the COO daily email for testing."""
    coo_to = os.getenv("COO_EMAIL_TO", "kerry@thegolffellowship.com")
    ok = send_coo_daily_email()
    if ok:
        return jsonify({"sent": True, "to": coo_to})
    # If Azure not configured, still return the HTML for preview
    try:
        subject, html_body = build_coo_email_html()
        return jsonify({"sent": False, "preview": True, "subject": subject, "html": html_body, "to": coo_to})
    except Exception as e:
        return jsonify({"sent": False, "error": str(e)}), 500


@app.route("/api/coo/agents")
@require_role("admin")
def api_coo_agents():
    return jsonify(get_coo_agents())


@app.route("/api/coo/agent-log")
@require_role("admin")
def api_coo_agent_log():
    return jsonify(get_agent_action_log(
        agent_name=request.args.get("agent_name"),
        date_from=request.args.get("date_from"),
        date_to=request.args.get("date_to"),
        limit=request.args.get("limit", 50, type=int),
    ))


# ═══════════════════════════════════════════════════════════════════════════
# Bank Reconciliation
# ═══════════════════════════════════════════════════════════════════════════

@app.route("/api/accounting/chart-of-accounts")
@require_role("admin")
def api_chart_of_accounts():
    return jsonify(get_chart_of_accounts())


@app.route("/api/accounting/ledger")
@require_role("admin")
def api_ledger_entries():
    return jsonify(get_ledger_entries(
        account_code=request.args.get("account_code"),
        date_from=request.args.get("date_from"),
        date_to=request.args.get("date_to"),
        reconciled=request.args.get("reconciled", type=int),
    ))


@app.route("/api/accounting/bank-import", methods=["POST"])
@require_role("admin")
def api_bank_import():
    """Upload and import a bank statement CSV."""
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400
    csv_text = request.files["file"].read().decode("utf-8", errors="replace")
    bank = request.form.get("bank", "Chase")
    account_last4 = request.form.get("account_last4", "")
    result = import_bank_statement(csv_text, bank, account_last4)
    return jsonify(result)


@app.route("/api/accounting/reconcile", methods=["POST"])
@require_role("admin")
def api_reconcile():
    """Run auto-match on imported bank rows."""
    d = request.json or {}
    result = run_bank_reconciliation(
        import_id=d.get("import_id"),
        account_last4=d.get("account_last4"),
        month=d.get("month"),
    )
    return jsonify(result)


@app.route("/api/accounting/reconcile/match", methods=["POST"])
@require_role("admin")
def api_reconcile_match():
    """Manually confirm a match between a bank row and a Tracker record."""
    d = request.json or {}
    bank_row_id = d.get("bank_row_id")
    matched_source = d.get("matched_source")
    matched_id = d.get("matched_id")
    if not bank_row_id:
        return jsonify({"error": "bank_row_id required"}), 400
    from email_parser.database import _connect
    with _connect() as conn:
        conn.execute(
            """UPDATE bank_statement_rows
               SET reconciled = 1, matched_source = ?, matched_id = ?
               WHERE id = ?""",
            (matched_source, matched_id, bank_row_id),
        )
        conn.commit()
    return jsonify({"status": "ok"})


@app.route("/api/accounting/reconciliation-summary")
@require_role("admin")
def api_reconciliation_summary():
    month = request.args.get("month")
    if not month:
        return jsonify({"error": "month parameter required (YYYY-MM)"}), 400
    return jsonify(get_reconciliation_summary(month))


@app.route("/api/accounting/close-period", methods=["POST"])
@require_role("admin")
def api_close_period():
    d = request.json or {}
    period = d.get("period")
    if not period:
        return jsonify({"error": "period required (YYYY-MM)"}), 400
    return jsonify(close_period(period))


# ── Bank Deposit Reconciliation Routes ──────────────────────────────────

@app.route("/accounting/reconcile")
@require_role("admin")
def page_reconcile():
    return render_template("reconcile.html")


@app.route("/accounting/cashflow")
@require_role("admin")
def page_cashflow():
    return render_template("cashflow.html")


@app.route("/accounting/money-flow")
@require_role("admin")
def page_money_flow():
    """Monthly Money Flow — pass-through vs TGF-keep waterfall (mailbox
    #242, Kerry approved). Kills the recurring 'we collected $30K but
    TGF's cut is only $5K — how?' question."""
    return render_template("moneyflow.html")


@app.route("/api/accounting/money-flow")
@require_role("admin")
def api_money_flow():
    from email_parser.database import get_monthly_money_flow
    month = (request.args.get("month") or "").strip()
    debug = request.args.get("debug") in ("1", "true")
    chapter = (request.args.get("chapter") or "").strip() or None
    ytd = request.args.get("ytd") in ("1", "true")
    try:
        return jsonify(get_monthly_money_flow(month, debug=debug,
                                              chapter=chapter, ytd=ytd))
    except ValueError as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/reconciliation/accounts")
@require_role("admin")
def api_recon_accounts():
    return jsonify(get_bank_accounts())


@app.route("/api/reconciliation/dashboard")
@require_role("admin")
def api_recon_dashboard():
    return jsonify(get_reconciliation_dashboard())


@app.route("/api/reconciliation/import", methods=["POST"])
@require_role("admin")
def api_recon_import():
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400
    f = request.files["file"]
    account_id = request.form.get("account_id", type=int)
    if not account_id:
        return jsonify({"error": "account_id required"}), 400
    file_bytes = f.read()
    # Detect Venmo statement format before default import path
    csv_text = file_bytes.decode("utf-8", errors="replace")
    first_line = csv_text.split("\n", 1)[0].strip()
    if "Account Statement" in first_line or first_line.lower().startswith("transaction id"):
        result = import_venmo_statement(csv_text, "Venmo")
        # Auto-match after import
        if result.get("imported", 0) > 0:
            match_result = run_deposit_auto_match()
            result["auto_match"] = match_result
        return jsonify(result)
    result = import_bank_deposits(file_bytes, f.filename or "upload.csv", account_id)
    # Auto-match after import
    if result.get("imported", 0) > 0:
        match_result = run_deposit_auto_match(account_id)
        result["auto_match"] = match_result
    return jsonify(result)


@app.route("/api/reconciliation/auto-match", methods=["POST"])
@require_role("admin")
def api_recon_auto_match():
    d = request.json or {}
    account_id = d.get("account_id")
    return jsonify(run_deposit_auto_match(account_id))


@app.route("/api/admin/run-recon-drift-fix", methods=["POST"])
@require_role("admin")
def api_admin_run_recon_drift_fix():
    """One-shot remediation for April-2026 reconciliation drift.

    Runs the same logic as scripts/fix_recon_drift_2026_04.py and
    returns a JSON report. Idempotent — safe to call repeatedly.
    """
    from email_parser.recon_drift_fix import apply_recon_drift_fix
    try:
        return jsonify(apply_recon_drift_fix())
    except Exception as e:  # noqa: BLE001
        logger.exception("recon-drift-fix failed")
        return jsonify({"error": str(e)}), 500


@app.route("/api/reconciliation/match", methods=["POST"])
@require_role("admin")
def api_recon_match():
    d = request.json or {}
    bank_deposit_id = d.get("bank_deposit_id")
    acct_transaction_id = d.get("acct_transaction_id")
    if not bank_deposit_id or not acct_transaction_id:
        return jsonify({"error": "bank_deposit_id and acct_transaction_id required"}), 400
    return jsonify(manual_match_deposit(bank_deposit_id, acct_transaction_id))


@app.route("/api/reconciliation/match-batch", methods=["POST"])
@require_role("admin")
def api_recon_match_batch():
    """Match multiple acct_transactions to a single bank deposit (1:many)."""
    d = request.json or {}
    bank_deposit_id = d.get("bank_deposit_id")
    acct_transaction_ids = d.get("acct_transaction_ids", [])
    if not bank_deposit_id or not acct_transaction_ids:
        return jsonify({"error": "bank_deposit_id and acct_transaction_ids required"}), 400
    return jsonify(batch_match_deposit(bank_deposit_id, acct_transaction_ids))


@app.route("/api/reconciliation/unmatch", methods=["POST"])
@require_role("admin")
def api_recon_unmatch():
    d = request.json or {}
    bank_deposit_id = d.get("bank_deposit_id")
    acct_transaction_id = d.get("acct_transaction_id")
    if not bank_deposit_id:
        return jsonify({"error": "bank_deposit_id required"}), 400
    return jsonify(unmatch_deposit(bank_deposit_id, acct_transaction_id))


@app.route("/api/reconciliation/create-entry", methods=["POST"])
@require_role("admin")
def api_recon_create_entry():
    """Create a ledger entry from an unmatched bank deposit and immediately reconcile it."""
    d = request.json or {}
    deposit_id = d.get("deposit_id")
    if not deposit_id:
        return jsonify({"error": "deposit_id required"}), 400
    result = create_entry_from_deposit(
        deposit_id=deposit_id,
        txn_type=d.get("txn_type", "expense"),
        category_name=d.get("category_name"),
        entity_name=d.get("entity_name"),
        notes=d.get("notes"),
        description=d.get("description"),
        date_override=d.get("date"),
        amount_override=d.get("amount"),
        event_name=d.get("event_name"),
        entry_type=d.get("entry_type"),
    )
    return jsonify(result)


@app.route("/api/reconciliation/deposits")
@require_role("admin")
def api_recon_deposits():
    account_id = request.args.get("account_id", type=int)
    status = request.args.get("status")
    month = request.args.get("month")
    include_dismissed = request.args.get("include_dismissed", "false").lower() == "true"
    return jsonify(get_bank_deposits(account_id, status, month, include_dismissed))


@app.route("/api/reconciliation/dismiss-deposit", methods=["POST"])
@require_role("admin")
def api_dismiss_deposit():
    d = request.json or {}
    deposit_id = d.get("deposit_id")
    reason = d.get("reason", "not_applicable")
    if not deposit_id:
        return jsonify({"error": "deposit_id required"}), 400
    return jsonify(dismiss_bank_deposit(deposit_id, reason))


@app.route("/api/reconciliation/record-transfer", methods=["POST"])
@require_role("admin")
def api_record_transfer():
    d = request.json or {}
    deposit_id = d.get("deposit_id")
    from_account = d.get("from_account", "TGF Checking")
    to_account = d.get("to_account", "Venmo")
    notes = d.get("notes", "")
    if not deposit_id:
        return jsonify({"error": "deposit_id required"}), 400
    return jsonify(record_internal_transfer(deposit_id, from_account, to_account, notes))


@app.route("/api/reconciliation/unreconciled")
@require_role("admin")
def api_recon_unreconciled():
    account = request.args.get("account")
    month = request.args.get("month")
    date_from = request.args.get("date_from")
    date_to = request.args.get("date_to")
    source = request.args.get("source")
    return jsonify(get_unreconciled_transactions(account, month,
                                                 date_from=date_from, date_to=date_to,
                                                 source=source))


@app.route("/api/reconciliation/suggestions/<int:deposit_id>")
@require_role("admin")
def api_recon_suggestions(deposit_id):
    return jsonify(get_match_suggestions(deposit_id))


@app.route("/api/reconciliation/matched/<int:deposit_id>")
@require_role("admin")
def api_recon_matched(deposit_id):
    """Return the acct_transactions matched to a specific bank deposit."""
    from email_parser.database import _connect
    with _connect() as conn:
        rows = conn.execute(
            """SELECT t.*, rm.match_confidence, rm.match_type
               FROM acct_transactions t
               JOIN reconciliation_matches rm ON rm.acct_transaction_id = t.id
               WHERE rm.bank_deposit_id = ?
               ORDER BY t.date""",
            (deposit_id,),
        ).fetchall()
        return jsonify([dict(r) for r in rows])


@app.route("/api/reconciliation/reconciled-items")
@require_role("admin")
def api_recon_reconciled_items():
    """Return item_id → bank_deposit_id mapping for reconciled items."""
    from email_parser.database import _connect
    with _connect() as conn:
        rows = conn.execute(
            """SELECT DISTINCT t.item_id, rm.bank_deposit_id
               FROM acct_transactions t
               JOIN reconciliation_matches rm ON rm.acct_transaction_id = t.id
               WHERE t.item_id IS NOT NULL"""
        ).fetchall()
        # Also get item_ids from order-level entries via splits
        split_rows = conn.execute(
            """SELECT DISTINCT s.item_id, rm.bank_deposit_id
               FROM godaddy_order_splits s
               JOIN reconciliation_matches rm ON rm.acct_transaction_id = s.transaction_id
               WHERE s.item_id IS NOT NULL"""
        ).fetchall()
        mapping = {}
        for r in rows:
            mapping[r["item_id"]] = r["bank_deposit_id"]
        for r in split_rows:
            mapping[r["item_id"]] = r["bank_deposit_id"]
        return jsonify(mapping)


@app.route("/api/reconciliation/monthly")
@require_role("admin")
def api_recon_monthly():
    month = request.args.get("month")
    if not month:
        return jsonify({"error": "month required (YYYY-MM)"}), 400
    return jsonify(get_monthly_reconciliation(month))


@app.route("/api/reconciliation/event/<event_name>")
@require_role("manager")
def api_recon_event(event_name):
    return jsonify(get_event_reconciliation_status(event_name))


@app.route("/api/reconciliation/cashflow")
@require_role("admin")
def api_cashflow():
    weeks = request.args.get("weeks", 13, type=int)
    return jsonify(get_cashflow_data(weeks))


@app.route("/api/reconciliation/migrate-to-order-level", methods=["POST"])
@require_role("admin")
def api_migrate_to_order_level():
    """Migrate old per-item GoDaddy entries to order-level format."""
    results = migrate_item_to_order_entries()
    return jsonify(results)


@app.route("/api/reconciliation/cleanup-godaddy-duplicates", methods=["POST"])
@require_role("admin")
def api_cleanup_godaddy_duplicates():
    """Reverse old per-item GoDaddy entries that coexist with newer order-level entries."""
    results = cleanup_duplicate_godaddy_entries()
    return jsonify(results)


@app.route("/api/reconciliation/merge-transactions", methods=["POST"])
@require_role("admin")
def api_merge_transactions():
    """Merge multiple GoDaddy order transactions into a single batch entry."""
    d = request.json or {}
    ids = d.get("acct_transaction_ids", [])
    if len(ids) < 2:
        return jsonify({"error": "Need at least 2 transaction IDs"}), 400
    return jsonify(merge_transactions(ids))


# ---------------------------------------------------------------------------
# TGF Payouts
# ---------------------------------------------------------------------------

@app.route("/tgf")
def tgf_page():
    # Payouts are admin-only (Kerry, 2026-07-08: managers must not make
    # prize payouts or see any payout tabs)
    if session.get("role") != "admin":
        return redirect("/events")
    return render_template("tgf.html")


@app.route("/api/tgf")
@require_role("admin")
def api_tgf_data():
    return jsonify(get_tgf_data())


@app.route("/api/tgf", methods=["POST"])
@require_role("admin")
def api_tgf_action():
    d = request.json or {}
    action = d.get("action")
    if action == "add_event":
        return jsonify(add_tgf_event(d))
    elif action == "add_golfer":
        return jsonify(add_tgf_golfer(d))
    elif action == "import_payouts":
        from email_parser.database import import_tgf_payouts
        if not d.get("event_id"):
            return jsonify({"error": "event_id required"}), 400
        return jsonify(import_tgf_payouts(d["event_id"], d.get("payouts", [])))
    elif action == "import_golfers":
        return jsonify(import_tgf_golfers(d.get("golfers", [])))
    elif action == "update_event":
        return jsonify(update_tgf_event(d["event_id"], d))
    elif action == "delete_event":
        return jsonify(delete_tgf_event(d["event_id"]))
    else:
        return jsonify({"error": f"Unknown action: {action}"}), 400


@app.route("/api/tgf/dedup-audit", methods=["GET"])
@require_role("admin")
def api_tgf_dedup_audit():
    """Diagnostic: list customers created by the tgf_golfers migration and
    find likely duplicates of existing customers (name-format mismatches).

    For each customer with acquisition_source IN ('tgf_payout','tgf_payout_migration'),
    try to find an existing customer whose name matches in reversed format
    ("LAST, First" ↔ "First Last") or normalized form.
    """
    from email_parser.database import _connect

    def normalize(s: str) -> str:
        return " ".join((s or "").strip().split()).lower()

    def parse_commaed(name: str) -> tuple[str, str] | None:
        """Parse 'LAST, First' → ('First', 'LAST'). Returns None if not that format."""
        if "," not in name:
            return None
        parts = [p.strip() for p in name.split(",", 1)]
        if len(parts) != 2 or not parts[0] or not parts[1]:
            return None
        # LAST, First → return (First, Last)
        return parts[1], parts[0]

    results = []
    with _connect() as conn:
        # All customers created from tgf_golfers migration
        migrated_rows = conn.execute(
            """SELECT customer_id, first_name, last_name, venmo_username, chapter,
                      acquisition_source, created_at
               FROM customers
               WHERE acquisition_source IN ('tgf_payout', 'tgf_payout_migration')
               ORDER BY customer_id"""
        ).fetchall()

        for mr in migrated_rows:
            mr_d = dict(mr)
            full_name = f"{mr_d['first_name']} {mr_d['last_name']}".strip()

            # How many payouts point to this customer?
            payout_stats = conn.execute(
                """SELECT COUNT(*) as cnt, COALESCE(SUM(amount), 0) as total
                   FROM tgf_payouts WHERE customer_id = ?""",
                (mr_d["customer_id"],),
            ).fetchone()

            # Try to find candidate matches in the customers table
            candidates = []

            # Case 1: stored as "LAST, First" with last_name containing comma format
            # e.g., first_name="CAMPOS" last_name="Roland" (because parser split incorrectly)
            # Try the reversed interpretation: treat last_name as first, first_name as last
            reversed_candidates = conn.execute(
                """SELECT customer_id, first_name, last_name, venmo_username, chapter,
                          acquisition_source
                   FROM customers
                   WHERE customer_id != ?
                     AND acquisition_source IS NOT 'tgf_payout_migration'
                     AND acquisition_source IS NOT 'tgf_payout'
                     AND (
                       (LOWER(first_name) = LOWER(?) AND LOWER(last_name) = LOWER(?))
                       OR (LOWER(first_name) = LOWER(?) AND LOWER(last_name) = LOWER(?))
                     )""",
                (
                    mr_d["customer_id"],
                    # interpretation 1: first_name/last_name swapped (comma form)
                    mr_d["last_name"], mr_d["first_name"],
                    # interpretation 2: same as stored (exact match)
                    mr_d["first_name"], mr_d["last_name"],
                ),
            ).fetchall()

            for rc in reversed_candidates:
                candidates.append({
                    "match_type": "reversed_exact",
                    **dict(rc),
                })

            # Case 2: fuzzy — same last name, either first name matches as initial or shared prefix
            if not candidates and mr_d["last_name"]:
                fuzzy = conn.execute(
                    """SELECT customer_id, first_name, last_name, venmo_username, chapter,
                              acquisition_source
                       FROM customers
                       WHERE customer_id != ?
                         AND acquisition_source IS NOT 'tgf_payout_migration'
                         AND acquisition_source IS NOT 'tgf_payout'
                         AND (
                           LOWER(last_name) = LOWER(?)
                           OR LOWER(first_name) = LOWER(?)
                         )
                       LIMIT 5""",
                    (mr_d["customer_id"], mr_d["last_name"], mr_d["last_name"]),
                ).fetchall()
                for f in fuzzy:
                    candidates.append({
                        "match_type": "fuzzy_lastname",
                        **dict(f),
                    })

            # Case 3: "LAST, First" parsing bug — the comma stayed in first_name.
            # The REAL last name is first_name with comma stripped. Search for that.
            first_stripped = (mr_d["first_name"] or "").rstrip(",").strip()
            if first_stripped and first_stripped != mr_d["last_name"]:
                comma_bug = conn.execute(
                    """SELECT customer_id, first_name, last_name, venmo_username, chapter,
                              acquisition_source
                       FROM customers
                       WHERE customer_id != ?
                         AND acquisition_source IS NOT 'tgf_payout_migration'
                         AND acquisition_source IS NOT 'tgf_payout'
                         AND LOWER(last_name) = LOWER(?)
                         AND LOWER(first_name) = LOWER(?)
                       LIMIT 5""",
                    (mr_d["customer_id"], first_stripped, mr_d["last_name"]),
                ).fetchall()
                for c in comma_bug:
                    # Avoid duplicates from previous cases
                    if not any(cand.get("customer_id") == c["customer_id"] for cand in candidates):
                        candidates.append({
                            "match_type": "comma_bug_reversed_exact",
                            **dict(c),
                        })

                # Broader: same real last name, any first name
                comma_bug_fuzzy = conn.execute(
                    """SELECT customer_id, first_name, last_name, venmo_username, chapter,
                              acquisition_source
                       FROM customers
                       WHERE customer_id != ?
                         AND acquisition_source IS NOT 'tgf_payout_migration'
                         AND acquisition_source IS NOT 'tgf_payout'
                         AND LOWER(last_name) = LOWER(?)
                       LIMIT 5""",
                    (mr_d["customer_id"], first_stripped),
                ).fetchall()
                for c in comma_bug_fuzzy:
                    if not any(cand.get("customer_id") == c["customer_id"] for cand in candidates):
                        candidates.append({
                            "match_type": "comma_bug_lastname_match",
                            **dict(c),
                        })

            # Also look for intra-migration duplicates (two customers both created by migration)
            intra_candidates = conn.execute(
                """SELECT customer_id, first_name, last_name, venmo_username, chapter,
                          acquisition_source
                   FROM customers
                   WHERE customer_id != ?
                     AND acquisition_source IN ('tgf_payout', 'tgf_payout_migration')
                     AND (
                       (LOWER(first_name) = LOWER(?) AND LOWER(last_name) = LOWER(?))
                       OR (LOWER(first_name) = LOWER(?) AND LOWER(last_name) = LOWER(?))
                     )""",
                (
                    mr_d["customer_id"],
                    mr_d["last_name"], mr_d["first_name"],
                    mr_d["first_name"], mr_d["last_name"],
                ),
            ).fetchall()
            for ic in intra_candidates:
                candidates.append({
                    "match_type": "intra_migration_reversed",
                    **dict(ic),
                })

            results.append({
                "migrated_customer_id": mr_d["customer_id"],
                "migrated_name": full_name,
                "first_name": mr_d["first_name"],
                "last_name": mr_d["last_name"],
                "venmo_username": mr_d["venmo_username"],
                "chapter": mr_d["chapter"],
                "acquisition_source": mr_d["acquisition_source"],
                "payout_count": payout_stats["cnt"],
                "payout_total": round(payout_stats["total"] or 0, 2),
                "candidate_matches": candidates,
            })

        summary = {
            "total_migrated_customers": len(migrated_rows),
            "with_candidates": sum(1 for r in results if r["candidate_matches"]),
            "without_candidates": sum(1 for r in results if not r["candidate_matches"]),
        }

    return jsonify({"summary": summary, "customers": results})


@app.route("/api/tgf/match-diagnostic", methods=["GET"])
@require_role("admin")
def api_tgf_match_diagnostic():
    """Diagnostic: for each pending payout, show WHY it didn't match a Venmo payment.

    Groups pending payouts by (customer_id, event_id) and for each group shows:
      - expected sum
      - event date & 7-day window
      - customer name as stored in customers table
      - candidate Venmo prize_payout transactions (all, not just matches)
      - reason for non-match (amount / customer / date / already linked)
    """
    from email_parser.database import _connect
    with _connect() as conn:
        # Get all pending payout groups (those with source='pending' acct_transaction)
        pending = conn.execute(
            """SELECT p.id as payout_id, p.event_id, p.customer_id, p.amount, p.category,
                      p.acct_transaction_id, t.source as txn_source,
                      e.event_date, e.name as event_name,
                      (c.first_name || ' ' || c.last_name) as customer_name
               FROM tgf_payouts p
               JOIN tgf_events e ON e.id = p.event_id
               JOIN customers c ON c.customer_id = p.customer_id
               LEFT JOIN acct_transactions t ON t.id = p.acct_transaction_id
               WHERE p.paid_at IS NULL"""
        ).fetchall()

        # Group by (event_id, customer_id)
        groups = {}
        for p in pending:
            key = (p["event_id"], p["customer_id"])
            if key not in groups:
                groups[key] = {
                    "event_name": p["event_name"],
                    "event_date": p["event_date"],
                    "customer_id": p["customer_id"],
                    "customer_name": p["customer_name"],
                    "payouts": [],
                }
            groups[key]["payouts"].append({
                "payout_id": p["payout_id"],
                "amount": p["amount"],
                "category": p["category"],
                "txn_source": p["txn_source"],
            })

        results = []
        for (event_id, customer_id), g in groups.items():
            group_sum = round(sum(p["amount"] for p in g["payouts"]), 2)
            event_date = g["event_date"]

            # List ALL Venmo prize_payouts for this customer (any amount/date)
            all_venmo = conn.execute(
                """SELECT id, date, amount, customer, description,
                          COALESCE(status, 'active') as status,
                          (SELECT existing.id FROM tgf_payouts existing
                           WHERE existing.acct_transaction_id = t.id LIMIT 1) as already_linked_to
                   FROM acct_transactions t
                   WHERE source = 'venmo' AND category = 'prize_payout'
                     AND LOWER(customer) = LOWER(?)
                   ORDER BY date DESC""",
                (g["customer_name"],),
            ).fetchall()

            candidates = []
            for v in all_venmo:
                v_dict = dict(v)
                # Compute why this venmo didn't match (or did)
                reasons = []
                if round(abs(v_dict["amount"]), 2) != group_sum:
                    reasons.append(f"amount: venmo ${abs(v_dict['amount']):.2f} vs needed ${group_sum:.2f}")
                if v_dict["status"] != "active":
                    reasons.append(f"status: {v_dict['status']}")
                if v_dict["already_linked_to"]:
                    reasons.append(f"already linked to payout {v_dict['already_linked_to']}")
                # Date check
                date_ok = conn.execute(
                    "SELECT DATE(?) >= DATE(?) AND DATE(?) <= DATE(?, '+7 days') as ok",
                    (v_dict["date"], event_date, v_dict["date"], event_date),
                ).fetchone()
                if not date_ok or not date_ok["ok"]:
                    reasons.append(f"date: {v_dict['date']} outside {event_date} to +7d")
                v_dict["match_blockers"] = reasons or ["would match ✓"]
                candidates.append(v_dict)

            # Also search for venmo transactions where customer name might differ
            name_variants = conn.execute(
                """SELECT id, date, amount, customer, description
                   FROM acct_transactions
                   WHERE source = 'venmo' AND category = 'prize_payout'
                     AND ROUND(ABS(amount), 2) = ?
                     AND DATE(date) >= DATE(?)
                     AND DATE(date) <= DATE(?, '+7 days')
                     AND LOWER(customer) != LOWER(?)""",
                (group_sum, event_date, event_date, g["customer_name"]),
            ).fetchall()

            results.append({
                "event_name": g["event_name"],
                "event_date": g["event_date"],
                "customer_name": g["customer_name"],
                "customer_id": g["customer_id"],
                "payout_sum": group_sum,
                "payout_count": len(g["payouts"]),
                "payouts": g["payouts"],
                "all_venmo_for_this_customer": candidates,
                "venmo_amount_matches_but_different_customer_name": [dict(v) for v in name_variants],
            })

    return jsonify({
        "total_pending_groups": len(results),
        "groups": results,
    })


@app.route("/api/tgf/mark-paid", methods=["POST"])
@require_role("admin")
def api_tgf_mark_paid():
    """Mark a group of payouts as paid via non-Venmo method.

    Body: {event_id, customer_id, payment_method, paid_date?, reference?}

    Creates a real acct_transaction (expense/prize_payout) with the
    specified source, reverses the pending placeholder, and links all
    matching tgf_payouts rows.
    """
    from email_parser.database import _connect
    d = request.json or {}
    event_id = d.get("event_id")
    customer_id = d.get("customer_id")
    payment_method = (d.get("payment_method") or "").strip().lower()
    paid_date = d.get("paid_date")  # YYYY-MM-DD, defaults to today
    reference = d.get("reference") or ""

    if not event_id or not customer_id or not payment_method:
        return jsonify({"error": "event_id, customer_id, and payment_method required"}), 400

    # Whitelist allowed sources
    ALLOWED_SOURCES = {"paypal", "cashapp", "cash", "check", "zelle", "other"}
    if payment_method not in ALLOWED_SOURCES:
        return jsonify({"error": f"payment_method must be one of {sorted(ALLOWED_SOURCES)}"}), 400

    if not paid_date:
        from datetime import date as _date
        paid_date = _date.today().isoformat()

    with _connect() as conn:
        # Find all pending payouts for this customer+event
        pending = conn.execute(
            """SELECT p.id as payout_id, p.amount, p.acct_transaction_id,
                      t.source as txn_source, t.status as txn_status,
                      e.name as event_name,
                      TRIM(c.first_name || ' ' || c.last_name || COALESCE(' ' || NULLIF(TRIM(c.suffix), ''), '')) as customer_name
               FROM tgf_payouts p
               JOIN tgf_events e ON e.id = p.event_id
               JOIN customers c ON c.customer_id = p.customer_id
               LEFT JOIN acct_transactions t ON t.id = p.acct_transaction_id
               WHERE p.event_id = ? AND p.customer_id = ?
                 AND (p.paid_at IS NULL OR t.source = 'pending')""",
            (event_id, customer_id),
        ).fetchall()

        if not pending:
            return jsonify({"error": "No pending payouts found for this customer+event"}), 404

        total_amount = round(sum(float(p["amount"]) for p in pending), 2)
        customer_name = pending[0]["customer_name"]
        event_name = pending[0]["event_name"]

        # Create the real acct_transaction for the payment
        description = f"{payment_method.upper()} payout: {customer_name} — {event_name}"
        if reference:
            description += f" (ref: {reference})"

        cur = conn.execute(
            """INSERT INTO acct_transactions
                   (date, description, total_amount, type, source, source_ref,
                    customer, order_id, entry_type, category, amount, account, status, event_name)
               VALUES (?, ?, ?, 'expense', ?, ?, ?, ?, 'expense', 'prize_payout',
                       ?, ?, 'active', ?)""",
            (
                paid_date, description, total_amount, payment_method,
                f"manual-payout-{event_id}-{customer_id}",
                customer_name,
                f"MANUAL-PAYOUT-{event_id}-{customer_id}",
                -total_amount,
                payment_method.capitalize(),
                event_name,
            ),
        )
        new_txn_id = cur.lastrowid

        # Reverse the pending placeholders + link payouts to the new entry
        linked = 0
        for p in pending:
            if p["acct_transaction_id"] and p["txn_source"] == "pending":
                conn.execute(
                    "UPDATE acct_transactions SET status = 'reversed' WHERE id = ?",
                    (p["acct_transaction_id"],),
                )
            conn.execute(
                "UPDATE tgf_payouts SET acct_transaction_id = ?, paid_at = ? WHERE id = ?",
                (new_txn_id, paid_date, p["payout_id"]),
            )
            linked += 1

        conn.commit()
        return jsonify({
            "ok": True, "linked_payouts": linked,
            "amount": total_amount, "payment_method": payment_method,
            "acct_transaction_id": new_txn_id,
        })


@app.route("/api/tgf/parse-screenshot", methods=["POST"])
@require_role("admin")
def api_tgf_parse_screenshot():
    """Accept a base64 image, send to Claude vision, return parsed payouts JSON."""
    d = request.json or {}
    image_data = d.get("image")
    if not image_data:
        return jsonify({"error": "No image data provided"}), 400

    api_key = os.getenv("ANTHROPIC_API_KEY")
    if not api_key:
        return jsonify({"error": "ANTHROPIC_API_KEY not configured"}), 500

    # Strip data URL prefix if present
    if "," in image_data:
        image_data = image_data.split(",", 1)[1]

    media_type = d.get("media_type", "image/png")

    client = _anthropic.Anthropic(api_key=api_key)
    try:
        resp = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=4096,
            messages=[{
                "role": "user",
                "content": [
                    {
                        "type": "image",
                        "source": {"type": "base64", "media_type": media_type, "data": image_data},
                    },
                    {
                        "type": "text",
                        "text": (
                            "Parse this Golf Genius payout screenshot. Return ONLY valid JSON with:\n"
                            "{\n"
                            '  "event": {"code": "s9.X ...", "name": "Event Name", "date": "YYYY-MM-DD", "course": "Course Name"},\n'
                            '  "payouts": [\n'
                            '    {"golferName": "First Last", "category": "team_net|individual_net|individual_gross|skins|closest_to_pin|hole_in_one|mvp|other", "amount": 12.50, "description": "Game description"}\n'
                            "  ]\n"
                            "}\n\n"
                            "Categories: team_net, individual_net, individual_gross, skins, closest_to_pin, hole_in_one, mvp, other.\n"
                            "Extract every payout line. Amount should be a number (no $ sign).\n"
                            "Return ONLY the JSON — no markdown, no explanation."
                        ),
                    },
                ],
            }],
        )
        text = resp.content[0].text.strip()
        # Try to parse JSON from the response
        if text.startswith("```"):
            text = text.split("\n", 1)[1].rsplit("```", 1)[0].strip()
        parsed = json.loads(text)
        return jsonify(parsed)
    except json.JSONDecodeError:
        return jsonify({"error": "Failed to parse AI response as JSON", "raw": text}), 422
    except Exception as e:
        logger.exception("Screenshot parse failed")
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# App startup
# ---------------------------------------------------------------------------
init_db()

# ── Run acct_transactions backfill once if unfilled entries exist ──
try:
    from email_parser.database import _connect as _startup_connect
    with _startup_connect() as _conn:
        _acct_count = _conn.execute(
            "SELECT COUNT(*) as cnt FROM acct_transactions WHERE entry_type IS NOT NULL"
        ).fetchone()["cnt"]
    if _acct_count == 0:
        _bf_result = backfill_acct_transactions()
        logger.info("Startup backfill: %s", _bf_result)
    else:
        logger.info("Accounting entries exist (%d), skipping full backfill", _acct_count)

    # Always run the targeted GoDaddy-order backfill — idempotent, only
    # touches orders that are missing an active ledger entry. Catches new
    # events whose order rows arrived after the one-shot full backfill.
    try:
        _gd_added = backfill_missing_godaddy_orders()
        if _gd_added:
            logger.info("Startup: backfilled %d missing GoDaddy order entries", _gd_added)
    except Exception:
        logger.warning("Startup GoDaddy order backfill failed", exc_info=True)

    # Heal +PAY children whose parent_item_id points at a row that no longer
    # exists or was reverted to rsvp_only — the residue from a credit-transfer
    # reversal that ran before the Venmo balance-due cascade was fixed.
    # Idempotent: only touches genuine orphans.
    try:
        _orphan_repair = repair_orphan_pay_children()
        if any(_orphan_repair.values()):
            logger.info("Startup: orphan +PAY repair %s", _orphan_repair)
    except Exception:
        logger.warning("Startup orphan +PAY repair failed", exc_info=True)

    # Surface order-time email variants as customer_aliases so historical
    # typos on items.customer_email become visible on the Customer Info page
    # (where the manager can review/keep/delete each one). Idempotent.
    try:
        _email_aliases = capture_email_aliases_from_items()
        if _email_aliases:
            logger.info("Startup: captured %d email aliases from items", _email_aliases)
    except Exception:
        logger.warning("Startup email-alias capture failed", exc_info=True)

    # Phase 1B: flatten items.customer_email/phone/chapter/first_name/last_name
    # to match the canonical customers / customer_emails record on every
    # boot. Idempotent — only updates rows that still differ. Belt-and-
    # suspenders behind the Phase-1A resolvers, in case any read site
    # forgets to use them.
    try:
        _heal = heal_items_from_customers()
        if any(v for v in _heal.values()):
            logger.info("Startup: items heal %s", _heal)
    except Exception:
        logger.warning("Startup items-heal failed", exc_info=True)

    # ── Auto-migrate old per-item GoDaddy entries to order-level ──
    try:
        with _startup_connect() as _mig_conn:
            _old_format_count = _mig_conn.execute(
                "SELECT COUNT(*) as cnt FROM acct_transactions WHERE source_ref LIKE 'godaddy-income-%' AND COALESCE(status, 'active') = 'active'"
            ).fetchone()["cnt"]
        if _old_format_count > 0:
            logger.info("Found %d old-format per-item GoDaddy entries — running migration", _old_format_count)
            _mig_result = migrate_item_to_order_entries()
            logger.info("Order-level migration: %s", _mig_result)
        else:
            logger.info("No old-format GoDaddy entries — order-level migration not needed")
    except Exception:
        logger.warning("Order-level migration check failed", exc_info=True)

    # ── Clean up duplicate GoDaddy per-item entries (coexisting with order entries) ──
    try:
        with _startup_connect() as _dup_conn:
            _dup_count = _dup_conn.execute(
                """SELECT COUNT(*) as cnt FROM acct_transactions
                   WHERE COALESCE(status, 'active') = 'active'
                   AND (
                       description LIKE 'GoDaddy registration:%'
                       OR description LIKE 'GoDaddy merchant fee:%'
                       OR source_ref LIKE 'godaddy-income-%'
                       OR source_ref LIKE 'godaddy-fee-%'
                   )"""
            ).fetchone()["cnt"]
        if _dup_count > 0:
            logger.info("Found %d duplicate GoDaddy per-item entries — running cleanup", _dup_count)
            _dup_result = cleanup_duplicate_godaddy_entries()
            logger.info("GoDaddy cleanup: %s", _dup_result)
        else:
            logger.info("No duplicate GoDaddy per-item entries found")
    except Exception:
        logger.warning("GoDaddy duplicate cleanup failed", exc_info=True)

    # ── One-time fix: recalculate merchant fees from 2.7% to 2.9% ──
    try:
        with _startup_connect() as _fee_conn:
            # Check if any orders still have the old 2.7% rate
            # Signature: old rate gives merchant_fee ≈ amount * 0.027 + 0.30
            # New rate gives merchant_fee ≈ amount * 0.029 + 0.30
            _sample = _fee_conn.execute(
                """SELECT id, amount, merchant_fee FROM acct_transactions
                   WHERE category = 'godaddy_order' AND merchant_fee IS NOT NULL
                   AND amount > 0 AND COALESCE(status, 'active') NOT IN ('reversed', 'merged')
                   LIMIT 1"""
            ).fetchone()
            if _sample:
                _expected_29 = round(_sample["amount"] * 0.029 + 0.30, 2)
                _expected_27 = round(_sample["amount"] * 0.027 + 0.30, 2)
                if abs(_sample["merchant_fee"] - _expected_27) < 0.02 and abs(_sample["merchant_fee"] - _expected_29) > 0.02:
                    logger.info("Detected old 2.7%% merchant fees — recalculating to 2.9%%")
                    _orders = _fee_conn.execute(
                        """SELECT id, amount FROM acct_transactions
                           WHERE category = 'godaddy_order' AND amount > 0
                           AND COALESCE(status, 'active') NOT IN ('reversed', 'merged')"""
                    ).fetchall()
                    _updated = 0
                    for _ord in _orders:
                        _new_fee = round(_ord["amount"] * 0.029 + 0.30, 2)
                        _new_net = round(_ord["amount"] - _new_fee, 2)
                        _fee_conn.execute(
                            "UPDATE acct_transactions SET merchant_fee = ?, net_deposit = ? WHERE id = ?",
                            (_new_fee, _new_net, _ord["id"]),
                        )
                        _updated += 1
                    # Also update merchant_fee splits in godaddy_order_splits
                    _splits = _fee_conn.execute(
                        """SELECT s.id, s.transaction_id, t.amount as order_amount, s.amount as split_amount
                           FROM godaddy_order_splits s
                           JOIN acct_transactions t ON t.id = s.transaction_id
                           WHERE s.split_type = 'merchant_fee'"""
                    ).fetchall()
                    for _sp in _splits:
                        _item_count = _fee_conn.execute(
                            "SELECT COUNT(*) as cnt FROM godaddy_order_splits WHERE transaction_id = ? AND split_type = 'registration'",
                            (_sp["transaction_id"],),
                        ).fetchone()["cnt"] or 1
                        _total_fee = round(_sp["order_amount"] * 0.029 + 0.30, 2)
                        _per_item = round(_total_fee / _item_count, 2)
                        _fee_conn.execute(
                            "UPDATE godaddy_order_splits SET amount = ? WHERE id = ?",
                            (-_per_item, _sp["id"]),
                        )
                    _fee_conn.commit()
                    logger.info("Recalculated merchant fees for %d orders (2.7%% → 2.9%%)", _updated)
                else:
                    logger.info("Merchant fees already at 2.9%% rate — no recalculation needed")
    except Exception:
        logger.warning("Merchant fee recalculation failed", exc_info=True)

    # ── One-time fix: recalculate doubled order totals for multi-item orders ──
    # Bug: _write_godaddy_order_entry() summed total_amount across all items,
    # but total_amount stores the FULL ORDER total on each item row.  Multi-item
    # orders got their amount doubled/tripled, causing wrong net_deposit values.
    try:
        from email_parser.database import _parse_dollar as _pd
        with _startup_connect() as _otf:
            _gd_orders = _otf.execute(
                """SELECT t.id, t.source_ref, t.amount
                   FROM acct_transactions t
                   WHERE t.category = 'godaddy_order'
                   AND COALESCE(t.status, 'active') NOT IN ('reversed', 'merged')
                   AND t.source_ref LIKE 'godaddy-order-%'"""
            ).fetchall()

            _recalc_count = 0
            for _gdo in _gd_orders:
                _oid = _gdo["source_ref"].replace("godaddy-order-", "")

                # Get items for this order (same filter as backfill)
                _order_items = _otf.execute(
                    """SELECT id, item_price, transaction_fees, total_amount,
                              item_name, customer, coupon_amount, coupon_code
                       FROM items
                       WHERE order_id = ?
                       AND COALESCE(transaction_status, 'active') NOT IN
                           ('rsvp_only', 'credited', 'refunded', 'transferred')
                       AND parent_item_id IS NULL
                       AND transferred_from_id IS NULL
                       ORDER BY item_index""",
                    (_oid,),
                ).fetchall()
                _order_items = [
                    dict(i) for i in _order_items if _pd(dict(i).get("item_price")) > 0
                ]

                if len(_order_items) < 2:
                    continue  # Single-item orders unaffected by the doubling bug

                # Correct order_total: total_amount from first item
                _first_ta = _pd(_order_items[0].get("total_amount"))
                _computed = sum(
                    _pd(i.get("item_price")) + _pd(i.get("transaction_fees"))
                    for i in _order_items
                )
                _correct_total = _first_ta if _first_ta > 0 else _computed

                # Skip if already correct (within $1)
                if abs(_gdo["amount"] - _correct_total) < 1.0:
                    continue

                # Update entry in-place (preserves ID → keeps reconciliation_matches)
                _new_mf = round(_correct_total * 0.029 + 0.30, 2)
                _new_nd = round(_correct_total - _new_mf, 2)
                _otf.execute(
                    "UPDATE acct_transactions SET amount = ?, merchant_fee = ?, net_deposit = ? WHERE id = ?",
                    (_correct_total, _new_mf, _new_nd, _gdo["id"]),
                )

                # Recreate splits with correct proportions
                _otf.execute(
                    "DELETE FROM godaddy_order_splits WHERE transaction_id = ?",
                    (_gdo["id"],),
                )
                for _oi in _order_items:
                    _ip = _pd(_oi.get("item_price"))
                    _tf = _pd(_oi.get("transaction_fees"))
                    _it = _ip + _tf  # per-item contribution

                    if _ip > 0:
                        _otf.execute(
                            """INSERT INTO godaddy_order_splits
                               (transaction_id, item_id, event_name, customer, split_type, amount)
                               VALUES (?, ?, ?, ?, 'registration', ?)""",
                            (_gdo["id"], _oi["id"], _oi.get("item_name", ""),
                             _oi.get("customer", ""), _ip),
                        )
                    if _tf > 0:
                        _otf.execute(
                            """INSERT INTO godaddy_order_splits
                               (transaction_id, item_id, event_name, customer, split_type, amount)
                               VALUES (?, ?, ?, ?, 'transaction_fee', ?)""",
                            (_gdo["id"], _oi["id"], _oi.get("item_name", ""),
                             _oi.get("customer", ""), _tf),
                        )
                    _coupon = _pd(_oi.get("coupon_amount"))
                    if _coupon > 0 and _oi.get("coupon_code"):
                        _otf.execute(
                            """INSERT INTO godaddy_order_splits
                               (transaction_id, item_id, event_name, customer, split_type, amount)
                               VALUES (?, ?, ?, ?, 'coupon', ?)""",
                            (_gdo["id"], _oi["id"], _oi.get("item_name", ""),
                             _oi.get("customer", ""), -_coupon),
                        )
                    if _it > 0 and _correct_total > 0:
                        _item_mf = round(_new_mf * _it / _computed, 2)
                        _otf.execute(
                            """INSERT INTO godaddy_order_splits
                               (transaction_id, item_id, event_name, customer, split_type, amount)
                               VALUES (?, ?, ?, ?, 'merchant_fee', ?)""",
                            (_gdo["id"], _oi["id"], _oi.get("item_name", ""),
                             _oi.get("customer", ""), -_item_mf),
                        )

                _recalc_count += 1

            if _recalc_count > 0:
                _otf.commit()
                logger.info("Fixed doubled order totals for %d multi-item orders", _recalc_count)
            else:
                logger.info("No doubled order totals found — all multi-item orders correct")
    except Exception:
        logger.warning("Order total recalculation failed", exc_info=True)

    # ── Verify s18.4 LANDA PARK numbers ──
    # Works with both old (registration + processing_fee) and new (godaddy_order) formats.
    try:
        with _startup_connect() as _vconn:
            _landa = _vconn.execute(
                """SELECT entry_type, category,
                          COALESCE(SUM(amount), 0) as total,
                          COALESCE(SUM(merchant_fee), 0) as total_merchant_fee,
                          COALESCE(SUM(net_deposit), 0) as total_net_deposit
                   FROM acct_transactions
                   WHERE event_name = 's18.4 LANDA PARK'
                   AND COALESCE(status, 'active') = 'active'
                   AND entry_type IS NOT NULL
                   GROUP BY entry_type, category""",
            ).fetchall()
            # Also check splits for multi-event orders where event_name on the
            # parent entry might differ from the item's event
            _landa_splits = _vconn.execute(
                """SELECT COALESCE(SUM(s.amount), 0) as reg_total,
                          COALESCE(SUM(CASE WHEN s.split_type = 'merchant_fee' THEN s.amount ELSE 0 END), 0) as mf_total
                   FROM godaddy_order_splits s
                   WHERE s.event_name = 's18.4 LANDA PARK'
                   AND s.split_type IN ('registration', 'merchant_fee')""",
            ).fetchone()

            if _landa:
                # Old format: separate registration income + processing_fee expense entries
                _old_income = sum(r["total"] for r in _landa if r["entry_type"] == "income" and r["category"] == "registration")
                _old_fees = sum(r["total"] for r in _landa if r["entry_type"] == "expense" and r["category"] == "processing_fee")

                # New format: godaddy_order entries with merchant_fee column
                _new_income = sum(r["total"] for r in _landa if r["entry_type"] == "income" and r["category"] == "godaddy_order")
                _new_merchant = sum(r["total_merchant_fee"] for r in _landa if r["category"] == "godaddy_order")
                _new_net = sum(r["total_net_deposit"] for r in _landa if r["category"] == "godaddy_order")

                _landa_income = _old_income + _new_income
                _landa_fees = _old_fees + _new_merchant
                _landa_refunds = sum(r["total"] for r in _landa if r["entry_type"] == "expense" and r["category"] == "refund")
                _landa_net_deposit = _new_net if _new_net > 0 else round(_landa_income - _landa_fees, 2)

                logger.info(
                    "LANDA PARK verification: income=$%.2f, merchant_fees=$%.2f, refunds=$%.2f, net_deposit=$%.2f",
                    _landa_income, _landa_fees, _landa_refunds, _landa_net_deposit,
                )
                for r in _landa:
                    logger.info("  %s/%s: amount=$%.2f merchant_fee=$%.2f net_deposit=$%.2f",
                                r["entry_type"], r["category"], r["total"],
                                r["total_merchant_fee"], r["total_net_deposit"])
    except Exception:
        logger.warning("LANDA PARK verification query failed", exc_info=True)
except Exception:
    logger.warning("Startup backfill failed", exc_info=True)

# Seed upcoming San Antonio events (idempotent — skips existing)
_SA_EVENTS = [
    {"item_name": "s9.1 The Quarry", "event_date": "2026-03-17", "course": "The Quarry", "chapter": "San Antonio"},
    {"item_name": "s9.2 Canyon Springs", "event_date": "2026-03-24", "course": "Canyon Springs", "chapter": "San Antonio"},
    {"item_name": "s9.3 Silverhorn", "event_date": "2026-03-31", "course": "Silverhorn", "chapter": "San Antonio"},
    {"item_name": "s9.4 The Quarry", "event_date": "2026-04-07", "course": "The Quarry", "chapter": "San Antonio"},
    {"item_name": "s18.4 LANDA PARK", "event_date": "2026-04-11", "course": "Landa Park", "chapter": "San Antonio"},
    {"item_name": "s9.5 Cedar Creek", "event_date": "2026-04-14", "course": "Cedar Creek", "chapter": "San Antonio"},
    {"item_name": "s9.6 The Quarry", "event_date": "2026-04-21", "course": "The Quarry", "chapter": "San Antonio"},
    {"item_name": "s9.7 Canyon Springs", "event_date": "2026-04-28", "course": "Canyon Springs", "chapter": "San Antonio"},
    {"item_name": "s18.5 WILLOW SPRINGS", "event_date": "2026-05-02", "course": "Willow Springs", "chapter": "San Antonio"},
    {"item_name": "s9.8 Silverhorn", "event_date": "2026-05-05", "course": "Silverhorn", "chapter": "San Antonio"},
    {"item_name": "s9.9 TPC San Antonio | Canyons", "event_date": "2026-05-12", "course": "TPC San Antonio - Canyons", "chapter": "San Antonio"},
    {"item_name": "HILL COUNTRY MATCHES | Comanche Trace", "event_date": "2026-05-16", "course": "Comanche Trace", "chapter": "San Antonio"},
    {"item_name": "s9.10 Brackenridge", "event_date": "2026-05-19", "course": "Brackenridge", "chapter": "San Antonio"},
    {"item_name": "s9.11 The Quarry", "event_date": "2026-05-26", "course": "The Quarry", "chapter": "San Antonio"},
    {"item_name": "s18.6 KISSING TREE", "event_date": "2026-05-30", "course": "Kissing Tree", "chapter": "San Antonio"},
    {"item_name": "s9.12 Canyon Springs", "event_date": "2026-06-02", "course": "Canyon Springs", "chapter": "San Antonio"},
]
_seed_result = seed_events(_SA_EVENTS)
if _seed_result["inserted"]:
    logger.info("Seeded %d SA events", _seed_result["inserted"])

# Seed TGF payout data — s9.4 The Quarry (April 7, 2026)
_s94_result = add_tgf_event({
    "code": "s9.4 The Quarry",
    "name": "The Quarry",
    "event_date": "2026-04-07",
    "course": "The Quarry",
    "chapter": "San Antonio",
    "total_purse": 894.00,
    "winners_count": 14,
    "payouts": [
        {"golferName": "Gilbert Ellis", "category": "mvp", "amount": 84.00, "description": "TGF MVP"},
        {"golferName": "Gilbert Ellis", "category": "individual_net", "amount": 65.25, "description": "Individual Net"},
        {"golferName": "Gilbert Ellis", "category": "other", "amount": 58.00, "description": "s9.4 MVP Net"},
        {"golferName": "Gilbert Ellis", "category": "individual_gross", "amount": 21.00, "description": "Individual Gross"},
        {"golferName": "Pat Youngs", "category": "skins", "amount": 37.80, "description": "Skins Gross"},
        {"golferName": "Pat Youngs", "category": "closest_to_pin", "amount": 31.00, "description": "Closest to Pin #16"},
        {"golferName": "Pat Youngs", "category": "individual_net", "amount": 26.10, "description": "Individual Net"},
        {"golferName": "Pat Youngs", "category": "individual_gross", "amount": 21.00, "description": "Individual Gross"},
        {"golferName": "Pat Youngs", "category": "team_net", "amount": 19.50, "description": "Team Net"},
        {"golferName": "Jeff Young", "category": "skins", "amount": 56.70, "description": "Skins Gross"},
        {"golferName": "Jeff Young", "category": "closest_to_pin", "amount": 31.00, "description": "Closest to Pin #12"},
        {"golferName": "Jeff Young", "category": "team_net", "amount": 19.50, "description": "Team Net"},
        {"golferName": "Roland Campos", "category": "individual_net", "amount": 65.25, "description": "Individual Net"},
        {"golferName": "Roland Campos", "category": "individual_gross", "amount": 21.00, "description": "Individual Gross"},
        {"golferName": "Roland Campos", "category": "team_net", "amount": 19.50, "description": "Team Net"},
        {"golferName": "Jeff Rideout", "category": "skins", "amount": 47.25, "description": "Skins Gross"},
        {"golferName": "Jeff Rideout", "category": "individual_gross", "amount": 21.00, "description": "Individual Gross"},
        {"golferName": "Fred Wicker", "category": "skins", "amount": 47.25, "description": "Skins Gross"},
        {"golferName": "Adam Baker", "category": "individual_net", "amount": 39.15, "description": "Individual Net"},
        {"golferName": "Rob Callaway", "category": "individual_net", "amount": 39.15, "description": "Individual Net"},
        {"golferName": "Joe Decker", "category": "individual_net", "amount": 26.10, "description": "Individual Net"},
        {"golferName": "Jordan Bastin", "category": "team_net", "amount": 19.50, "description": "Team Net"},
        {"golferName": "Eric Taft", "category": "team_net", "amount": 19.50, "description": "Team Net"},
        {"golferName": "Steven Hunt", "category": "team_net", "amount": 19.50, "description": "Team Net"},
        {"golferName": "Jeff Greenwell", "category": "team_net", "amount": 19.50, "description": "Team Net"},
        {"golferName": "Brian Thompson", "category": "team_net", "amount": 19.50, "description": "Team Net"},
    ],
})
if "event_id" in _s94_result and "error" not in _s94_result:
    logger.info("Seeded TGF payout event s9.4 The Quarry (event_id=%d)", _s94_result["event_id"])

# Only start the scheduler in one Gunicorn worker (or in dev mode).
# Gunicorn's --preload flag shares module-level state, but with forked workers
# each gets its own scheduler.  We use a PID-based guard so only one runs.
_scheduler_lock = threading.Lock()
with _scheduler_lock:
    _scheduler_pid = os.getenv("_SCHEDULER_STARTED_PID")
    _is_main_worker = _scheduler_pid is None or _scheduler_pid == str(os.getpid())
    if os.getenv("EMAIL_ADDRESS") and _is_main_worker:
        os.environ["_SCHEDULER_STARTED_PID"] = str(os.getpid())
        start_scheduler()
    elif not os.getenv("EMAIL_ADDRESS"):
        logger.info("Email not configured — scheduler not started. Set up .env to enable auto-checking.")

if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
